from __future__ import annotations

import copy
import re
import docx
import transaction
from pytest import approx

from decimal import Decimal
from datetime import date

from freezegun import freeze_time
from io import BytesIO
from onegov.user.models.user import User
from onegov.core.utils import module_path
from onegov.translator_directory.models.ticket import (
    AccreditationTicket,
)
from onegov.core.crypto.password import hash_password
from onegov.translator_directory.models.translator import Translator
from os.path import basename
from onegov.file import FileCollection
from onegov.file import File
from tests.onegov.translator_directory.shared import iter_block_items
from onegov.gis import Coordinates
from onegov.pdf import Pdf
from onegov.translator_directory.collections.translator import (
    TranslatorCollection,
)
from onegov.translator_directory.models.time_report import TranslatorTimeReport
from onegov.user import UserCollection, UserGroup, UserGroupCollection
from openpyxl import load_workbook
from pdftotext import PDF  # type: ignore[import-not-found]
from tests.onegov.translator_directory.shared import (
    translator_data, create_languages, create_certificates)
from tests.shared.utils import decode_map_value, encode_map_value
from unittest.mock import patch
from webtest import Upload


from typing import Any, TYPE_CHECKING
if TYPE_CHECKING:
    from onegov.core.types import EmailJsonDict
    from tests.shared.client import ExtendedResponse
    from unittest.mock import MagicMock
    from .conftest import Client


class FakeResponse:
    def __init__(
        self,
        json_data: Any | None = None,
        status_code: int = 200
    ) -> None:
        self.status_code = status_code
        self.json_data = json_data or {}

    def json(self) -> Any:
        return self.json_data


def upload_pdf(filename: str) -> Upload:
    file = BytesIO()
    pdf = Pdf(file)
    pdf.init_report()
    pdf.p(filename)
    pdf.generate()
    file.seek(0)

    return Upload(filename, file.read(), 'application/pdf')


def check_pdf(page: ExtendedResponse, filename: str, link: str) -> None:
    response = page.click(link, index=0)
    headers = dict(response.headers)
    assert filename in headers['Content-Disposition']
    assert headers['Content-Type'] == 'application/pdf'
    assert filename in ''.join(PDF(BytesIO(response.body)))


def upload_file(
    filename: str,
    client: Client,
    content_type: str | None = None
) -> None:
    with open(filename, 'rb') as f:
        page = client.get('/files')
        page.form['file'] = [
            Upload(basename(filename), f.read(), content_type)
        ]
        page.form.submit()


def get_accountant_email(client: Client) -> str:
    """Get accountant email from user group for testing."""
    session = client.app.session()
    user_group = (
        session.query(UserGroup)
        .filter(UserGroup.meta['finanzstelle'].astext.isnot(None))
        .first()
    )
    assert user_group is not None, 'No user group with finanzstelle found'
    emails = user_group.meta.get('accountant_emails', [])
    assert emails, 'No accountant emails in user group'
    return emails[0]


def collect_emails(client: Client) -> list['EmailJsonDict']:
    """Collect all emails from client and flush the queue."""
    all_emails = []
    for i in range(10):
        try:
            email = client.get_email(i)
            if email:
                all_emails.append(email)
        except IndexError:
            break
    # client.flush_email_queue()
    return all_emails


def filter_emails_by_recipient(
    emails: list['EmailJsonDict'], recipient: str
) -> list['EmailJsonDict']:
    """Filter emails to only those sent to the specified recipient."""
    return [e for e in emails if recipient in e['To']]


def extract_ticket_link_from_email(
    emails: list['EmailJsonDict'], recipient: str
) -> str:
    """Extract ticket link from email sent to recipient."""
    matching_emails = [e for e in emails if recipient in e['To']]
    assert len(matching_emails) >= 1, f'No email found for {recipient}'

    mail = matching_emails[0]
    link_match = re.search(
        r'<a href="([^"]+)">Zeiterfassung anzeigen</a>',
        mail['HtmlBody'],
    )
    assert link_match is not None, 'No ticket link found in email'
    return link_match.group(1)


def test_view_translator(client: Client) -> None:
    session = client.app.session()
    languages = create_languages(session)
    certs = create_certificates(session)
    cert_ids = [str(cert.id) for cert in certs]
    cert_names = [cert.name for cert in certs]
    language_ids = [str(lang.id) for lang in languages]
    language_names = [lang.name for lang in languages]
    transaction.commit()

    client.login_editor()
    client.get('/translators/new', status=403)
    client.logout()

    with freeze_time('2021-01-01'):
        client.login_admin()
        page = client.get('/translators/new')

        # check choices
        assert 'Männlich' in page
        assert language_names[0] in page
        assert cert_names[0] in page
        assert 'Simultandolmetschen' in page
        assert 'Human- und Sozialwissenschaften' in page
        assert not decode_map_value(page.form['coordinates'].value)

        page.form['pers_id'] = 978654
        page.form['first_name'] = 'Uncle'
        page.form['last_name'] = 'Bob'
        page.form['nationalities'] = ['CH']
        page.form['social_sec_number'] = 'xxxx'
        page.form['zip_code'] = 'xxxx'
        page.form['iban'] = 'xxxx'
        page = page.form.submit()
        assert "Ungültige AHV-Nummer" in page
        assert "Postleitzahl muss aus 4 Ziffern bestehen" in page
        assert "Ungültige Eingabe" in page

        # input required fields
        page.form['social_sec_number'] = '756.1234.5678.97'
        page.form['tel_mobile'] = '079 700 80 97'
        page.form['agency_references'] = 'All okay'
        page.form['zip_code'] = '7890'
        page.form['mother_tongues_ids'] = [language_ids[3]]

        # non required fields
        page.form['hometown'] = 'Gersau'
        page.form['email'] = 'test@test.com'
        page.form['spoken_languages_ids'] = [language_ids[0], language_ids[1]]
        page.form['written_languages_ids'] = [language_ids[2]]
        page.form['certificates_ids'] = [cert_ids[0]]
        page.form['iban'] = 'DE07 1234 1234 1234 1234 12'
        page.form.get('expertise_professional_guilds', index=0).checked = True
        page.form['expertise_professional_guilds_other'] = ['Psychologie']
        page.form.get('expertise_interpreting_types', index=0).checked = True
        page.form['coordinates'] = encode_map_value({
            'lat': 46, 'lon': 7, 'zoom': 12
        })
        drive_distance = 111.11
        with patch(
                'onegov.gis.utils.MapboxRequests.directions',
                return_value=FakeResponse({
                    'code': 'Ok',
                    'routes': [{'distance': drive_distance * 1000}]})
        ):
            page = page.form.submit()
            assert 'Der eigene Standort ist nicht konfiguriert' in page
            settings = client.get('/directory-settings')
            settings.form['coordinates'] = encode_map_value({
                'lat': 46, 'lon': 7, 'zoom': 12
            })
            settings.form.submit()
            page = page.form.submit().follow()

        translator_url = page.request.url
        assert 'Uncle' in page
        assert 'BOB' in page
        assert '<a href="mailto:test@test.com">test@test.com</a>' in page
        values = {
            dl.find('dt').text_content().strip():
            dl.find('dd').text_content().strip()
            for dl in page.pyquery('dl')
        }
        assert len(values) == 27
        assert values['Nachname'] == 'BOB'
        assert values['Vorname'] == 'Uncle'
        assert values['Personal Nr.'] == '978654'
        assert values['Zulassung'] == ('nicht akkreditiert / Einsatz '
                                       'Dringlichkeit')
        assert values['Quellensteuer'] == 'Nein'
        assert values['Selbständig'] == 'Nein'
        assert values['Geschlecht'] == 'Männlich'
        assert values['PLZ'] == '7890'
        assert values['Wegberechnung'] == f'{round(drive_distance, 1)} km'
        assert values['AHV-Nr.'] == '756.1234.5678.97'
        assert values['IBAN'] == 'DE07 1234 1234 1234 1234 12'
        assert values['E-Mail'] == 'test@test.com'
        assert values['Telefon Mobile'] == '079 700 80 97'
        assert values['Fachkenntnisse nach Dolmetscherart'] == (
            'Simultandolmetschen')
        assert 'Ernährung' in values['Fachkenntnisse nach Berufssparte']
        assert 'Psychologie' in values['Fachkenntnisse nach Berufssparte']
        assert values['Muttersprachen'] == language_names[3]
        assert language_names[0] in values['Arbeitssprache - Wort']
        assert language_names[1] in values['Arbeitssprache - Wort']
        assert values['Arbeitssprache - Schrift'] == language_names[2]
        assert values['Referenzen Behörden'] == 'All okay'
        assert values['Ausbildung Dolmetscher'] == 'Nein'
        assert values['Versteckt'] == 'Nein'
        assert values['Zertifikate'] == cert_names[0]
        assert values['Heimatort'] == 'Gersau'
        assert values['Nationalität(en)'] == 'Schweiz'

        # test user account created and activation mail sent
        user = UserCollection(session).by_username('test@test.com')
        assert user is not None
        assert user.translator.title == 'BOB, Uncle'  # type: ignore[attr-defined]
        assert user.active is True
        assert user.role == 'translator'

        mail = client.get_email(0, flush_queue=True)
        assert mail['To'] == 'test@test.com'
        assert mail['Subject'] == 'Ein Konto wurde für Sie erstellt'

        # test translator can login and view his own data
        client.logout()
        reset_password_url = re.search(  # type: ignore[union-attr]
            r'(http://localhost/auth/reset-password[^)]+)',
            mail['TextBody']
        ).group()
        page = client.get(reset_password_url)
        page.form['email'] = 'test@test.com'
        page.form['password'] = 'known_very_secure_password'
        page.form.submit()

    with freeze_time('2021-12-31'):
        page = client.login(
            'test@test.com', 'known_very_secure_password').maybe_follow()
        assert 'Sind ihre Daten noch aktuell? Bitte überprüfen Sie' not in page
        assert '978654' in page
        assert 'Uncle' in page
        assert 'BOB' in page

    page = client.login(
        'test@test.com', 'known_very_secure_password').maybe_follow()
    assert 'Sind ihre Daten noch aktuell? Bitte überprüfen Sie' in page
    assert '978654' in page
    assert 'Uncle' in page
    assert 'BOB' in page
    assert '<a href="mailto:test@test.com">test@test.com</a>' in page
    assert '756.1234.5678.97' in page
    assert 'All okay' in page
    assert '7890' in page
    assert 'DE07 1234 1234 1234 1234 12' in page
    assert 'Ernährung und Landwirtschaft' in page
    assert 'Psychologie' in page
    assert 'Simultandolmetschen' in page
    assert 'Versteckt' not in page
    assert str(round(drive_distance, 1)) in page
    assert language_names[3] in page
    assert language_names[0] in page
    assert language_names[1] in page
    assert language_names[2] in page

    page = page.click('Aktuelle Daten bestätigen').follow()
    assert 'Ihre Daten wurden bestätigt' in page
    client.logout()

    # test editors access on the edit view
    client.login_editor()
    page = client.get(translator_url)
    assert '978654' in page
    page = page.click('Bearbeiten')
    page.form['pers_id'] = 123456
    page.form['contract_number'] = 'CN-123'
    page = page.form.submit().follow()
    assert '123456' in page
    assert 'CN-123' in page
    client.logout()

    # edit some key attribute
    client.login_admin()
    page = client.get(translator_url).click('Bearbeiten')
    assert 'Zulassung' in page
    assert decode_map_value(page.form['coordinates'].value) == Coordinates(
        lat=46, lon=7, zoom=12
    )
    drive_distance = 60.01
    tel_mobile = '044 123 50 50'

    page.form['first_name'] = 'Aunt'
    page.form['last_name'] = 'Maggie'
    page.form['email'] = 'aunt.maggie@translators.com'
    page.form['iban'] = 'CH5604835012345678009'
    page.form['pers_id'] = 234567
    page.form['contract_number'] = 'CN-456'
    page.form['admission'] = 'in_progress'
    page.form['withholding_tax'] = True
    page.form['self_employed'] = True
    page.form['gender'] = 'F'
    page.form['date_of_birth'] = '2019-01-01'
    page.form['nationalities'] = ['PE']  # Peru
    page.form['address'] = 'Somestreet'
    page.form['zip_code'] = '4052'
    page.form['city'] = 'Somecity'
    page.form['drive_distance'] = drive_distance
    page.form['social_sec_number'] = '756.1111.1111.13'
    page.form['bank_name'] = 'Abank'
    page.form['bank_address'] = 'AB Address'
    page.form['account_owner'] = 'AccountOwner'
    page.form['tel_mobile'] = tel_mobile
    page.form['tel_private'] = '044 123 50 51'
    page.form['tel_office'] = '044 123 50 52'
    page.form['availability'] = 'always 24h'
    page.form['confirm_name_reveal'] = False
    page.form['date_of_application'] = '2015-01-01'
    page.form['date_of_decision'] = '2016-01-01'
    page.form['proof_of_preconditions'] = 'ZHCW'
    page.form['agency_references'] = 'Kt. ZG'
    page.form['education_as_interpreter'] = True
    page.form['comments'] = 'My Comments'
    page.form['operation_comments'] = 'operational'
    page.form['profession'] = 'Baker'
    page.form['occupation'] = 'Salesman'
    page.form['for_admins_only'] = True
    page.form.get('expertise_professional_guilds', index=0).checked = False
    page.form.get('expertise_professional_guilds', index=1).checked = True
    page.form['expertise_professional_guilds_other'] = ['Religion']
    page.form.get('expertise_interpreting_types', index=0).checked = False
    page.form.get('expertise_interpreting_types', index=1).checked = True
    page.form['mother_tongues_ids'] = [language_ids[1]]
    page.form['spoken_languages_ids'] = [language_ids[2]]
    page.form['written_languages_ids'] = [language_ids[3]]
    page.form['certificates_ids'] = [cert_ids[1]]

    new_drive_distance = 250.666
    # when old and new coords are not same, we update the driving_distance
    page.form['coordinates'] = encode_map_value({
        'lat': 47, 'lon': 8, 'zoom': 12
    })
    with patch(
            'onegov.gis.utils.MapboxRequests.directions',
            return_value=FakeResponse({
                'code': 'Ok',
                'routes': [{'distance': new_drive_distance * 1000}]})
    ):
        page = page.form.submit().follow()

    assert 'Ihre Änderungen wurden gespeichert' in page
    assert 'Aunt' in page
    assert 'MAGGIE' in page
    assert f'<a href="tel:{tel_mobile}">{tel_mobile}</a>' in page
    values = {
        dl.find('dt').text_content().strip():
        dl.find('dd').text_content().strip()
        for dl in page.pyquery('dl')
    }
    assert len(values) == 43
    assert values['Nachname'] == 'MAGGIE'
    assert values['Vorname'] == 'Aunt'
    assert values['AHV-Nr.'] == '756.1111.1111.13'
    assert values['Anschrift'] == 'Somestreet'
    assert values['Ausbildung Dolmetscher'] == 'Ja'
    assert values['Bank Adresse'] == 'AB Address'
    assert values['Bank Konto lautend auf'] == 'AccountOwner'
    assert values['Bank Name'] == 'Abank'
    assert values['Bemerkungen'] == 'My Comments'
    assert values['Besondere Hinweise Einsatzmöglichkeiten'] == 'operational'
    assert values['Bewerbung Datum'] == '01.01.2015'
    assert values['E-Mail'] == 'aunt.maggie@translators.com'
    assert values['Entscheid Datum'] == '01.01.2016'
    assert values['Erreich- und Verfügbarkeit'] == 'always 24h'
    assert 'Wirtschaft' in values['Fachkenntnisse nach Berufssparte']
    assert 'Religion' in values['Fachkenntnisse nach Berufssparte']
    assert values['Fachkenntnisse nach Dolmetscherart'] == (
        'Konsektutivdolmetschen')
    assert values['Geburtsdatum'] == '01.01.2019'
    assert values['Geschlecht'] == 'Weiblich'
    assert values['IBAN'] == 'CH5604835012345678009'
    assert values['Nachweis der Voraussetzung'] == 'ZHCW'
    assert values['Nationalität(en)'] == 'Peru'
    assert values['Ort'] == 'Somecity'
    assert values['PLZ'] == '4052'
    assert values['Personal Nr.'] == '234567'
    assert values['Vertragsnummer'] == 'CN-456'
    assert values['Quellensteuer'] == 'Ja'
    assert values['Referenzen Behörden'] == 'Kt. ZG'
    assert values['Selbständig'] == 'Ja'
    assert values['Telefon Geschäft'] == '044 123 50 52'
    assert values['Telefon Mobile'] == '044 123 50 50'
    assert values['Telefon Privat'] == '044 123 50 51'
    assert values['Erlernter Beruf'] == 'Baker'
    assert values['Aktuelle berufliche Tatigkeit'] == 'Salesman'
    assert values['Versteckt'] == 'Ja'
    assert values['Wegberechnung'] == f'{round(new_drive_distance, 1)} km'
    assert values['Zulassung'] == 'im Zulassungsverfahren'
    assert values['Muttersprachen'] == language_names[1]
    assert values['Arbeitssprache - Wort'] == language_names[2]
    assert values['Arbeitssprache - Schrift'] == language_names[3]
    assert values['Zertifikate'] == cert_names[1]

    # test user account updated
    users = UserCollection(session)
    assert not users.by_username('test@test.com')
    user = users.by_username('aunt.maggie@translators.com')
    assert user is not None
    assert user.translator.title == 'MAGGIE, Aunt'  # type: ignore[attr-defined]
    assert user.active is True
    assert user.role == 'translator'

    # try adding another with same email
    page = client.get('/translators/new')
    page.form['first_name'] = 'Uncle'
    page.form['last_name'] = 'Bob'
    page.form['agency_references'] = 'All okay'
    page.form['email'] = 'aunt.maggie@translators.com'

    page = page.form.submit()
    assert 'Ein(e) Übersetzer/in mit dieser E-Mail existiert bereits' in page

    # Test if the for_admins_only works
    client.logout()
    client.login_editor()
    client.get(translator_url, status=403)
    client.logout()


def test_view_languages(client: Client) -> None:
    create_languages(client.app.session())
    transaction.commit()

    client.get('/languages', status=403)

    client.login_member()
    client.get('/languages', status=403)

    client.login_editor()
    page = client.get('/languages')
    assert 'Italian' in page
    assert 'French' in page
    assert 'Arabic' in page


def test_manage_language(client: Client) -> None:
    client.login_editor()
    client.get('/languages/new', status=403)
    client.login_admin()
    page = client.get('/languages/new')

    page.form['name'] = '     '
    page = page.form.submit()
    assert 'Dieses Feld wird benötigt' in page

    page.form['name'] = ' English   '
    page = page.form.submit().follow()
    assert 'Sprache English hinzugefügt' in page
    assert 'English' in page

    page = client.get('/languages/new')
    page.form['name'] = 'English'
    page = page.form.submit()
    assert 'English existiert bereits' in page

    page = client.get('/languages').click('English')
    page.form['name'] = 'English (British)'
    page = page.form.submit().follow()
    assert 'English (British)' in page

    page = client.get('/languages').click('English').click('Löschen')
    page = client.get('/languages')
    assert 'English' not in page
    assert 'English (British)' not in page


def test_view_search_translator(client: Client) -> None:
    """
    - test excluding hidden ones for non-admins
    """
    client.get('/translators', status=403)
    client.login_member()
    page = client.get('/translators')
    assert 'Keine Ergebnisse gefunden' in page
    assert 'Suche in Vor- und Nachname' in page

    session = client.app.session()
    languages = create_languages(session)
    lang_ids = [str(lang.id) for lang in languages]
    translators = TranslatorCollection(client.app)

    data = copy.deepcopy(translator_data)
    mail = 'first@test.com'
    data['email'] = mail
    data['spoken_languages'] = [languages[0]]
    data['written_languages'] = [languages[1]]
    data['first_name'] = 'Sebastian Stefan'
    data['last_name'] = 'Hugentobler Meeringer'

    translators.add(**data)

    data = copy.deepcopy(translator_data)
    hide_mail = 'hidden@test.com'
    data['email'] = hide_mail
    data['spoken_languages'] = []
    data['written_languages'] = []
    data['first_name'] = 'Maryan'
    data['last_name'] = 'Sitkova Lavrova'
    translators.add(**data)

    transaction.commit()

    page = client.get('/translators')
    assert mail in page
    page.form['spoken_langs'] = [lang_ids[0]]
    page = page.form.submit().follow()
    assert mail in page
    assert hide_mail not in page

    # Check condition both must be fulfilled (AND not OR)
    page.form['spoken_langs'] = [lang_ids[0]]
    page.form['written_langs'] = [lang_ids[0]]
    page = page.form.submit().follow()
    assert 'Keine Ergebnisse gefunden' in page

    page.form['spoken_langs'] = [lang_ids[0], lang_ids[1]]
    page.form['written_langs'] = []
    page = page.form.submit().follow()
    assert 'Keine Ergebnisse gefunden' in page

    page.form['spoken_langs'] = []

    # Test search simple search, the rest is covered in the collection tests
    page.form['search'] = 'xxx Lavrov'
    page = page.form.submit().follow()
    assert 'Sitkova Lavrova'.upper() in page
    assert 'Hugentobler'.upper() not in page


def test_view_export_translators(client: Client) -> None:
    session = client.app.session()
    languages = create_languages(session)
    translators = TranslatorCollection(client.app)

    data = copy.deepcopy(translator_data)
    data['spoken_languages'] = [languages[0]]
    data['written_languages'] = [languages[1]]
    data['monitoring_languages'] = [languages[2]]
    data['expertise_professional_guilds'] = ['economy']
    data['expertise_professional_guilds_other'] = ['Psychologie', 'Religion']
    data['expertise_interpreting_types'] = ['simultaneous', 'whisper']

    translators.add(**data)
    transaction.commit()

    client.login_admin()
    response = client.get('/translators').click('Export')
    sheet = load_workbook(BytesIO(response.body)).worksheets[0]
    assert sheet.cell(2, 1).value == 1234
    assert sheet.cell(2, 2).value == (
        'nicht akkreditiert / Einsatz Dringlichkeit'
    )
    assert sheet.cell(2, 3).value == 0
    assert sheet.cell(2, 4).value == 0
    assert sheet.cell(2, 5).value == 'Benito'
    assert sheet.cell(2, 6).value == 'Hugo'
    assert sheet.cell(2, 7).value == 'Männlich'
    assert sheet.cell(2, 8).value == data['date_of_birth'].isoformat()
    assert sheet.cell(2, 9).value == 'Schweiz, Österreich'
    # assert sheet.cell(2, 10).value == '{"lon":null,"zoom":null,"lat":null}'
    assert sheet.cell(2, 11).value == 'Downing Street 5'
    assert sheet.cell(2, 12).value == '4000'
    assert sheet.cell(2, 13).value == 'Luzern'
    assert sheet.cell(2, 14).value is None
    assert sheet.cell(2, 15).value == '756.1234.4568.94'
    assert sheet.cell(2, 16).value == 'R-BS'
    assert sheet.cell(2, 17).value == 'Bullstreet 5'
    assert sheet.cell(2, 18).value == 'Hugo Benito'
    assert sheet.cell(2, 19).value is None
    assert sheet.cell(2, 20).value == 'hugo@benito.com'
    assert sheet.cell(2, 21).value is None
    assert sheet.cell(2, 22).value == '079 000 00 00'
    assert sheet.cell(2, 23).value == '041 444 44 44'
    assert sheet.cell(2, 24).value == 'always'
    assert sheet.cell(2, 25).value is None
    assert sheet.cell(2, 26).value == 0
    assert sheet.cell(2, 27).value == data['date_of_application'].isoformat()
    assert sheet.cell(2, 28).value == data['date_of_decision'].isoformat()
    assert sheet.cell(2, 29).value is None
    assert sheet.cell(2, 30).value == 'German'
    assert sheet.cell(2, 31).value == 'French'
    assert sheet.cell(2, 32).value == 'Italian'
    assert sheet.cell(2, 33).value == 'craftsman'
    assert sheet.cell(2, 34).value == 'baker'
    assert sheet.cell(2, 35).value == 'Wirtschaft|Psychologie|Religion'
    assert sheet.cell(2, 36).value == 'Simultandolmetschen|Flüsterdolmetschen'
    assert sheet.cell(2, 37).value == 'all okay'
    assert sheet.cell(2, 38).value == 'Some ref'
    assert sheet.cell(2, 39).value == 0
    assert sheet.cell(2, 40).value is None
    assert sheet.cell(2, 41).value is None


def test_view_export_translators_with_filters(client: Client) -> None:
    session = client.app.session()
    languages = create_languages(session)
    lang_ids = [str(lang.id) for lang in languages]
    translators = TranslatorCollection(client.app)

    # Add translator 1
    data1 = copy.deepcopy(translator_data)
    data1['pers_id'] = 1111
    data1['first_name'] = 'Filtered'
    data1['last_name'] = 'One'
    data1['email'] = 'filtered.one@example.com'
    data1['spoken_languages'] = [languages[0], languages[1]]  # German, French
    data1['written_languages'] = [languages[2]]  # Italian
    translators.add(**data1)

    # Add translator 2
    data2 = copy.deepcopy(translator_data)
    data2['pers_id'] = 2222
    data2['first_name'] = 'Filtered'
    data2['last_name'] = 'Two'
    data2['email'] = 'filtered.two@example.com'
    data2['spoken_languages'] = [languages[0]]  # German
    data2['written_languages'] = [languages[3]]  # Arabic
    translators.add(**data2)

    # Add translator 3 (should be filtered out)
    data3 = copy.deepcopy(translator_data)
    data3['pers_id'] = 3333
    data3['first_name'] = 'NotFiltered'
    data3['last_name'] = 'Three'
    data3['email'] = 'not.filtered@example.com'
    data3['spoken_languages'] = [languages[1]]  # French
    data3['written_languages'] = [languages[2]]  # Italian
    translators.add(**data3)

    transaction.commit()

    client.login_admin()
    page = client.get('/translators')

    # Fill the search form
    page.form['spoken_langs'] = [lang_ids[0]]  # German
    page.form['written_langs'] = [lang_ids[2]]  # Italian
    page = page.form.submit().follow()

    # Check that only translator 1 is shown in the results table
    results_table = page.pyquery('#search-results-table')[0].text_content()
    assert 'filtered.one@example.com' in results_table
    assert 'filtered.two@example.com' not in results_table
    assert 'filtered.three@example.com' not in results_table

    # Click the export button (which now includes filters)
    response = page.click('Export')
    sheet = load_workbook(BytesIO(response.body)).worksheets[0]
    assert sheet.max_row == 2
    header: dict[str, int] = {cell.value: cell.column for cell in sheet[1]}  # type: ignore[misc]
    # Check data of the exported translator (should be translator 1) using
    # column names from the header
    data_row = 2
    assert sheet.cell(
        row=data_row, column=header['Personal Nr.']).value == 1111
    assert sheet.cell(
        row=data_row, column=header['Vorname']).value == 'Filtered'
    assert sheet.cell(
        row=data_row, column=header['Nachname']).value == 'One'
    assert sheet.cell(
        row=data_row, column=header['Arbeitssprache - Wort']
    ).value == 'German|French'
    assert sheet.cell(
        row=data_row, column=header['Arbeitssprache - Schrift']
    ).value == 'Italian'


def test_view_export_translators_with_filters_two_langs(
    client: Client
) -> None:
    session = client.app.session()
    languages = create_languages(session)
    lang_ids = [str(lang.id) for lang in languages]
    translators = TranslatorCollection(client.app)

    # search two languages
    # Create two translators, both having two languages
    data4 = copy.deepcopy(translator_data)
    data4['pers_id'] = 4444
    data4['first_name'] = 'two'
    data4['last_name'] = 'langs'
    data4['email'] = 'two.langs@example.com'
    data4['spoken_languages'] = [languages[0], languages[1]]  # German, French
    translators.add(**data4)

    data5 = copy.deepcopy(translator_data)
    data5['pers_id'] = 5555
    data5['first_name'] = 'alsotwo'
    data5['last_name'] = 'alsotwo'
    data5['email'] = 'alsotwo.alsotwo@example.com'
    data5['spoken_languages'] = [languages[0], languages[1]]  # German, French
    translators.add(**data5)
    transaction.commit()

    client.login_admin()

    page = client.get('/translators')
    page.form['spoken_langs'] = [lang_ids[0], lang_ids[1]]  # German, French
    page = page.form.submit().follow()

    # Check that both are found
    results_table = page.pyquery('#search-results-table')[0].text_content()
    assert 'two.langs@example.com' in results_table
    assert 'alsotwo.alsotwo@example.com' in results_table
    # Click the export button (which now includes filters)
    response = page.click('Export')
    sheet = load_workbook(BytesIO(response.body)).worksheets[0]
    assert sheet.max_row == 3
    header: dict[str, int] = {cell.value: cell.column for cell in sheet[1]}  # type: ignore[misc]

    data_row = 2
    # first one will be the 5555  (ordered by lastname
    assert sheet.cell(
        row=data_row, column=header['Personal Nr.']).value == 5555
    assert sheet.cell(
        row=data_row, column=header['Vorname']).value == 'alsotwo'
    assert sheet.cell(
        row=data_row, column=header['Nachname']).value == 'alsotwo'
    assert sheet.cell(
        row=data_row, column=header['Arbeitssprache - Wort']
    ).value == 'German|French'

    data_row = 3
    assert sheet.cell(
        row=data_row, column=header['Personal Nr.']).value == 4444
    assert sheet.cell(
        row=data_row, column=header['Vorname']).value == 'two'
    assert sheet.cell(
        row=data_row, column=header['Nachname']).value == 'langs'
    assert sheet.cell(
        row=data_row, column=header['Arbeitssprache - Wort']
    ).value == 'German|French'

    # Test descending order
    page = client.get('/translators')
    page.form['spoken_langs'] = [lang_ids[0], lang_ids[1]]  # German, French
    page.form['order_desc'] = '1'
    page = page.form.submit().follow()

    # Check that both are found and order is reversed in the table
    results_table = page.pyquery('#search-results-table')[0].text_content()
    assert 'two.langs@example.com' in results_table
    assert 'alsotwo.alsotwo@example.com' in results_table
    assert results_table.find('two.langs@example.com') < results_table.find(
        'alsotwo.alsotwo@example.com'
    )

    # Click the export button again
    response = page.click('Export')
    sheet = load_workbook(BytesIO(response.body)).worksheets[0]
    assert sheet.max_row == 3
    header = {cell.value: cell.column for cell in sheet[1]}  # type: ignore[misc]

    # Check order is reversed in export (4444 first, then 5555)
    data_row = 2
    assert sheet.cell(
        row=data_row, column=header['Personal Nr.']).value == 4444
    assert sheet.cell(
        row=data_row, column=header['Nachname']).value == 'langs'

    data_row = 3
    assert sheet.cell(
        row=data_row, column=header['Personal Nr.']).value == 5555
    assert sheet.cell(
        row=data_row, column=header['Nachname']).value == 'alsotwo'



def test_file_security(client: Client) -> None:
    translators = TranslatorCollection(client.app)
    trs_id = translators.add(**translator_data).id
    transaction.commit()

    def content_disposition(file: str, filename: str) -> bool:
        return filename in client.get(file).headers['Content-Disposition']

    # Add a published general, an unpublished general and a translator file
    client.login_admin()
    page = client.get('/files')
    page.form['file'] = [upload_pdf('p.pdf')]
    page = page.form.submit()
    url = page.pyquery('div[ic-get-from]')[0].attrib['ic-get-from']
    published_file = url.replace('/details', '')
    assert 'Öffentlich' in client.get(url)
    assert 'Private' not in client.get(url)
    assert content_disposition(published_file, 'p.pdf')

    page = client.get('/files')
    page.form['file'] = [upload_pdf('u.pdf')]
    page = page.form.submit()
    url = page.pyquery('div[ic-get-from]')[0].attrib['ic-get-from']
    unpublished_file = url.replace('/details', '')
    assert content_disposition(unpublished_file, 'u.pdf')
    page = client.get(url)
    page = client.post(page.pyquery('a.is-published')[0].attrib['ic-post-to'])
    assert 'Öffentlich' not in client.get(url)
    assert 'Privat' in client.get(url)

    # A bug stemmed from conflicting 'cache-control' headers ('private'
    # and 'public'), causing ambiguous caching rules. For private files, merely
    # setting a 'private' Cache-Control header isn't sufficient.
    # We have to override the 'public' Cache-Control directive.
    headers = client.get(unpublished_file).headers
    cache_header_values = [
        i[1] for i in headers.items() if i[0].lower() == 'cache-control'
    ]
    for header_val in cache_header_values:
        assert 'private' in header_val
        assert 'public' not in header_val

    page = client.get(f'/translator/{trs_id}').click('Dokumente')
    page.form['file'] = [upload_pdf('t.pdf')]
    page = page.form.submit()
    translator_file = page.pyquery('div[ic-get-from]')[0].attrib['ic-get-from']
    translator_file = translator_file.replace('/details', '')
    assert content_disposition(translator_file, 't.pdf')
    client.logout()

    # Editors can't manage and can see general files but not translator files
    client.login_editor()
    page = client.get(f'/translator/{trs_id}')
    assert 'Dokumente' not in page
    client.get('/files', status=403)
    assert content_disposition(published_file, 'p.pdf')
    assert content_disposition(unpublished_file, 'u.pdf')
    client.get(f'/documents/{trs_id}', status=403)
    client.get(translator_file, status=403)
    client.logout()

    # Members can't manage and can see general files but not translator files
    client.login_member()
    page = client.get(f'/translator/{trs_id}')
    assert 'Dokumente' not in page
    client.get('/files', status=403)
    assert content_disposition(published_file, 'p.pdf')
    assert content_disposition(unpublished_file, 'u.pdf')
    client.get(f'/documents/{trs_id}', status=403)
    client.get(translator_file, status=403)
    client.logout()

    # Anonymous can only view published general files
    client.get('/files', status=403)
    assert content_disposition(published_file, 'p.pdf')
    client.get(unpublished_file, status=403)
    client.get(f'/documents/{trs_id}', status=403)
    client.get(translator_file, status=403)


def test_translator_file_access_forbidden(client: Client) -> None:
    client.login_admin()
    page = client.get('/files')
    page.form['file'] = [upload_pdf('test.pdf')]
    page = page.form.submit()
    url = page.pyquery('div[ic-get-from]')[0].attrib['ic-get-from']
    file_url = url.replace('/details', '')
    client.logout()

    client.login_translator()
    client.get(file_url, status=403)
    client.get(f'{file_url}/thumbnail', status=403)
    client.get(f'{file_url}/small', status=403)
    client.get(f'{file_url}/medium', status=403)
    client.get(file_url, status=403, extra_environ={'REQUEST_METHOD': 'HEAD'})
    thumbnail_url = f'{file_url}/thumbnail'
    client.get(
        thumbnail_url,
        status=403,
        extra_environ={'REQUEST_METHOD': 'HEAD'}
    )
    client.logout()


def test_translator_directory_settings(client: Client) -> None:
    client.login_admin()
    settings = client.get('/').follow().click('Verzeichniseinstellungen')

    def map_value(page: ExtendedResponse) -> Any:
        return decode_map_value(page.form['coordinates'].value)

    assert not map_value(settings)

    settings = client.get('/directory-settings')
    assert map_value(settings) == Coordinates()
    assert not settings.form['declaration_link'].value

    settings.form['coordinates'] = encode_map_value({
        'lat': 46, 'lon': 7, 'zoom': 12
    })
    settings.form['declaration_link'].value = 'https://t.ch/file.pdf'

    page = settings.form.submit().follow()
    assert 'Ihre Änderungen wurden gespeichert' in page
    settings = client.get('/directory-settings')
    assert map_value(settings) == Coordinates(lat=46, lon=7, zoom=12)
    assert settings.form['declaration_link'].value == 'https://t.ch/file.pdf'

    assert 'https://t.ch/file.pdf' in client.get('/request-accreditation')


def test_view_redirects(client: Client) -> None:
    # Create a translator
    languages = create_languages(client.app.session())
    language_id = str(languages[0].id)
    transaction.commit()

    client.login_admin()
    page = client.get('/translators/new')
    page.form['pers_id'] = 123456
    page.form['first_name'] = 'First'
    page.form['last_name'] = 'Last'
    page.form['email'] = 'translator@example.org'
    page.form['tel_mobile'] = '+41791234567'
    page.form['nationalities'] = ['CH']
    page.form['social_sec_number'] = '756.1234.5678.97'
    page.form['agency_references'] = 'OK'
    page.form['mother_tongues_ids'] = [language_id]
    page = page.form.submit().follow()
    assert 'Übersetzer/in hinzugefügt' in page
    translator_url = page.request.url
    client.logout()

    mail = client.get_email(0, flush_queue=True)['TextBody']
    reset_password_url = re.search(  # type: ignore[union-attr]
        r'(http://localhost/auth/reset-password[^)]+)', mail
    ).group()
    page = client.get(reset_password_url)
    page.form['email'] = 'translator@example.org'
    page.form['password'] = 'known_very_secure_password'
    page.form.submit()

    # Test redirects
    urls = {
        'translator@example.org': {
            'homepage': translator_url,
            'login': translator_url,
            'logout': 'http://localhost/auth/login',
            'password': 'known_very_secure_password',
            'to': 'http://localhost/topics/informationen'
        },
        'member@example.org': {
            'homepage': 'http://localhost/translators',
            'login': 'http://localhost/translators',
            'logout': 'http://localhost/auth/login',
            'password': 'hunter2',
            'to': translator_url
        }
    }
    urls['editor@example.org'] = urls['member@example.org']
    urls['admin@example.org'] = urls['member@example.org']

    for user, data in urls.items():
        page = client.login(user, data['password']).maybe_follow()
        assert page.request.url == data['login']

        page = client.login(user, data['password'], data['to']).maybe_follow()
        assert page.request.url == data['to']

        page = client.get('/').maybe_follow()
        assert page.request.url == data['homepage']

        page = client.logout().maybe_follow()
        assert page.request.url == data['logout']


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_view_translator_mutation(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client
) -> None:

    session = client.app.session()
    languages = create_languages(session)
    certs = create_certificates(session)
    cert_ids = [str(cert.id) for cert in certs]
    language_ids = [str(lang.id) for lang in languages]
    transaction.commit()

    client.login_admin()

    settings = client.get('/directory-settings')
    settings.form['coordinates'] = encode_map_value({
        'lat': 46, 'lon': 7, 'zoom': 12
    })
    settings.form.submit()

    # Create a new translator
    page = client.get('/translators/new')
    # fields visible by members
    page.form['first_name'] = 'Uncle'
    page.form['last_name'] = 'Bob'
    page.form['pers_id'] = 978654
    page.form['admission'] = 'uncertified'
    page.form['gender'] = 'M'
    page.form['withholding_tax'] = False
    page.form['self_employed'] = False
    page.form['date_of_birth'] = '1970-01-01'
    page.form['nationalities'] = ['CH']
    page.form['coordinates'] = encode_map_value({
        'lat': 46, 'lon': 7, 'zoom': 12
    })
    page.form['address'] = 'Fakestreet 123'
    page.form['zip_code'] = '6000'
    page.form['city'] = 'Luzern'
    page.form['email'] = 'test@test.com'
    page.form['tel_private'] = '+41412223344'
    page.form['tel_mobile'] = '+41412223345'
    page.form['tel_office'] = '+41412223346'
    page.form['availability'] = 'Always'
    page.form['mother_tongues_ids'] = language_ids[0:1]
    page.form['spoken_languages_ids'] = language_ids[1:2]
    page.form['written_languages_ids'] = language_ids[2:3]
    page.form['monitoring_languages_ids'] = language_ids[3:4]
    page.form.get('expertise_professional_guilds', index=0).checked = True
    page.form['expertise_professional_guilds_other'] = ['Psychologie']
    page.form.get('expertise_interpreting_types', index=0).checked = True
    # additional editor fields
    page.form['social_sec_number'] = '756.1234.5678.97'
    page.form['bank_name'] = 'Luzerner Bank'
    page.form['bank_address'] = 'Bankplatz Luzern'
    page.form['account_owner'] = 'Oncle Bob'
    page.form['iban'] = 'DE07 1234 1234 1234 1234 12'
    # additional fields visible by translators
    page.form['operation_comments'] = 'No operation comments'
    page.form['confirm_name_reveal'] = False
    page.form['date_of_application'] = '2020-01-01'
    page.form['profession'] = 'Handwerker'
    page.form['occupation'] = 'Bäcker'
    page.form['agency_references'] = 'All okay'
    page.form['education_as_interpreter'] = False
    page.form['certificates_ids'] = cert_ids[0:1]
    page.form['comments'] = 'No comments'
    with patch(
            'onegov.gis.utils.MapboxRequests.directions',
            return_value=FakeResponse({
                'code': 'Ok',
                'routes': [{'distance': 1000}]
            })
    ):
        assert 'hinzugefügt' in page.form.submit().follow()

    client.logout()
    reset_password_url = re.search(  # type: ignore[union-attr]
        r'(http://localhost/auth/reset-password[^)]+)',
        client.get_email(0, flush_queue=True)['TextBody']
    ).group()
    page = client.get(reset_password_url)
    page.form['email'] = 'test@test.com'
    page.form['password'] = 'known_very_secure_password'
    page.form.submit()

    # Report change as editor
    client.login_editor()
    page = client.get('/').maybe_follow().click('BOB, Uncle')
    page = page.click('Mutation melden')
    page.form['submitter_message'] = 'Hallo!'
    page.form['first_name'] = 'Aunt'
    page.form['last_name'] = 'Anny'
    page.form['pers_id'] = 123456
    page.form['admission'] = 'in_progress'
    page.form['gender'] = 'F'
    page.form['withholding_tax'] = True
    page.form['self_employed'] = True
    page.form['date_of_birth'] = '1960-01-01'
    page.form['nationalities'] = ['DE']
    page.form['coordinates'] = encode_map_value({
        'lat': 47, 'lon': 8, 'zoom': 13
    })
    page.form['address'] = 'Fakestreet 321'
    page.form['zip_code'] = '6010'
    page.form['city'] = 'Kriens'
    page.form['hometown'] = 'Weggis'
    page.form['tel_private'] = '+41412223347'
    page.form['tel_mobile'] = '+41412223348'
    page.form['tel_office'] = '+41412223349'
    page.form['availability'] = 'Nie'
    page.form['mother_tongues'] = language_ids[1:3]
    page.form['spoken_languages'] = language_ids[0:2]
    page.form['written_languages'] = language_ids[2:4]
    page.form['monitoring_languages'] = language_ids[0:4]
    page.form['expertise_interpreting_types'].select_multiple([
        'consecutive', 'negotiation'
    ])
    page.form['expertise_professional_guilds'].select_multiple([
        'economy', 'art_leisure'
    ])
    page.form['expertise_professional_guilds_other'] = ['Exorzismus']
    # additional editor fields
    page.form['social_sec_number'] = '756.9217.0769.85'
    page.form['bank_name'] = 'Krienser Bank'
    page.form['bank_address'] = 'Bankplatz Kriens'
    page.form['account_owner'] = 'Aunt Anny'
    page.form['iban'] = 'CH5604835012345678009'
    with patch(
            'onegov.gis.utils.MapboxRequests.directions',
            return_value=FakeResponse({
                'code': 'Ok',
                'routes': [{'distance': 2000}]
            })
    ):
        page = page.form.submit().follow()
        assert 'Ihre Anfrage wird in Kürze bearbeitet' in page

    mail = client.get_email(0, flush_queue=True)
    assert mail['To'] == 'editor@example.org'
    assert 'BOB, Uncle: Ihr Ticket wurde eröffnet' in mail['Subject']

    assert connect.call_count == 1
    assert authenticate.call_count == 1
    assert broadcast.call_count == 1
    assert broadcast.call_args[0][3]['event'] == 'browser-notification'
    assert broadcast.call_args[0][3]['title'] == 'Neues Ticket'
    assert broadcast.call_args[0][3]['created']

    client.logout()
    client.login_admin()
    page = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert 'Hallo!' in page
    assert 'Vorname: Aunt' in page
    assert 'Nachname: Anny' in page
    assert 'Personal Nr.: 123456' in page
    assert 'Zulassung: im Zulassungsverfahren' in page
    assert 'Geschlecht: Weiblich' in page
    assert 'Quellensteuer: Ja' in page
    assert 'Selbständig: Ja' in page
    assert 'Geburtsdatum: 1960-01-01' in page
    assert 'Nationalität(en): Deutschland' in page
    assert 'Standort: 47, 8' in page
    assert 'Strasse und Hausnummer: Fakestreet 321' in page
    assert 'PLZ: 6010' in page
    assert 'Ort: Kriens' in page
    assert 'Weggis' in page
    assert 'Fahrdistanz (km): 2.0' in page
    assert 'Telefon Privat: +41412223347' in page
    assert 'Telefon Mobile: +41412223348' in page
    assert 'Telefon Geschäft: +41412223349' in page
    assert 'Erreich- und Verfügbarkeit: Nie' in page
    assert 'Muttersprachen: French, Italian' in page
    assert 'Arbeitssprache - Wort: French, German' in page
    assert 'Arbeitssprache - Schrift: Arabic, Italian' in page
    assert (
        'Arbeitssprache - Kommunikationsüberwachung: '
        'Arabic, French, German, Italian'
    ) in page
    assert (
        'Fachkenntnisse nach Dolmetscherart: Konsektutivdolmetschen, '
        'Verhandlungsdolmetschen'
    ) in page
    assert (
        'Fachkenntnisse nach Berufssparte: Wirtschaft, Kunst und Freizeit'
    ) in page
    assert 'Fachkenntnisse nach Berufssparte: andere: Exorzismus' in page
    assert 'AHV-Nr.: 756.9217.0769.85' in page
    assert 'Bank Name: Krienser Bank' in page
    assert 'Bank Adresse: Bankplatz Kriens' in page
    assert 'Bank Konto lautend auf: Aunt Anny' in page
    assert 'IBAN: CH5604835012345678009' in page
    page.click('Ticket abschliessen')

    mail = client.get_email(0, flush_queue=True)
    assert mail['To'] == 'editor@example.org'
    assert 'BOB, Uncle: Ihre Anfrage wurde abgeschlossen' in mail['Subject']

    # Report change as translator
    client.logout()
    client.login('test@test.com', 'known_very_secure_password', None)
    page = client.get('/').maybe_follow()
    page = page.click('Mutation melden')
    page.form['submitter_message'] = 'Hallo!'
    page.form['first_name'] = 'Aunt'
    page.form['last_name'] = 'Anny'
    page.form['pers_id'] = 123456
    page.form['admission'] = 'in_progress'
    page.form['gender'] = 'F'
    page.form['withholding_tax'] = True
    page.form['self_employed'] = True
    page.form['date_of_birth'] = '1960-01-01'
    page.form['nationalities'] = ['SE', 'DE']
    page.form['coordinates'] = encode_map_value({
        'lat': 47, 'lon': 8, 'zoom': 13
    })
    page.form['address'] = 'Fakestreet 321'
    page.form['zip_code'] = '6010'
    page.form['city'] = 'Kriens'
    page.form['tel_private'] = '+41412223347'
    page.form['tel_mobile'] = '+41412223348'
    page.form['tel_office'] = '+41412223349'
    page.form['availability'] = 'Nie'
    page.form['mother_tongues'] = language_ids[1:3]
    page.form['spoken_languages'] = language_ids[0:2]
    page.form['written_languages'] = language_ids[2:4]
    page.form['monitoring_languages'] = language_ids[0:4]
    page.form['expertise_interpreting_types'].select_multiple([
        'consecutive', 'negotiation'
    ])
    page.form['expertise_professional_guilds'].select_multiple([
        'economy', 'art_leisure'
    ])
    page.form['expertise_professional_guilds_other'] = ['Exorzismus']
    # additional editor fields
    page.form['social_sec_number'] = '756.9217.0769.85'
    page.form['bank_name'] = 'Krienser Bank'
    page.form['bank_address'] = 'Bankplatz Kriens'
    page.form['account_owner'] = 'Aunt Anny'
    page.form['iban'] = 'CH5604835012345678009'
    # additional translator fields
    page.form['operation_comments'] = 'Keine'
    page.form['confirm_name_reveal'] = True
    page.form['date_of_application'] = '2021-01-01'
    page.form['date_of_decision'] = '2021-02-02'
    page.form['profession'] = 'Hochbauzeichner'
    page.form['occupation'] = 'Bauarbeiter'
    page.form['proof_of_preconditions'] = 'Keine'
    page.form['agency_references'] = 'Kanton LU'
    page.form['education_as_interpreter'] = True
    page.form['certificates'] = cert_ids[1:3]
    page.form['comments'] = 'Kein Kommentar'
    with patch(
            'onegov.gis.utils.MapboxRequests.directions',
            return_value=FakeResponse({
                'code': 'Ok',
                'routes': [{'distance': 2000}]
            })
    ):
        page = page.form.submit().follow()
        assert 'Ihre Anfrage wird in Kürze bearbeitet' in page

    mail = client.get_email(0, flush_queue=True)
    assert mail['To'] == 'test@test.com'
    assert 'BOB, Uncle: Ihr Ticket wurde eröffnet' in mail['Subject']

    assert connect.call_count == 2
    assert authenticate.call_count == 2
    assert broadcast.call_count == 2
    assert broadcast.call_args[0][3]['event'] == 'browser-notification'
    assert broadcast.call_args[0][3]['title'] == 'Neues Ticket'
    assert broadcast.call_args[0][3]['created']

    client.logout()
    client.login_admin()
    page = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert 'Hallo!' in page
    assert 'Vorname: Aunt' in page
    assert 'Nachname: Anny' in page
    assert 'Personal Nr.: 123456' in page
    assert 'Zulassung: im Zulassungsverfahren' in page
    assert 'Geschlecht: Weiblich' in page
    assert 'Quellensteuer: Ja' in page
    assert 'Selbständig: Ja' in page
    assert 'Geburtsdatum: 1960-01-01' in page
    assert 'Nationalität(en): Deutschland, Schweden' in page
    assert 'Standort: 47, 8' in page
    assert 'Strasse und Hausnummer: Fakestreet 321' in page
    assert 'PLZ: 6010' in page
    assert 'Ort: Kriens' in page
    assert 'Fahrdistanz (km): 2.0' in page
    assert 'Telefon Privat: +41412223347' in page
    assert 'Telefon Mobile: +41412223348' in page
    assert 'Telefon Geschäft: +41412223349' in page
    assert 'Erreich- und Verfügbarkeit: Nie' in page
    assert 'Muttersprachen: French, Italian' in page
    assert 'Arbeitssprache - Wort: French, German' in page
    assert 'Arbeitssprache - Schrift: Arabic, Italian' in page
    assert 'Arbeitssprache - Wort: French, German' in page
    assert 'Arbeitssprache - Schrift: Arabic, Italian' in page
    assert (
        'Arbeitssprache - Kommunikationsüberwachung: '
        'Arabic, French, German, Italian'
    ) in page
    assert (
        'Fachkenntnisse nach Dolmetscherart: Konsektutivdolmetschen, '
        'Verhandlungsdolmetschen'
    ) in page
    assert (
        'Fachkenntnisse nach Berufssparte: Wirtschaft, Kunst und Freizeit'
    ) in page
    assert 'Fachkenntnisse nach Berufssparte: andere: Exorzismus' in page
    assert 'AHV-Nr.: 756.9217.0769.85' in page
    assert 'Bank Name: Krienser Bank' in page
    assert 'Bank Adresse: Bankplatz Kriens' in page
    assert 'Bank Konto lautend auf: Aunt Anny' in page
    assert 'IBAN: CH5604835012345678009' in page
    assert 'Besondere Hinweise Einsatzmöglichkeiten: Keine' in page
    assert 'Zustimmung Namensbekanntgabe: Ja' in page
    assert 'Bewerbung Datum: 2021-01-01' in page
    assert 'Entscheid Datum: 2021-02-02' in page
    assert 'Erlernter Beruf: Hochbauzeichner' in page
    assert 'Aktuelle berufliche Tatigkeit: Bauarbeiter' in page
    assert 'Nachweis der Voraussetzung: Keine' in page
    assert 'Referenzen Behörden: Kanton LU' in page
    assert 'Ausbildung Dolmetscher: Ja' in page
    assert 'Zertifikate: BBBB, CCCC' in page
    assert 'Bemerkungen: Kein Kommentar' in page

    page = page.click('Vorgeschlagene Änderungen übernehmen')
    page.form.get('changes', index=40).checked = False
    page = page.form.submit().follow()
    assert (
        'Vorgeschlagene Änderungen übernommen: '
        'Vorname, Nachname, Personal Nr., Zulassung, Quellensteuer, '
        'Selbständig, Geschlecht, Geburtsdatum, Nationalität(en), '
        'Standort, Strasse und Hausnummer, PLZ, Ort, Fahrdistanz (km), '
        'AHV-Nr., Bank Name, Bank Adresse, Bank Konto lautend auf, IBAN, '
        'Telefon Privat, Telefon Mobile, Telefon Geschäft, '
        'Erreich- und Verfügbarkeit, '
        'Fachkenntnisse nach Dolmetscherart, '
        'Fachkenntnisse nach Berufssparte, '
        'Fachkenntnisse nach Berufssparte: andere, '
        'Besondere Hinweise Einsatzmöglichkeiten, '
        'Zustimmung Namensbekanntgabe, Bewerbung Datum, Entscheid Datum, '
        'Muttersprachen, Arbeitssprache - Wort, Arbeitssprache - Schrift, '
        'Arbeitssprache - Kommunikationsüberwachung, '
        'Erlernter Beruf, '
        'Aktuelle berufliche Tatigkeit, '
        'Nachweis der Voraussetzung, Referenzen Behörden, '
        'Ausbildung Dolmetscher, Zertifikate.'
    ) in page

    page.click('Ticket abschliessen')

    mail = client.get_email(0, flush_queue=True)
    assert mail['To'] == 'test@test.com'
    assert 'ANNY, Aunt: Ihre Anfrage wurde abgeschlossen' in mail['Subject']

    page = client.get('/').follow().click('ANNY, Aunt')
    assert 'Aunt' in page
    assert 'Anny' in page
    assert '123456' in page
    assert 'im Zulassungsverfahren' in page
    assert 'Weiblich' in page
    assert 'Ja' in page
    assert 'Ja' in page
    assert '01.01.1960' in page
    assert 'Deutschland, Schweden' in page
    assert 'Fakestreet 321' in page
    assert '6010' in page
    assert 'Kriens' in page
    assert '2.0' in page
    assert '+41412223347' in page
    assert '+41412223348' in page
    assert '+41412223349' in page
    assert 'Nie' in page
    assert 'Arabic' in page
    assert 'French' in page
    assert 'German' in page
    assert 'Italian' in page
    assert 'Konsektutivdolmetschen' in page
    assert 'Verhandlungsdolmetschen' in page
    assert 'Wirtschaft' in page
    assert 'Kunst und Freizeit' in page
    assert 'Exorzismus' in page
    assert '756.9217.0769.85' in page
    assert 'Krienser Bank' in page
    assert 'Bankplatz Kriens' in page
    assert 'Aunt Anny' in page
    assert 'CH5604835012345678009' in page
    assert 'Keine' in page
    assert 'Ja' in page
    assert '01.01.2021' in page
    assert '02.02.2021' in page
    assert 'French' in page
    assert 'German' in page
    assert 'Arabic' in page
    assert 'Italian' in page
    assert 'Hochbauzeichner' in page
    assert 'Bauarbeiter' in page
    assert 'Keine' in page
    assert 'Kanton LU' in page
    assert 'Ja' in page
    assert 'BBBB' in page
    assert 'CCCC' in page
    assert 'Kein Kommentar' not in page


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_translator_mutation_with_document_upload(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client
) -> None:

    session = client.app.session()
    languages = create_languages(session)
    certs = create_certificates(session)
    cert_ids = [str(cert.id) for cert in certs]
    language_ids = [str(lang.id) for lang in languages]
    transaction.commit()

    client.login_admin()

    settings = client.get('/directory-settings')
    settings.form['coordinates'] = encode_map_value({
        'lat': 46, 'lon': 7, 'zoom': 12
    })
    settings.form.submit()

    # Create a new translator
    page = client.get('/translators/new')
    page.form['first_name'] = 'Uncle'
    page.form['last_name'] = 'Bob'
    page.form['pers_id'] = 978654
    page.form['admission'] = 'uncertified'
    page.form['gender'] = 'M'
    page.form['withholding_tax'] = False
    page.form['self_employed'] = False
    page.form['date_of_birth'] = '1970-01-01'
    page.form['nationalities'] = ['CH']
    page.form['coordinates'] = encode_map_value({
        'lat': 46, 'lon': 7, 'zoom': 12
    })
    page.form['address'] = 'Fakestreet 123'
    page.form['zip_code'] = '6000'
    page.form['city'] = 'Luzern'
    page.form['email'] = 'translator@test.com'
    page.form['tel_private'] = '+41412223344'
    page.form['tel_mobile'] = '+41412223345'
    page.form['tel_office'] = '+41412223346'
    page.form['availability'] = 'Always'
    page.form['mother_tongues_ids'] = language_ids[0:1]
    page.form['spoken_languages_ids'] = language_ids[1:2]
    page.form['written_languages_ids'] = language_ids[2:3]
    page.form['monitoring_languages_ids'] = language_ids[3:4]
    page.form.get('expertise_professional_guilds', index=0).checked = True
    page.form['expertise_professional_guilds_other'] = ['Psychologie']
    page.form.get('expertise_interpreting_types', index=0).checked = True
    page.form['social_sec_number'] = '756.1234.5678.97'
    page.form['bank_name'] = 'Luzerner Bank'
    page.form['bank_address'] = 'Bankplatz Luzern'
    page.form['account_owner'] = 'Uncle Bob'
    page.form['iban'] = 'DE07 1234 1234 1234 1234 12'
    page.form['operation_comments'] = 'No operation comments'
    page.form['confirm_name_reveal'] = False
    page.form['date_of_application'] = '2020-01-01'
    page.form['profession'] = 'Handwerker'
    page.form['occupation'] = 'Bäcker'
    page.form['agency_references'] = 'All okay'
    page.form['education_as_interpreter'] = False
    page.form['certificates_ids'] = cert_ids[0:1]
    page.form['comments'] = 'No comments'
    with patch(
        'onegov.gis.utils.MapboxRequests.directions',
        return_value=FakeResponse({
            'code': 'Ok',
            'routes': [{'distance': 1000}]
        })
    ):
        assert 'hinzugefügt' in page.form.submit().follow()

    initial_file_count = session.query(File).count()
    client.logout()
    client.login_editor()
    page = client.get('/').maybe_follow().click('BOB, Uncle')
    page = page.click('Mutation melden')
    page.form['submitter_message'] = 'Uploading new certificate!'

    assert 'uploaded_certificates' in page.form.fields
    page.form['uploaded_certificates'] = upload_pdf('certificate.pdf')
    page = page.form.submit().follow()
    assert 'Ihre Anfrage wird in Kürze bearbeitet' in page

    client.logout()
    client.login_admin()
    page = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert 'Uploading new certificate!' in page
    assert 'Dokumente' in page
    assert '.pdf' in page

    session = client.app.session()
    from onegov.ticket import TicketCollection

    new_file_count = session.query(File).count()
    assert new_file_count > initial_file_count

    tickets = TicketCollection(session).by_handler_code('TRN')
    ticket = tickets[-1] if tickets else None
    assert ticket is not None, "No ticket found"

    handler_data = ticket.handler_data.get('handler_data', {})
    file_ids = handler_data.get('file_ids', [])
    assert len(file_ids) > 0, "No file_ids found in ticket"

    uploaded_file = session.query(File).filter_by(id=file_ids[0]).first()
    assert uploaded_file is not None
    assert uploaded_file.name.endswith('.pdf')


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_view_accreditation(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client
) -> None:

    session = client.app.session()
    language_ids = [str(lang.id) for lang in create_languages(session)]
    transaction.commit()

    client.login_admin()

    settings = client.get('/directory-settings')
    settings.form['coordinates'] = encode_map_value({
        'lat': 46, 'lon': 7, 'zoom': 12
    })
    settings.form.submit()

    websocket_messages = 0

    def request_accreditation() -> ExtendedResponse:
        client.logout()

        page = client.get('/request-accreditation')
        page.form['last_name'] = 'Benito'
        page.form['first_name'] = 'Hugo'
        page.form['gender'] = 'M'
        page.form['date_of_birth'] = '1970-01-01'
        page.form['hometown'] = 'Zug'
        page.form['nationalities'] = ['CH']
        page.form['marital_status'] = 'verheiratet'
        page.form['coordinates'] = encode_map_value({
            'lat': 1, 'lon': 2, 'zoom': 12
        })
        page.form['address'] = 'Downing Street 5'
        page.form['zip_code'] = '4000'
        page.form['city'] = 'Luzern'
        page.form['drive_distance'] = '1.1'
        page.form['withholding_tax'] = False
        page.form['self_employed'] = True
        page.form['social_sec_number'] = '756.1234.4568.94'
        page.form['bank_name'] = 'R-BS'
        page.form['bank_address'] = 'Bullstreet 5'
        page.form['account_owner'] = 'Hugo Benito'
        page.form['iban'] = 'CH9300762011623852957'
        page.form['email'] = 'hugo.benito@translators.com'
        page.form['tel_private'] = '041 444 44 45'
        page.form['tel_office'] = '041 444 44 44'
        page.form['tel_mobile'] = '079 000 00 00'
        page.form['availability'] = '24h'
        page.form['confirm_name_reveal'] = True
        page.form['profession'] = 'Baker'
        page.form['occupation'] = 'Reporter'
        page.form['education_as_interpreter'] = False
        page.form['mother_tongues_ids'] = language_ids[0:1]
        page.form['spoken_languages_ids'] = language_ids[1:2]
        page.form['written_languages_ids'] = language_ids[2:3]
        page.form['monitoring_languages_ids'] = language_ids[3:4]
        page.form['expertise_interpreting_types'].select_multiple([
            'consecutive', 'negotiation'
        ])
        page.form['expertise_professional_guilds'].select_multiple([
            'economy', 'art_leisure'
        ])
        page.form['expertise_professional_guilds_other'] = ['Psychologie']
        page.form['agency_references'] = 'Some ref'
        page.form['admission_course_completed'] = False
        page.form['admission_course_agreement'] = True
        page.form['declaration_of_authorization'] = upload_pdf('1.pdf')
        page.form['letter_of_motivation'] = upload_pdf('2.pdf')
        page.form['resume'] = upload_pdf('3.pdf')
        page.form['certificates'] = upload_pdf('4.pdf')
        page.form['social_security_card'] = upload_pdf('5.pdf')
        page.form['passport'] = upload_pdf('6.pdf')
        page.form['passport_photo'] = upload_pdf('7.pdf')
        page.form['debt_collection_register_extract'] = upload_pdf('8.pdf')
        page.form['criminal_register_extract'] = upload_pdf('9.pdf')
        page.form['certificate_of_capability'] = upload_pdf('A.pdf')
        page.form['confirmation_compensation_office'] = upload_pdf('B.pdf')
        page.form['remarks'] = 'Some remarks'
        page.form['confirm_submission'] = True

        with patch(
                'onegov.gis.utils.MapboxRequests.directions',
                return_value=FakeResponse({
                    'code': 'Ok',
                    'routes': [{'distance': 2000}]
                })
        ):
            page = page.form.submit().follow()
            assert 'Ihre Anfrage wird in Kürze bearbeitet' in page

        mail = client.get_email(0, flush_queue=True)
        assert mail['To'] == 'hugo.benito@translators.com'
        assert 'BENITO, Hugo: Ihr Ticket wurde eröffnet' in mail['Subject']

        nonlocal websocket_messages
        websocket_messages += 1
        assert connect.call_count == websocket_messages
        assert authenticate.call_count == websocket_messages
        assert broadcast.call_count == websocket_messages
        assert broadcast.call_args[0][3]['event'] == 'browser-notification'
        assert broadcast.call_args[0][3]['title'] == 'Neues Ticket'
        assert broadcast.call_args[0][3]['created']

        client.login_admin()
        page = client.get('/tickets/ALL/open').click('Annehmen').follow()
        assert 'Benito' in page
        assert 'Hugo' in page
        assert 'Männlich' in page
        assert '01.01.1970' in page
        assert 'Zug' in page
        assert 'CH' in page
        assert 'verheiratet' in page
        assert '2.0 km' in page
        assert 'Downing Street 5' in page
        assert '4000' in page
        assert 'Luzern' in page
        assert '1.1' in page
        assert '"withholding-tax">Nein' in page
        assert '"self-employed">Ja' in page
        assert '756.1234.4568.94' in page
        assert 'R-BS' in page
        assert 'Bullstreet 5' in page
        assert 'Hugo Benito' in page
        assert 'CH9300762011623852957' in page
        assert 'hugo.benito@translators.com' in page
        assert '041 444 44 45' in page
        assert '041 444 44 44' in page
        assert '079 000 00 00' in page
        assert '24h' in page
        assert '"confirm-name-reveal">Ja' in page
        assert 'Baker' in page
        assert 'Reporter' in page
        assert '"education-as-interpreter">Nein' in page
        assert 'German' in page
        assert 'French' in page
        assert 'Italian' in page
        assert 'Arabic' in page
        assert 'Wirtschaft' in page
        assert 'Kunst und Freizeit' in page
        assert 'Psychologie' in page
        assert 'Konsektutivdolmetschen' in page
        assert 'Verhandlungsdolmetschen' in page
        assert 'Some ref' in page
        assert '"admission-course-completed">Nein' in page
        assert '"admission-course-agreement">Ja' in page
        assert 'Some remarks' in page

        check_pdf(page, '1.pdf', 'Unterschriebene Einverständniserklärung.pdf')
        check_pdf(page, '2.pdf', 'Kurzes Motivationsschreiben.pdf')
        check_pdf(page, '3.pdf', 'Lebenslauf.pdf')
        check_pdf(page, '4.pdf', 'Zertifikate.pdf')
        check_pdf(page, '5.pdf', 'AHV-Ausweis.pdf')
        check_pdf(page, '6.pdf', 'ID, Pass oder Ausländerausweis.pdf')
        check_pdf(page, '7.pdf', 'Aktuelles Passfoto.pdf')
        check_pdf(page, '8.pdf',
                  'Aktueller Auszug aus dem Betreibungsregister.pdf')
        check_pdf(page, '9.pdf',
                  'Aktueller Auszug aus dem Zentralstrafregister.pdf')
        check_pdf(page, 'A.pdf', 'Handlungsfähigkeitszeugnis.pdf')
        check_pdf(page, 'B.pdf',
                  'Bestätigung der Ausgleichskasse betreffend '
                  'Selbständigkeit.pdf')

        return page

    # Request accreditation
    page = request_accreditation()

    assert 'Briefvorlagen' in page
    mail_templates_link = page.pyquery('a.envelope')[0].attrib['href']
    resp = client.request(mail_templates_link)
    assert resp.status_code == 200

    # Refuse admission
    page = page.click('Zulassung verweigern').form.submit().follow()
    assert 'Der hinterlegte Datensatz wurde entfernt' in page

    page.click('Ticket abschliessen')

    mail = client.get_email(0, flush_queue=True)
    assert mail['To'] == 'hugo.benito@translators.com'
    assert 'BENITO, Hugo: Ihre Anfrage wurde abgeschlossen' in mail['Subject']

    # Request accredtitation
    page = request_accreditation()

    # Grant admission
    page = page.click('Zulassung erteilen').form.submit().follow()
    assert 'Zulassung erteilt' in page
    assert 'Aktivierungs-Email verschickt' in page

    mail = client.get_email(0, flush_queue=True)
    assert mail['To'] == 'hugo.benito@translators.com'
    assert mail['Subject'] == 'Ein Konto wurde für Sie erstellt'

    reset_password_url = re.search(  # type: ignore[union-attr]
        r'(http://localhost/auth/reset-password[^)]+)',
        mail['TextBody']
    ).group()

    page.click('Ticket abschliessen')

    mail = client.get_email(0, flush_queue=True)
    assert mail['To'] == 'hugo.benito@translators.com'
    assert 'BENITO, Hugo: Ihre Anfrage wurde abgeschlossen' in mail['Subject']

    page = client.get('/').follow()
    assert 'hugo.benito@translators.com' in page

    page = page.click('BENITO, Hugo')
    assert 'BENITO, Hugo' in page
    assert '756.1234.4568.94' in page

    client.logout()

    # Login as translator
    page = client.get(reset_password_url)
    page.form['email'] = 'hugo.benito@translators.com'
    page.form['password'] = 'known_very_secure_password'
    page.form.submit()

    page = client.login(
        'hugo.benito@translators.com', 'known_very_secure_password')
    page = page.maybe_follow()
    assert 'BENITO, Hugo' in page
    assert '756.1234.4568.94' in page


@patch(
    'onegov.gis.utils.MapboxRequests.directions',
    return_value=FakeResponse({'code': 'Ok', 'routes': [{'distance': 2000}]})
)
def test_view_accreditation_errors(
    directions: MagicMock,
    client: Client
) -> None:

    session = client.app.session()
    language_ids = [str(lang.id) for lang in create_languages(session)]
    transaction.commit()

    client.login_admin()
    settings = client.get('/directory-settings')
    settings.form['coordinates'] = encode_map_value({
        'lat': 46, 'lon': 7, 'zoom': 12
    })
    settings.form.submit()
    client.logout()

    # first try
    page = client.get('/request-accreditation')
    page.form['last_name'] = 'Benito'
    page.form['declaration_of_authorization'] = upload_pdf('1.pdf')
    page = page.form.submit()
    assert 'Das Formular enthält Fehler' in page
    assert page.form['last_name'].value == 'Benito'
    upload = page.form.fields['declaration_of_authorization']
    assert upload[0].value == 'keep'
    assert upload[2].value == '1.pdf'
    assert upload[3].value

    # second try
    page.form['date_of_birth'] = '1970-01-01'
    page.form['letter_of_motivation'] = upload_pdf('2.pdf')
    page.form['resume'] = upload_pdf('3.pdf')
    page = page.form.submit()
    assert 'Das Formular enthält Fehler' in page
    assert page.form['last_name'].value == 'Benito'
    assert page.form['date_of_birth'].value == '1970-01-01'
    upload = page.form.fields['declaration_of_authorization']
    assert upload[0].value == 'keep'
    assert upload[2].value == '1.pdf'
    assert upload[3].value
    upload = page.form.fields['letter_of_motivation']
    assert upload[0].value == 'keep'
    assert upload[2].value == '2.pdf'
    assert upload[3].value
    upload = page.form.fields['resume']
    assert upload[0].value == 'keep'
    assert upload[2].value == '3.pdf'
    assert upload[3].value

    # final try
    page.form['first_name'] = 'Hugo'
    page.form['gender'] = 'M'
    page.form['hometown'] = 'Zug'
    page.form['nationalities'] = ['CH']
    page.form['marital_status'] = 'verheiratet'
    page.form['coordinates'] = encode_map_value({
        'lat': 1, 'lon': 2, 'zoom': 12
    })
    page.form['address'] = 'Downing Street 5'
    page.form['zip_code'] = '4000'
    page.form['city'] = 'Luzern'
    page.form['drive_distance'] = '1.1'
    page.form['withholding_tax'] = False
    page.form['self_employed'] = False
    page.form['social_sec_number'] = '756.1234.4568.94'
    page.form['bank_name'] = 'R-BS'
    page.form['bank_address'] = 'Bullstreet 5'
    page.form['account_owner'] = 'Hugo Benito'
    page.form['iban'] = 'CH9300762011623852957'
    page.form['email'] = 'hugo.benito@translators.com'
    page.form['tel_private'] = '041 444 44 45'
    page.form['tel_office'] = '041 444 44 44'
    page.form['tel_mobile'] = '079 000 00 00'
    page.form['availability'] = '24h'
    page.form['confirm_name_reveal'] = True
    page.form['profession'] = 'Baker'
    page.form['occupation'] = 'Reporter'
    page.form['education_as_interpreter'] = False
    page.form['mother_tongues_ids'] = language_ids[0:1]
    page.form['spoken_languages_ids'] = language_ids[1:2]
    page.form['written_languages_ids'] = language_ids[2:3]
    page.form['monitoring_languages_ids'] = language_ids[3:4]
    page.form['expertise_interpreting_types'].select_multiple([
        'consecutive', 'negotiation'
    ])
    page.form['expertise_professional_guilds'].select_multiple([
        'economy', 'art_leisure'
    ])
    page.form['expertise_professional_guilds_other'] = ['Psychologie']
    page.form['agency_references'] = 'Some ref'
    page.form['admission_course_completed'] = False
    page.form['admission_course_agreement'] = True
    page.form.get('resume', 0).select('replace')
    page.form.get('resume', 1).value = upload_pdf('3_new.pdf')
    page.form['certificates'] = upload_pdf('4.pdf')
    page.form['social_security_card'] = upload_pdf('5.pdf')
    page.form['passport'] = upload_pdf('6.pdf')
    page.form['passport_photo'] = upload_pdf('7.pdf')
    page.form['debt_collection_register_extract'] = upload_pdf('8.pdf')
    page.form['criminal_register_extract'] = upload_pdf('9.pdf')
    page.form['certificate_of_capability'] = upload_pdf('A.pdf')
    page.form['remarks'] = 'Some remarks'
    page.form['confirm_submission'] = True

    page = page.form.submit().follow()
    assert 'Ihre Anfrage wird in Kürze bearbeitet' in page

    client.login_admin()
    page = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert 'Benito' in page
    assert 'Hugo' in page
    assert 'Männlich' in page
    assert '01.01.1970' in page
    assert 'Zug' in page
    assert 'CH' in page
    assert 'verheiratet' in page
    assert '2.0 km' in page
    assert 'Downing Street 5' in page
    assert '4000' in page
    assert 'Luzern' in page
    assert '1.1' in page
    assert '"withholding-tax">Nein' in page
    assert '"self-employed">Nein' in page
    assert '756.1234.4568.94' in page
    assert 'R-BS' in page
    assert 'Bullstreet 5' in page
    assert 'Hugo Benito' in page
    assert 'CH9300762011623852957' in page
    assert 'hugo.benito@translators.com' in page
    assert '041 444 44 45' in page
    assert '041 444 44 44' in page
    assert '079 000 00 00' in page
    assert '24h' in page
    assert '"confirm-name-reveal">Ja' in page
    assert 'Baker' in page
    assert 'Reporter' in page
    assert '"education-as-interpreter">Nein' in page
    assert 'German' in page
    assert 'French' in page
    assert 'Italian' in page
    assert 'Arabic' in page
    assert 'Wirtschaft' in page
    assert 'Kunst und Freizeit' in page
    assert 'Psychologie' in page
    assert 'Konsektutivdolmetschen' in page
    assert 'Verhandlungsdolmetschen' in page
    assert 'Some ref' in page
    assert '"admission-course-completed">Nein' in page
    assert '"admission-course-agreement">Ja' in page
    assert 'Some remarks' in page

    check_pdf(page, '1.pdf', 'Unterschriebene Einverständniserklärung.pdf')
    check_pdf(page, '2.pdf', 'Kurzes Motivationsschreiben.pdf')
    check_pdf(page, '3_new.pdf', 'Lebenslauf.pdf')
    check_pdf(page, '4.pdf', 'Zertifikate.pdf')
    check_pdf(page, '5.pdf', 'AHV-Ausweis.pdf')
    check_pdf(page, '6.pdf', 'ID, Pass oder Ausländerausweis.pdf')
    check_pdf(page, '7.pdf', 'Aktuelles Passfoto.pdf')
    check_pdf(page, '8.pdf',
              'Aktueller Auszug aus dem Betreibungsregister.pdf')
    check_pdf(page, '9.pdf',
              'Aktueller Auszug aus dem Zentralstrafregister.pdf')
    check_pdf(page, 'A.pdf', 'Handlungsfähigkeitszeugnis.pdf')


def test_view_mail_template(client: Client) -> None:
    session = client.app.session()
    translator = Translator(**translator_data)
    translators = TranslatorCollection(client.app)
    trs_id = translators.add(**translator_data).id
    transaction.commit()

    docx_path = module_path(
        "tests.onegov.translator_directory", "fixtures/Vorlage.docx"
    )
    signature_path = module_path(
        "tests.onegov.translator_directory",
        "fixtures/Unterschrift__DOJO__Adj_mV_John_Doe__Stv_Dienstchef.jpg",
    )
    client.login_admin()

    upload_file(docx_path, client, content_type='application/vnd.ms-office')
    upload_file(signature_path, client)
    files = FileCollection(session).query().all()

    assert files[0].name == basename(docx_path)
    assert files[1].name == basename(signature_path)

    # User.realname has to exist since this is required for signature
    user = UserCollection(session).by_username('admin@example.org')
    assert user is not None
    assert user.realname is not None
    first_name, last_name = user.realname.split(" ")
    assert first_name == 'John'
    assert last_name == 'Doe'

    # Now we have everything set up, go to the mail templates and generate one
    page = client.get(f'/translator/{trs_id}').click('Briefvorlagen')
    page.form['templates'] = basename(docx_path)
    resp = page.form.submit()

    found_variables_in_docx = set()
    expected_variables_in_docx = {
        'Sehr geehrter Herr',
        translator.address,
        translator.zip_code,
        translator.city,
        translator.first_name,
        translator.last_name,
        first_name,
        last_name
    }

    doc = docx.Document(BytesIO(resp.body))
    for target in expected_variables_in_docx:
        for block in iter_block_items(doc):
            line = block.text
            # make sure all variables have been rendered
            assert '{{' not in line and '}}' not in line
            if target and target in line:
                found_variables_in_docx.add(target)

    assert expected_variables_in_docx == found_variables_in_docx


def test_mail_templates_with_hometown_and_ticket_nr(client: Client) -> None:
    session = client.app.session()
    translator_data_copy = copy.deepcopy(translator_data)
    translator_data_copy['city'] = 'SomeOtherTown'
    translator_data_copy['hometown'] = 'Luzern'

    translators = TranslatorCollection(client.app)
    trs_id = translators.add(**translator_data_copy).id

    ticket = AccreditationTicket(
        number='AKK-1000-0000',
        title='AKK-1000-0000',
        group='AKK-1000-0000',
        handler_id='1',
        handler_data={
            'handler_data': {
                'id': str(trs_id),
                'submitter_email': 'translator@example.org',
            }
        }
    )
    # prev_ticket_id = ticket.id
    session.add(ticket)
    transaction.commit()

    docx_path = module_path(
        "tests.onegov.translator_directory",
        "fixtures/Vorlage_hometown_ticket_number.docx",
    )
    signature_path = module_path(
        "tests.onegov.translator_directory",
        "fixtures/Unterschrift__DOJO__Adj_mV_John_Doe__Stv_Dienstchef.jpg",
    )
    client.login_admin()
    upload_file(docx_path, client, content_type='application/vnd.ms-office')
    upload_file(signature_path, client)
    user = UserCollection(session).by_username('admin@example.org')
    assert user is not None
    assert user.realname is not None
    first_name, last_name = user.realname.split(" ")

    # Now we have everything set up, go to the mail templates and generate one
    page = client.get(f'/translator/{trs_id}').click('Briefvorlagen')
    page.form['templates'] = basename(docx_path)
    resp = page.form.submit()

    found_variables_in_docx = set()
    expected_variables_in_docx = {
        'AKK-1000-0000',
        'Sehr geehrter Herr',
        'Luzern',  # hometown
        first_name,
        last_name
    }

    doc = docx.Document(BytesIO(resp.body))
    for target in expected_variables_in_docx:
        for block in iter_block_items(doc):
            line = block.text
            # make sure all variables have been rendered
            assert '{{' not in line and '}}' not in line, line
            if target in line:
                found_variables_in_docx.add(target)

    assert expected_variables_in_docx == found_variables_in_docx


def test_basic_search(client_with_fts: Client) -> None:
    client = client_with_fts
    client.login_admin()
    anom = client.spawn()

    assert 'Resultate' in client.get('/search?q=test')
    assert client.get('/search/suggest?q=test').json == []
    assert 'Resultate' in anom.get('/search?q=test')
    assert anom.get('/search/suggest?q=test').json == []


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_member_cannot_submit_mutation(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client
) -> None:

    session = client.app.session()
    languages = create_languages(session)
    certs = create_certificates(session)
    cert_ids = [str(cert.id) for cert in certs]
    language_ids = [str(lang.id) for lang in languages]
    transaction.commit()

    client.login_admin()

    settings = client.get('/directory-settings')
    settings.form['coordinates'] = encode_map_value(
        {'lat': 46, 'lon': 7, 'zoom': 12}
    )
    settings.form.submit()

    page = client.get('/translators/new')
    page.form['first_name'] = 'Uncle'
    page.form['last_name'] = 'Bob'
    page.form['pers_id'] = 978654
    page.form['admission'] = 'uncertified'
    page.form['gender'] = 'M'
    page.form['withholding_tax'] = False
    page.form['self_employed'] = False
    page.form['date_of_birth'] = '1970-01-01'
    page.form['nationalities'] = ['CH']
    page.form['coordinates'] = encode_map_value(
        {'lat': 46, 'lon': 7, 'zoom': 12}
    )
    page.form['address'] = 'Fakestreet 123'
    page.form['zip_code'] = '6000'
    page.form['city'] = 'Luzern'
    page.form['email'] = 'member@test.com'
    page.form['tel_private'] = '+41412223344'
    page.form['tel_mobile'] = '+41412223345'
    page.form['tel_office'] = '+41412223346'
    page.form['availability'] = 'Always'
    page.form['mother_tongues_ids'] = language_ids[0:1]
    page.form['spoken_languages_ids'] = language_ids[1:2]
    page.form['written_languages_ids'] = language_ids[2:3]
    page.form['monitoring_languages_ids'] = language_ids[3:4]
    page.form.get('expertise_professional_guilds', index=0).checked = True
    page.form['expertise_professional_guilds_other'] = ['Psychologie']
    page.form.get('expertise_interpreting_types', index=0).checked = True
    page.form['social_sec_number'] = '756.1234.5678.97'
    page.form['bank_name'] = 'Luzerner Bank'
    page.form['bank_address'] = 'Bankplatz Luzern'
    page.form['account_owner'] = 'Uncle Bob'
    page.form['iban'] = 'DE07 1234 1234 1234 1234 12'
    page.form['operation_comments'] = 'No operation comments'
    page.form['confirm_name_reveal'] = False
    page.form['date_of_application'] = '2020-01-01'
    page.form['profession'] = 'Handwerker'
    page.form['occupation'] = 'Bäcker'
    page.form['agency_references'] = 'All okay'
    page.form['education_as_interpreter'] = False
    page.form['certificates_ids'] = cert_ids[0:1]
    page.form['comments'] = 'No comments'
    with patch(
        'onegov.gis.utils.MapboxRequests.directions',
        return_value=FakeResponse(
            {'code': 'Ok', 'routes': [{'distance': 1000}]}
        ),
    ):
        assert 'hinzugefügt' in page.form.submit().follow()

    client.logout()
    reset_password_url = re.search(  # type: ignore[union-attr]
        r'(http://localhost/auth/reset-password[^)]+)',
        client.get_email(0, flush_queue=True)['TextBody'],
    ).group()
    page = client.get(reset_password_url)
    page.form['email'] = 'member@test.com'
    page.form['password'] = 'known_very_secure_password'
    page.form.submit()

    client.login_member()
    page = client.get('/').maybe_follow().click('BOB, Uncle')

    assert 'Report change' not in page
    assert 'Mutation melden' not in page

    session = client.app.session()
    translator = session.query(Translator).filter_by(last_name='Bob').first()
    assert translator is not None

    client.get(f'/translator/{translator.id.hex}/report-change', status=403)


def test_view_time_reports(client: Client) -> None:

    session = client.app.session()
    translators = TranslatorCollection(client.app)
    translator = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
    )

    report = TranslatorTimeReport(
        translator_id=translator.id,
        assignment_type='consecutive',
        finanzstelle='finanzstelle',
        duration=90,
        case_number='CASE-001',
        assignment_date=date(2025, 1, 15),
        hourly_rate=Decimal('90.0'),
        travel_compensation=Decimal('50.0'),
        total_compensation=Decimal('162.75'),
        status='pending',
    )
    session.add(report)
    session.flush()
    report_id = report.id
    transaction.commit()

    client.login_admin()
    page = client.get('/time-reports')
    assert '162.75' in page


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_time_report_workflow(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client
) -> None:
    """Test editor submitting time report."""
    session = client.app.session()
    languages = create_languages(session)
    translators = TranslatorCollection(client.app)
    translator_id = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
        drive_distance=35.0,
    ).id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(name='migrationsamt_und_passbuero')
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org']
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    assert 'Zeit erfassen' in page

    page = page.click('Zeit erfassen')
    page.form['assignment_type'] = 'on-site'
    page.form['assignment_location'] = 'obergericht'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-11'
    page.form['start_time'] = '09:00'
    page.form['end_date'] = '2025-01-11'
    page.form['end_time'] = '10:30'
    page.form['case_number'] = 'CASE-123'
    page.form['is_urgent'] = False
    page.form['notes'] = 'Test notes'
    page = page.form.submit().follow()

    assert 'Zeiterfassung zur Überprüfung eingereicht' in page
    assert '135' in page  # base compensation

    mail_to_submitter = client.get_email(0)
    assert 'TRANSLATOR, Test' in mail_to_submitter['Subject']

    accountant_email = get_accountant_email(client)
    accountant_emails = filter_emails_by_recipient(
        collect_emails(client), accountant_email
    )
    assert len(accountant_emails) >= 1

    assert accountant_emails[0]['To'] == accountant_email
    mail_to_accountant = accountant_emails[0]
    assert 'TRANSLATOR, Test' in mail_to_accountant['Subject']
    assert (
        'Eine neue Zeiterfassung wurde zur Überprüfung eingereicht.'
        in mail_to_accountant['TextBody']
    )
    link_match = re.search(
        r'<a href="([^"]+)">Zeiterfassung anzeigen</a>',
        mail_to_accountant['HtmlBody'],
    )
    assert link_match is not None
    # save ticket link for later
    ticket_link = link_match.group(1)

    translator_emails = filter_emails_by_recipient(
        collect_emails(client), 'translator@example.org'
    )
    assert len(translator_emails) >= 1
    mail_to_translator = translator_emails[0]
    assert mail_to_translator['To'] == 'translator@example.org'
    assert (
        'Eine Zeiterfassung wurde für Sie eingereicht'
        in mail_to_translator['Subject']
    )
    assert (
        'Eine Zeiterfassung wurde für Sie eingereicht und wird nun '
        'geprüft.' in mail_to_translator['TextBody']
    )
    assert (
        'Bitte sorgfältig überprüfen und Unstimmigkeiten umgehend melden'
        in mail_to_translator['TextBody']
    )

    translator_link_match = re.search(
        r'<a href="([^"]+)">Zeiterfassung anzeigen</a>',
        mail_to_translator['HtmlBody'],
    )
    assert translator_link_match is not None
    translator_status_link = translator_link_match.group(1)
    assert '/status' in translator_status_link
    assert accountant_email in mail_to_translator['TextBody']

    client.login_translator()
    status_page = client.get(translator_status_link)
    assert status_page.status_code == 200
    assert 'TRP-' in status_page

    translator = session.query(Translator).filter_by(id=translator_id).one()
    assert len(translator.time_reports) == 1
    report = translator.time_reports[0]
    assert report.duration == 90
    assert report.hourly_rate == 90.0
    assert report.surcharge_types == ['weekend_holiday']  # Saturday
    assert report.travel_compensation == Decimal('50.00')
    assert report.case_number == 'CASE-123'
    assert report.status == 'pending'

    client.login_editor()
    ticket_page = client.get(ticket_link)
    assert 'Test notes' in ticket_page
    # Accept ticket
    ticket_page = ticket_page.click('Ticket annehmen').follow()

    # Test edit functionality as editor
    # Should see edit link since status is pending and user is editor
    edit_links = [
        link
        for link in ticket_page.pyquery('a')
        if 'bearbeiten' in link.text_content().lower()
    ]
    assert len(edit_links) > 0, 'No edit link found'
    edit_url = edit_links[0].attrib['href']
    edit_page = client.get(edit_url)
    assert 'Zeiterfassung bearbeiten' in edit_page or 'Edit' in edit_page

    # Edit the time report
    assert edit_page.form['start_date'].value == '2025-01-11'
    assert edit_page.form['start_time'].value
    assert edit_page.form['end_time'].value
    edit_page.form['end_time'] = '12:00'
    edit_page.form['case_number'] = 'CASE-456-UPDATED'
    edit_page.form['is_urgent'] = True
    edit_page = edit_page.form.submit().follow()
    assert (
        'Zeiterfassung erfolgreich aktualisiert' in edit_page
    )

    # Verify updated values in database
    report = (
        session.query(Translator)
        .filter_by(id=translator_id)
        .one()
        .time_reports[0]
    )
    assert report.duration == 180  # 09:00 to 12:00 = 3 hours = 180 minutes
    assert report.case_number == 'CASE-456-UPDATED'
    assert report.surcharge_types is not None
    assert 'urgent' in report.surcharge_types
    breakdown = report.calculate_compensation_breakdown()
    assert breakdown['subtotal'] == approx(
        Decimal(421.87), abs=Decimal('0.01')
    )

    # Test that member can also edit (Personal permission)
    client.login_member()
    member_edit_page = client.get(edit_url)
    assert member_edit_page.status_code == 200

    # Accept time report as editor
    client.login_editor()
    ticket_page = client.get(ticket_link)
    accept_url = ticket_page.pyquery('a.accept-link')[0].attrib['ic-post-to']
    page = client.post(accept_url).follow()
    assert 'Zeiterfassung akzeptiert' in page

    # Verify ticket was closed
    report = (
        session.query(Translator)
        .filter_by(id=translator_id)
        .one()
        .time_reports[0]
    )
    ticket = report.get_ticket(session)
    assert ticket is not None
    assert ticket.state == 'closed'
    assert ticket.closed_on is not None

    # Test that edit is not available after confirmation
    report = (
        session.query(Translator)
        .filter_by(id=translator_id)
        .one()
        .time_reports[0]
    )
    assert report.status == 'confirmed'
    ticket_page = client.get(ticket_link)
    edit_links = ticket_page.pyquery('a.edit-link')
    assert len(edit_links) == 0

    emails = collect_emails(client)
    translator_emails = filter_emails_by_recipient(
        emails, 'translator@example.org'
    )
    acceptance_emails = [
        e
        for e in translator_emails
        if 'Zeiterfassung akzeptiert' in e['Subject']
    ]
    assert len(acceptance_emails) > 0
    mail_to_translator = acceptance_emails[0]

    assert 'TRANSLATOR, Test' in mail_to_translator['Subject']
    assert 'translator@example.org' in mail_to_translator['To']
    assert 'Zeiterfassung akzeptiert' in mail_to_translator['Subject']

    accepted_link_match = re.search(
        r'<a href="([^"]+)">Zeiterfassung anzeigen</a>',
        mail_to_translator['HtmlBody'],
    )
    assert accepted_link_match is not None
    accepted_status_link = accepted_link_match.group(1)
    assert '/status' in accepted_status_link

    attachments = mail_to_translator.get('Attachments', [])
    assert len(attachments) == 1
    attachment = attachments[0]
    assert attachment['Name'].startswith('Zeiterfassung_')
    assert attachment['Name'].endswith('.pdf')
    assert attachment['ContentType'] == 'application/pdf'
    assert len(attachment['Content']) > 0

    report = session.query(Translator).filter_by(
        id=translator_id).one().time_reports[0]
    assert report.status == 'confirmed'


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_time_report_skip_travel_calculation(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client,
) -> None:
    """Test skip_travel_calculation checkbox disables travel."""
    session = client.app.session()
    languages = create_languages(session)
    translators = TranslatorCollection(client.app)
    translator_id = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
        drive_distance=35.0,
    ).id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(name='migrationsamt_und_passbuero')
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org'],
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')

    page.form['assignment_type'] = 'on-site'
    page.form['assignment_location'] = 'obergericht'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-11'
    page.form['start_time'] = '09:00'
    page.form['end_date'] = '2025-01-11'
    page.form['end_time'] = '10:30'
    page.form['case_number'] = 'CASE-123'
    page.form['is_urgent'] = False
    page.form['skip_travel_calculation'] = True
    page = page.form.submit().follow()

    assert 'Zeiterfassung zur Überprüfung eingereicht' in page

    translator = session.query(Translator).filter_by(id=translator_id).one()
    assert len(translator.time_reports) == 1
    report = translator.time_reports[0]
    assert report.duration == 90
    assert report.travel_compensation == Decimal('0')
    assert report.travel_distance == 0.0

    breakdown = report.calculate_compensation_breakdown()
    assert breakdown['travel'] == Decimal('0')
    assert breakdown['total'] == breakdown['adjusted_subtotal']


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_time_report_edit_toggle_skip_travel(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client,
) -> None:
    """Test editing time report to toggle skip_travel_calculation."""
    session = client.app.session()
    languages = create_languages(session)
    translators = TranslatorCollection(client.app)
    translator_id = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
        drive_distance=35.0,
    ).id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(name='migrationsamt_und_passbuero')
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org'],
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')

    page.form['assignment_type'] = 'on-site'
    page.form['assignment_location'] = 'obergericht'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-11'
    page.form['start_time'] = '09:00'
    page.form['end_date'] = '2025-01-11'
    page.form['end_time'] = '10:30'
    page.form['case_number'] = 'CASE-123'
    page.form['is_urgent'] = False
    page = page.form.submit().follow()

    translator = session.query(Translator).filter_by(id=translator_id).one()
    report = translator.time_reports[0]
    assert report.travel_compensation == Decimal('50')
    assert report.travel_distance == 35.0

    accountant_email = get_accountant_email(client)
    accountant_emails = filter_emails_by_recipient(
        collect_emails(client), accountant_email
    )
    assert len(accountant_emails) >= 1

    mail_to_accountant = accountant_emails[0]
    link_match = re.search(
        r'<a href="([^"]+)">Zeiterfassung anzeigen</a>',
        mail_to_accountant['HtmlBody'],
    )
    assert link_match is not None
    ticket_link = link_match.group(1)

    client.login_editor()
    page = client.get(ticket_link)
    page = page.click('Ticket annehmen').follow()

    edit_links = [
        link
        for link in page.pyquery('a')
        if 'bearbeiten' in link.text_content().lower()
    ]
    assert len(edit_links) > 0
    edit_url = edit_links[0].attrib['href']

    client.login_member()
    edit_page = client.get(edit_url)
    assert edit_page.form['skip_travel_calculation'].checked is False
    edit_page.form['skip_travel_calculation'] = True
    edit_page = edit_page.form.submit().follow()

    session.expire_all()
    report = (
        session.query(Translator)
        .filter_by(id=translator_id)
        .one()
        .time_reports[0]
    )
    assert report.travel_compensation == Decimal('0')
    assert report.travel_distance == 0.0

    breakdown = report.calculate_compensation_breakdown()
    assert breakdown['travel'] == Decimal('0')


def extract_total_from_ticket_html(html: str) -> tuple[str, str]:
    total_match = re.search(
        r'<dt><strong>Total</strong>\s*\(([^)]+)\)</dt>',
        html,
        re.DOTALL
    )
    assert total_match is not None, 'Total not found in ticket HTML'
    calculation_formula = total_match.group(1).strip()

    total_amount_match = re.search(
        r'<dt><strong>Total</strong>.*?</dt>\s*<dd><strong>([^<]+)',
        html,
        re.DOTALL
    )
    assert (
        total_amount_match is not None
    ), 'Total amount not found in ticket HTML'
    total_amount = total_amount_match.group(1).strip()

    return calculation_formula, total_amount


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_time_report_skip_travel_ticket_html_unchanged_after_edit(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client,
) -> None:
    """Test that total cost of time report unchanged after empty
    edit and submit"""
    session = client.app.session()
    languages = create_languages(session)
    translators = TranslatorCollection(client.app)
    translator_id = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
        drive_distance=35.0,
    ).id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(
        name='migrationsamt_und_passbuero'
    )
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org'],
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')

    page.form['assignment_type'] = 'on-site'
    page.form['assignment_location'] = 'obergericht'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-11'
    page.form['start_time'] = '09:00'
    page.form['end_date'] = '2025-01-11'
    page.form['end_time'] = '10:30'
    page.form['case_number'] = 'CASE-123'
    page.form['is_urgent'] = False
    page.form['skip_travel_calculation'] = True
    page = page.form.submit().follow()

    assert 'Zeiterfassung zur Überprüfung eingereicht' in page

    accountant_email = get_accountant_email(client)
    ticket_link = extract_ticket_link_from_email(
        collect_emails(client), accountant_email
    )

    client.login_editor()
    ticket_page = client.get(ticket_link)
    ticket_page = ticket_page.click('Ticket annehmen').follow()

    initial_html = str(ticket_page)
    initial_formula, initial_amount = extract_total_from_ticket_html(
        initial_html
    )

    edit_page = ticket_page.click('Bearbeiten')
    page = edit_page.form.submit().follow()

    final_html = str(page)
    final_formula, final_amount = extract_total_from_ticket_html(final_html)

    assert initial_formula == final_formula, (
        f'Calculation formula changed: {initial_formula} -> {final_formula}'
    )
    assert initial_amount == final_amount, (
        f'Total amount changed: {initial_amount} -> {final_amount}'
    )
    assert final_amount == 'CHF 168.75'


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_time_report_workflow_self_employed(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client,
) -> None:
    """Test time report with self-employed translator generates QR bill."""
    session = client.app.session()
    languages = create_languages(session)
    translators = TranslatorCollection(client.app)
    translator_id = translators.add(
        first_name='Hans',
        last_name='Muster',
        admission='certified',
        email='self-employed@example.org',
        drive_distance=35.0,
        self_employed=True,
        iban='CH93 0076 2011 6238 5295 7',
        address='Musterstrasse 123',
        zip_code='8000',
        city='Zürich',
    ).id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(name='migrationsamt_und_passbuero')
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org']
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')
    page.form['assignment_type'] = 'on-site'
    page.form['assignment_location'] = 'obergericht'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-11'
    page.form['start_time'] = '09:00'
    page.form['end_date'] = '2025-01-11'
    page.form['end_time'] = '10:30'
    page.form['case_number'] = 'CASE-123'
    page.form['is_urgent'] = False
    page.form['notes'] = 'Test notes'
    page = page.form.submit().follow()

    assert 'Zeiterfassung zur Überprüfung eingereicht' in page

    accountant_email = get_accountant_email(client)
    ticket_link = extract_ticket_link_from_email(
        collect_emails(client), accountant_email
    )

    session = client.app.session()
    translator = session.query(Translator).filter_by(id=translator_id).one()
    assert translator.self_employed is True
    assert translator.iban == 'CH93 0076 2011 6238 5295 7'
    assert len(translator.time_reports) == 1

    client.login_editor()
    time_report = translator.time_reports[0]
    ticket = time_report.get_ticket(session)
    assert ticket is not None

    ticket_page = client.get(ticket_link)
    ticket_page = ticket_page.click('Ticket annehmen').follow()

    accept_url = ticket_page.pyquery('a.accept-link')[0].attrib['ic-post-to']
    page = client.post(accept_url).follow()
    assert 'Zeiterfassung akzeptiert' in page

    emails = collect_emails(client)
    translator_emails = filter_emails_by_recipient(
        emails, 'self-employed@example.org'
    )
    acceptance_emails = [
        e
        for e in translator_emails
        if 'Zeiterfassung akzeptiert' in e['Subject']
    ]
    assert len(acceptance_emails) > 0
    mail_to_translator = acceptance_emails[0]

    assert 'MUSTER, Hans' in mail_to_translator['Subject']
    assert 'self-employed@example.org' in mail_to_translator['To']
    assert 'Zeiterfassung akzeptiert' in mail_to_translator['Subject']

    attachments = mail_to_translator.get('Attachments', [])
    assert len(attachments) == 1

    time_report_pdf = next(
        (a for a in attachments if a['Name'].startswith('Zeiterfassung_')),
        None,
    )
    assert time_report_pdf is not None
    assert time_report_pdf['ContentType'] == 'application/pdf'
    assert len(time_report_pdf['Content']) > 0

    ticket_page = client.get(ticket_link)
    qr_bill_link = ticket_page.pyquery('a[href*="qr-bill-pdf"]')
    assert len(qr_bill_link) == 1

    qr_bill_response = client.get(qr_bill_link[0].attrib['href'])
    assert qr_bill_response.content_type == 'application/pdf'
    assert qr_bill_response.content_disposition is not None
    assert 'QR_Rechnung_Muster_' in qr_bill_response.content_disposition
    assert len(qr_bill_response.body) > 0


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_time_report_workflow_self_employed_missing_iban(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client,
) -> None:
    """Test acceptance blocked for self-employed without IBAN."""
    session = client.app.session()
    languages = create_languages(session)
    translators = TranslatorCollection(client.app)
    translator_id = translators.add(
        first_name='Hans',
        last_name='Muster',
        admission='certified',
        email='no-iban@example.org',
        drive_distance=35.0,
        self_employed=True,
        iban=None,
        address='Musterstrasse 123',
        zip_code='8000',
        city='Zürich',
    ).id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(name='migrationsamt_und_passbuero')
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org']
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')
    page.form['assignment_type'] = 'on-site'
    page.form['assignment_location'] = 'obergericht'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-11'
    page.form['start_time'] = '09:00'
    page.form['end_date'] = '2025-01-11'
    page.form['end_time'] = '10:30'
    page.form['case_number'] = 'CASE-123'
    page.form['is_urgent'] = False
    page = page.form.submit().follow()

    accountant_email = get_accountant_email(client)
    ticket_link = extract_ticket_link_from_email(
        collect_emails(client), accountant_email
    )

    session = client.app.session()
    translator = session.query(Translator).filter_by(id=translator_id).one()
    time_report = translator.time_reports[0]

    client.login_editor()
    ticket_page = client.get(ticket_link)
    ticket_page = ticket_page.click('Ticket annehmen').follow()
    accept_url = ticket_page.pyquery('a.accept-link')[0].attrib['ic-post-to']
    page = client.post(accept_url).maybe_follow()
    # It will refuse, as no IBAN was set on translator
    assert time_report.status == 'pending'


def test_user_groups(client: Client) -> None:
    """Test creating and editing user groups with finanzstelle."""

    transaction.begin()
    client.app.session().add(
        User(
            username='editor2@example.org',
            password_hash=hash_password('hunter2'),
            role='editor',
        )
    )
    transaction.commit()

    client.login_admin()
    page = client.get('/usergroups/new')

    page.form['name'] = 'Accountants'
    page.form['finanzstelle'] = 'polizei'
    page.form['users'].select_multiple(texts=['member@example.org'])
    page.form['accountant_emails'].select_multiple(
        texts=['editor@example.org']
    )

    page = page.form.submit().follow()
    assert 'Accountants' in page

    session = client.app.session()
    group = session.query(UserGroup).filter_by(name='Accountants').one()

    assert group.meta['finanzstelle'] == 'polizei'
    assert group.meta['accountant_emails'] == ['editor@example.org']

    page = client.get(f'/user-groups/{str(group.id)}/edit')
    assert 'Benutzergruppe bearbeiten' in page
    assert page.form['name'].value == 'Accountants'
    assert page.form['finanzstelle'].value == 'polizei'

    page.form['finanzstelle'] = 'staatsanwaltschaft'
    page.form['name'].value = 'Accountants2'
    page.form['accountant_emails'].select_multiple(
        texts=['editor2@example.org', 'editor@example.org']
    )
    page = page.form.submit().follow()

    session = client.app.session()
    group = session.query(UserGroup).filter_by(name='Accountants2').one()
    assert group.meta['finanzstelle'] == 'staatsanwaltschaft'
    assert set(group.meta['accountant_emails']) == {
        'editor2@example.org',
        'editor@example.org',
    }
    # empty re-submit
    # This used to raise 'Uniqueness violation in ticket permissions'
    page = client.get(f'/user-groups/{str(group.id)}/edit')
    page = page.form.submit().follow()

    # Create second user group with different finanzstelle
    page = client.get('/usergroups/new')
    page.form['name'] = 'Staatsanwaltschaft Group'
    page.form['finanzstelle'] = 'staatsanwaltschaft'
    page.form['users'].select_multiple(texts=['editor@example.org'])
    page.form['accountant_emails'] = ['editor@example.org']
    page = page.form.submit().follow()
    assert 'Staatsanwaltschaft Group' in page

    session = client.app.session()
    group2 = (
        session.query(UserGroup)
        .filter_by(name='Staatsanwaltschaft Group')
        .one()
    )
    assert group2.meta['finanzstelle'] == 'staatsanwaltschaft'
    assert group2.meta['accountant_emails'] == ['editor@example.org']

    # Verify both groups exist with correct permissions
    all_groups = (
        session.query(UserGroup)
        .filter(
            UserGroup.name.in_(['Accountants2', 'Staatsanwaltschaft Group'])
        )
        .all()
    )
    assert len(all_groups) == 2


def test_time_report_telephonic_no_location(client: Client) -> None:
    """Test that telephonic time reports have no assignment_location."""
    session = client.app.session()
    create_languages(session)
    translators = TranslatorCollection(client.app)
    translator_id = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
        drive_distance=35.0,
    ).id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(name='test_group')
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org'],
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')

    # Side quest: verify location field is required
    # (Pretend we forgot to set  set location (real use case))
    page.form['assignment_type'] = 'on-site'
    page.form['finanzstelle'] = 'polizei'
    page.form['start_date'] = '2025-01-15'
    page.form['start_time'] = '10:00'
    page.form['end_date'] = '2025-01-15'
    page.form['end_time'] = '11:30'
    page.form['case_number'] = 'PHONE-123'
    page.form['notes'] = 'Telephonic interpretation'
    page = page.form.submit()
    assert 'Bitte wählen Sie einen Standort' in page

    # proceed testing
    page.form['assignment_type'] = 'telephonic'
    page.form['assignment_location'] = 'migrationsamt'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-15'
    page.form['start_time'] = '10:00'
    page.form['end_date'] = '2025-01-15'
    page.form['end_time'] = '11:30'
    page.form['case_number'] = 'PHONE-123'
    page.form['notes'] = 'Telephonic interpretation'
    page = page.form.submit().follow()

    assert 'Zeiterfassung zur Überprüfung eingereicht' in page

    time_report = (
        session.query(TranslatorTimeReport)
        .filter_by(translator_id=translator_id)
        .first()
    )
    assert time_report is not None
    assert time_report.assignment_type == 'telephonic'
    assert time_report.assignment_location is None

    accountant_email = get_accountant_email(client)
    ticket_link = extract_ticket_link_from_email(
        collect_emails(client), accountant_email
    )

    client.login_editor()
    ticket_page = client.get(ticket_link)
    ticket_page = ticket_page.click('Ticket annehmen').follow()
    # Einsatzort should be hidden as there it has not been set
    assert 'Einsatzort' not in ticket_page


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_delete_time_report(
    broadcast: 'MagicMock',
    authenticate: 'MagicMock',
    connect: 'MagicMock',
    client: 'Client',
) -> None:
    """Test that admins and accountants can delete time reports."""
    session = client.app.session()
    create_languages(session)
    translators = TranslatorCollection(client.app)
    translator = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
        drive_distance=35.0,
    )
    translator_id = translator.id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(name='migrationsamt_und_passbuero')
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org'],
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')
    page.form['assignment_type'] = 'telephonic'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-15'
    page.form['start_time'] = '10:00'
    page.form['end_date'] = '2025-01-15'
    page.form['end_time'] = '11:30'
    page.form['case_number'] = 'DELETE-TEST'
    page = page.form.submit().follow()

    time_report = (
        session.query(TranslatorTimeReport)
        .filter_by(translator_id=translator_id)
        .first()
    )
    assert time_report is not None
    report_id = time_report.id

    client.login_member()
    page = client.get('/time-reports')
    assert 'Delete</a>' not in str(page)

    client.login_editor()
    page = client.get('/time-reports')
    csrf_token_match = re.search(
        r'\?csrf-token=([a-zA-Z0-9\._\-]+)', str(page)
    )
    assert csrf_token_match is not None
    csrf_token = csrf_token_match.group(1)

    response = client.delete(
        f'/time-report/{report_id}?csrf-token={csrf_token}',
    )
    assert response.status_code == 200

    session.expire_all()
    time_report = session.query(TranslatorTimeReport).get(report_id)
    assert time_report is None


@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_delete_time_report_admin(
    broadcast: 'MagicMock',
    authenticate: 'MagicMock',
    connect: 'MagicMock',
    client: 'Client',
) -> None:
    """Test that admin can delete time reports."""
    session = client.app.session()
    create_languages(session)
    translators = TranslatorCollection(client.app)
    translator = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
        drive_distance=35.0,
    )
    translator_id = translator.id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(name='polizei')
    user_group.meta = {
        'finanzstelle': 'polizei',
        'accountant_emails': ['other@example.org'],
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')
    page.form['assignment_type'] = 'telephonic'
    page.form['finanzstelle'] = 'polizei'
    page.form['start_date'] = '2025-01-15'
    page.form['start_time'] = '10:00'
    page.form['end_date'] = '2025-01-15'
    page.form['end_time'] = '11:30'
    page.form['case_number'] = 'ADMIN-DELETE'
    page = page.form.submit().follow()

    time_report = (
        session.query(TranslatorTimeReport)
        .filter_by(translator_id=translator_id)
        .first()
    )
    assert time_report is not None
    report_id = time_report.id

    client.login_admin()
    page = client.get('/time-reports')
    csrf_token_match = re.search(
        r'\?csrf-token=([a-zA-Z0-9\._\-]+)', str(page)
    )
    assert csrf_token_match is not None
    csrf_token = csrf_token_match.group(1)

    response = client.delete(
        f'/time-report/{report_id}?csrf-token={csrf_token}',
    )
    assert response.status_code == 200

    session.expire_all()
    time_report = session.query(TranslatorTimeReport).get(report_id)
    assert time_report is None


def test_export_time_reports(
    client: 'Client',
) -> None:
    """Test exporting confirmed time reports as CSV."""
    session = client.app.session()
    create_languages(session)
    translators = TranslatorCollection(client.app)
    translator = translators.add(
        first_name='Test',
        last_name='Translator',
        admission='certified',
        email='translator@example.org',
        drive_distance=35.0,
    )
    translator.pers_id = 12345
    translator_id = translator.id

    user_group_collection = UserGroupCollection(session)
    user_group = user_group_collection.add(
        name='migrationsamt_und_passbuero'
    )
    user_group.meta = {
        'finanzstelle': 'migrationsamt_und_passbuero',
        'accountant_emails': ['editor@example.org'],
    }
    transaction.commit()

    client.login_member()
    page = client.get(f'/translator/{translator_id}')
    page = page.click('Zeit erfassen')
    page.form['assignment_type'] = 'telephonic'
    page.form['finanzstelle'] = 'migrationsamt_und_passbuero'
    page.form['start_date'] = '2025-01-15'
    page.form['start_time'] = '10:00'
    page.form['end_date'] = '2025-01-15'
    page.form['end_time'] = '11:30'
    page.form['case_number'] = 'EXPORT-TEST'
    page = page.form.submit().follow()

    report = (
        session.query(TranslatorTimeReport)
        .filter_by(translator_id=translator_id)
        .first()
    )
    assert report is not None
    report_id = report.id
    assert report.status == 'pending'
    assert report.exported is False

    # Claim and accept the time report to confirm it
    client.login_editor()
    ticket = report.get_ticket(session)
    assert ticket is not None
    ticket_url = (
        f'/ticket/{ticket.handler_code}/{ticket.id.hex}'
    )
    ticket_page = client.get(ticket_url)
    ticket_page = ticket_page.click('Ticket annehmen').follow()
    accept_url = ticket_page.pyquery(
        'a.accept-link'
    )[0].attrib['ic-post-to']
    client.post(accept_url)

    session.expire_all()
    report = session.query(TranslatorTimeReport).get(report_id)
    assert report is not None
    assert report.status == 'confirmed'
    assert report.exported is False

    # Verify export button is shown
    page = client.get('/time-reports')
    assert 'TRANSLATOR, Test' in page
    assert 'export-accounting' in page

    # Extract csrf-protected export URL from ic-post-to
    page_text = str(page)
    export_match = re.search(
        r'ic-post-to="([^"]*export-accounting[^"]*)"',
        page_text,
    )
    assert export_match is not None, (
        'No export intercooler link found'
    )
    export_url = export_match.group(1)

    response = client.post(export_url)
    assert response.headers['Content-Type'] == 'text/csv; charset=iso-8859-1'
    assert response.content_disposition is not None
    assert 'translator_export_' in response.content_disposition

    csv_content = response.body.decode('iso-8859-1')
    assert '12345' in csv_content

    # Verify report is marked as exported with timestamp and batch id
    session.expire_all()
    report = session.query(TranslatorTimeReport).get(report_id)
    assert report is not None
    assert report.exported is True
    assert report.exported_at is not None
    assert report.export_batch_id is not None
    batch_id = report.export_batch_id

    # Default view no longer shows exported reports
    page = client.get('/time-reports')
    assert 'Keine Zeiterfassungen gefunden' in page

    # Archive view shows exported reports
    page = client.get('/time-reports?archive=true')

    table = page.pyquery('.time-reports-list')[0].text_content()
    assert 'Exportiert' in table
    assert 'TRANSLATOR, Test' in table

    # Batch id is preserved
    report = session.query(TranslatorTimeReport).get(report_id)
    assert report is not None
    assert report.export_batch_id == batch_id
