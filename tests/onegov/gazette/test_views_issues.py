from datetime import datetime
from freezegun import freeze_time
from io import BytesIO
from onegov.pdf.utils import extract_pdf_info
from openpyxl import load_workbook
from pyquery import PyQuery as pq
from tests.onegov.gazette.common import login_editor_1
from tests.onegov.gazette.common import login_publisher
from webtest import TestApp as Client


def test_view_issues(gazette_app):
    with freeze_time("2017-11-01 12:00"):
        client = Client(gazette_app)
        login_publisher(client)

        # Test data:
        # 2017:
        #     40: 2017-10-06 / 2017-10-04T12:00:00
        #     41: 2017-10-13 / 2017-10-11T12:00:00
        #     42: 2017-10-20 / 2017-10-18T12:00:00
        #     43: 2017-10-27 / 2017-10-25T12:00:00
        # >>>>>
        #     44: 2017-11-03 / 2017-11-01T12:00:00
        #     45: 2017-11-10 / 2017-11-08T12:00:00
        #     46: 2017-11-17 / 2017-11-15T12:00:00
        #     47: 2017-11-24 / 2017-11-22T12:00:00
        #     48: 2017-12-01 / 2017-11-29T12:00:00
        #     49: 2017-12-08 / 2017-12-06T12:00:00
        #     50: 2017-12-15 / 2017-12-13T12:00:00
        #     51: 2017-12-22 / 2017-12-20T12:00:00
        #     52: 2017-12-29 / 2017-12-27T12:00:00
        # 2018:
        #     1: 2018-01-05 / 2018-01-03T12:00:00

        # add a issue
        manage = client.get('/issues')
        manage = manage.click('Neu')
        manage.form['number'] = '1'
        manage.form['date_'] = '2019-01-02'
        manage.form['deadline'] = '2019-01-01T12:00'
        manage = manage.form.submit().maybe_follow()
        assert 'Ausgabe hinzugefügt.' in manage
        assert '2019-1' in manage
        upcoming_issues = [
            [td.text.strip() for td in pq(tr)('td')]
            for tr in manage.pyquery('#panel_upcoming table.issues tbody tr')
        ]
        past_issues = [
            [td.text.strip() for td in pq(tr)('td')]
            for tr in manage.pyquery('#panel_past table.issues tbody tr')
        ]
        assert upcoming_issues == [
            ['2017-44', '03.11.2017', 'Mittwoch 01.11.2017 13:00', '', ''],
            ['2017-45', '10.11.2017', 'Mittwoch 08.11.2017 13:00', '', ''],
            ['2017-46', '17.11.2017', 'Mittwoch 15.11.2017 13:00', '', ''],
            ['2017-47', '24.11.2017', 'Mittwoch 22.11.2017 13:00', '', ''],
            ['2017-48', '01.12.2017', 'Mittwoch 29.11.2017 13:00', '', ''],
            ['2017-49', '08.12.2017', 'Mittwoch 06.12.2017 13:00', '', ''],
            ['2017-50', '15.12.2017', 'Mittwoch 13.12.2017 13:00', '', ''],
            ['2017-51', '22.12.2017', 'Mittwoch 20.12.2017 13:00', '', ''],
            ['2017-52', '29.12.2017', 'Mittwoch 27.12.2017 13:00', '', ''],
            ['2018-1', '05.01.2018', 'Mittwoch 03.01.2018 13:00', '', ''],
            ['2019-1', '02.01.2019', 'Dienstag 01.01.2019 12:00', '', '']
        ]
        assert past_issues == [
            ['2017-43', '27.10.2017', 'Mittwoch 25.10.2017 14:00', '', ''],
            ['2017-42', '20.10.2017', 'Mittwoch 18.10.2017 14:00', '', ''],
            ['2017-41', '13.10.2017', 'Mittwoch 11.10.2017 14:00', '', ''],
            ['2017-40', '06.10.2017', 'Mittwoch 04.10.2017 14:00', '', ''],
        ]

        # use the first available issue in a notice
        manage = client.get('/notices/drafted/new-notice')
        manage.form['title'] = 'Titel'
        manage.form['organization'] = '200'
        manage.form['category'] = '13'
        manage.form['issues'] = ['2017-44']
        manage.form['text'] = 'Text'
        manage.form['author_place'] = 'Govikon'
        manage.form['author_name'] = 'State Chancellerist'
        manage.form['author_date'] = '2019-01-01'
        manage = manage.form.submit().maybe_follow()
        assert '<h2>Titel</h2>' in manage
        assert 'Nr. 44, 03.11.2017' in manage

        # edit the issue
        manage = client.get('/issues')
        manage = manage.click('Bearbeiten', index=0)
        manage.form['date_'] = '2017-11-02'
        manage.form['deadline'] = '2017-11-01T12:00'
        manage = manage.form.submit().maybe_follow()
        assert 'Ausgabe geändert.' in manage

        upcoming_issues = [
            [td.text.strip() for td in pq(tr)('td')]
            for tr in manage.pyquery('#panel_upcoming table.issues tbody tr')
        ]
        past_issues = [
            [td.text.strip() for td in pq(tr)('td')]
            for tr in manage.pyquery('#panel_past table.issues tbody tr')
        ]
        assert upcoming_issues == [
            ['2017-44', '02.11.2017', 'Mittwoch 01.11.2017 12:00', '', ''],
            ['2017-45', '10.11.2017', 'Mittwoch 08.11.2017 13:00', '', ''],
            ['2017-46', '17.11.2017', 'Mittwoch 15.11.2017 13:00', '', ''],
            ['2017-47', '24.11.2017', 'Mittwoch 22.11.2017 13:00', '', ''],
            ['2017-48', '01.12.2017', 'Mittwoch 29.11.2017 13:00', '', ''],
            ['2017-49', '08.12.2017', 'Mittwoch 06.12.2017 13:00', '', ''],
            ['2017-50', '15.12.2017', 'Mittwoch 13.12.2017 13:00', '', ''],
            ['2017-51', '22.12.2017', 'Mittwoch 20.12.2017 13:00', '', ''],
            ['2017-52', '29.12.2017', 'Mittwoch 27.12.2017 13:00', '', ''],
            ['2018-1', '05.01.2018', 'Mittwoch 03.01.2018 13:00', '', ''],
            ['2019-1', '02.01.2019', 'Dienstag 01.01.2019 12:00', '', '']
        ]
        assert past_issues == [
            ['2017-43', '27.10.2017', 'Mittwoch 25.10.2017 14:00', '', ''],
            ['2017-42', '20.10.2017', 'Mittwoch 18.10.2017 14:00', '', ''],
            ['2017-41', '13.10.2017', 'Mittwoch 11.10.2017 14:00', '', ''],
            ['2017-40', '06.10.2017', 'Mittwoch 04.10.2017 14:00', '', ''],
        ]

        # check if the notice has been updated
        manage = client.get('/notice/titel')
        assert 'Nr. 44, 03.11.2017' not in manage
        assert 'Nr. 44, 02.11.2017' in manage

        # delete all but one (unused) issues
        manage = client.get('/issues')
        manage.click('Löschen', index=1).form.submit()
        manage.click('Löschen', index=2).form.submit()
        manage.click('Löschen', index=3).form.submit()
        manage.click('Löschen', index=4).form.submit()
        manage.click('Löschen', index=5).form.submit()
        manage.click('Löschen', index=6).form.submit()
        manage.click('Löschen', index=7).form.submit()
        manage.click('Löschen', index=8).form.submit()
        manage.click('Löschen', index=9).form.submit()
        manage.click('Löschen', index=10).form.submit()
        manage.click('Löschen', index=11).form.submit()
        manage.click('Löschen', index=12).form.submit()
        manage.click('Löschen', index=13).form.submit()
        manage.click('Löschen', index=14).form.submit()

        manage = client.get('/issues')
        issues = [
            [td.text.strip() for td in pq(tr)('td')]
            for tr in manage.pyquery('table.issues tbody tr')
        ]
        assert issues == [
            ['2017-44', '02.11.2017', 'Mittwoch 01.11.2017 12:00', '', '']
        ]

        # Try to delete the used issue
        manage = client.get('/issues')
        manage = manage.click('Löschen')
        assert 'Es können nur unbenutzte Ausgaben gelöscht werden.' in manage
        assert not manage.forms


def test_view_issues_permissions(gazette_app):
    with freeze_time("2017-10-20 12:00"):
        client = Client(gazette_app)

        login_publisher(client)
        manage = client.get('/issues').click('Neu')
        manage.form['number'] = '1'
        manage.form['date_'] = '2017-01-02'
        manage.form['deadline'] = '2017-01-01T12:00'
        manage = manage.form.submit().maybe_follow()
        edit_link = manage.click('Bearbeiten', index=0).request.url
        delete_link = manage.click('Löschen', index=0).request.url

        login_editor_1(client)
        client.get('/issues', status=403)
        client.get(edit_link, status=403)
        client.get(delete_link, status=403)


def test_view_issues_publish(gazette_app):
    with freeze_time("2017-11-01 12:00"):
        client = Client(gazette_app)
        login_publisher(client)

        for number, issues, print_only, accept in (
            (0, (44, 45), False, True),
            (1, (45, 46), True, True),
            (2, (45,), False, False),
        ):
            slug = 'notice-{}'.format(number)
            manage = client.get('/notices/drafted/new-notice')
            manage.form['title'] = slug
            manage.form['organization'] = '200'
            manage.form['category'] = '13'
            manage.form['issues'] = ['2017-{}'.format(i) for i in issues]
            manage.form['text'] = 'Text'
            manage.form['author_place'] = 'Govikon'
            manage.form['author_name'] = 'State Chancellerist'
            manage.form['author_date'] = '2019-01-01'
            manage.form['print_only'] = print_only
            manage = manage.form.submit()
            client.get('/notice/{}/submit'.format(slug)).form.submit()
            if accept:
                client.get('/notice/{}/accept'.format(slug)).form.submit()

        # publish 44
        manage = client.get('/issues').click('Veröffentlichen', index=0)
        manage = manage.form.submit().maybe_follow()
        assert "Ausgabe veröffentlicht." in manage
        assert "2017-44.pdf" in manage
        assert "2017-44.pdf (Stopp Internet)" in manage

        manage = manage.click(href='2017-44.pdf')
        assert manage.content_type == 'application/pdf'
        _, pdf = extract_pdf_info(BytesIO(manage.body))
        assert 'Amtsblatt Nr. 44, 03.11.2017' in pdf
        assert 'notice-0' in pdf

        manage = client.get('/issues')
        manage = manage.click('2017-44.pdf', href='print-only-pdf')
        assert manage.content_type == 'application/pdf'
        _, pdf = extract_pdf_info(BytesIO(manage.body))
        assert 'Amtsblatt Nr. 44, 03.11.2017' in pdf
        assert 'notice-0' not in pdf

        notice_0 = client.get('/notice/notice-0')
        notice_1 = client.get('/notice/notice-1')
        notice_2 = client.get('/notice/notice-2')
        assert '<li>Nr. 44, 03.11.2017 / 1</li>' in notice_0
        assert '<li>Nr. 45, 10.11.2017</li>' in notice_0
        assert '<li>Nr. 45, 10.11.2017</li>' in notice_1
        assert '<li>Nr. 46, 17.11.2017</li>' in notice_1
        assert '<li>Nr. 45, 10.11.2017</li>' in notice_2

        # publish 46
        manage = client.get('/issues').click('Veröffentlichen', index=2)
        manage = manage.form.submit().maybe_follow()
        assert "Ausgabe veröffentlicht." in manage
        assert "2017-46.pdf" in manage
        assert "2017-46.pdf (Stopp Internet)" in manage

        manage = manage.click(href='2017-46.pdf')
        assert manage.content_type == 'application/pdf'
        _, pdf = extract_pdf_info(BytesIO(manage.body))
        assert 'Amtsblatt Nr. 46, 17.11.2017' in pdf
        assert 'schützenswerten' in pdf
        assert 'Papierversion' in pdf

        manage = client.get('/issues')
        manage = manage.click('2017-46.pdf', href='print-only-pdf')
        assert manage.content_type == 'application/pdf'
        _, pdf = extract_pdf_info(BytesIO(manage.body))
        assert 'Amtsblatt Nr. 46, 17.11.2017' in pdf
        assert 'schützenswerten' in pdf
        assert 'Papierversion' not in pdf

        notice_0 = client.get('/notice/notice-0')
        notice_1 = client.get('/notice/notice-1')
        notice_2 = client.get('/notice/notice-2')
        assert '<li>Nr. 44, 03.11.2017 / 1</li>' in notice_0
        assert '<li>Nr. 45, 10.11.2017</li>' in notice_0
        assert '<li>Nr. 45, 10.11.2017</li>' in notice_1
        assert '<li>Nr. 46, 17.11.2017 / 2</li>' in notice_1
        assert '<li>Nr. 45, 10.11.2017</li>' in notice_2

        # publish 45
        manage = client.get('/issues').click('Veröffentlichen', index=1)
        assert "Diese Ausgabe hat eingereichte Meldungen!" in manage
        manage = manage.form.submit().maybe_follow()
        assert "Ausgabe veröffentlicht." in manage
        assert "2017-45.pdf" in manage
        assert "2017-45.pdf (Stopp Internet)" in manage

        manage = manage.click(href='2017-45.pdf')
        assert manage.content_type == 'application/pdf'
        assert manage.content_type == 'application/pdf'
        _, pdf = extract_pdf_info(BytesIO(manage.body))
        assert 'Amtsblatt Nr. 45, 10.11.2017' in pdf
        assert 'schützenswerten' in pdf
        assert 'Papierversion' in pdf

        manage = client.get('/issues')
        manage = manage.click('2017-45.pdf', href='print-only-pdf')
        assert manage.content_type == 'application/pdf'
        _, pdf = extract_pdf_info(BytesIO(manage.body))
        assert 'Amtsblatt Nr. 45, 10.11.2017' in pdf
        assert 'schützenswerten' in pdf
        assert 'notice-1' in pdf

        notice_0 = client.get('/notice/notice-0')
        notice_1 = client.get('/notice/notice-1')
        notice_2 = client.get('/notice/notice-2')
        assert '<li>Nr. 44, 03.11.2017 / 1</li>' in notice_0
        assert '<li>Nr. 45, 10.11.2017 / 2</li>' in notice_0
        assert '<li>Nr. 45, 10.11.2017 / 3</li>' in notice_1
        assert '<li>Nr. 46, 17.11.2017 / 2</li>' in notice_1
        assert '<li>Nr. 45, 10.11.2017</li>' in notice_2  # submitted

        # re-publish 46
        manage = client.get('/issues').click('Veröffentlichen', index=2)
        assert "Diese Ausgabe hat bereits ein PDF!" in manage
        assert "Diese Ausgabe hat bereits Publikationsnummern!" in manage
        manage = manage.form.submit().maybe_follow()
        assert "Ausgabe veröffentlicht." in manage
        assert "2017-46.pdf" in manage
        assert "2017-46.pdf (Stopp Internet)" in manage
        assert "Publikationsnummern haben geändert" in manage

        manage = manage.click(href='2017-46.pdf')
        assert manage.content_type == 'application/pdf'
        _, pdf = extract_pdf_info(BytesIO(manage.body))
        assert 'Amtsblatt Nr. 46, 17.11.2017' in pdf
        assert 'schützenswerten' in pdf

        manage = client.get('/issues')
        manage = manage.click('2017-46.pdf', href='print-only-pdf')
        assert manage.content_type == 'application/pdf'
        _, pdf = extract_pdf_info(BytesIO(manage.body))
        assert 'Amtsblatt Nr. 46, 17.11.2017' in pdf

        notice_0 = client.get('/notice/notice-0')
        notice_1 = client.get('/notice/notice-1')
        notice_2 = client.get('/notice/notice-2')
        assert '<li>Nr. 44, 03.11.2017 / 1</li>' in notice_0
        assert '<li>Nr. 45, 10.11.2017 / 2</li>' in notice_0
        assert '<li>Nr. 45, 10.11.2017 / 3</li>' in notice_1
        assert '<li>Nr. 46, 17.11.2017 / 4</li>' in notice_1
        assert '<li>Nr. 45, 10.11.2017</li>' in notice_2  # submitted


def test_view_issues_publish_disabled(gazette_app):
    client = Client(gazette_app)
    login_publisher(client)

    manage = client.get('/issues')
    assert "PDF" in manage
    assert "Veröffentlichen" in manage

    manage = client.get('/issue/2017-40/publish')
    assert "Veröffentlichung ist deaktiviert." not in manage
    assert 'form' in manage

    principal = gazette_app.principal
    principal.publishing = False
    gazette_app.cache.set('principal', principal)

    manage = client.get('/issues')
    assert "PDF" not in manage
    assert "Veröffentlichen" not in manage

    manage = client.get('/issue/2017-40/publish')
    assert "Veröffentlichung ist deaktiviert." in manage
    assert 'form' not in manage


def test_view_issues_export(gazette_app):
    client = Client(gazette_app)

    client.get('/issues/export', status=403)

    login_editor_1(client)
    client.get('/issues/export', status=403)

    login_publisher(client)
    response = client.get('/issues/export')

    book = load_workbook(BytesIO(response.body))
    assert len(book.worksheets) == 1

    sheet = book.worksheets[0]
    assert sheet.max_column == 4
    assert sheet.max_row == 15

    assert sheet.cell(1, 1).value == 'Jahr'
    assert sheet.cell(1, 2).value == 'Nummer'
    assert sheet.cell(1, 3).value == 'Datum'
    assert sheet.cell(1, 4).value == 'Eingabeschluss'

    assert sheet.cell(2, 1).value == 2017
    assert sheet.cell(2, 2).value == 40
    assert sheet.cell(2, 3).value == datetime(2017, 10, 6)
    assert sheet.cell(2, 4).value == datetime(2017, 10, 4, 14)

    assert sheet.cell(6, 1).value == 2017
    assert sheet.cell(6, 2).value == 44
    assert sheet.cell(6, 3).value == datetime(2017, 11, 3)
    assert sheet.cell(6, 4).value == datetime(2017, 11, 1, 13)

    assert sheet.cell(14, 1).value == 2017
    assert sheet.cell(14, 2).value == 52
    assert sheet.cell(14, 3).value == datetime(2017, 12, 29)
    assert sheet.cell(14, 4).value == datetime(2017, 12, 27, 13)

    assert sheet.cell(15, 1).value == 2018
    assert sheet.cell(15, 2).value == 1
    assert sheet.cell(15, 3).value == datetime(2018, 1, 5)
    assert sheet.cell(15, 4).value == datetime(2018, 1, 3, 13)
