from freezegun import freeze_time
from io import BytesIO
from onegov.gazette.models import Organization
from openpyxl import load_workbook
from pyquery import PyQuery as pq
from tests.onegov.gazette.common import login_editor_1
from tests.onegov.gazette.common import login_publisher
from webtest import TestApp as Client


def test_view_organizations(gazette_app):
    with freeze_time("2017-10-20 12:00"):
        client = Client(gazette_app)
        login_publisher(client)

        # Test data:
        # 100 / State Chancellery / active
        # 200 / Civic Community / active
        # 300 / Municipality / active
        # 400 / Evangelical Reformed Parish / active
        # 510 / Sikh Community / inactive
        # 500 / Catholic Parish / active
        # 600 / Corporation / active

        # add a organization
        manage = client.get('/organizations')
        manage = manage.click('Neu')
        manage.form['title'] = 'Organisation XY'
        manage.form['active'] = True
        manage = manage.form.submit().maybe_follow()
        assert 'Organisation hinzugefügt.' in manage
        assert 'Organisation XY' in manage
        organizations = [
            [
                ''.join((td.text_content(), td.attrib['class']))
                for td in pq(tr)('td')[:1]
            ][0]
            for tr in manage.pyquery('table.organizations tbody tr')
        ]
        assert organizations == [
            'State Chancellery (100)',
            'Civic Community (200)',
            'Municipality (300)',
            'Churches (400)',
            'Evangelical Reformed Parish (410)child ',
            'Sikh Community (420)child inactive',
            'Catholic Parish (430)child ',
            'Corporation (500)',
            'Organisation XY (501)'
        ]

        # use the first organization in a notice
        manage = client.get('/notices/drafted/new-notice')
        manage.form['title'] = 'Titel'
        manage.form['organization'] = '100'
        manage.form['category'] = '13'
        manage.form['issues'] = ['2017-44']
        manage.form['text'] = 'Text'
        manage.form['author_place'] = 'Govikon'
        manage.form['author_name'] = 'State Chancellerist'
        manage.form['author_date'] = '2019-01-01'
        manage = manage.form.submit().maybe_follow()
        assert '<h2>Titel</h2>' in manage
        assert 'State Chancellery' in manage

        # edit the first organization
        manage = client.get('/organizations')
        manage = manage.click('Bearbeiten', index=0)
        manage.form['title'] = 'Organisation Z'
        manage.form['active'] = False
        manage = manage.form.submit().maybe_follow()
        assert 'Organisation geändert.' in manage
        assert 'State Chancellery' not in manage
        assert 'Organisation Z' in manage

        organizations = [
            [
                ''.join((td.text_content(), td.attrib['class']))
                for td in pq(tr)('td')[:1]
            ][0]
            for tr in manage.pyquery('table.organizations tbody tr')
        ]
        assert organizations == [
            'Organisation Z (100)inactive',
            'Civic Community (200)',
            'Municipality (300)',
            'Churches (400)',
            'Evangelical Reformed Parish (410)child ',
            'Sikh Community (420)child inactive',
            'Catholic Parish (430)child ',
            'Corporation (500)',
            'Organisation XY (501)'
        ]

        # check if the notice has been updated
        manage = client.get('/notice/titel')
        assert 'State Chancellery' not in manage
        assert 'Organisation Z' in manage

        # try to delete the organization with suborganizations
        manage = client.get('/organizations')
        manage = manage.click('Löschen', index=3)
        assert (
            'Nur unbenutzte Organisationen ohne Unterorganisationen können '
            'gelöscht werden.'
        ) in manage
        assert not manage.forms

        # try to delete the used organization
        manage = client.get('/organizations')
        manage = manage.click('Löschen', index=3)
        assert (
            'Nur unbenutzte Organisationen ohne Unterorganisationen können '
            'gelöscht werden.'
        ) in manage
        assert not manage.forms

        # delete all but one (unused) organizations
        manage = client.get('/organizations')
        manage.click('Löschen', index=8).form.submit()
        manage.click('Löschen', index=7).form.submit()
        manage.click('Löschen', index=6).form.submit()
        manage.click('Löschen', index=5).form.submit()
        manage.click('Löschen', index=4).form.submit()
        manage.click('Löschen', index=3).form.submit()
        manage.click('Löschen', index=2).form.submit()
        manage.click('Löschen', index=1).form.submit()

        manage = client.get('/organizations')
        assert 'Organisation Z' in manage
        assert 'Civic Community' not in manage
        assert 'Municipality' not in manage
        assert 'Evangelical Reformed Parish' not in manage
        assert 'Sikh Community' not in manage
        assert 'Catholic Parish' not in manage
        assert 'Corporation' not in manage
        assert 'Organisation XY' not in manage


def test_view_organizations_permissions(gazette_app):
    client = Client(gazette_app)

    login_publisher(client)
    manage = client.get('/organizations').click('Neu')
    manage.form['title'] = 'XY'
    manage = manage.form.submit().maybe_follow()
    edit_link = manage.click('Bearbeiten', index=0).request.url
    delete_link = manage.click('Löschen', index=0).request.url

    login_editor_1(client)
    client.get('/organizations', status=403)
    client.get(edit_link, status=403)
    client.get(delete_link, status=403)


def test_view_organizations_order(gazette_app):
    client = Client(gazette_app)
    login_publisher(client)

    manage = client.get('/organizations')
    manage = manage.click('Ordnen')
    organizations = [
        t.text.strip() for t in manage.pyquery('ul[data-sortable] li')
    ]
    assert organizations == [
        'State Chancellery',
        'Civic Community',
        'Municipality',
        'Churches',
        'Evangelical Reformed Parish',
        'Sikh Community',
        'Catholic Parish',
        'Corporation',
    ]
    url = manage.pyquery('ul[data-sortable]')[0].attrib['data-sortable-url']
    expected = (
        '/move/organization/%7Bsubject_id%7D/%7Bdirection%7D/%7Btarget_id%7D'
        '?csrf-token='
    )
    assert expected in url

    # Move items
    session = gazette_app.session()
    query = session.query(Organization)

    subject = query.filter_by(title='State Chancellery').one().id
    target = query.filter_by(title='Municipality').one().id
    move = url.replace('%7Bsubject_id%7D', str(subject))
    move = move.replace('%7Btarget_id%7D', str(target))
    move = move.replace('%7Bdirection%7D', 'below')
    client.put(move)

    subject = query.filter_by(title='Catholic Parish').one().id
    target = query.filter_by(title='Sikh Community').one().id
    move = url.replace('%7Bsubject_id%7D', str(subject))
    move = move.replace('%7Btarget_id%7D', str(target))
    move = move.replace('%7Bdirection%7D', 'above')
    client.put(move)

    manage = client.get('/organizations')
    manage = manage.click('Ordnen')
    organizations = [
        t.text.strip() for t in manage.pyquery('ul[data-sortable] li')
    ]
    assert organizations == [
        'Civic Community',
        'Municipality',
        'State Chancellery',
        'Churches',
        'Evangelical Reformed Parish',
        'Catholic Parish',
        'Sikh Community',
        'Corporation',
    ]


def test_view_organizations_export(gazette_app):
    client = Client(gazette_app)

    client.get('/organizations/export', status=403)

    login_editor_1(client)
    client.get('/organizations/export', status=403)

    login_publisher(client)
    response = client.get('/organizations/export')

    book = load_workbook(BytesIO(response.body))
    assert len(book.worksheets) == 1

    sheet = book.worksheets[0]
    assert sheet.max_column == 6
    assert sheet.max_row == 9

    assert sheet.cell(1, 1).value == 'ID'
    assert sheet.cell(1, 2).value == 'Name'
    assert sheet.cell(1, 3).value == 'Titel'
    assert sheet.cell(1, 4).value == 'Aktiv'
    assert sheet.cell(1, 5).value == 'Externe ID'
    assert sheet.cell(1, 6).value == 'Übergeordnete Organisation'

    assert sheet.cell(2, 2).value == '100'
    assert sheet.cell(2, 3).value == 'State Chancellery'
    assert sheet.cell(2, 4).value == 1
    assert sheet.cell(2, 5).value is None
    assert sheet.cell(2, 6).value is None

    assert sheet.cell(3, 2).value == '200'
    assert sheet.cell(3, 3).value == 'Civic Community'
    assert sheet.cell(3, 4).value == 1
    assert sheet.cell(3, 5).value is None
    assert sheet.cell(3, 6).value is None

    assert sheet.cell(4, 2).value == '300'
    assert sheet.cell(4, 3).value == 'Municipality'
    assert sheet.cell(4, 4).value == 1
    assert sheet.cell(4, 5).value is None
    assert sheet.cell(4, 6).value is None

    assert sheet.cell(5, 2).value == '400'
    assert sheet.cell(5, 3).value == 'Churches'
    assert sheet.cell(5, 4).value == 1
    assert sheet.cell(5, 5).value is None
    assert sheet.cell(5, 6).value is None

    assert sheet.cell(6, 2).value == '410'
    assert sheet.cell(6, 3).value == 'Evangelical Reformed Parish'
    assert sheet.cell(6, 4).value == 1
    assert sheet.cell(6, 5).value is None
    assert sheet.cell(6, 6).value == sheet.cell(5, 1).value

    assert sheet.cell(7, 2).value == '420'
    assert sheet.cell(7, 3).value == 'Sikh Community'
    assert sheet.cell(7, 4).value == 0
    assert sheet.cell(7, 5).value == '4'
    assert sheet.cell(7, 6).value == sheet.cell(5, 1).value

    assert sheet.cell(8, 2).value == '430'
    assert sheet.cell(8, 3).value == 'Catholic Parish'
    assert sheet.cell(8, 4).value == 1
    assert sheet.cell(8, 5).value is None
    assert sheet.cell(8, 6).value == sheet.cell(5, 1).value

    assert sheet.cell(9, 2).value == '500'
    assert sheet.cell(9, 3).value == 'Corporation'
    assert sheet.cell(9, 4).value == 1
    assert sheet.cell(9, 5).value is None
    assert sheet.cell(9, 6).value is None
