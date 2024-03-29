from freezegun import freeze_time
from io import BytesIO
from openpyxl import load_workbook
from pyquery import PyQuery as pq
from tests.onegov.gazette.common import login_editor_1
from tests.onegov.gazette.common import login_publisher
from webtest import TestApp as Client


def test_view_categories(gazette_app):
    with freeze_time("2017-10-20 12:00"):
        client = Client(gazette_app)
        login_publisher(client)

        # Test data:
        # 10 / Complaints / inactive
        # 11 / Education / active
        # 12 / Submissions / active
        # 13 / Commercial Register / active
        # 14 / Elections / active

        # add a category
        manage = client.get('/categories')
        manage = manage.click('Neu')
        manage.form['title'] = 'Rubrik XY'
        manage.form['active'] = True
        manage = manage.form.submit().maybe_follow()
        assert 'Rubrik hinzugefügt.' in manage
        assert 'Rubrik XY' in manage
        categories = [
            [
                ''.join((td.text_content(), td.attrib['class']))
                for td in pq(tr)('td')[:1]
            ][0]
            for tr in manage.pyquery('table.categories tbody tr')
        ]
        assert categories == [
            'Commercial Register (13)',
            'Complaints (10)inactive',
            'Education (11)',
            'Elections (14)',
            'Rubrik XY (15)',
            'Submissions (12)'
        ]

        # use the first category in a notice
        manage = client.get('/notices/drafted/new-notice')
        manage.form['title'] = 'Titel'
        manage.form['organization'] = '200'
        manage.form['category'] = '13'
        manage.form['issues'] = ['2017-44']
        manage.form['text'] = 'Text'
        manage.form['author_place'] = 'Govikon'
        manage.form['author_name'] = 'State Chancellerist'
        manage.form['author_date'] = '2019-01-01'
        manage = manage.form.submit().maybe_follow()
        assert '<h2>Titel</h2>' in manage
        assert 'Commercial Register' in manage

        # edit the first category
        manage = client.get('/categories')
        manage = manage.click('Bearbeiten', index=0)
        manage.form['title'] = 'Rubrik Z'
        manage.form['active'] = False
        manage = manage.form.submit().maybe_follow()
        assert 'Rubrik geändert.' in manage
        assert 'Commercial Register' not in manage

        categories = [
            [
                ''.join((td.text_content(), td.attrib['class']))
                for td in pq(tr)('td')[:1]
            ][0]
            for tr in manage.pyquery('table.categories tbody tr')
        ]
        assert categories == [
            'Complaints (10)inactive',
            'Education (11)',
            'Elections (14)',
            'Rubrik XY (15)',
            'Rubrik Z (13)inactive',
            'Submissions (12)'
        ]

        # check if the notice has been updated
        manage = client.get('/notice/titel')
        assert 'Commercial Register' not in manage
        assert 'Rubrik Z' in manage

        # delete all but one (unused) categories
        manage = client.get('/categories')
        manage.click('Löschen', index=0).form.submit()
        manage.click('Löschen', index=1).form.submit()
        manage.click('Löschen', index=2).form.submit()
        manage.click('Löschen', index=3).form.submit()
        manage.click('Löschen', index=5).form.submit()

        manage = client.get('/categories')
        assert 'Complaints' not in manage
        assert 'Education' not in manage
        assert 'Elections' not in manage
        assert 'Rubrik XY' not in manage
        assert 'Rubrik Z' in manage
        assert 'Submissions' not in manage

        # Try to delete the used category
        manage = client.get('/categories')
        manage = manage.click('Löschen')
        assert 'Es können nur unbenutzte Rubriken gelöscht werden.' in manage
        assert not manage.forms


def test_view_categories_permissions(gazette_app):
    client = Client(gazette_app)

    login_publisher(client)
    manage = client.get('/categories').click('Neu')
    manage.form['title'] = 'XY'
    manage = manage.form.submit().maybe_follow()
    edit_link = manage.click('Bearbeiten', index=0).request.url
    delete_link = manage.click('Löschen', index=0).request.url

    login_editor_1(client)
    client.get('/categories', status=403)
    client.get(edit_link, status=403)
    client.get(delete_link, status=403)


def test_view_categories_export(gazette_app):
    client = Client(gazette_app)

    client.get('/categories/export', status=403)

    login_editor_1(client)
    client.get('/categories/export', status=403)

    login_publisher(client)
    response = client.get('/categories/export')

    book = load_workbook(BytesIO(response.body))
    assert len(book.worksheets) == 1

    sheet = book.worksheets[0]
    assert sheet.max_column == 4
    assert sheet.max_row == 6

    assert sheet.cell(1, 1).value == 'ID'
    assert sheet.cell(1, 2).value == 'Name'
    assert sheet.cell(1, 3).value == 'Titel'
    assert sheet.cell(1, 4).value == 'Aktiv'

    assert sheet.cell(2, 2).value == '13'
    assert sheet.cell(2, 3).value == 'Commercial Register'
    assert sheet.cell(2, 4).value == 1

    assert sheet.cell(3, 2).value == '10'
    assert sheet.cell(3, 3).value == 'Complaints'
    assert sheet.cell(3, 4).value == 0

    assert sheet.cell(4, 2).value == '11'
    assert sheet.cell(4, 3).value == 'Education'
    assert sheet.cell(4, 4).value == 1

    assert sheet.cell(5, 2).value == '14'
    assert sheet.cell(5, 3).value == 'Elections'
    assert sheet.cell(5, 4).value == 1

    assert sheet.cell(6, 2).value == '12'
    assert sheet.cell(6, 3).value == 'Submissions'
    assert sheet.cell(6, 4).value == 1
