from __future__ import annotations

import os
import re
import transaction

from datetime import datetime
from freezegun import freeze_time
from io import BytesIO
from onegov.core.csv import convert_list_of_dicts_to_xlsx
from onegov.core.utils import Bunch
from onegov.newsletter import RecipientCollection, NewsletterCollection
from onegov.user import UserCollection
from openpyxl import load_workbook
from sedate import replace_timezone
from tests.shared.utils import find_link_by_href_end
from webtest.forms import Upload


from typing import Any, TYPE_CHECKING
if TYPE_CHECKING:
    from .conftest import Client


def test_show_newsletter(client: Client) -> None:
    client.login_admin().follow()

    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page = page.form.submit().follow()

    page = client.get('/news')
    page = page.click('Nachricht')
    page.form['title'] = "We have a new homepage"
    page.form['lead'] = "It is very good"
    page.form['text'] = "It is lots of fun"
    page = page.form.submit().follow()

    page = client.get('/news')
    assert "Newsletter" in page.text

    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = False
    page = page.form.submit().follow()

    page = client.get('/')
    assert "Newsletter" not in page.text

    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page = page.form.submit().follow()

    page = client.get('/news')
    assert "Newsletter" in page.text

    page = client.get('/newsletters')
    assert "Newsletter" in page.text


def test_newsletter_disabled(client: Client) -> None:

    anon = client.spawn()

    client.login_admin()

    assert anon.get('/newsletters', expect_errors=True).status_code == 404
    assert client.get('/newsletters').status_code == 200

    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form.submit().follow()
    client.logout()

    assert anon.get('/newsletters').status_code == 200
    assert client.get('/newsletters').status_code == 200


def test_unsubscribe_link(client: Client) -> None:
    request: Any = Bunch(
        identity_secret=client.app.identity_secret,
        app=client.app
    )

    session = client.app.session()
    user = UserCollection(session).by_username('editor@example.org')
    assert user is not None
    assert not user.data

    # valid token
    token = client.app.request_class.new_url_safe_token(
        request, {'user': 'editor@example.org'},
        salt='unsubscribe'
    )
    url = '/unsubscribe?token={}'.format(token)

    assert "abgemeldet" in client.get(url)
    client.post(url)

    user = UserCollection(session).by_username('editor@example.org')
    assert user is not None
    assert user.data['ticket_statistics'] == 'never'

    # unknown user
    token = client.app.request_class.new_url_safe_token(
        request, {'user': 'unknown@example.org'},
        salt='unsubscribe'
    )
    url = '/unsubscribe?token={}'.format(token)

    assert client.get(url, expect_errors=True).status_code == 403
    client.post('/unsubscribe?token={}'.format(token))

    # invalid token
    token = client.app.request_class.new_url_safe_token(
        request, {'user': 'editor@example.org'},
        salt='foobar'
    )
    url = '/unsubscribe?token={}'.format(token)

    assert client.get(url, expect_errors=True).status_code == 403
    client.post(url)


def test_newsletters_crud(client: Client) -> None:

    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form.submit().follow()
    client.logout()

    client.login_editor()

    newsletter = client.get('/newsletters')
    assert 'Es wurden noch keine Newsletter versendet' in newsletter

    new_link = find_link_by_href_end(newsletter, '/newsletters/new')
    assert new_link is not None
    new = newsletter.click(href=new_link['href'])
    new.form['title'] = "Our town is AWESOME"
    new.form['lead'] = "Like many of you, I just love our town..."
    new.select_checkbox("occurrences", "150 Jahre Govikon")
    new.select_checkbox("occurrences", "Gemeinsames Turnen")
    newsletter = new.form.submit().follow()

    assert newsletter.pyquery('h1').text() == "Our town is AWESOME"
    assert "Like many of you" in newsletter
    assert "Gemeinsames Turnen" in newsletter
    assert "Turnhalle" in newsletter
    assert "150 Jahre Govikon" in newsletter
    assert "Sportanlage" in newsletter

    edit = newsletter.click("Bearbeiten")
    edit.form['title'] = "I can't even"
    edit.select_checkbox("occurrences", "150 Jahre Govikon", checked=False)

    newsletter = edit.form.submit().follow()

    assert newsletter.pyquery('h1').text() == "I can't even"
    assert "Like many of you" in newsletter
    assert "Gemeinsames Turnen" in newsletter
    assert "Turnhalle" in newsletter
    assert "150 Jahre Govikon" not in newsletter
    assert "Sportanlage" not in newsletter

    newsletters = client.get('/newsletters')
    assert "I can't even" in newsletters
    assert "Noch nicht gesendet." in newsletters

    # not sent, therefore not visible to the public
    assert "noch keine Newsletter" in client.spawn().get('/newsletters')

    delete_link = newsletter.pyquery('a.delete-link').attr('ic-delete-from')
    client.delete(delete_link)

    newsletters = client.get('/newsletters')
    assert "noch keine Newsletter" in newsletters


def test_newsletter_secret_private_content(client: Client) -> None:
    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form['secret_content_allowed'] = False
    page.form.submit().follow()
    client.logout()

    client.login_editor()
    page = client.get('/news').click('Nachricht')
    page.form['title'] = 'Public Information'
    page.form['lead'] = 'Public Info'
    page.form['text'] = 'Public Info Text'
    page.form['access'] = 'public'
    page.form.submit()

    page = client.get('/news').click('Nachricht')
    page.form['title'] = 'Secret Information'
    page.form['lead'] = 'Secret Info'
    page.form['text'] = 'Secret Info Text'
    page.form['access'] = 'secret'
    page.form.submit()

    page = client.get('/news').click('Nachricht')
    page.form['title'] = 'Private Information'
    page.form['lead'] = 'Private Info'
    page.form['text'] = 'Private Info Text'
    page.form['access'] = 'private'
    page.form.submit()

    newsletter = client.get('/newsletters')
    new_link = find_link_by_href_end(newsletter, '/newsletters/new')
    assert new_link is not None
    new = newsletter.click(href=new_link['href'])
    new.form['title'] = "Information"
    new.form['lead'] = ("We love information about our town!")
    new.select_checkbox("news", "Public Information")
    new.select_checkbox("news", "Secret Information")
    new.select_checkbox("news", "Private Information")
    newsletter = new.form.submit().follow()

    assert "Public Information" in newsletter
    assert "Secret Information" in newsletter
    assert "Private Information" in newsletter
    assert "Sie haben 'geheime' Inhalte für Ihren Newsletter" in newsletter
    assert "Sie haben 'privaten' Inhalt für Ihren Newsletter" in newsletter

    # render newsletter before sending. Preview shows the content of the
    # iframe in '/newsletter/information/send'
    preview = client.get('/newsletter/information/preview')
    assert "Public Information" in preview
    assert "Secret Information" not in preview
    assert "Private Information" not in preview

    # enable setting for secret content
    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form['secret_content_allowed'] = True
    page.form.submit().follow()
    client.logout()

    client.login_editor()
    newsletter = client.get('/newsletter/information')
    assert "Public Information" in newsletter
    assert "Secret Information" in newsletter
    assert "Private Information" in newsletter
    assert "Sie haben 'geheime' Inhalte für Ihren Newsletter" not in newsletter
    assert "Sie haben 'privaten' Inhalt für Ihren Newsletter" in newsletter

    # render newsletter before sending (including secret content)
    preview = client.get('/newsletter/information/preview')
    assert "Public Information" in preview
    assert "Secret Information" in preview
    assert "Private Information" not in preview
    client.logout()

    # anybody can see the public content only
    newsletter = client.get('/newsletter/information')
    assert "Public Information" in newsletter
    assert "Secret Information" not in newsletter
    assert "Private Information" not in newsletter


def test_newsletter_signup(client: Client) -> None:

    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form.submit().follow()
    client.logout()

    page = client.get('/newsletters')
    page.form['address'] = 'asdf'
    page = page.form.submit()

    assert 'Ungültig' in page

    page.form['address'] = 'info@example.org'
    page.form.submit()

    assert len(os.listdir(client.app.maildir)) == 1

    # make sure double submissions don't result in multiple e-mails
    page.form.submit()
    assert len(os.listdir(client.app.maildir)) == 1

    message = client.get_email(0)['TextBody']
    assert 'Mit freundlichen Grüssen' not in message
    assert 'Das OneGov Cloud Team' not in message

    confirm = re.search(r'Anmeldung bestätigen\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]
    unsubscribe = re.search(r'abzumelden.\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]
    assert confirm.split('/confirm')[0] == unsubscribe.split('/unsubscribe')[0]

    # unsubscribing before the opt-in does nothing, no emails are sent
    assert "vom Newsletter abgemeldet" in client.get(unsubscribe).follow()

    # try an illegal token first
    illegal_confirm = confirm.split('/confirm')[0] + 'x/confirm'
    assert "falsches Token" in client.get(illegal_confirm).follow()

    # make sure double calls work
    assert "info@example.org wurde erfolgreich" in client.get(confirm).follow()
    assert "info@example.org wurde erfolgreich" in client.get(confirm).follow()

    # subscribing still works the same, but there's still no email sent
    page.form.submit()
    assert len(os.listdir(client.app.maildir)) == 1

    # unsubscribing does not result in an e-mail either
    illegal_unsub = unsubscribe.split('/unsubscribe')[0] + 'x/unsubscribe'
    assert "falsches Token" in client.get(illegal_unsub).follow()
    assert "vom Newsletter abgemeldet" in client.get(unsubscribe).follow()

    # no e-mail is sent when unsubscribing
    assert len(os.listdir(client.app.maildir)) == 1

    # however, we can now signup again
    page.form.submit()
    assert len(os.listdir(client.app.maildir)) == 2


def test_newsletter_signup_for_categories(client: Client) -> None:
    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form['newsletter_categories'] = """
    - News
    - Aktivitäten:
      - Anlässe
      - Sport
    """
    page.form.submit().follow()
    client.logout()

    page = client.get('/newsletters')
    assert 'News' in page
    assert 'Aktivitäten' in page
    assert 'Anlässe' in page
    assert 'Sport' in page
    page.form['address'] = 'info@example.org'
    page.form['subscribed_categories'] = ['News', 'Anlässe']
    page = page.form.submit().follow()
    assert ("Erfolg! Wir senden eine E-Mail zur Bestätigung Ihres "
            "Abonnements an info@example.org" in page)
    assert "Ihre abonnierten Kategorien sind News, Anlässe." in page

    # test confirmation email
    assert len(os.listdir(client.app.maildir)) == 1
    message = client.get_email(0)['TextBody']
    assert 'Sie haben folgende Newsletter-Kategorien abonniert:' in message
    assert 'abonniert: News, Anlässe' in message
    assert 'Link um Ihre Anmeldung zu best\u00e4tigen' in message
    assert ('Um Ihre Abonnementkategorien zu aktualisieren, klicken Sie hier'
            in message)
    assert 'Klicken Sie hier, um sich abzumelden' in message
    update_link = re.search(r'aktualisieren\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]
    assert update_link.endswith('newsletters/update')

    # test recipient
    recipients = RecipientCollection(client.app.session())
    recipient = recipients.by_address('info@example.org')
    assert recipient is not None
    assert recipient.subscribed_categories == ['News', 'Anlässe']
    assert recipient.confirmed is False

    # update subscription topics
    page = client.get(update_link)
    page.form['address'] = 'info@example.org'
    page.form['subscribed_categories'] = ['Sport']
    page = page.form.submit().follow()
    assert "Erfolg! Wir haben Ihre abonnierten Kategorien aktualisiert" in page

    recipients = RecipientCollection(client.app.session())
    recipient = recipients.by_address('info@example.org')
    assert recipient is not None
    assert recipient.subscribed_categories == ['Sport']
    assert recipient.confirmed is False


def test_newsletter_rfc8058(client: Client) -> None:

    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form.submit().follow()
    client.logout()

    page = client.get('/newsletters')
    page.form['address'] = 'info@example.org'
    page.form.submit()

    assert len(os.listdir(client.app.maildir)) == 1

    # make sure double submissions don't result in multiple e-mails
    page.form.submit()
    assert len(os.listdir(client.app.maildir)) == 1

    email = client.get_email(0)
    message = email['TextBody']
    assert 'Mit freundlichen Grüssen' not in message
    assert 'Das OneGov Cloud Team' not in message
    headers = {h['Name']: h['Value'] for h in email['Headers']}
    assert 'List-Unsubscribe' in headers
    assert 'List-Unsubscribe-Post' in headers
    unsubscribe = headers['List-Unsubscribe'].strip('<>')

    confirm = re.search(r'Anmeldung bestätigen\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]
    assert confirm.split('/confirm')[0] == unsubscribe.split('/unsubscribe')[0]

    # unsubscribing before the opt-in does nothing, no emails are sent
    client.post(unsubscribe)

    # try an illegal token first
    illegal_confirm = confirm.split('/confirm')[0] + 'x/confirm'
    assert "falsches Token" in client.get(illegal_confirm).follow()

    # make sure double calls work
    assert "info@example.org wurde erfolgreich" in client.get(confirm).follow()
    assert "info@example.org wurde erfolgreich" in client.get(confirm).follow()

    # subscribing still works the same, but there's still no email sent
    page.form.submit()
    assert len(os.listdir(client.app.maildir)) == 1

    # unsubscribing does not result in an e-mail either
    client.post(unsubscribe)

    # no e-mail is sent when unsubscribing
    assert len(os.listdir(client.app.maildir)) == 1

    # however, we can now signup again
    page.form.submit()
    assert len(os.listdir(client.app.maildir)) == 2


def test_newsletter_subscribers_and_edit_bar(client: Client) -> None:
    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form.submit().follow()
    client.logout()

    admin = client.spawn()
    admin.login_admin()
    editor = client.spawn()
    editor.login_editor()

    # only managers can see the subscribers and edit bar
    for current_client in (admin, editor):
        assert current_client.get('/subscribers').status_code == 200
        page = current_client.get('/newsletters')
        assert 'Abonnenten' in page
        assert page.pyquery('a.manage-subscribers')
        assert page.pyquery('a.new-newsletter')

    member = client.spawn()
    member.login_member()
    anom = client.spawn()

    for current_client in (member, anom):
        assert current_client.get(
            '/subscribers', expect_errors=True).status_code == 403
        page = current_client.get('/newsletters')
        assert 'Abonnenten' not in page
        assert not page.pyquery('a.manage-subscribers')
        assert not page.pyquery('a.new-newsletter')


def test_newsletter_subscribers_management(client: Client) -> None:

    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form.submit().follow()
    client.logout()

    page = client.get('/newsletters')
    page.form['address'] = 'info@example.org'
    page.form.submit()

    assert len(os.listdir(client.app.maildir)) == 1

    message = client.get_email(0)['TextBody']

    confirm = re.search(r'Anmeldung bestätigen\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]
    assert "info@example.org wurde erfolgreich" in client.get(confirm).follow()

    client.login_editor()

    subscribers = client.get('/subscribers')
    assert "info@example.org" in subscribers

    unsubscribe = subscribers.pyquery('a[ic-get-from]').attr('ic-get-from')
    result = client.get(unsubscribe).follow()
    assert "info@example.org erfolgreich vom Newsletter abgemeldet" in result


def test_newsletter_subscribers_management_by_manager(client: Client) -> None:
    # a manager (editor or admin) adds a new subscriber

    def subscribe_by_manager(client: Client) -> None:
        page = client.get('/newsletters')
        page.form['address'] = 'info@govikon.org'
        page.form['confirmed'] = True
        page = page.form.submit().follow()
        assert ('Wir haben info@govikon.org zur Liste der Empfänger '
                'hinzugefügt.' in page)

        assert len(os.listdir(client.app.maildir)) == 0  # no emails sent

        subscribers = client.get('/subscribers')
        assert "info@govikon.org" in subscribers

        recipient = RecipientCollection(client.app.session()).query().first()
        assert recipient is not None
        assert recipient.confirmed is True

        unsubscribe = subscribers.pyquery('a[ic-get-from]').attr('ic-get-from')
        result = client.get(unsubscribe).follow()
        assert "info@govikon.org erfolgreich vom Newsletter" in result

    client.login_admin()
    subscribe_by_manager(client)
    client.logout()

    client.login_editor()
    subscribe_by_manager(client)
    client.logout()


def test_newsletter_creation_limited_to_logged_in_users(
    client: Client
) -> None:
    # verify adding a new newsletter view is set to private

    # enable the newsletter
    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form['newsletter_categories'] = ''
    page.form.submit().follow()

    admin = client.spawn()
    admin.login_admin()
    editor = client.spawn()
    editor.login_editor()

    for current_client in (admin, editor):
        page = current_client.get('/newsletters/new')
        assert 'Neuer Newsletter' in page
        assert page.status_code == 200

    # member and anonymous users can't create newsletters
    anom = client.spawn()
    member = client.spawn()
    member.login_member()

    for current_client in (member, anom):
        assert current_client.get(
            '/newsletters/new', expect_errors=True).status_code == 403


def test_newsletter_send(client: Client) -> None:

    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form.submit().follow()
    client.logout()

    anon = client.spawn()

    client.login_editor()
    page = client.get('/news').click('Nachricht')
    page.form['title'] = 'Testnews'
    page.form['lead'] = 'My Lead Text'
    page.form['text'] = '<p>My Html editor text</p>'
    page.form['text_in_newsletter'] = True
    page.form.submit().follow()

    # add a newsletter
    newsletters = client.get('/newsletters')
    new_link = find_link_by_href_end(newsletters, '/newsletters/new')
    assert new_link is not None
    new = newsletters.click(href=new_link['href'])
    new.form['title'] = "Our town is AWESOME"
    new.form['lead'] = "Like many of you, I just love our town..."

    new.select_checkbox("news", "Testnews")
    new.select_checkbox("occurrences", "150 Jahre Govikon")
    new.select_checkbox("occurrences", "Gemeinsames Turnen", limit=3)

    new.form['closing_remark'] = '<p>Closing Remarks</p>'

    newsletter = new.form.submit().follow()

    # add some recipients the quick wqy
    recipients = RecipientCollection(client.app.session())
    recipients.add('one@example.org', confirmed=True)
    recipients.add('two@example.org', confirmed=True)
    recipients.add('xxx@example.org', confirmed=False)

    transaction.commit()

    assert "2 Abonnenten registriert" in client.get('/newsletters')

    # send the newsletter
    send = newsletter.click('Senden')
    assert "Dieser Newsletter wurde noch nicht gesendet." in send
    assert "one@example.org" not in send
    assert "two@example.org" not in send
    assert "xxx@example.org" not in send

    # # set newsletter reporting categories (no categories selects all for
    # # backward compatibility)
    # send.form['categories'] = []
    newsletter = send.form.submit().follow()
    assert '"Our town is AWESOME" wurde an 2 Empfänger gesendet' in newsletter
    assert '<p>Closing Remarks</p>' in newsletter

    page = anon.get('/newsletters')
    assert "gerade eben" in page

    # the send form should now look different
    send = newsletter.click('Senden')

    assert "Zum ersten Mal gesendet gerade eben." in send
    assert "Dieser Newsletter wurde an 2 Abonnenten gesendet." in send
    assert "one@example.org" in send
    assert "two@example.org" in send
    assert "xxx@example.org" not in send

    assert len(send.pyquery('.previous-recipients li')) == 2

    # make sure the mail was sent correctly
    assert len(os.listdir(client.app.maildir)) == 1

    mail = client.get_email(0, 0)
    message = mail['TextBody']

    assert "Our town is AWESOME" in message
    assert "Like many of you" in message

    web = re.search(r'Web-Version anzuzeigen.\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]
    assert web.endswith('/newsletter/our-town-is-awesome')

    # make sure the unconfirm link is different for each mail
    unconfirm_1 = re.search(r'abzumelden.\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]

    mail = client.get_email(0, 1)
    message = mail['TextBody']
    unconfirm_2 = re.search(r'abzumelden.\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]

    assert unconfirm_1 and unconfirm_2
    assert unconfirm_1 != unconfirm_2

    # make sure the unconfirmed link actually works
    anon.get(unconfirm_1)
    assert recipients.query().count() == 2

    anon.get(unconfirm_2)
    assert recipients.query().count() == 1

    # check content of mail
    assert 'Like many of you,' in message
    assert '150 Jahre Govikon' in message
    assert message.count('Gemeinsames Turnen') == 3
    assert 'Testnews' in message
    assert 'My Lead Text' in message
    assert 'My Html editor text' in message
    assert 'Closing Remarks' in message


def test_newsletter_send_with_categories(client: Client) -> None:

    client.login_admin()
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form['newsletter_categories'] = """
      - News
      - Aktivitäten:
        - Anlässe
        - Sport
    """
    page.form.submit().follow()
    client.logout()

    anon = client.spawn()

    client.login_editor()
    page = client.get('/news').click('Nachricht')
    page.form['title'] = 'Testnews'
    page.form['lead'] = 'My Lead Text'
    page.form['text'] = '<p>My Html editor text</p>'
    page.form['text_in_newsletter'] = True
    page.form.submit().follow()

    # add a newsletter
    newsletters = client.get('/newsletters')
    new_link = find_link_by_href_end(newsletters, '/newsletters/new')
    assert new_link is not None
    new = newsletters.click(href=new_link['href'])
    new.form['title'] = "Our town is AWESOME"
    new.form['lead'] = "Like many of you, I just love our town..."

    new.select_checkbox("news", "Testnews")
    new.select_checkbox("occurrences", "150 Jahre Govikon")
    new.select_checkbox("occurrences", "Gemeinsames Turnen")

    newsletter = new.form.submit().follow()

    # add some recipients the quick wqy
    recipients = RecipientCollection(client.app.session())
    recipients.add('one@example.org', confirmed=True,
                   subscribed_categories=['News', 'Sport'])
    recipients.add('two@example.org', confirmed=True,
                   subscribed_categories=['Aktivitäten', 'Sport'])
    recipients.add('three@example.org', confirmed=True,
                   subscribed_categories=None)
    recipients.add('xxx@example.org', confirmed=False,
                   subscribed_categories=['News', 'Aktivitäten', 'Anlässe',
                                          'Sport'])

    transaction.commit()

    assert "3 Abonnenten registriert" in client.get('/newsletters')

    # send the newsletter
    send = newsletter.click('Senden')
    assert "Dieser Newsletter wurde noch nicht gesendet." in send
    assert "one@example.org" not in send
    assert "two@example.org" not in send
    assert "three@example.org" not in send
    assert "xxx@example.org" not in send

    send.select_checkbox("categories", "News")
    newsletter = send.form.submit().follow()
    assert '"Our town is AWESOME" wurde an 2 Empfänger gesendet' in newsletter

    page = anon.get('/newsletters')
    assert "gerade eben" in page

    # the send form should now look different
    send = newsletter.click('Senden')

    assert "Zum ersten Mal gesendet gerade eben." in send
    assert "Dieser Newsletter wurde an 2 Abonnenten gesendet." in send
    assert "one@example.org" in send
    assert "two@example.org" not in send
    assert "three@example.org" in send
    assert "xxx@example.org" not in send

    assert len(send.pyquery('.previous-recipients li')) == 2

    # make sure the mail was sent correctly
    assert len(os.listdir(client.app.maildir)) == 1

    # add a second newsletter
    newsletters = client.get('/newsletters')
    new_link = find_link_by_href_end(newsletters, '/newsletters/new')
    assert new_link is not None
    new = newsletters.click(href=new_link['href'])
    new.form['title'] = "Sport Update"
    new.form['lead'] = "Bla bla blupp..."

    new.select_checkbox("occurrences", "Gemeinsames Turnen")
    newsletter = new.form.submit().follow()

    # send the newsletter
    send = newsletter.click('Senden')
    assert "Dieser Newsletter wurde noch nicht gesendet." in send
    assert "one@example.org" not in send
    assert "two@example.org" not in send
    assert "three@example.org" not in send
    assert "xxx@example.org" not in send

    send.select_checkbox("categories", "Sport")
    newsletter = send.form.submit().follow()
    assert '"Sport Update" wurde an 3 Empfänger gesendet' in newsletter

    page = anon.get('/newsletters')
    assert "gerade eben" in page

    # the send form should now look different
    send = newsletter.click('Senden')

    assert "Zum ersten Mal gesendet gerade eben." in send
    assert "Dieser Newsletter wurde an 3 Abonnenten gesendet." in send
    assert "one@example.org" in send
    assert "two@example.org" in send
    assert "xxx@example.org" not in send

    assert len(send.pyquery('.previous-recipients li')) == 3

    # make sure the mails were sent correctly
    assert len(os.listdir(client.app.maildir)) == 2


def test_newsletter_schedule(client: Client) -> None:
    client.login_editor()

    # add a newsletter
    newsletters = client.get('/newsletters')
    new_link = find_link_by_href_end(newsletters, '/newsletters/new')
    assert new_link is not None
    new = newsletters.click(href=new_link['href'])
    new.form['title'] = "Our town is AWESOME"
    new.form['lead'] = "Like many of you, I just love our town..."

    new.select_checkbox("news", "Willkommen bei OneGov")
    new.select_checkbox("occurrences", "150 Jahre Govikon")

    newsletter = new.form.submit().follow()

    # add some recipients the quick wqy
    recipients = RecipientCollection(client.app.session())
    recipients.add('one@example.org', confirmed=True)
    recipients.add('two@example.org', confirmed=True)

    transaction.commit()

    send = newsletter.click('Senden')
    send.form['send'] = 'specify'

    # schedule the newsletter too close to execute
    time = replace_timezone(datetime(2018, 5, 31, 11, 55, 1), 'Europe/Zurich')

    with freeze_time(time):
        send.form['time'] = '2018-05-31 12:00'
        assert '5 Minuten in der Zukunft' in send.form.submit()

        # schedule the newsletter outside the hour
        send.form['time'] = '2018-05-31 12:55'
        assert 'nur zur vollen Stunde' in send.form.submit()

    # schedule the newsletter at a valid time
    time = replace_timezone(datetime(2018, 5, 31, 11, 55, 0), 'Europe/Zurich')

    with freeze_time(time):
        send.form['time'] = '2018-05-31 12:00'
        send.form.submit().follow()


def test_newsletter_test_delivery(client: Client) -> None:
    client.login_editor()

    # add a newsletter
    newsletters = client.get('/newsletters')
    new_link = find_link_by_href_end(newsletters, '/newsletters/new')
    assert new_link is not None
    new = newsletters.click(href=new_link['href'])
    new.form['title'] = "Our town is AWESOME"
    new.form['lead'] = "Like many of you, I just love our town..."

    new.select_checkbox("news", "Willkommen bei OneGov")
    new.select_checkbox("occurrences", "150 Jahre Govikon")

    newsletter_page = new.form.submit().follow()

    # add some recipients the quick wqy
    recipients = RecipientCollection(client.app.session())
    recipients.add('one@example.org', confirmed=True)
    recipients.add('two@example.org', confirmed=True)

    recipient = recipients.query().first().id.hex  # type: ignore[union-attr]

    transaction.commit()

    send = newsletter_page.click('Test')
    send.form['selected_recipient'] = recipient
    send.form.submit().follow()

    assert len(os.listdir(client.app.maildir)) == 1

    send = newsletter_page.click('Test')
    send.form['selected_recipient'] = recipient
    send.form.submit().follow()

    assert len(os.listdir(client.app.maildir)) == 2

    newsletter = NewsletterCollection(client.app.session()).query().one()
    assert newsletter.sent is None
    assert not newsletter.recipients


def test_import_export_subscribers(client: Client) -> None:
    with freeze_time("2018-05-31 12:00"):
        session = client.app.session()
        client.login_admin()

        # add a newsletter
        newsletters = client.get('/newsletters')
        new_link = find_link_by_href_end(newsletters, '/newsletters/new')
        assert new_link is not None
        new = newsletters.click(href=new_link['href'])
        new.form['title'] = "Our town is AWESOME"
        new.form['lead'] = "Like many of you, I just love our town..."

        new.select_checkbox("news", "Willkommen bei OneGov")
        new.select_checkbox("occurrences", "150 Jahre Govikon")

        new.form.submit().follow()

        # add some recipients the quick way
        recipients = RecipientCollection(session)
        recipients.add('one@example.org', confirmed=True)
        recipients.add('two@example.org', confirmed=True)
        # this recipient will not show up because they are unconfirmed
        recipients.add('xxx@example.org', confirmed=False)

        transaction.commit()

        # perform export
        page = client.get('/subscribers/export-newsletter-recipients')
        page.form['file_format'] = 'xlsx'

        response = page.form.submit()

        wb = load_workbook(BytesIO(response.body), data_only=True)
        sheet = tuple(wb[wb.sheetnames[0]].rows)
        assert sheet[0][0].value == 'Adresse'
        assert sheet[1][0].value == 'one@example.org'
        assert sheet[2][0].value == 'two@example.org'

        page.form['file_format'] = 'json'
        response = page.form.submit().json
        assert response == [
            {'Adresse': 'one@example.org',
             'Abonniert am': '2018-05-31T12:00:00+00:00'},
            {'Adresse': 'two@example.org',
             'Abonniert am': '2018-05-31T12:00:00+00:00'},
        ]

        page.form['file_format'] = 'xlsx'

        more_recipients = [
            {'Adresse': 'three@example.org',
             'Abonniert am': '2018-05-31T12:00:00+00:00'},
            {'Adresse': 'four@example.org',
             'Abonniert am': '2018-05-31T12:00:00+00:00'},
        ]
        file = Upload(
            'file',
            convert_list_of_dicts_to_xlsx(more_recipients),
            'application/vnd.openxmlformats-'
            'officedocument.spreadsheetml.sheet',
        )

        # Import (Dry run)
        page = client.get('/subscribers/import-newsletter-recipients')
        page.form['dry_run'] = True
        page.form['file'] = file
        page = page.form.submit()
        assert "Importvorschau: Imported: 2" in page

        # Import
        page = client.get('/subscribers/import-newsletter-recipients')
        page.form['dry_run'] = False
        page.form['file'] = file
        page = page.form.submit().follow()
        assert "Import abgeschlossen: Imported: 2" in page
        assert recipients.query().count() == 5
        assert recipients.query().filter_by(confirmed=True).count() == 4


def test_admin_receives_email_notification_on_unsubscription(
    client: Client
) -> None:

    def extract_unsubscription_link(client: Client, index: int) -> str:
        message = client.get_email(index)['TextBody']
        unsubscribe = re.search(r'abzumelden.\]\(([^\)]+)', message).group(1)  # type: ignore[union-attr]
        return unsubscribe

    client.login_admin()

    # newsletter settings no admin notification
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form['notify_on_unsubscription'] = []
    page.form.submit().follow()

    # add a newsletter
    newsletters = client.get('/newsletters')
    new_link = find_link_by_href_end(newsletters, '/newsletters/new')
    assert new_link is not None
    new = newsletters.click(href=new_link['href'])
    new.form['title'] = "Our town is AWESOME"
    new.form['lead'] = "Like many of you, I just love our town..."
    new.select_checkbox("news", "Willkommen bei OneGov")
    new.select_checkbox("occurrences", "150 Jahre Govikon")
    new.form.submit().follow()

    # add some recipients the quick way
    recipients = RecipientCollection(client.app.session())
    recipients.add('one@example.org', confirmed=True)
    recipients.add('two@example.org', confirmed=True)

    transaction.commit()

    # send out newsletter
    newsletter = client.get('/newsletter/our-town-is-awesome')
    preview = newsletter.click('Senden')
    preview.form.submit().follow()

    # verify newsletter was sent
    assert len(os.listdir(client.app.maildir)) == 1

    # extract unsubscription link from email and unsubscribe
    unsubscribe = extract_unsubscription_link(client, 0)
    result = client.get(unsubscribe).follow()
    assert "one@example.org erfolgreich vom Newsletter abgemeldet" in result

    # verify admin received NO email notification
    assert len(os.listdir(client.app.maildir)) == 1  # no new email

    # newsletter settings enable admin notification
    page = client.get('/newsletter-settings')
    page.form['show_newsletter'] = True
    page.form['notify_on_unsubscription'].select_multiple(texts=[
        'admin@example.org'])
    page.form.submit().follow()

    # add another newsletter
    newsletters = client.get('/newsletters')
    new_link = find_link_by_href_end(newsletters, '/newsletters/new')
    assert new_link is not None
    new = newsletters.click(href=new_link['href'])
    new.form['title'] = "Our town is AWESOME 2"
    new.form['lead'] = "Event reminder"
    new.select_checkbox("occurrences", "150 Jahre Govikon")
    new.form.submit().follow()

    # send newsletter
    newsletter = client.get('/newsletter/our-town-is-awesome-2')
    preview = newsletter.click('Senden')
    preview.form.submit().follow()

    assert len(os.listdir(client.app.maildir)) == 2

    # extract unsubscription link from email and unsubscribe
    unsubscribe = extract_unsubscription_link(client, 1)
    result = client.get(unsubscribe).follow()
    assert "two@example.org erfolgreich vom Newsletter abgemeldet" in result

    # verify admin received email notification
    assert len(os.listdir(client.app.maildir)) == 3
    assert 'two@example.org' in client.get_email(2)['TextBody']
