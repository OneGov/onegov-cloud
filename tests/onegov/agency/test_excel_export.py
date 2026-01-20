from __future__ import annotations

from onegov.agency.excel_export import column_mapper
from onegov.agency.excel_export import export_person_xlsx
from onegov.agency.excel_export import extract_person_data
from onegov.agency.models import ExtendedPerson, ExtendedAgency
from openpyxl import load_workbook


from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from sqlalchemy.orm import Session


def test_excel_export(session: Session) -> None:
    agency_x = ExtendedAgency(title='Agency X', name='x')
    agency_y = ExtendedAgency(title='Agency Y', name='y')

    person_a = ExtendedPerson(
        salutation='Herr',
        academic_title='Dr.',
        last_name='A',
        first_name='B',
        email='a@b.com',
        phone='+4444',
        phone_direct='+4445',
        born='01.01.2019',
        profession='Funktionär',
        function='Funktionär',
        political_party='PDS',
        parliamentary_group='Parlamentarische Gruppe',
        website='a.com',
        website_2='c.com',
        location_address='',
        location_code_city='',
        postal_address='Funktionärstrasse 5',
        postal_code_city='4321 Rüsslikon',
        notes='Notizen',
    )

    person_aa = ExtendedPerson(
        last_name='A',
        first_name='A',
    )

    person_ac = ExtendedPerson(
        last_name='A',
        first_name='C',
    )

    person_b = ExtendedPerson(
        salutation='Herr',
        last_name='B',
        first_name='C',
        email='b@c.com',
        phone='+4444',
        phone_direct='+4445',
        born='01.01.2019',
        profession='Funktionär',
        function='Funktionär',
        political_party='PDS',
        parliamentary_group='Parlamentarische Gruppe',
        website='a.com',
        website_2='b.com',
        location_address='',
        location_code_city='',
        postal_address='Funktionärstrasse 5',
        postal_code_city='4321 Rüsslikon',
        notes='Notizen',
    )

    session.add_all(
        [agency_x, agency_y, person_a, person_b, person_aa, person_ac])
    session.flush()

    agency_x.add_person(person_a.id, 'a in x')
    agency_x.add_person(person_b.id, 'b in x')
    agency_y.add_person(person_a.id, 'a in y')

    # Test data extraction
    write_out = extract_person_data(session)
    assert len(write_out) == 4

    file = export_person_xlsx(session)
    workbook = load_workbook(file)
    sheet = workbook['Personen']
    titles = [cell.value for cell in tuple(sheet.rows)[0]]

    # Test titles
    assert titles == list(column_mapper.values())

    # Test ordering in sheet sorted by last_name and first_name
    ix_ln = list(column_mapper.values()).index('Nachname')
    ix_fn = list(column_mapper.values()).index('Vorname')

    order = [f"{row[ix_ln].value}{row[ix_fn].value}" for row in sheet.rows]
    assert order == ['NachnameVorname', 'AA', 'AB', 'AC', 'BC']

    # Test a row
    def none_to_empty_string(value: object) -> object:
        if value is None:
            return ''
        return value

    row_num = 2
    expected: object
    for ix, (class_attrib, col_name) in enumerate(column_mapper.items()):
        if class_attrib == 'memberships':
            expected = f"{agency_x.title} - a in x\n{agency_y.title} - a in y"
        else:
            expected = none_to_empty_string(getattr(person_a, class_attrib))
        assert sheet.cell(row_num + 1, ix + 1).value == expected
