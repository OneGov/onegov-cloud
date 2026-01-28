from __future__ import annotations

import json
import os
import pytest
import re
import tempfile
import textwrap
import transaction
import warnings

from base64 import b64decode
from datetime import datetime, date, timedelta
from decimal import Decimal
from freezegun import freeze_time
from io import BytesIO
from libres.db.models import Reservation
from onegov.core.utils import module_path, normalize_for_url
from onegov.file import FileCollection
from onegov.form import FormSubmission
from onegov.org.models import ResourceRecipientCollection
from onegov.pay import Payment, PaymentCollection, InvoiceCollection
from onegov.pdf.utils import extract_pdf_info
from onegov.reservation import Resource, ResourceCollection
from onegov.ticket import TicketCollection, TicketInvoice
from onegov.user import UserCollection, User
from openpyxl import load_workbook
from pathlib import Path
from sqlalchemy import exc
from sqlalchemy.orm.session import close_all_sessions
from tests.shared.utils import add_reservation
from unittest.mock import patch
from urllib.parse import quote
from uuid import uuid4
from webtest import Upload


from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from unittest.mock import MagicMock
    from .conftest import Client


def test_resource_slots(client: Client) -> None:
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.add("Foo", 'Europe/Zurich')

    scheduler = resource.get_scheduler(client.app.libres_context)
    scheduler.allocate(
        dates=[
            (datetime(2015, 8, 4), datetime(2015, 8, 4)),
            (datetime(2015, 8, 5), datetime(2015, 8, 5))
        ],
        whole_day=True
    )
    scheduler.allocate(
        dates=[
            (datetime(2015, 8, 6, 12, 0), datetime(2015, 8, 6, 16, 0)),
        ],
        whole_day=False
    )
    scheduler.approve_reservations(
        scheduler.reserve(
            'info@example.org',
            (datetime(2015, 8, 6, 12, 0), datetime(2015, 8, 6, 16, 0)),
        )
    )

    transaction.commit()

    url = '/resource/foo/slots'
    assert client.get(url).json == []

    url = '/resource/foo/slots?start=2015-08-04&end=2015-08-05'
    result = client.get(url).json

    assert len(result) == 2

    assert result[0]['start'] == '2015-08-04T00:00:00+02:00'
    assert result[0]['end'] == '2015-08-05T00:00:00+02:00'
    assert result[0]['classNames'] == ['event-in-past', 'event-available']
    assert result[0]['title'] == "Ganztägig \nVerfügbar"

    assert result[1]['start'] == '2015-08-05T00:00:00+02:00'
    assert result[1]['end'] == '2015-08-06T00:00:00+02:00'
    assert result[1]['classNames'] == ['event-in-past', 'event-available']
    assert result[1]['title'] == "Ganztägig \nVerfügbar"

    url = '/resource/foo/slots?start=2015-08-06&end=2015-08-06'
    result = client.get(url).json

    assert len(result) == 1
    assert result[0]['classNames'] == ['event-in-past', 'event-unavailable']
    assert result[0]['title'] == "12:00 - 16:00 \nBesetzt"


def test_resources(client: Client) -> None:
    client.login_admin()

    resources = client.get('/resources')

    new_item = resources.click('Gegenstand')
    new_item.form['title'] = 'Beamer'
    resource = new_item.form.submit().follow()
    assert 'Beamer' in resource
    edit = resource.click('Bearbeiten')
    edit.form['title'] = 'Beamers'
    edit.form.submit()

    new = resources.click('Raum')
    new.form['title'] = 'Meeting Room 1'
    new.form['group'] = 'Office'
    new.form['subgroup'] = 'Big Meeting Room'
    resource = new.form.submit().follow()

    assert 'calendar' in resource
    assert 'Meeting Room 1' in resource

    edit = resource.click('Bearbeiten')
    edit.form['title'] = 'Besprechungsraum 1'
    edit.form.submit()

    new = resources.click('Raum')
    new.form['title'] = 'Meeting Room 2'
    new.form['group'] = 'Office'
    new.form['subgroup'] = 'Big Meeting Room'
    resource = new.form.submit()

    new = resources.click('Raum')
    new.form['title'] = 'Meeting Room 3'
    new.form['group'] = 'Office'
    new.form['subgroup'] = 'Big Meeting Room'
    resource = new.form.submit()

    # name collisions between titles and subgroups are fine
    new = resources.click('Raum')
    new.form['title'] = 'Big Meeting Room'
    new.form['group'] = 'Office'
    resource = new.form.submit()

    new = resources.click('Raum')
    new.form['title'] = 'Concert Hall'
    resource = new.form.submit()

    resources = client.get('/resources')
    assert 'Besprechungsraum 1' in resources
    assert 'Meeting Room 2' in resources
    assert 'Meeting Room 3' in resources
    assert 'Big Meeting Room' in resources
    assert 'Office' in resources
    assert 'Allgemein' in resources
    assert 'Beamers' in resources

    # Check warning duplicate
    duplicate = resources.click('Raum')
    duplicate.form['title'] = 'Meeting Room 1'
    page = new.form.submit()
    assert "Eine Resource mit diesem Namen existiert bereits" in page

    resource = client.get('/resource/meeting-room-1')
    delete_link = resource.pyquery('a.delete-link').attr('ic-delete-from')

    assert client.delete(delete_link, status=200)


def test_resources_person_link_extension(client: Client) -> None:
    client.login_admin()

    # add person
    people_page = client.get('/people')
    new_person_page = people_page.click('Person')
    assert "Neue Person" in new_person_page

    new_person_page.form['first_name'] = 'Franz'
    new_person_page.form['last_name'] = 'Müller'
    new_person_page.form['function'] = 'Gemeindeschreiber'

    page = new_person_page.form.submit()
    assert page.location is not None
    person_uuid = page.location.split('/')[-1]
    page = page.follow()
    assert 'Müller Franz' in page

    # add resource
    resources = client.get('/resources')
    new_item = resources.click('Gegenstand')
    new_item.form['title'] = 'Dorf Bike'
    new_item.form['western_name_order'] = False
    new_item.form['people-0-person'] = person_uuid
    resource = new_item.form.submit().follow()
    assert 'Dorf Bike' in resource
    assert 'Müller Franz' in resource

    edit_resource = resource.click('Bearbeiten')
    edit_resource.form['western_name_order'] = True
    resource = edit_resource.form.submit().follow()
    assert 'Franz Müller' in resource


def test_resources_explicitly_link_referenced_files(client: Client) -> None:
    admin = client.spawn()
    admin.login_admin()

    path = module_path('tests.onegov.org', 'fixtures/sample.pdf')
    with open(path, 'rb') as f:
        page = admin.get('/files')
        page.form['file'] = [Upload('Sample.pdf', f.read(), 'application/pdf')]
        page.form.submit()

    pdf_url = (
        admin.get('/files')
        .pyquery('[ic-trigger-from="#button-1"]')
        .attr('ic-get-from')
        .removesuffix('/details')
    )
    pdf_link = f'<a href="{pdf_url}">Sample.pdf</a>'

    editor = client.spawn()
    editor.login_editor()

    # add resource
    resources = editor.get('/resources')
    new_item = resources.click('Gegenstand')
    new_item.form['title'] = 'Dorf Bike'
    new_item.form['text'] = pdf_link
    resource_page = new_item.form.submit().follow()
    assert 'Dorf Bike' in resource_page

    session = client.app.session()
    pdf = FileCollection(session).query().one()
    resource = (
        ResourceCollection(client.app.libres_context).query()
        .filter(Resource.title == 'Dorf Bike').one()
    )
    assert resource.files == [pdf]
    assert pdf.access == 'public'

    resource.access = 'mtan'  # type: ignore[attr-defined]
    session.flush()
    assert pdf.access == 'mtan'

    # link removed
    resource.files = []
    session.flush()
    assert pdf.access == 'secret'


@freeze_time("2020-01-01", tick=True)
def test_find_your_spot(client: Client) -> None:
    client.login_admin()

    resources = client.get('/resources')
    new = resources.click('Raum')
    new.form['title'] = 'Meeting 1'
    new.form['group'] = 'Meeting Rooms'
    new.form.submit().follow()

    # with only one room in the group there should be no room filter
    find_your_spot = client.get('/find-your-spot?group=Meeting+Rooms')
    assert 'Räume' not in find_your_spot
    assert 'An Feiertagen' not in find_your_spot
    assert 'Während Schulferien' not in find_your_spot

    new.form['title'] = 'Meeting 2'
    new.form['group'] = 'Meeting Rooms'
    new.form.submit().follow()

    find_your_spot = client.get('/find-your-spot?group=Meeting+Rooms')
    assert 'Räume' in find_your_spot
    assert 'An Feiertagen' not in find_your_spot
    assert 'Während Schulferien' not in find_your_spot

    # enable holidays
    page = client.get('/holiday-settings')
    page.select_checkbox('cantonal_holidays', "Zürich")
    page.form.submit()

    find_your_spot = client.get('/find-your-spot?group=Meeting+Rooms')
    assert 'Räume' in find_your_spot
    assert 'An Feiertagen' in find_your_spot
    assert 'Während Schulferien' not in find_your_spot

    # enable school holidays
    page = client.get('/holiday-settings')
    page.form['school_holidays'] = '02.01.2020 - 03.01.2020'
    page.form.submit()

    find_your_spot = client.get('/find-your-spot?group=Meeting+Rooms')
    assert 'Räume' in find_your_spot
    assert 'An Feiertagen' in find_your_spot
    assert 'Während Schulferien' in find_your_spot
    find_your_spot.form['start'] = '2020-01-01'
    find_your_spot.form['end'] = '2020-01-06'
    result = find_your_spot.form.submit()

    # the result should skip holidays, i.e. new years day
    # and the user defined holidays from 2nd to 3th of January
    # but then also the weekend of the 4th and 5th of January
    assert '01.01.2020' not in result
    assert '02.01.2020' not in result
    assert '03.01.2020' not in result
    assert '04.01.2020' not in result
    assert '05.01.2020' not in result
    assert '06.01.2020' in result

    find_your_spot.select_checkbox('weekdays', 'Sa')
    find_your_spot.select_checkbox('weekdays', 'So')
    result = find_your_spot.form.submit()
    assert '01.01.2020' not in result
    assert '02.01.2020' not in result
    assert '03.01.2020' not in result
    assert '04.01.2020' in result
    assert '05.01.2020' in result
    assert '06.01.2020' in result

    find_your_spot.form['on_holidays'] = 'yes'
    result = find_your_spot.form.submit()
    assert '01.01.2020' in result
    assert '02.01.2020' not in result
    assert '03.01.2020' not in result
    assert '04.01.2020' in result
    assert '05.01.2020' in result
    assert '06.01.2020' in result

    find_your_spot.form['during_school_holidays'] = 'yes'
    result = find_your_spot.form.submit()
    assert '01.01.2020' in result
    assert '02.01.2020' in result
    assert '03.01.2020' in result
    assert '04.01.2020' in result
    assert '05.01.2020' in result
    assert '06.01.2020' in result

    # create a blocked and an unblocked allocation
    transaction.begin()

    scheduler_1 = (
        ResourceCollection(client.app.libres_context)  # type: ignore[union-attr]
        .by_name('meeting-1')
        .get_scheduler(client.app.libres_context)
    )
    scheduler_1.allocate(
        dates=(
            (datetime(2020, 1, 1), datetime(2020, 1, 1)),
            (datetime(2020, 1, 2), datetime(2020, 1, 2)),
            (datetime(2020, 1, 3), datetime(2020, 1, 3)),
            (datetime(2020, 1, 4), datetime(2020, 1, 4)),
        ),
        whole_day=True,
        partly_available=True
    )

    scheduler_2 = (
        ResourceCollection(client.app.libres_context)  # type: ignore[union-attr]
        .by_name('meeting-2')
        .get_scheduler(client.app.libres_context)
    )
    scheduler_2.allocate(
        dates=(
            (datetime(2020, 1, 1), datetime(2020, 1, 1)),
            (datetime(2020, 1, 2), datetime(2020, 1, 2)),
            (datetime(2020, 1, 3), datetime(2020, 1, 3)),
            (datetime(2020, 1, 5), datetime(2020, 1, 5)),
        ),
        whole_day=True,
        partly_available=True
    )

    transaction.commit()

    # force client to open a fresh session that includes our
    # newly added allocations
    close_all_sessions()

    find_your_spot.form['auto_reserve_available_slots'] = 'for_first_day'
    result = find_your_spot.form.submit()

    result = client.get(
        '/find-your-spot/reservations?group=Meeting+Rooms'
    ).json
    assert len(result['reservations']) == 1
    reservation = result['reservations'][0]
    assert reservation['resource'] == 'meeting-1'
    assert reservation['date'].startswith('2020-01-01')
    assert reservation['time'] == '07:00 - 08:00'

    # resubmitting doesn't add a reservation or change the existing one
    result = find_your_spot.form.submit()

    result = client.get(
        '/find-your-spot/reservations?group=Meeting+Rooms'
    ).json
    assert len(result['reservations']) == 1
    reservation = result['reservations'][0]
    assert reservation['resource'] == 'meeting-1'
    assert reservation['date'].startswith('2020-01-01')
    assert reservation['time'] == '07:00 - 08:00'

    find_your_spot.form['auto_reserve_available_slots'] = 'for_every_day'
    result = find_your_spot.form.submit()

    result = client.get(
        '/find-your-spot/reservations?group=Meeting+Rooms'
    ).json
    assert len(result['reservations']) == 5
    for idx, reservation in enumerate(result['reservations'][:-1]):
        assert reservation['resource'] == 'meeting-1'
        assert reservation['date'].startswith(f'2020-01-0{idx+1}')
        assert reservation['time'] == '07:00 - 08:00'
    # the final reservation only works for the other room
    reservation = result['reservations'][-1]
    assert reservation['resource'] == 'meeting-2'
    assert reservation['date'].startswith('2020-01-05')
    assert reservation['time'] == '07:00 - 08:00'

    # resubmitting doesn't change anything
    result = find_your_spot.form.submit()

    result = client.get(
        '/find-your-spot/reservations?group=Meeting+Rooms'
    ).json
    assert len(result['reservations']) == 5
    for idx, reservation in enumerate(result['reservations'][:-1]):
        assert reservation['resource'] == 'meeting-1'
        assert reservation['date'].startswith(f'2020-01-0{idx+1}')
        assert reservation['time'] == '07:00 - 08:00'
    # the final reservation only works for the other room
    reservation = result['reservations'][-1]
    assert reservation['resource'] == 'meeting-2'
    assert reservation['date'].startswith('2020-01-05')
    assert reservation['time'] == '07:00 - 08:00'

    find_your_spot.form['auto_reserve_available_slots'] = 'for_every_room'
    result = find_your_spot.form.submit()

    result = client.get(
        '/find-your-spot/reservations?group=Meeting+Rooms'
    ).json
    assert len(result['reservations']) == 8
    for idx, reservation in enumerate(result['reservations'][::2]):
        assert reservation['resource'] == 'meeting-1'
        assert reservation['date'].startswith(f'2020-01-0{idx+1}')
        assert reservation['time'] == '07:00 - 08:00'

    for idx, reservation in enumerate(result['reservations'][1:-1:2]):
        assert reservation['resource'] == 'meeting-2'
        assert reservation['date'].startswith(f'2020-01-0{idx+1}')
        assert reservation['time'] == '07:00 - 08:00'
    reservation = result['reservations'][-1]
    assert reservation['resource'] == 'meeting-2'
    assert reservation['date'].startswith('2020-01-05')
    assert reservation['time'] == '07:00 - 08:00'

    # resubmitting also doesn't change anything here
    result = find_your_spot.form.submit()

    result = client.get(
        '/find-your-spot/reservations?group=Meeting+Rooms'
    ).json
    assert len(result['reservations']) == 8
    for idx, reservation in enumerate(result['reservations'][::2]):
        assert reservation['resource'] == 'meeting-1'
        assert reservation['date'].startswith(f'2020-01-0{idx+1}')
        assert reservation['time'] == '07:00 - 08:00'

    for idx, reservation in enumerate(result['reservations'][1:-1:2]):
        assert reservation['resource'] == 'meeting-2'
        assert reservation['date'].startswith(f'2020-01-0{idx+1}')
        assert reservation['time'] == '07:00 - 08:00'
    reservation = result['reservations'][-1]
    assert reservation['resource'] == 'meeting-2'
    assert reservation['date'].startswith('2020-01-05')
    assert reservation['time'] == '07:00 - 08:00'


def test_resource_room_deletion(client: Client) -> None:
    # TicketMessage.create(ticket, request, 'opened')
    resources = ResourceCollection(client.app.libres_context)
    foyer = resources.add('Foyer', 'Europe/Zurich', type='room')

    # Adds allocations and reservations in the past
    add_reservation(
        foyer,
        client.app.session(),
        datetime(2017, 1, 6, 12),
        datetime(2017, 1, 6, 16),
    )
    assert foyer.deletable  # type: ignore[attr-defined]
    transaction.commit()

    client.login_admin()
    page = client.get('/resources').click('Foyer')
    delete_link = page.pyquery('a.delete-link').attr('ic-delete-from')
    assert delete_link
    assert client.delete(delete_link, status=200)

    # check if the tickets have been closed
    tickets = TicketCollection(client.app.session())
    ticket = tickets.query().one()
    assert ticket.state == 'closed'


def test_resource_room_deletion_with_future_reservation_file_and_payment(
    client: Client,
) -> None:
    session = client.app.session()
    resources = ResourceCollection(client.app.libres_context)
    tickets = TicketCollection(session)
    users = UserCollection(session)
    admin = users.query().filter(User.role == 'admin').first()
    assert admin
    files = FileCollection(session)

    daypass = resources.add('Daypass', 'Europe/Zurich', type='daypass')
    add_reservation(
        daypass,
        session,
        datetime(2037, 11, 6, 12),
        datetime(2037, 11, 6, 16),
    )
    readme_file = files.add('readme', b'content.', published=True)
    daypass.files = [readme_file]

    hall = resources.add('Town Hall', 'Europe/Zurich', type='room')
    hall.pricing_method = 'per_item'
    hall.price_per_hour = 999.00
    hall.payment_method = 'manual'
    add_reservation(
        hall,
        session,
        datetime(2038, 11, 6, 12),
        datetime(2038, 11, 6, 16),
    )

    lopper = resources.add('Lopper', 'Europe/Zurich', type='daily-item')
    add_reservation(
        lopper,
        session,
        datetime(2039, 11, 6, 12),
        datetime(2039, 11, 6, 16),
    )

    transaction.commit()

    assert resources.query().count() == 5  # as two are by default
    resource_titles = [r.title for r in resources.query().all()]
    assert 'Daypass' in resource_titles
    assert 'Town Hall' in resource_titles
    assert 'Lopper' in resource_titles
    resource = resources.query().filter_by(title='Daypass').first()
    assert resource and resource.files[0].name == 'readme'

    for ticket in tickets.query().all():
        assert ticket.state == 'open'
        ticket.accept_ticket(admin)

        if 'Town Hall' in ticket.group:
            # ticket.payment = ManualPayment(
            ticket.payment = Payment(
                source='manual',
                amount=Decimal(999),
                currency='CHF',
                state='paid'
            )
            invoice = TicketInvoice(id=uuid4())
            session.add(invoice)
            ticket.invoice = invoice
            invoice_item = ticket.invoice.add(
                text='Town Hall Cleaning',
                group='manual',
                kind='surcharge',
                cost_object='Town Hall',
                unit=Decimal(100)
            )
            invoice_item.payment_date = date.today()
            invoice_item.payments.append(ticket.payment)
            invoice_item.paid = ticket.payment.state == 'paid'

        if 'Daypass' in ticket.group:
            ticket.payment = Payment(
                source='stripe_connect',
                amount=Decimal(10),
                currency='CHF',
                state='open'
            )
            invoice = TicketInvoice(id=uuid4())
            session.add(invoice)
            ticket.invoice = invoice
            invoice_item = ticket.invoice.add(
                text='Daypass Discount',
                group='manual',
                kind='discount',
                cost_object='Daypass',
                unit=Decimal(1)
            )
            invoice_item.payment_date = date.today()
            invoice_item.payments.append(ticket.payment)
            invoice_item.paid = ticket.payment.state == 'paid'

    transaction.commit()

    assert PaymentCollection(session).query().count() == 2
    assert InvoiceCollection(session).query().count() == 2
    assert InvoiceCollection(session).query_items().count() == 2

    tickets = TicketCollection(session)
    for index, ticket in enumerate(tickets.query().all()):
        assert ticket.state == 'pending'
        assert not ticket.snapshot

        if 'Daypass' in ticket.group:
            assert ticket.payment and ticket.payment.paid is False
        if 'Town Hall' in ticket.group:
            assert ticket.payment and ticket.payment.paid is True
        if 'Lopper' in ticket.group:
            assert ticket.payment is None

    # delete resource
    client.login_admin()
    for resource_title in ['Daypass', 'Town Hall', 'Lopper']:
        page = client.get('/resources').click(resource_title)
        delete_link = page.pyquery('a.delete-link').attr('ic-delete-from')
        assert delete_link
        assert client.delete(delete_link, status=200)
        page = client.get('/resources')
        assert resource_title not in page.pyquery('a.list-title').text()

    assert resources.query().count() == 2  # as two are by default

    # check if the tickets have been closed
    tickets = TicketCollection(client.app.session())
    for index, ticket in enumerate(tickets.query().all()):
        assert ticket.state == 'closed'
        assert ticket.snapshot is not None

        if 'Daypass' in ticket.group:
            assert ticket.payment is None
        if 'Town Hall' in ticket.group:
            assert ticket.payment is None
        if 'Lopper' in ticket.group:
            assert ticket.payment is None

    assert PaymentCollection(session).query().count() == 0
    assert InvoiceCollection(session).query().count() == 0
    assert InvoiceCollection(session).query_items().count() == 0


def test_reserved_resources_fields(client: Client) -> None:
    client.login_admin()

    room = client.get('/resources').click('Raum')
    room.form['title'] = 'Meeting Room'
    room.form['definition'] = "Email *= @@@"
    room = room.form.submit()

    assert "'Email' ist ein reservierter Name" in room

    # fieldsets act as a namespace for field names
    room.form['definition'] = "# Title\nEmail *= @@@"
    room = room.form.submit().follow()

    assert "calendar" in room
    assert "Meeting Room" in room


def test_allocations(client: Client) -> None:
    client.login_admin()
    items = client.get('/resources').click('Gegenstand')
    items.form['title'] = 'Beamer'
    items.form.submit().follow()

    # create new beamer allocation
    new = client.get((
        '/resource/beamer/new-rule'
    ))
    new.form['title'] = 'Period 1'
    new.form['start'] = '2015-08-04'
    new.form['end'] = '2015-08-05'
    new.form['items'] = 1
    new.form['item_limit'] = 1
    new.form.submit()

    # view the beamers
    slots = client.get((
        '/resource/beamer/slots'
        '?start=2015-08-04&end=2015-08-05'
    ))

    assert len(slots.json) == 2
    assert slots.json[0]['title'] == "Ganztägig \nVerfügbar"

    # change the beamers
    edit = client.get(client.extract_href(slots.json[0]['actions'][0]))
    edit.form['items'] = 2
    edit.form.submit()

    slots = client.get((
        '/resource/beamer/slots'
        '?start=2015-08-04&end=2015-08-04'
    ))

    assert len(slots.json) == 1
    assert slots.json[0]['title'] == "Ganztägig \n2 Verfügbar"

    # create a new daypass allocation
    new = client.get((
        '/resource/tageskarte/new-rule'
    ))
    new.form['title'] = 'Period 2'
    new.form['start'] = '2015-08-04'
    new.form['end'] = '2015-08-05'
    new.form['daypasses'] = 1
    new.form['daypasses_limit'] = 1
    new.form.submit()

    # view the daypasses
    slots = client.get((
        '/resource/tageskarte/slots'
        '?start=2015-08-04&end=2015-08-05'
    ))

    assert len(slots.json) == 2
    assert slots.json[0]['title'] == "Ganztägig \nVerfügbar"

    # change the daypasses
    edit = client.get(client.extract_href(slots.json[0]['actions'][0]))
    edit.form['daypasses'] = 2
    edit.form.submit()

    slots = client.get((
        '/resource/tageskarte/slots'
        '?start=2015-08-04&end=2015-08-04'
    ))

    assert len(slots.json) == 1
    assert slots.json[0]['title'] == "Ganztägig \n2 Verfügbar"

    # move the existing allocations
    slots = client.get((
        '/resource/tageskarte/slots'
        '?start=2015-08-04&end=2015-08-05'
    ))

    edit = client.get(client.extract_href(slots.json[0]['actions'][0]))
    edit.form['date'] = '2015-08-06'
    edit.form.submit()

    edit = client.get(client.extract_href(slots.json[1]['actions'][0]))
    edit.form['date'] = '2015-08-07'
    edit.form.submit()

    # get the new slots
    slots = client.get((
        '/resource/tageskarte/slots'
        '?start=2015-08-06&end=2015-08-07'
    ))

    assert len(slots.json) == 2

    # delete an allocation
    client.delete(client.extract_href(slots.json[0]['actions'][2]))

    # get the new slots
    slots = client.get((
        '/resource/tageskarte/slots'
        '?start=2015-08-06&end=2015-08-07'
    ))

    assert len(slots.json) == 1

    # delete an allocation
    client.delete(client.extract_href(slots.json[0]['actions'][2]))

    # get the new slots
    slots = client.get((
        '/resource/tageskarte/slots'
        '?start=2015-08-06&end=2015-08-07'
    ))

    assert len(slots.json) == 0


def test_allocation_times(client: Client) -> None:
    client.login_admin()

    new = client.get('/resources').click('Raum')
    new.form['title'] = 'Meeting Room'
    new.form.submit()

    # 12:00 - 00:00
    new = client.get('/resource/meeting-room/new-rule')
    new.form['title'] = 'Period 1'
    new.form['start'] = '2015-08-20'
    new.form['end'] = '2015-08-20'
    new.form['start_time'] = '12:00'
    new.form['end_time'] = '00:00'
    new.form['is_partly_available'] = 'no'
    new.form.submit()

    slots = client.get(
        '/resource/meeting-room/slots?start=2015-08-20&end=2015-08-20'
    )

    assert len(slots.json) == 1
    assert slots.json[0]['start'] == '2015-08-20T12:00:00+02:00'
    assert slots.json[0]['end'] == '2015-08-21T00:00:00+02:00'

    # 00:00 - 02:00
    new = client.get('/resource/meeting-room/new-rule')
    new.form['title'] = 'Period 2'
    new.form['start'] = '2015-08-22'
    new.form['end'] = '2015-08-22'
    new.form['start_time'] = '00:00'
    new.form['end_time'] = '02:00'
    new.form['is_partly_available'] = 'no'
    new.form.submit()

    slots = client.get(
        '/resource/meeting-room/slots?start=2015-08-22&end=2015-08-22'
    )

    assert len(slots.json) == 1
    assert slots.json[0]['start'] == '2015-08-22T00:00:00+02:00'
    assert slots.json[0]['end'] == '2015-08-22T02:00:00+02:00'

    # 12:00 - 00:00 over two days
    new = client.get('/resource/meeting-room/new-rule')
    new.form['title'] = 'Period 3'
    new.form['start'] = '2015-08-24'
    new.form['end'] = '2015-08-25'
    new.form['start_time'] = '12:00'
    new.form['end_time'] = '00:00'
    new.form['is_partly_available'] = 'no'
    new.form.submit()

    slots = client.get(
        '/resource/meeting-room/slots?start=2015-08-24&end=2015-08-25'
    )

    assert len(slots.json) == 2
    assert slots.json[0]['start'] == '2015-08-24T12:00:00+02:00'
    assert slots.json[0]['end'] == '2015-08-25T00:00:00+02:00'
    assert slots.json[1]['start'] == '2015-08-25T12:00:00+02:00'
    assert slots.json[1]['end'] == '2015-08-26T00:00:00+02:00'


def test_allocation_visibility(client: Client) -> None:
    client.login_admin()

    new = client.get('/resources').click('Raum')
    new.form['title'] = 'Meeting Room'
    new.form.submit()

    # 12:00 - 14:00 private
    new = client.get('/resource/meeting-room/new-rule')
    new.form['title'] = 'Period 1'
    new.form['start'] = '2015-08-20'
    new.form['end'] = '2015-08-20'
    new.form['start_time'] = '12:00'
    new.form['end_time'] = '14:00'
    new.form['is_partly_available'] = 'no'
    new.form['access'] = 'private'
    new.form.submit()

    # 14:00 - 16:00 member
    new = client.get('/resource/meeting-room/new-rule')
    new.form['title'] = 'Period 2'
    new.form['start'] = '2015-08-20'
    new.form['end'] = '2015-08-20'
    new.form['start_time'] = '14:00'
    new.form['end_time'] = '16:00'
    new.form['is_partly_available'] = 'no'
    new.form['access'] = 'member'
    new.form.submit()

    # 16:00 - 18:00 public
    new = client.get('/resource/meeting-room/new-rule')
    new.form['title'] = 'Period 3'
    new.form['start'] = '2015-08-20'
    new.form['end'] = '2015-08-20'
    new.form['start_time'] = '16:00'
    new.form['end_time'] = '18:00'
    new.form['is_partly_available'] = 'no'
    new.form['access'] = 'public'
    new.form.submit()

    slots = client.get(
        '/resource/meeting-room/slots?start=2015-08-20&end=2015-08-20'
    )

    assert len(slots.json) == 3
    assert slots.json[0]['start'] == '2015-08-20T12:00:00+02:00'
    assert slots.json[0]['end'] == '2015-08-20T14:00:00+02:00'
    assert slots.json[1]['start'] == '2015-08-20T14:00:00+02:00'
    assert slots.json[1]['end'] == '2015-08-20T16:00:00+02:00'
    assert slots.json[2]['start'] == '2015-08-20T16:00:00+02:00'
    assert slots.json[2]['end'] == '2015-08-20T18:00:00+02:00'

    client.logout()
    client.login_member()

    slots = client.get(
        '/resource/meeting-room/slots?start=2015-08-20&end=2015-08-20'
    )

    assert len(slots.json) == 2
    assert slots.json[0]['start'] == '2015-08-20T14:00:00+02:00'
    assert slots.json[0]['end'] == '2015-08-20T16:00:00+02:00'
    assert slots.json[1]['start'] == '2015-08-20T16:00:00+02:00'
    assert slots.json[1]['end'] == '2015-08-20T18:00:00+02:00'

    client.logout()

    slots = client.get(
        '/resource/meeting-room/slots?start=2015-08-20&end=2015-08-20'
    )

    assert len(slots.json) == 1
    assert slots.json[0]['start'] == '2015-08-20T16:00:00+02:00'
    assert slots.json[0]['end'] == '2015-08-20T18:00:00+02:00'


def test_allocation_holidays(client: Client) -> None:
    client.login_admin()

    page = client.get('/holiday-settings')
    page.select_checkbox('cantonal_holidays', "Zürich")
    page.form.submit()

    # allocations that are made during holidays
    page = client.get('/resources').click('Raum')
    page.form['title'] = 'Foo'
    page.form.submit()

    new = client.get('/resource/foo/new-rule')
    new.form['title'] = 'Period 1'
    new.form['start'] = '2019-07-30'
    new.form['end'] = '2019-08-02'
    new.form['start_time'] = '07:00'
    new.form['end_time'] = '12:00'
    new.form['on_holidays'] = 'yes'
    new.form['is_partly_available'] = 'no'
    new.form.submit()

    slots = client.get('/resource/foo/slots?start=2019-07-29&end=2019-08-03')

    assert len(slots.json) == 4
    assert slots.json[0]['start'].startswith('2019-07-30')
    assert slots.json[1]['start'].startswith('2019-07-31')
    assert slots.json[2]['start'].startswith('2019-08-01')
    assert slots.json[3]['start'].startswith('2019-08-02')

    # allocations that are not made during holidays
    page = client.get('/resources').click('Raum')
    page.form['title'] = 'Bar'
    page.form.submit()

    new = client.get('/resource/bar/new-rule')
    new.form['title'] = 'Period 2'
    new.form['start'] = '2019-07-30'
    new.form['end'] = '2019-08-02'
    new.form['start_time'] = '07:00'
    new.form['end_time'] = '12:00'
    new.form['on_holidays'] = 'no'
    new.form['is_partly_available'] = 'no'
    new.form.submit()

    slots = client.get('/resource/bar/slots?start=2019-07-29&end=2019-08-03')

    assert len(slots.json) == 3
    assert slots.json[0]['start'].startswith('2019-07-30')
    assert slots.json[1]['start'].startswith('2019-07-31')
    assert slots.json[2]['start'].startswith('2019-08-02')


def test_allocation_school_holidays(client: Client) -> None:
    client.login_admin()

    page = client.get('/holiday-settings')
    page.form['school_holidays'] = '31.07.2019 - 01.08.2019'
    page.form.submit()

    # allocations that are made during holidays
    page = client.get('/resources').click('Raum')
    page.form['title'] = 'Foo'
    page.form.submit()

    new = client.get('/resource/foo/new-rule')
    new.form['title'] = 'Period 1'
    new.form['start'] = '2019-07-30'
    new.form['end'] = '2019-08-02'
    new.form['start_time'] = '07:00'
    new.form['end_time'] = '12:00'
    new.form['during_school_holidays'] = 'yes'
    new.form['is_partly_available'] = 'no'
    new.form.submit()

    slots = client.get('/resource/foo/slots?start=2019-07-29&end=2019-08-03')

    assert len(slots.json) == 4
    assert slots.json[0]['start'].startswith('2019-07-30')
    assert slots.json[1]['start'].startswith('2019-07-31')
    assert slots.json[2]['start'].startswith('2019-08-01')
    assert slots.json[3]['start'].startswith('2019-08-02')

    # allocations that are not made during holidays
    page = client.get('/resources').click('Raum')
    page.form['title'] = 'Bar'
    page.form.submit()

    new = client.get('/resource/bar/new-rule')
    new.form['title'] = 'Period 2'
    new.form['start'] = '2019-07-30'
    new.form['end'] = '2019-08-02'
    new.form['start_time'] = '07:00'
    new.form['end_time'] = '12:00'
    new.form['during_school_holidays'] = 'no'
    new.form['is_partly_available'] = 'no'
    new.form.submit()

    slots = client.get('/resource/bar/slots?start=2019-07-29&end=2019-08-03')

    assert len(slots.json) == 2
    assert slots.json[0]['start'].startswith('2019-07-30')
    assert slots.json[1]['start'].startswith('2019-08-02')


@freeze_time("2015-08-28", tick=True)
def test_auto_accept_reservations(client: Client) -> None:
    # prepare the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.definition = 'Note = ___'
    resource.pick_up = 'You can pick it up at the counter'
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True,
        quota=4,
        quota_limit=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    admin_client = client
    admin_client.login_admin()
    settings = admin_client.get('/ticket-settings')
    settings.form['ticket_auto_accepts'] = ['RSV']
    settings.form.submit()

    # create a reservation
    result = reserve(quota=4, whole_day=True)
    assert result.json == {'success': True}

    # and fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    formular.form['note'] = 'Foobar'

    page = formular.form.submit().follow().form.submit().follow()

    assert 'Ihr Anliegen wurde abgeschlossen' in page
    assert 'Die Reservationen wurden angenommen' in page
    assert len(os.listdir(client.app.maildir)) == 1
    message = client.get_email(0)
    assert message is not None
    assert 'Ihre Reservationen wurden bestätigt' in message['Subject']
    assert 'Foobar' in message['TextBody']
    assert message['Attachments']
    _, pdf_content = extract_pdf_info(BytesIO(
        b64decode(message['Attachments'][0]['Content'])
    ))
    assert 'Tageskarte' in pdf_content
    assert '28.08.2015' in pdf_content
    assert 'Ganztägig' in pdf_content

    # close the ticket and check not email is sent
    tickets = client.get('/tickets/ALL/closed')
    assert 'RSV-' in tickets

    # Test display of status page of ticket
    # Generic message, shown when ticket is open or closed
    assert 'Ihre Anfrage wurde erfolgreich abgeschlossen' not in page
    assert 'You can pick it up at the counter' in page


@freeze_time("2015-08-28", tick=True)
def test_blocker_views(client: Client) -> None:
    # prepare the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.definition = 'Note = ___'
    resource.pick_up = 'You can pick it up at the counter'
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 29), datetime(2015, 8, 29)),
        whole_day=True,
        partly_available=True
    )

    add_blocker = client.bound_add_blocker(allocations[0])
    transaction.commit()

    client.login_admin()

    # create a blocker
    result = add_blocker(whole_day=True, reason='Cleaning')
    assert result.json == {'success': True}

    blocker = scheduler.managed_blockers().one()
    assert blocker.reason == 'Cleaning'

    # change the reason
    client.post(
        f'/reservation-blocker/{blocker.resource}/{blocker.id}/set-reason'
        f'?blocker-id={blocker.id}&reason=No+reason'
        f'&csrf-token={client.csrf_token}'
    )

    blocker = scheduler.managed_blockers().one()
    assert blocker.reason == 'No reason'

    # adjust the time
    new_start = blocker.display_start() + timedelta(hours=1)
    new_end = blocker.display_end() - timedelta(hours=1)
    result = client.post(
        f'/reservation-blocker/{blocker.resource}/{blocker.id}/adjust'
        f'?blocker-id={blocker.id}&start={quote(new_start.isoformat())}'
        f'&end={quote(new_end.isoformat())}&csrf-token={client.csrf_token}'
    )
    assert result.json == {'success': True}

    blocker = scheduler.managed_blockers().one()
    assert blocker.display_start() == new_start
    assert blocker.display_end() == new_end

    # delete the blocker
    client.delete(
        f'/reservation-blocker/{blocker.resource}/{blocker.id}'
        f'?csrf-token={client.csrf_token}'
    )
    assert scheduler.managed_blockers().one_or_none() is None


@pytest.mark.parametrize(
    'reject_type', ['reject-all', 'reject-all-with-message']
)
@freeze_time("2015-08-28", tick=True)
@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_reserve_allocation(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client,
    reject_type: str
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.definition = 'Note = ___'
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True,
        quota=4,
        quota_limit=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    result = reserve(quota=4, whole_day=True)
    assert result.json == {'success': True}
    assert result.headers['X-IC-Trigger'] == 'rc-reservations-changed'

    # and fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    formular.form['note'] = 'Foobar'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # make sure the resulting reservation has no session_id set
    ids = [r.session_id for r in scheduler.managed_reservations()]
    assert not any(ids)

    # try to create another reservation the same time
    result = reserve(quota=4, whole_day=True)
    assert result.json == {
        'success': False,
        'message': 'Der gewünschte Zeitraum ist nicht mehr verfügbar.'
    }

    assert result.headers['X-IC-Trigger'] == 'rc-reservation-error'
    assert json.loads(result.headers['X-IC-Trigger-Data']) == result.json

    # try deleting the allocation with the existing reservation
    client.login_admin()

    slots = client.get((
        '/resource/tageskarte/slots'
        '?start=2015-08-28&end=2015-08-28'
    ))

    assert len(slots.json) == 1

    response = client.delete(client.extract_href(slots.json[0]['actions'][3]))
    result = json.loads(response.headers['X-IC-Trigger-Data'])
    assert result == {
        'success': False,
        'message': (
            "Eine bestehende Reservation wäre von der gewünschten Änderung "
            "betroffen."
        )
    }

    # open the created ticket
    client.login_supporter()
    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    assert 'Foobar' in ticket
    assert '28. August 2015' in ticket
    assert '4' in ticket

    # accept it
    assert 'Alle Reservationen annehmen' in ticket
    ticket = ticket.click('Alle Reservationen annehmen').follow()

    assert 'Alle Reservationen annehmen' not in ticket
    assert len(os.listdir(client.app.maildir)) == 2

    message = client.get_email(1)['TextBody']
    assert 'Tageskarte' in message
    assert '28. August 2015' in message
    assert '4' in message

    assert connect.call_count == 1
    assert authenticate.call_count == 1
    assert broadcast.call_count == 1
    assert broadcast.call_args[0][3]['event'] == 'browser-notification'
    assert broadcast.call_args[0][3]['title'] == 'Neues Ticket'
    assert broadcast.call_args[0][3]['created']

    # edit its details
    details = ticket.click('Details bearbeiten')
    details.form['note'] = '0xdeadbeef'
    ticket = details.form.submit().follow()

    assert '0xdeadbeef' in ticket

    # reject it
    assert client.app.session().query(Reservation).count() == 1
    assert client.app.session().query(FormSubmission).count() == 1

    if reject_type == 'reject-all':
        link = ticket.pyquery('a.delete-link')[0].attrib['ic-get-from']
        ticket = client.get(link).follow()
    elif reject_type == 'reject-all-with-message':
        link = ticket.pyquery('a.delete-link')[1].attrib['href']
        comment = client.get(link)
        comment.form['text'] = 'Sorry!'
        ticket = comment.form.submit().follow()
    else:
        raise AssertionError('Unknown reject type')

    assert client.app.session().query(Reservation).count() == 0
    assert client.app.session().query(FormSubmission).count() == 0

    assert "Der hinterlegte Datensatz wurde entfernt" in ticket
    assert '28. August 2015' in ticket
    assert '4' in ticket
    assert '0xdeadbeef' in ticket

    assert len(os.listdir(client.app.maildir)) == 3

    message = client.get_email(2)['TextBody']
    assert 'Tageskarte' in message
    assert '28. August 2015' in message
    assert '4' in message

    # close the ticket
    ticket.click('Ticket abschliessen')

    assert len(os.listdir(client.app.maildir)) == 4


@freeze_time("2015-08-28", tick=True)
@patch('onegov.websockets.integration.connect')
@patch('onegov.websockets.integration.authenticate')
@patch('onegov.websockets.integration.broadcast')
def test_reserve_with_tag(
    broadcast: MagicMock,
    authenticate: MagicMock,
    connect: MagicMock,
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.definition = 'Note = ___'
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True,
        quota=4,
        quota_limit=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    client.login_admin()

    settings = client.get('/ticket-settings')
    settings.form['ticket_tags'] = (
        '- Test 1:\n'
        '    Note: Default Note\n'
        '    Foo: Bar\n'
        '- Test 2:\n'
        '    Foo: Baz\n'
    )
    settings.form.submit().follow()

    # create a reservation
    result = reserve(quota=4, whole_day=True)
    assert result.json == {'success': True}
    assert result.headers['X-IC-Trigger'] == 'rc-reservations-changed'

    # and fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['ticket_tag'] = 'Test 1'
    formular.form['email'] = 'info@example.org'
    formular.form['note'] = 'Foobar'
    formular.form.submit().follow().form.submit().follow()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert 'RSV-' in ticket.text
    assert '<dt>Foo</dt><dd>Bar</dd>' in ticket.text
    assert '<dt>Foo</dt><dd>Baz</dd>' not in ticket.text
    assert 'Test 1' in ticket.text
    assert 'Test 2' not in ticket.text
    assert 'Foobar' in ticket.text
    assert 'Default Note' not in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # make sure the resulting reservation has no session_id set
    ids = [r.session_id for r in scheduler.managed_reservations()]
    assert not any(ids)

    # change the selected tag
    change_tag = ticket.click(href='change-tag')
    change_tag.form['tag'] = 'Test 2'
    ticket = change_tag.form.submit().follow()
    assert '<dt>Foo</dt><dd>Bar</dd>' not in ticket.text
    assert '<dt>Foo</dt><dd>Baz</dd>' in ticket.text
    assert 'Test 1' not in ticket.text
    assert 'Test 2' in ticket.text
    assert 'Foobar' in ticket.text
    assert 'Default Note' not in ticket.text


@freeze_time("2015-08-28", tick=True)
def test_reserve_allocation_partially(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28, 10), datetime(2015, 8, 28, 14)),
        whole_day=False,
        partly_available=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve('10:00', '12:00').json == {'success': True}

    # fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # open the created ticket
    client.login_admin()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    assert "info@example.org" in ticket
    assert "10:00" in ticket
    assert "12:00" in ticket

    # accept it
    ticket = ticket.click('Alle Reservationen annehmen').follow()

    message = client.get_email(1)['TextBody']
    assert "Tageskarte" in message
    assert "28. August 2015" in message
    assert "10:00" in message
    assert "12:00" in message

    # see if the slots are partitioned correctly
    url = '/resource/tageskarte/slots?start=2015-08-01&end=2015-08-30'
    slots = client.get(url).json
    assert slots[0]['partitions'] == [[50.0, True], [50.0, False]]


@freeze_time("2015-08-28", tick=True)
def test_reserve_allocation_change_email(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28, 10), datetime(2015, 8, 28, 14)),
        whole_day=False,
        partly_available=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve('10:00', '12:00').json == {'success': True}

    # fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # open the created ticket
    client.login_admin()
    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert ticket.pyquery('.ticket-submitter-email a').text() == (
        'info@example.org')

    # change the email
    client.post(
        ticket.pyquery('.ticket-submitter-email').attr('data-edit'),
        {'email': 'new@example.org'}
    )
    ticket = client.get(ticket.request.url)
    assert 'E-Mail Adresse des Antragstellers geändert' in ticket
    assert ticket.pyquery('.ticket-submitter-email a').text() == (
        'new@example.org')


# NOTE: We're at UTC+2 in summertime in Europe/Zurich
@freeze_time("2015-08-28 07:15:00", tick=True)
def test_reserve_allocation_adjustment_pre_acceptance(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28, 9), datetime(2015, 8, 28, 13)),
        whole_day=False,
        partly_available=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve('10:00', '12:00').json == {'success': True}

    # fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # open the created ticket
    client.login_admin()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    assert "info@example.org" in ticket
    assert "10:00" in ticket
    assert "12:00" in ticket
    assert "Anpassen" in ticket

    # try an invalid adjustment
    adjust = ticket.click('Anpassen', index=0)
    adjust.form['end_time'] = '13:05'
    adjust = adjust.form.submit()
    assert 'Zeitraum liegt ausserhalb' in adjust

    # try another invalid adjustment
    adjust.form['start_time'] = '09:00'
    adjust.form['end_time'] = '12:00'
    adjust = adjust.form.submit()
    assert 'kann nicht in die Vergangenheit verschoben werden' in adjust

    # adjust it (valid this time)
    adjust.form['start_time'] = '10:00'
    adjust.form['end_time'] = '11:00'
    ticket = adjust.form.submit().follow()
    assert "10:00" in ticket
    assert "11:00" in ticket

    # accept it
    ticket = ticket.click('Alle Reservationen annehmen').follow()

    message = client.get_email(1)['TextBody']
    assert "Tageskarte" in message
    assert "28. August 2015" in message
    assert "10:00" in message
    assert "11:00" in message

    # see if the slots are partitioned correctly
    url = '/resource/tageskarte/slots?start=2015-08-01&end=2015-08-30'
    slots = client.get(url).json
    assert slots[0]['partitions'] == [
        [25.0, False], [25.0, True], [50.0, False]
    ]


@freeze_time("2015-08-28", tick=True)
def test_reserve_allocation_adjustment_post_acceptance(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28, 10), datetime(2015, 8, 28, 14)),
        whole_day=False,
        partly_available=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve('10:00', '12:00').json == {'success': True}

    # fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # open the created ticket
    client.login_admin()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    assert "info@example.org" in ticket
    assert "10:00" in ticket
    assert "12:00" in ticket
    assert "Anpassen" in ticket

    # accept it
    ticket = ticket.click('Alle Reservationen annehmen').follow()

    # it can still be adjusted
    assert 'Anpassen' in ticket
    adjust = ticket.click('Anpassen', index=0)
    # but there is a warning
    assert 'Diese Reservation wurde bereits angenommen' in adjust
    adjust.form['start_time'] = '10:00'
    adjust.form['end_time'] = '11:00'
    ticket = adjust.form.submit().follow()
    assert "10:00" in ticket
    assert "11:00" in ticket

    # see if the slots are partitioned correctly
    url = '/resource/tageskarte/slots?start=2015-08-01&end=2015-08-30'
    slots = client.get(url).json
    assert slots[0]['partitions'] == [[25.0, True], [75.0, False]]


@freeze_time("2015-08-28", tick=True)
def test_reserve_allocation_adjustment_invoice_change(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_hour'
    resource.price_per_hour = 10.00
    resource.payment_method = 'manual'
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28, 10), datetime(2015, 8, 28, 14)),
        whole_day=False,
        partly_available=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve('10:00', '12:00').json == {'success': True}

    # fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # open the created ticket
    client.login_admin()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    assert "info@example.org" in ticket
    assert "10:00" in ticket
    assert "12:00" in ticket
    assert "Anpassen" in ticket
    assert "20.00" in ticket

    # adjust it
    adjust = ticket.click('Anpassen', index=0)
    adjust.form['start_time'] = '10:00'
    adjust.form['end_time'] = '11:00'
    ticket = adjust.form.submit().follow()
    assert "10:00" in ticket
    assert "11:00" in ticket
    assert "10.00" in ticket

    # mark as paid
    assert ticket.pyquery('.payment-state').text() == "Offen"
    client.post(ticket.pyquery('.mark-as-paid').attr('ic-post-to'))
    ticket = client.get(ticket.request.url)

    # try to adjust it again (invalid since it changes the price)
    adjust = ticket.click('Anpassen', index=0)
    adjust.form['start_time'] = '10:00'
    adjust.form['end_time'] = '12:00'
    adjust = adjust.form.submit()
    assert 'die Zahlung ist nicht mehr offen' in adjust

    # try a valid adjustment that doesn't change the price
    adjust.form['start_time'] = '11:00'
    adjust.form['end_time'] = '12:00'
    ticket = adjust.form.submit().follow()
    assert "11:00" in ticket
    assert "12:00" in ticket
    assert "10.00" in ticket


@freeze_time("2015-08-28", tick=True)
def test_reserve_allocation_add_reservation_invoice_change(
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('konferenzraum')
    assert resource is not None
    resource.pricing_method = 'per_hour'
    resource.price_per_hour = 10.00
    resource.payment_method = 'manual'
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=[
            (datetime(2015, 8, 28, 10), datetime(2015, 8, 28, 12)),
            (datetime(2015, 8, 28, 12), datetime(2015, 8, 28, 14)),
            (datetime(2015, 8, 28, 14), datetime(2015, 8, 28, 16)),
        ],
        whole_day=False,
        partly_available=False
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve('10:00', '12:00').json == {'success': True}

    # fill out the form
    formular = client.get('/resource/konferenzraum/form')
    formular.form['email'] = 'info@example.org'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # open the created ticket
    client.login_admin()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert "info@example.org" in ticket
    assert "10:00" in ticket
    assert "12:00" in ticket
    assert "20.00" in ticket

    # accept it
    ticket = ticket.click('Alle Reservationen annehmen').follow()

    # try to add a second reservation (invalid time range)
    add = ticket.click('Reservation hinzufügen', index=0)
    add.form['date'] = '2015-08-28'
    add.form['whole_day'] = 'yes'
    assert 'kann nicht teilweise reserviert werden' in add.form.submit()

    # actually add a second reservation
    add = ticket.click('Reservation hinzufügen', index=0)
    add.form['date'] = '2015-08-28'
    add.form['whole_day'] = 'no'
    add.form['start_time'] = '12:00'
    add.form['end_time'] = '14:00'
    ticket = add.form.submit().follow()
    assert "10:00" in ticket
    assert "12:00" in ticket
    assert "14:00" in ticket
    assert "40.00" in ticket
    # the new reservation should already have been accepted
    assert 'Alle Reservationen annehmen' not in ticket

    # mark as paid
    assert ticket.pyquery('.payment-state').text() == "Offen"
    client.post(ticket.pyquery('.mark-as-paid').attr('ic-post-to'))
    ticket = client.get(ticket.request.url)

    # try to add a third reservation
    add = ticket.click('Reservation hinzufügen', index=0)
    add.form['date'] = '2015-08-28'
    add.form['whole_day'] = 'no'
    add.form['start_time'] = '14:00'
    add.form['end_time'] = '16:00'
    assert 'die Zahlung ist nicht mehr offen' in add.form.submit()


@freeze_time("2015-08-28", tick=True)
@patch('onegov.pay.models.payment.Payment.sync')
def test_reject_reservation_price_change(
    sync: MagicMock,
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_item'
    resource.price_per_item = 10.00
    resource.payment_method = 'manual'
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28, 10), datetime(2015, 8, 28, 14)),
        whole_day=False,
        partly_available=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create three reservations
    assert reserve('10:00', '10:30').json == {'success': True}
    assert reserve('11:00', '11:30').json == {'success': True}
    assert reserve('12:00', '12:30').json == {'success': True}

    # and fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # open the created ticket
    client.login_supporter()
    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert '28. August 2015' in ticket
    assert '10:00' in ticket
    assert '10:30' in ticket
    assert '11:00' in ticket
    assert '11:30' in ticket
    assert '12:00' in ticket
    assert '12:30' in ticket
    assert '30.00' in ticket

    # reject the final reservation
    assert client.app.session().query(Reservation).count() == 3
    link = ticket.pyquery('a.delete-link')[-1].attrib['ic-get-from']
    ticket = client.get(link).follow()

    assert client.app.session().query(Reservation).count() == 2
    assert '10:00' in ticket
    assert '10:30' in ticket
    assert '11:00' in ticket
    assert '11:30' in ticket
    assert '20.00' in ticket

    # mark as paid
    assert ticket.pyquery('.payment-state').text() == "Offen"
    client.post(ticket.pyquery('.mark-as-paid').attr('ic-post-to'))
    ticket = client.get(ticket.request.url)

    # try to delete again (payment needs to be open or refunded)
    link = ticket.pyquery('a.delete-link')[-1].attrib['ic-get-from']
    ticket = client.get(link).follow()
    assert 'Zahlung muss rückgängig gemacht werden' in ticket

    # but if it changes the price even an open payment should
    # result it in an error, if it's not a manual payment
    payment = client.app.session().query(Payment).one()
    # HACK: change type of payment to generic so it isn't treated
    #       like a manual payment, even though it's also not an
    #       online payment, we monkeypatch `Payment.sync` so it
    #       doesn't raise. We would still need to monkeypatch
    #       the `sync` of `StripePayment` & co. so this seems safer
    payment.source = 'generic'
    payment.state = 'open'
    with warnings.catch_warnings():
        warnings.filterwarnings(
            action='ignore',
            message=r'Flushing object',
            category=exc.SAWarning
        )
        transaction.commit()
    close_all_sessions()
    link = ticket.pyquery('a.delete-link')[-1].attrib['ic-get-from']
    ticket = client.get(link).follow()
    assert 'die Zahlung ist nicht mehr offen' in ticket


@freeze_time("2015-08-28", tick=True)
def test_send_reservation_summary(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(
            (datetime(2015, 8, 28, 10), datetime(2015, 8, 28, 14)),
            (datetime(2015, 8, 29, 10), datetime(2015, 8, 29, 14)),
        ),
        whole_day=False,
        partly_available=True
    )

    reserve1 = client.bound_reserve(allocations[0])
    reserve2 = client.bound_reserve(allocations[1])
    transaction.commit()

    # create a reservation
    assert reserve1('10:00', '12:00').json == {'success': True}
    assert reserve2('10:30', '12:00').json == {'success': True}

    # fill out the form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    # open the created ticket
    client.login_admin()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    assert "info@example.org" in ticket
    assert "28. August 2015" in ticket
    assert "29. August 2015" in ticket
    assert "10:00" in ticket
    assert "12:00" in ticket
    assert "Anpassen" in ticket

    # adjust the first reservation
    adjust = ticket.click('Anpassen', index=0)
    adjust.form['start_time'] = '10:00'
    adjust.form['end_time'] = '11:00'
    ticket = adjust.form.submit().follow()
    assert "10:00" in ticket
    assert "11:00" in ticket

    # then adjust it back to the original state
    adjust = ticket.click('Anpassen', index=0)
    adjust.form['start_time'] = '10:00'
    adjust.form['end_time'] = '12:00'
    ticket = adjust.form.submit().follow()

    # reject the second reservation
    ticket = ticket.click('Absagen', index=1).follow()
    assert len(os.listdir(client.app.maildir)) == 2

    # send a summary
    ticket = client.get(
        ticket.pyquery('a.envelope')[0].attrib['ic-get-from']
    ).follow()
    assert 'Erfolgreich 1 E-Mails gesendet' in ticket
    assert len(os.listdir(client.app.maildir)) == 3

    message = client.get_email(2)['TextBody']
    assert "Tageskarte" in message
    assert "10:00 - 12:00  | → |" not in message
    assert "10:30 - 12:00  | → | Abgelehnt" in message


@freeze_time("2015-08-28", tick=True)
def test_reserve_no_definition_pick_up_hint(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True,
        quota=4,
        quota_limit=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    result = reserve(quota=4)
    assert result.json == {'success': True}

    # fill out the reservation form
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'

    ticket = formular.form.submit().follow().form.submit().follow()

    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1


@freeze_time("2022-10-30", tick=True)
def test_reserve_allocation_dst_to_st_transition(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2022, 10, 30, 0), datetime(2022, 10, 30, 10)),
        whole_day=False,
        partly_available=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve('02:00', '03:00').json == {'success': True}

    # see if the reservation was created correctly
    url = '/resource/tageskarte/reservations?start=2022-10-30&end=2022-10-30'
    reservations = client.get(url).json

    assert len(reservations['reservations']) == 1
    reservation = reservations['reservations'][0]
    assert reservation['time'] == '02:00 - 03:00'
    # for this ambiguous time we should get ST and not DST
    assert reservation['date'] == '2022-10-30T02:00:00+01:00'


@freeze_time("2022-03-27", tick=True)
def test_reserve_allocation_st_to_dst_transition(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2022, 3, 27, 0), datetime(2022, 3, 27, 10)),
        whole_day=False,
        partly_available=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve('02:00', '03:00').json == {
        'message': "Die ausgewählte Zeit existiert nicht an diesem Datum "
                   "aufgrund der Sommerzeitumstellung.",
        'success': False
    }


def test_reserve_in_past(client: Client) -> None:
    admin = client.spawn()
    admin.login_admin()

    editor = client.spawn()
    editor.login_editor()

    transaction.begin()

    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2019, 4, 3), datetime(2019, 4, 3)),
        whole_day=True,
        quota=4,
        quota_limit=4
    )

    reserve_as_anonymous = client.bound_reserve(allocations[0])
    reserve_as_admin = admin.bound_reserve(allocations[0])
    reserve_as_editor = editor.bound_reserve(allocations[0])

    transaction.commit()

    # create a reservation
    assert reserve_as_anonymous().json == {
        'message': "Dieses Datum liegt in der Vergangenheit",
        'success': False
    }

    assert reserve_as_admin().json == {'success': True}
    assert reserve_as_editor().json == {'success': True}


@freeze_time("2015-08-28", tick=True)
def test_reserve_confirmation_no_definition(client: Client) -> None:
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True,
        quota=4,
        quota_limit=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve(quota=4).json == {'success': True}

    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = "info@example.org"

    confirmation = formular.form.submit().follow()

    assert "Bestätigen Sie Ihre Reservation" in confirmation
    assert "Ganztägig" in confirmation
    assert "4" in confirmation
    assert "info@example.org" in confirmation

    formular = confirmation.click("Bearbeiten")
    assert "info@example.org" in formular
    assert "4" in formular

    formular.form['email'] = "changed@example.org"
    confirmation = formular.form.submit().follow()

    assert "Bestätigen Sie Ihre Reservation" in confirmation
    assert "Ganztägig" in confirmation
    assert "changed@example.org" in confirmation


@freeze_time("2015-08-28", tick=True)
def test_reserve_confirmation_with_definition(client: Client) -> None:
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.definition = "Vorname *= ___\nNachname *= ___"

    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28, 10), datetime(2015, 8, 28, 12)),
        whole_day=False,
        partly_available=True
    )
    reserve = client.bound_reserve(allocations[0])

    transaction.commit()

    # create a reservation
    assert reserve("10:30", "12:00").json == {'success': True}

    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = "info@example.org"
    formular.form['vorname'] = "Thomas"
    formular.form['nachname'] = "Anderson"

    confirmation = formular.form.submit().follow()
    assert "10:30" in confirmation
    assert "12:00" in confirmation
    assert "Thomas" in confirmation
    assert "Anderson" in confirmation

    # edit the reservation
    formular = confirmation.click("Bearbeiten")
    formular.form['vorname'] = "Elliot"
    formular.form['nachname'] = "Alderson"

    confirmation = formular.form.submit().follow()
    assert "10:30" in confirmation
    assert "12:00" in confirmation
    assert "Elliot" in confirmation
    assert "Alderson" in confirmation


@freeze_time("2015-08-28", tick=True)
def test_reserve_session_bound(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True,
        quota=4,
        quota_limit=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve(quota=4).json == {'success': True}

    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'

    confirm = formular.form.submit().follow()
    complete_url = confirm.pyquery('form:last').attr('action')

    # make sure the finalize step can only be called by the original client
    c2 = client.spawn()

    assert c2.post(complete_url, expect_errors=True).status_code == 403
    assert client.post(complete_url).follow().status_code == 200


@freeze_time("2015-08-28", tick=True)
def test_delete_reservation_anonymous(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True,
        quota=4,
        quota_limit=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve(quota=4).json == {'success': True}

    # get the delete url
    reservations_url = '/resource/tageskarte/reservations'

    reservations = client.get(reservations_url).json['reservations']
    url = reservations[0]['delete']

    # other clients still can't use the link
    assert client.spawn().delete(url, status=403)
    assert len(client.get(reservations_url).json['reservations']) == 1

    # only the original client can
    client.delete(url)
    assert len(client.get(reservations_url).json['reservations']) == 0


@freeze_time("2015-08-28", tick=True)
def test_reserve_in_parallel(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True
    )

    c1 = client.spawn()
    c2 = client.spawn()

    c1_reserve = c1.bound_reserve(allocations[0])
    c2_reserve = c2.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert c1_reserve().json == {'success': True}
    formular = c1.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    f1 = formular.form.submit().follow()

    # create a parallel reservation
    assert c2_reserve().json == {'success': True}
    formular = c2.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    f2 = formular.form.submit().follow()

    # one will win, one will lose
    assert f1.form.submit().status_code == 302
    assert ("Der gewünschte Zeitraum ist nicht mehr verfügbar."
        ) in f2.form.submit().follow()


@freeze_time("2015-08-28", tick=True)
def test_occupancy_view(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    client.login_admin()

    # create a reservation
    assert reserve().json == {'success': True}
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    formular.form.submit().follow().form.submit()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    occupancy = client.get('/resource/tageskarte/occupancy?date=20150828')
    assert occupancy.status_code == 200
    occupancy = client.get(
        '/resource/tageskarte/occupancy-json?start=2015-08-28&end=2015-08-29'
    )
    assert occupancy.status_code == 200
    occupancy = client.get(
        '/resource/tageskarte/occupancy-stats?start=2015-08-28&end=2015-08-29'
    )
    assert occupancy.status_code == 200
    data = occupancy.json
    assert data['count'] == 1
    assert data['pending'] == 0
    assert data['utilization'] == 100.0


def test_occupancy_view_member_access(client: Client) -> None:
    # setup a resource that's visible to members
    client.login_admin()

    resources = client.get('/resources')
    new = resources.click('Raum')
    new.form['title'] = 'test'
    new.form['access'] = 'member'
    new.form.submit().follow()

    # by default members aren't allowed to view occupancy
    client.login_member()
    occupancy = client.get('/resource/test/occupancy', expect_errors=True)
    assert occupancy.status_code == 403

    # but if we explicitly enable it on the resource
    client.login_admin()
    edit = client.get('/resource/test/edit')
    edit.form['occupancy_is_visible_to_members'] = True
    edit.form.submit().maybe_follow()

    # now members should be able to access it
    client.login_member()
    occupancy = client.get('/resource/test/occupancy')
    assert occupancy.status_code == 200

    occupancy = client.get('/resource/test/occupancy-json')
    assert occupancy.status_code == 200
    occupancy = client.get(
        '/resource/test/occupancy-stats?start=2015-08-28&end=2015-08-29'
    )
    assert occupancy.status_code == 200


@freeze_time("2015-08-28", tick=True)
def test_reservation_export_view(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.definition = "Vorname *= ___\nNachname *= ___"

    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    client.login_admin()

    # create a reservation
    assert reserve().json == {'success': True}
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    formular.form['vorname'] = 'Charlie'
    formular.form['nachname'] = 'Carson'
    formular.form.submit().follow().form.submit()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    # at this point, the reservation won't show up in the export
    export = client.get('/resource/tageskarte/export')
    export.form['start'] = date(2015, 8, 28)
    export.form['end'] = date(2015, 8, 28)
    export.form['file_format'] = 'json'
    assert not export.form.submit().json

    # until we confirm the reservation
    ticket.click('Alle Reservationen annehmen')
    charlie = export.form.submit().json[0]

    assert charlie['email'] == 'info@example.org'
    assert charlie['title'] == 'info@example.org, Charlie, Carson'
    assert charlie['start'] == '2015-08-28T00:00:00+02:00'
    assert charlie['end'] == '2015-08-29T00:00:00+02:00'
    assert charlie['ticket'].startswith('RSV-')
    assert charlie['quota'] == 1
    assert charlie['form'] == {'vorname': 'Charlie', 'nachname': 'Carson'}


@freeze_time("2022-09-07", tick=True)
def test_export_all_default_date_range(client: Client) -> None:
    """ Date range in the export form is the current week. (from
    monday to friday)
    """
    client.login_admin()

    export = client.get('/resources/export-all')

    start = export.form['start']
    end = export.form['end']

    actual_start_date = datetime.strptime(start.value__get(), '%Y-%m-%d')
    assert actual_start_date == datetime(2022, 9, 7, 0, 0)
    actual_end_date = datetime.strptime(end.value__get(), '%Y-%m-%d')
    assert actual_end_date == datetime(2022, 9, 9, 0, 0)


@freeze_time("2022-09-05", tick=True)
def test_export_all_default_date_range_from_start_of_week(
    client: Client
) -> None:
    client.login_admin()
    export = client.get('/resources/export-all')
    start = export.form['start']
    end = export.form['end']

    actual_start_date = datetime.strptime(start.value__get(), '%Y-%m-%d')
    assert actual_start_date == datetime(2022, 9, 5, 0, 0)
    actual_end_date = datetime.strptime(end.value__get(), '%Y-%m-%d')
    assert actual_end_date == datetime(2022, 9, 9, 0, 0)


@freeze_time("2022-09-09", tick=True)
def test_export_all_default_date_range_from_end_of_week(
    client: Client
) -> None:
    client.login_admin()
    export = client.get('/resources/export-all')
    start = export.form['start']
    end = export.form['end']

    actual_start_date = datetime.strptime(start.value__get(), '%Y-%m-%d')
    assert actual_start_date == datetime(2022, 9, 9, 0, 0)
    actual_end_date = datetime.strptime(end.value__get(), '%Y-%m-%d')
    assert actual_end_date == datetime(2022, 9, 9, 0, 0)


@freeze_time("2023-08-28", tick=True)
def test_reservation_export_all_view(client: Client) -> None:
    """ Create reservations with two different resources.
        Then export everything to Excel.
        It should create one Worksheet per resource.
    """
    resources = ResourceCollection(client.app.libres_context)
    daypass_resource = resources.by_name('tageskarte')
    assert daypass_resource is not None
    daypass_resource.definition = "Vorname *= ___\nNachname *= ___"

    scheduler = daypass_resource.get_scheduler(client.app.libres_context)
    daypass_allocations = scheduler.allocate(
        dates=(datetime(2023, 8, 28, 12, 0), datetime(2023, 8, 28, 13, 0)),
        whole_day=False
    )

    reserve_daypass = client.bound_reserve(daypass_allocations[0])

    resources.add(
        "Conference room",
        'Europe/Zurich',
        type='room',
        name='conference-room'
    )

    room_resource = resources.by_name('conference-room')
    assert room_resource is not None
    room_resource.definition = "title *= ___"

    room_allocations = room_resource.scheduler.allocate(
        dates=(datetime(2023, 8, 28), datetime(2023, 8, 28)),
        whole_day=True
    )

    reserve_room = client.bound_reserve(room_allocations[0])
    transaction.commit()
    client.login_admin()

    # create all reservations
    assert reserve_daypass().json == {'success': True}
    assert reserve_room().json == {'success': True}

    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    formular.form['vorname'] = 'Charlie'
    formular.form['nachname'] = 'Carson'
    formular.form.submit().follow().form.submit()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    ticket.click('Alle Reservationen annehmen')

    formular = client.get('/resource/conference-room/form')
    formular.form['title'] = 'Room'
    formular.form.submit().follow().form.submit()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    ticket.click('Alle Reservationen annehmen')

    export = client.get('/resources/export-all')
    export.form['start'] = date(2023, 8, 28)
    export.form['end'] = date(2023, 8, 28)

    response = export.form.submit()
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        tmp.write(response.body)

        wb = load_workbook(Path(tmp.name))

        daypass_sheet_name = wb.sheetnames[1]
        daypass_sheet = wb[daypass_sheet_name]

        tab_2 = tuple(daypass_sheet.rows)
        assert tab_2, "Sheet should not be empty"

        assert tab_2[0][0].value == "start"
        assert tab_2[0][1].value == "end"
        assert tab_2[0][2].value == "quota"
        assert tab_2[0][3].value == "email"
        assert tab_2[0][4].value == "ticket"
        assert tab_2[0][5].value == "title"
        assert tab_2[0][6].value == "form_nachname"
        assert tab_2[0][7].value == "form_vorname"

        assert tab_2[1][0].value == "28.08.2023 12:00"
        assert tab_2[1][1].value == "28.08.2023 13:00"
        assert tab_2[1][2].value == int("1")
        assert tab_2[1][3].value == "info@example.org"

        room_sheet_name = wb.sheetnames[0]
        room_sheet = wb[room_sheet_name]

        tab_1 = tuple(room_sheet.rows)
        assert tab_1, "Sheet should not be empty"

        assert tab_1[0][0].value == "start"
        assert tab_1[0][1].value == "end"
        assert tab_1[0][2].value == "quota"
        assert tab_1[0][3].value == "email"
        assert tab_1[0][4].value == "ticket"
        assert tab_1[0][5].value == "title"
        assert tab_1[1][0].value == "28.08.2023 00:00"
        assert tab_1[1][1].value == "29.08.2023 00:00"
        assert tab_1[1][2].value == int("1")
        assert "RSV-" in tab_1[1][4].value  # type: ignore[operator]
        assert "Room" in tab_1[1][5].value  # type: ignore[operator]


@freeze_time("2023-08-28", tick=True)
def test_reservation_export_all_view_normalizes_sheet_names(
    client: Client
) -> None:
    """ Names of Excel worksheets have to be valid.
        Limited to 31 characters, no special characters.
        Duplicate titles will be incremented numerically.
    """

    duplicate_title = normalize_for_url("Gemeindeverwaltung Sitzungszimmer "
                                        "gross (2. OG)")

    resources = ResourceCollection(client.app.libres_context)
    daypass_resource = resources.by_name('tageskarte')
    assert daypass_resource is not None
    daypass_resource.definition = "Vorname *= ___\nNachname *= ___"
    daypass_resource.title = duplicate_title

    scheduler = daypass_resource.get_scheduler(client.app.libres_context)
    daypass_allocations = scheduler.allocate(
        dates=(datetime(2023, 8, 28, 12, 0), datetime(2023, 8, 28, 13, 0)),
        whole_day=False
    )

    reserve_daypass = client.bound_reserve(daypass_allocations[0])

    resources.add(
        "Conference room",
        'Europe/Zurich',
        type='room',
        name='conference-room'
    )

    room_resource = resources.by_name('conference-room')
    assert room_resource is not None
    room_resource.definition = "Name *= ___"
    room_resource.title = duplicate_title

    room_allocations = room_resource.scheduler.allocate(
        dates=(datetime(2023, 8, 28), datetime(2023, 8, 28)),
        whole_day=True
    )

    reserve_room = client.bound_reserve(room_allocations[0])
    transaction.commit()
    client.login_admin()

    # create all reservations
    assert reserve_daypass().json == {'success': True}
    assert reserve_room().json == {'success': True}

    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    formular.form['vorname'] = 'Charlie'
    formular.form['nachname'] = 'Carson'

    formular.form.submit().follow().form.submit()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    ticket.click('Alle Reservationen annehmen')

    formular = client.get('/resource/conference-room/form')
    formular.form['name'] = 'Name'
    formular.form.submit().follow().form.submit()

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    ticket.click('Alle Reservationen annehmen')

    export = client.get('/resources/export-all')
    export.form['start'] = date(2023, 8, 28)
    export.form['end'] = date(2023, 8, 28)

    response = export.form.submit()

    with tempfile.NamedTemporaryFile(suffix='.xlsx') as tmp:
        tmp.write(response.body)
        wb = load_workbook(Path(tmp.name))
        actual_sheet_name_room = wb.sheetnames[0]
        actual_sheet_name_daypass = wb.sheetnames[1]
        assert "sitzungszimmer-gross-2-og" == actual_sheet_name_room
        assert "sitzungszimmer-gross-2-og_1" == actual_sheet_name_daypass


@freeze_time("2022-09-07", tick=True)
def test_reservation_export_all_with_no_resources(client: Client) -> None:
    client.login_admin()

    export = client.get('/resources/export-all')
    export.form['start'] = date(2022, 9, 7)
    export.form['end'] = date(2022, 9, 7)

    # should not export anything, if there's nothing to export
    response = export.form.submit().follow()
    assert "Keine Reservierungen für den angegebenen Zeitraum gefunden"\
           in response


@freeze_time("2016-04-28", tick=True)
def test_reserve_session_separation(client: Client) -> None:
    c1 = client.spawn()
    c1.login_admin()

    c2 = client.spawn()
    c2.login_admin()

    reserve = []

    # check both for separation by resource and by client
    for room in ('meeting-room', 'gym'):
        new = c1.get('/resources').click('Raum')
        new.form['title'] = room
        new.form.submit()

        resource = client.app.libres_resources.by_name(room)
        assert resource is not None
        allocations = resource.scheduler.allocate(
            dates=(datetime(2016, 4, 28, 12, 0), datetime(2016, 4, 28, 13, 0)),
            whole_day=False
        )

        reserve.append(c1.bound_reserve(allocations[0]))
        reserve.append(c2.bound_reserve(allocations[0]))
        transaction.commit()

    c1_reserve_room, c2_reserve_room, c1_reserve_gym, c2_reserve_gym = reserve

    assert c1_reserve_room().json == {'success': True}
    assert c1_reserve_gym().json == {'success': True}
    assert c2_reserve_room().json == {'success': True}
    assert c2_reserve_gym().json == {'success': True}

    for room in ('meeting-room', 'gym'):
        result = c1.get('/resource/{}/reservations'.format(room)).json
        assert len(result['reservations']) == 1

        result = c2.get('/resource/{}/reservations'.format(room)).json
        assert len(result['reservations']) == 1

    # check combined reservations for rooms without a group
    result = c1.get('/find-your-spot/reservations').json
    assert len(result['reservations']) == 2
    assert result['reservations'][0]['resource'] == 'gym'
    assert result['reservations'][1]['resource'] == 'meeting-room'

    result = c2.get('/find-your-spot/reservations').json
    assert len(result['reservations']) == 2
    assert result['reservations'][0]['resource'] == 'gym'
    assert result['reservations'][1]['resource'] == 'meeting-room'

    formular = c1.get('/resource/meeting-room/form')
    formular.form['email'] = 'info@example.org'
    formular.form.submit()

    # we should get the same form by following the group link
    group_formular = c1.get('/find-your-spot/form').follow()
    assert 'meeting-room' in group_formular

    # make sure if we confirm one reservation, only one will be written
    next_form = formular.form.submit().follow().form.submit().follow()

    resource = client.app.libres_resources.by_name('meeting-room')
    assert resource is not None
    assert resource.scheduler.managed_reserved_slots().count() == 1

    result = c1.get('/resource/meeting-room/reservations').json
    assert len(result['reservations']) == 0

    result = c1.get('/resource/gym/reservations').json
    assert len(result['reservations']) == 1

    result = c2.get('/resource/meeting-room/reservations').json
    assert len(result['reservations']) == 1

    result = c2.get('/resource/gym/reservations').json
    assert len(result['reservations']) == 1

    # next_form should now be gym, since we had another pending
    # reservation in the same group (no group i.e. general)
    assert 'Bitte fahren Sie fort mit Ihrer Reservation für gym' in next_form
    # but e-mail shoud be pre-filled so we can just submit twice to reserve
    tickets = next_form.form.submit().follow().form.submit().follow()

    result = c1.get('/resource/meeting-room/reservations').json
    assert len(result['reservations']) == 0

    result = c1.get('/resource/gym/reservations').json
    assert len(result['reservations']) == 0

    result = c2.get('/resource/meeting-room/reservations').json
    assert len(result['reservations']) == 1

    result = c2.get('/resource/gym/reservations').json
    assert len(result['reservations']) == 1

    # we should have a ticket for each room we reserved
    assert 'Eingereichte Anfragen' in tickets
    assert 'meeting-room' in tickets
    assert 'gym' in tickets


@freeze_time("2016-04-28", tick=True)
def test_reserve_related_tickets_email_threading(client: Client) -> None:
    client.login_admin()

    reserve = []

    # check both for separation by resource and by client
    for room in ('meeting-room', 'gym'):
        new = client.get('/resources').click('Raum')
        new.form['title'] = room
        new.form.submit()

        resource = client.app.libres_resources.by_name(room)
        assert resource is not None
        allocations = resource.scheduler.allocate(
            dates=(datetime(2016, 4, 28, 12, 0), datetime(2016, 4, 28, 13, 0)),
            whole_day=False
        )

        reserve.append(client.bound_reserve(allocations[0]))
        transaction.commit()

    reserve_room, reserve_gym = reserve

    assert reserve_room().json == {'success': True}
    assert reserve_gym().json == {'success': True}

    formular = client.get('/resource/meeting-room/form')
    formular.form['email'] = 'info@example.org'
    formular.form.submit()

    # we should get the same form by following the group link
    group_formular = client.get('/find-your-spot/form').follow()
    assert 'meeting-room' in group_formular

    # confirm the first reservation
    next_form = formular.form.submit().follow().form.submit().follow()
    assert len(os.listdir(client.app.maildir)) == 1
    message = client.get_email(0)
    assert message is not None
    assert 'Headers' in message
    headers = {h['Name']: h['Value'] for h in message['Headers']}
    assert 'Message-ID' in headers
    message_id1 = headers['Message-ID']

    # confirm the second reservation
    assert 'Bitte fahren Sie fort mit Ihrer Reservation für gym' in next_form
    # but e-mail shoud be pre-filled so we can just submit twice to reserve
    tickets = next_form.form.submit().follow().form.submit().follow()

    assert len(os.listdir(client.app.maildir)) == 2
    message = client.get_email(1)
    assert message is not None
    assert 'Headers' in message
    headers = {h['Name']: h['Value'] for h in message['Headers']}
    assert 'Message-ID' in headers
    message_id2 = headers['Message-ID']
    assert message_id2 > message_id1
    assert headers['In-Reply-To'] == message_id1
    assert headers['References'] == message_id1

    # we should have a ticket for each room we reserved
    assert 'Eingereichte Anfragen' in tickets
    assert 'meeting-room' in tickets
    assert 'gym' in tickets

    # the two tickets should be linked
    open_tickets = client.get('/tickets/ALL/open')
    ticket = open_tickets.click('Annehmen', index=0).follow()
    assert 'Verknüpfte Tickets' in ticket
    # we can create a multi-ticket pdf
    pdf = ticket.click('Mit verknüpften Tickets')
    # we can mute all the related tickets
    ticket = ticket.click(href='/mute-related').follow()
    # and then unmute all the related tickets
    ticket = ticket.click(href='/unmute-related').follow()

    # trigger a third email in the thread
    ticket = ticket.click('Alle Reservationen annehmen').follow()
    assert len(os.listdir(client.app.maildir)) == 3
    message = client.get_email(2)
    assert 'Headers' in message
    headers = {h['Name']: h['Value'] for h in message['Headers']}
    assert 'Message-ID' in headers
    message_id3 = headers['Message-ID']
    assert message_id3 > message_id2
    assert headers['In-Reply-To'] == message_id2
    assert headers['References'] == f'{message_id1} {message_id2}'


def test_reserve_reservation_prediction(client: Client) -> None:
    client.login_admin()

    new = client.get('/resources').click('Raum')
    new.form['title'] = 'Gym'
    new.form.submit()

    transaction.begin()

    resource = client.app.libres_resources.by_name('gym')
    assert resource is not None

    a1 = resource.scheduler.allocate(
        dates=(datetime(2017, 1, 1, 12, 0), datetime(2017, 1, 1, 13, 0)),
        whole_day=False
    )[0]
    a2 = resource.scheduler.allocate(
        dates=(datetime(2017, 1, 2, 12, 0), datetime(2017, 1, 2, 13, 0)),
        whole_day=False
    )[0]

    reserve_a1 = client.bound_reserve(a1)
    reserve_a2 = client.bound_reserve(a2)

    transaction.commit()

    reserve_a1()
    reserve_a2()

    reservations_url = '/resource/gym/reservations'
    assert not client.get(reservations_url).json['prediction']

    transaction.begin()

    resource = client.app.libres_resources.by_name('gym')
    assert resource is not None
    a3 = resource.scheduler.allocate(
        dates=(datetime(2017, 1, 3, 12, 0), datetime(2017, 1, 3, 13, 0)),
        whole_day=False
    )[0]
    resource.scheduler.allocate(
        dates=(datetime(2017, 1, 4, 12, 0), datetime(2017, 1, 4, 13, 0)),
        whole_day=False
    )

    reserve_a3 = client.bound_reserve(a3)
    transaction.commit()

    reserve_a3()

    prediction = client.get(reservations_url).json['prediction']

    assert prediction['start'] == '2017-01-04T12:00:00+01:00'
    assert prediction['end'] == '2017-01-04T13:00:00+01:00'
    assert prediction['quota'] == 1
    assert prediction['time'] == '12:00 - 13:00'
    assert prediction['url'].endswith('/reserve')
    assert prediction['wholeDay'] is False


def test_reserve_multiple_allocations(client: Client) -> None:
    client.login_admin()

    transaction.begin()

    resource = client.app.libres_resources.by_name('tageskarte')
    assert resource is not None
    thursday = resource.scheduler.allocate(
        dates=(datetime(2016, 4, 28), datetime(2016, 4, 28)),
        whole_day=True
    )[0]
    friday = resource.scheduler.allocate(
        dates=(datetime(2016, 4, 29), datetime(2016, 4, 29)),
        whole_day=True
    )[0]

    reserve_thursday = client.bound_reserve(thursday)
    reserve_friday = client.bound_reserve(friday)

    transaction.commit()

    assert reserve_thursday().json == {'success': True}
    assert reserve_friday().json == {'success': True}

    formular = client.get('/resource/tageskarte/form')
    assert "28. April 2016" in formular
    assert "29. April 2016" in formular
    formular.form['email'] = "info@example.org"

    confirmation = formular.form.submit().follow()
    assert "28. April 2016" in confirmation
    assert "29. April 2016" in confirmation

    ticket = confirmation.form.submit().follow()
    assert 'RSV-' in ticket.text
    assert len(os.listdir(client.app.maildir)) == 1

    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()
    assert "info@example.org" in ticket
    assert "28. April 2016" in ticket
    assert "29. April 2016" in ticket

    # accept it
    ticket.click('Alle Reservationen annehmen')

    message = client.get_email(1)['TextBody']
    assert "Tageskarte" in message
    assert "28. April 2016" in message
    assert "29. April 2016" in message

    # make sure the reservations are no longer pending
    resource = client.app.libres_resources.by_name('tageskarte')
    assert resource is not None

    reservations = resource.scheduler.managed_reservations()
    assert reservations.filter(Reservation.status == 'approved').count() == 2
    assert reservations.filter(Reservation.status == 'pending').count() == 0
    assert resource.scheduler.managed_reserved_slots().count() == 2

    # now deny them
    client.get(ticket.pyquery('a.delete-link')[0].attrib['ic-get-from'])

    message = client.get_email(2)['TextBody']
    assert "Tageskarte" in message
    assert "28. April 2016" in message
    assert "29. April 2016" in message

    # make sure the reservations are now gone, together with the reserved slots
    reservations = resource.scheduler.managed_reservations()
    assert reservations.filter(Reservation.status == 'approved').count() == 0
    assert reservations.filter(Reservation.status == 'pending').count() == 0
    assert resource.scheduler.managed_reserved_slots().count() == 0


def test_reserve_and_deny_multiple_dates(client: Client) -> None:
    client.login_admin()

    transaction.begin()

    resource = client.app.libres_resources.by_name('tageskarte')
    assert resource is not None
    wednesday = resource.scheduler.allocate(
        dates=(datetime(2016, 4, 27), datetime(2016, 4, 27)),
        whole_day=True
    )[0]
    thursday = resource.scheduler.allocate(
        dates=(datetime(2016, 4, 28), datetime(2016, 4, 28)),
        whole_day=True
    )[0]
    friday = resource.scheduler.allocate(
        dates=(datetime(2016, 4, 29), datetime(2016, 4, 29)),
        whole_day=True
    )[0]

    reserve_wednesday = client.bound_reserve(wednesday)
    reserve_thursday = client.bound_reserve(thursday)
    reserve_friday = client.bound_reserve(friday)

    transaction.commit()

    assert reserve_wednesday().json == {'success': True}
    assert reserve_thursday().json == {'success': True}
    assert reserve_friday().json == {'success': True}

    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = "info@example.org"

    confirmation = formular.form.submit().follow()
    ticket = confirmation.form.submit().follow()
    ticket = client.get('/tickets/ALL/open').click('Annehmen').follow()

    # the resource needs to be refetched after the commit
    resource = client.app.libres_resources.by_name('tageskarte')
    assert resource is not None
    assert resource.scheduler.managed_reserved_slots().count() == 3

    # deny the last reservation
    client.get(ticket.pyquery('a.delete-link')[-1].attrib['ic-get-from'])
    assert resource.scheduler.managed_reserved_slots().count() == 2

    message = client.get_email(1)['TextBody']
    assert "abgesagt" in message
    assert "29. April 2016" in message

    # accept the others
    ticket = ticket.click('Alle Reservationen annehmen').follow()
    assert resource.scheduler.managed_reserved_slots().count() == 2

    message = client.get_email(2)['TextBody']
    assert "bestätigt" in message
    assert "27. April 2016" in message
    assert "28. April 2016" in message

    # deny the reservations that were accepted one by one
    client.get(ticket.pyquery('a.delete-link')[-1].attrib['ic-get-from'])
    assert resource.scheduler.managed_reserved_slots().count() == 1

    message = client.get_email(3)['TextBody']
    assert "abgesagt" in message
    assert "27. April 2016" not in message
    assert "28. April 2016" in message

    ticket = client.get(ticket.request.url)
    client.get(ticket.pyquery('a.delete-link')[-1].attrib['ic-get-from'])
    assert resource.scheduler.managed_reserved_slots().count() == 0

    message = client.get_email(4)['TextBody']
    assert "abgesagt" in message
    assert "27. April 2016" in message
    assert "28. April 2016" not in message

    ticket = client.get(ticket.request.url)
    assert "Der hinterlegte Datensatz wurde entfernt" in ticket
    assert "27. April 2016" in message
    assert "28. April 2016" not in message
    assert "29. April 2016" not in message


def test_reserve_failing_multiple(client: Client) -> None:
    c1 = client.spawn()
    c1.login_admin()

    c2 = client.spawn()
    c2.login_admin()

    transaction.begin()

    resource = client.app.libres_resources.by_name('tageskarte')
    assert resource is not None
    thursday = resource.scheduler.allocate(
        dates=(datetime(2016, 4, 28), datetime(2016, 4, 28)),
        whole_day=True
    )[0]
    friday = resource.scheduler.allocate(
        dates=(datetime(2016, 4, 29), datetime(2016, 4, 29)),
        whole_day=True
    )[0]

    c1_reserve_thursday = c1.bound_reserve(thursday)
    c1_reserve_friday = c1.bound_reserve(friday)
    c2_reserve_thursday = c2.bound_reserve(thursday)
    c2_reserve_friday = c2.bound_reserve(friday)

    transaction.commit()

    assert c1_reserve_thursday().json == {'success': True}
    assert c1_reserve_friday().json == {'success': True}
    assert c2_reserve_thursday().json == {'success': True}
    assert c2_reserve_friday().json == {'success': True}

    # accept the first reservation session
    formular = c1.get('/resource/tageskarte/form')
    formular.form['email'] = "info@example.org"
    formular.form.submit().follow().form.submit().follow()

    ticket = c1.get('/tickets/ALL/open').click('Annehmen').follow()
    ticket.click('Alle Reservationen annehmen')

    # then try to accept the second one
    formular = c2.get('/resource/tageskarte/form')
    formular.form['email'] = "info@example.org"
    confirmation = formular.form.submit().follow()
    confirmation = confirmation.form.submit().follow()

    assert 'failed_reservations' in confirmation.request.url
    assert 'class="reservation failed"' in confirmation


def test_cleanup_allocations(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(
            (datetime(2015, 8, 28), datetime(2015, 8, 28)),
            (datetime(2015, 8, 29), datetime(2015, 8, 29)),
            (datetime(2015, 8, 30), datetime(2015, 8, 30)),
        ),
        whole_day=True
    )
    scheduler.reserve(
        'info@example.org', (allocations[0]._start, allocations[0]._end))

    transaction.commit()

    # clean up the data
    client.login_admin()

    cleanup = client.get('/resource/tageskarte').click("Aufräumen")
    cleanup.form['start'] = date(2015, 8, 31)
    cleanup.form['end'] = date(2015, 8, 1)
    cleanup = cleanup.form.submit()

    assert "Das End-Datum muss nach dem Start-Datum liegen" in cleanup

    cleanup.form['start'] = date(2015, 8, 1)
    cleanup.form['end'] = date(2015, 8, 31)
    # only remove fridays and sundays, which excludes the middle allocation
    cleanup.form['weekdays'] = [4, 6]
    resource_page = cleanup.form.submit().follow()

    assert "1 Verfügbarkeiten wurden erfolgreich entfernt" in resource_page

    allocations = scheduler.managed_allocations().order_by('id').all()
    assert len(allocations) == 2
    # not removed due to existing reservation
    assert allocations[0].display_start().date() == date(2015, 8, 28)
    # not removed due to not being on a friday or sunday
    assert allocations[1].display_start().date() == date(2015, 8, 29)


@freeze_time("2017-07-09", tick=True)
def test_manual_reservation_payment_with_one_off_extra(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_item'
    resource.price_per_item = 15.00
    resource.extras_pricing_method = 'one_off'
    resource.payment_method = 'manual'
    resource.definition = textwrap.dedent("""
        Donation =
            (x) Yes (10 CHF)
            ( ) No
    """)

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9),
            datetime(2017, 7, 9)
        ),
        whole_day=True,
        quota=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    reserve(quota=2, whole_day=True)

    page = client.get('/resource/tageskarte/form')
    page.form['email'] = 'info@example.org'

    page.form['donation'] = 'No'
    assert '30.00' in page.form.submit().follow()

    page.form['donation'] = 'Yes'
    assert '40.00' in page.form.submit().follow()

    ticket = page.form.submit().follow().form.submit().follow()
    assert 'RSV-' in ticket.text

    # mark it as paid
    client.login_editor()
    page = client.get('/tickets/ALL/open').click("Annehmen").follow()

    assert page.pyquery('.payment-state').text() == "Offen"

    client.post(page.pyquery('.mark-as-paid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Bezahlt"

    client.post(page.pyquery('.mark-as-unpaid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Offen"

    invoice = page.click('Rechnung anzeigen')
    assert '30.00' in invoice
    assert '10.00' in invoice
    assert '40.00' in invoice
    assert invoice.pyquery('.payment-state').text() == "Offen"

    payments = client.get('/payments')
    assert "RSV-" in payments
    assert "Manuell" in payments
    assert "info@example.org" in payments
    assert "40.00" in payments
    assert "Offen" in payments

    invoices = client.get('/invoices')
    assert "RSV-" in invoices
    assert "info@example.org" in invoices
    assert "40.00" in invoices
    assert "Unbezahlt" in invoices
    assert "Unfakturiert" in invoices

    # add a manual discount
    item = invoice.click('Abzug / Zuschlag')
    item.form['booking_text'] = 'Gratis'
    item.form['discount'] = '40.00'
    invoice = item.form.submit().follow()
    assert '30.00' in invoice
    assert '10.00' in invoice
    assert '40.00' in invoice
    assert '-40.00' in invoice
    assert not invoice.pyquery('.payment-state').text()

    # delete the manual discount
    client.delete(
        invoice.pyquery('.remove-invoice-item').attr('ic-delete-from')
    )
    invoice = client.get(invoice.request.url)
    assert '30.00' in invoice
    assert '10.00' in invoice
    assert '40.00' in invoice
    assert 'Zahlungsmethode' in invoice
    assert invoice.pyquery('.payment-state').text() == "Offen"

    # reject the final reservation (invoice should get deleted)
    delete_url = page.pyquery('a.delete-link')[-1].attrib['ic-get-from']
    page = client.get(delete_url).follow()
    assert 'Rechnung anzeigen' not in page


@freeze_time("2017-07-09", tick=True)
def test_manual_reservation_payment_with_per_item_extra(
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_item'
    resource.price_per_item = 15.00
    resource.extras_pricing_method = 'per_item'
    resource.payment_method = 'manual'
    resource.definition = textwrap.dedent("""
        Donation =
            (x) Yes (10 CHF)
            ( ) No
    """)

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9),
            datetime(2017, 7, 9)
        ),
        whole_day=True,
        quota=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    reserve(quota=2, whole_day=True)

    page = client.get('/resource/tageskarte/form')
    page.form['email'] = 'info@example.org'

    page.form['donation'] = 'No'
    assert '30.00' in page.form.submit().follow()

    page.form['donation'] = 'Yes'
    assert '50.00' in page.form.submit().follow()

    ticket = page.form.submit().follow().form.submit().follow()
    assert 'RSV-' in ticket.text

    client.login_editor()
    page = client.get('/tickets/ALL/open').click("Annehmen").follow()

    assert page.pyquery('.payment-state').text() == "Offen"

    # change the submission (works until it is marked as paid)
    edit_page = page.click('Details bearbeiten')
    edit_page.form['donation'] = 'No'
    page = edit_page.form.submit().follow()
    assert '30.00' in page

    # the invoice should change accordingly
    invoice = page.click('Rechnung anzeigen')
    assert '15.00' in invoice
    assert '2.00' in invoice
    assert '30.00' in invoice

    # let's add a manual invoice item
    item = invoice.click('Abzug / Zuschlag')
    item.form['booking_text'] = 'Preisreduktion'
    item.form['discount'] = '10.00'
    invoice = item.form.submit().follow()
    assert '15.00' in invoice
    assert '2.00' in invoice
    assert '-10.00' in invoice
    assert '20.00' in invoice

    # mark it as paid
    client.post(invoice.pyquery('.mark-as-paid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Bezahlt"

    # try to change it again (it should not work)
    edit_page = page.click('Details bearbeiten')
    edit_page.form['donation'] = 'Yes'
    assert 'die Zahlung ist nicht mehr offen' in edit_page.form.submit()

    # try to add a new reservation (it should not work)
    add_page = page.click('Reservation hinzufügen')
    add_page.form['date'] = '2017-07-09'
    add_page.form['quota_other'] = '2'
    assert 'die Zahlung ist nicht mehr offen' in add_page.form.submit()

    # mark it as unpaid again
    client.post(page.pyquery('.mark-as-unpaid').attr('ic-post-to'))
    page = client.get(page.request.url)

    # and change it back
    edit_page = page.click('Details bearbeiten')
    edit_page.form['donation'] = 'Yes'
    page = edit_page.form.submit().follow()
    assert '40.00' in page
    assert page.pyquery('.payment-state').text() == "Offen"

    # add a new reservation
    add_page = page.click('Reservation hinzufügen')
    add_page.form['date'] = '2017-07-09'
    add_page.form['quota_other'] = '2'
    page = add_page.form.submit().follow()
    assert '90.00' in page

    invoice = page.click('Rechnung anzeigen')
    assert '15.00' in invoice
    assert '10.00' in invoice
    assert '-10.00' in invoice
    assert '2.00' in invoice
    assert '90.00' in invoice

    # delete the manual invoice item
    client.delete(
        invoice.pyquery('.remove-invoice-item').attr('ic-delete-from')
    )
    invoice = client.get(invoice.request.url)
    assert '15.00' in invoice
    assert '10.00' in invoice
    assert '2.00' in invoice
    assert '100.00' in invoice

    page = client.get(page.request.url)
    assert '100.00' in page

    payments = client.get('/payments')
    assert "RSV-" in payments
    assert "Manuell" in payments
    assert "info@example.org" in payments
    assert "100.00" in payments
    assert "Offen" in payments

    invoices = client.get('/invoices')
    assert "RSV-" in invoices
    assert "info@example.org" in invoices
    assert "100.00" in invoices
    assert "Unbezahlt" in invoices
    assert "Unfakturiert" in invoices

    # adding another new reservation should fail
    # since we used up all of the quota
    add_page = page.click('Reservation hinzufügen')
    add_page.form['date'] = '2017-07-09'
    add_page.form['quota_other'] = '1'
    assert ('Der gewünschte Zeitraum ist nicht mehr verfügbar'
        ) in add_page.form.submit()


@freeze_time("2017-07-09", tick=True)
def test_manual_reservation_payment_with_per_hour_extra(
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_item'
    resource.price_per_item = 15.00
    resource.extras_pricing_method = 'per_hour'
    resource.payment_method = 'manual'
    resource.definition = textwrap.dedent("""
        Donation =
            (x) Yes (1 CHF)
            ( ) No
    """)

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9),
            datetime(2017, 7, 9)
        ),
        whole_day=True,
        quota=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    reserve(quota=1, whole_day=True)

    page = client.get('/resource/tageskarte/form')
    page.form['email'] = 'info@example.org'

    page.form['donation'] = 'No'
    assert '15.00' in page.form.submit().follow()

    page.form['donation'] = 'Yes'
    assert '39.00' in page.form.submit().follow()

    ticket = page.form.submit().follow().form.submit().follow()
    assert 'RSV-' in ticket.text

    # mark it as paid
    client.login_editor()
    page = client.get('/tickets/ALL/open').click("Annehmen").follow()

    assert page.pyquery('.payment-state').text() == "Offen"

    client.post(page.pyquery('.mark-as-paid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Bezahlt"

    client.post(page.pyquery('.mark-as-unpaid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Offen"

    payments = client.get('/payments')
    assert "RSV-" in payments
    assert "Manuell" in payments
    assert "info@example.org" in payments
    assert "39.00" in payments
    assert "Offen" in payments

    invoices = client.get('/invoices')
    assert "RSV-" in invoices
    assert "info@example.org" in invoices
    assert "39.00" in invoices
    assert "Unbezahlt" in invoices
    assert "Unfakturiert" in invoices


@freeze_time("2017-07-09", tick=True)
def test_manual_reservation_payment_without_extra(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_hour'
    resource.price_per_hour = 10.00
    resource.payment_method = 'manual'

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9, 10),
            datetime(2017, 7, 9, 12)
        )
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    reserve()

    page = client.get('/resource/tageskarte/form')
    page.form['email'] = 'info@example.org'
    assert '20.00' in page.form.submit().follow()

    ticket = page.form.submit().follow().form.submit().follow()
    assert 'RSV-' in ticket.text

    # mark it as paid
    client.login_editor()
    page = client.get('/tickets/ALL/open').click("Annehmen").follow()

    assert page.pyquery('.payment-state').text() == "Offen"

    client.post(page.pyquery('.mark-as-paid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Bezahlt"

    client.post(page.pyquery('.mark-as-unpaid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Offen"

    payments = client.get('/payments')
    assert "RSV-" in payments
    assert "Manuell" in payments
    assert "info@example.org" in payments
    assert "20.00" in payments
    assert "Offen" in payments

    invoices = client.get('/invoices')
    assert "RSV-" in invoices
    assert "info@example.org" in invoices
    assert "20.00" in invoices
    assert "Unbezahlt" in invoices
    assert "Unfakturiert" in invoices


@freeze_time("2017-07-09", tick=True)
def test_manual_reservation_payment_allocation_override(
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_hour'
    resource.price_per_hour = 10.00
    resource.payment_method = 'manual'

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations_inherit = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9, 10),
            datetime(2017, 7, 9, 12)
        )
    )
    allocations_free = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9, 12),
            datetime(2017, 7, 9, 14)
        ),
        data={'pricing_method': 'free'}
    )
    allocations_per_item = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9, 14),
            datetime(2017, 7, 9, 16)
        ),
        data={
            'pricing_method': 'per_item',
            'price_per_item': 23.00,
            'currency': 'CHF'
        }
    )
    allocations_per_hour = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9, 16),
            datetime(2017, 7, 9, 18)
        ),
        data={
            'pricing_method': 'per_hour',
            'price_per_hour': 12.00,
            'currency': 'CHF'
        }
    )

    reserve_inherit = client.bound_reserve(allocations_inherit[0])
    reserve_free = client.bound_reserve(allocations_free[0])
    reserve_per_item = client.bound_reserve(allocations_per_item[0])
    reserve_per_hour = client.bound_reserve(allocations_per_hour[0])
    transaction.commit()
    close_all_sessions()

    # create a reservation of each type
    reserve_inherit()
    reserve_free()
    reserve_per_item()
    reserve_per_hour()

    page = client.get('/resource/tageskarte/form')
    page.form['email'] = 'info@example.org'
    confirmation_page = page.form.submit().follow()
    assert '20.00' in confirmation_page
    assert '24.00' in confirmation_page
    assert '23.00' in confirmation_page
    assert '67.00' in confirmation_page

    ticket = confirmation_page.form.submit().follow()
    assert 'RSV-' in ticket.text

    # mark it as paid
    client.login_editor()
    page = client.get('/tickets/ALL/open').click("Annehmen").follow()

    assert page.pyquery('.payment-state').text() == "Offen"

    client.post(page.pyquery('.mark-as-paid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Bezahlt"

    client.post(page.pyquery('.mark-as-unpaid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Offen"

    payments = client.get('/payments')
    assert "RSV-" in payments
    assert "Manuell" in payments
    assert "info@example.org" in payments
    assert "67.00" in payments
    assert "Offen" in payments

    invoices = client.get('/invoices')
    assert "RSV-" in invoices
    assert "info@example.org" in invoices
    assert "67.00" in invoices
    assert "Unbezahlt" in invoices
    assert "Unfakturiert" in invoices


@freeze_time("2017-07-09", tick=True)
def test_manual_reservation_payment_with_resource_discount(
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_item'
    resource.price_per_item = 10.00
    resource.payment_method = 'manual'
    resource.extras_pricing_method = 'one_off'
    resource.discount_method = 'resource'
    resource.definition = textwrap.dedent("""
        Shipping =
            (x) Yes (6 CHF)
            ( ) No (0 CHF)
        Discount =
            (x) Yes (50%)
            ( ) No
    """)

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9),
            datetime(2017, 7, 9)
        ),
        whole_day=True,
        quota=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    reserve(quota=2, whole_day=True)

    page = client.get('/resource/tageskarte/form')
    page.form['email'] = 'info@example.org'

    page.form['shipping'] = 'Yes'
    page.form['discount'] = 'No'
    result = page.form.submit().follow()
    assert '20.00' in result
    assert '6.00' in result
    assert '26.00' in result
    assert '-10.00' not in result

    page.form['shipping'] = 'Yes'
    page.form['discount'] = 'Yes'
    result = page.form.submit().follow()
    assert '20.00' in result
    assert '6.00' in result
    assert '16.00' in result
    assert '-10.00' in result

    ticket = page.form.submit().follow().form.submit().follow()
    assert 'RSV-' in ticket.text

    # mark it as paid
    client.login_editor()
    page = client.get('/tickets/ALL/open').click("Annehmen").follow()

    assert page.pyquery('.payment-state').text() == "Offen"

    client.post(page.pyquery('.mark-as-paid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Bezahlt"

    client.post(page.pyquery('.mark-as-unpaid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Offen"

    payments = client.get('/payments')
    assert "RSV-" in payments
    assert "Manuell" in payments
    assert "info@example.org" in payments
    assert "16.00" in payments
    assert "Offen" in payments

    invoices = client.get('/invoices')
    assert "RSV-" in invoices
    assert "info@example.org" in invoices
    assert "16.00" in invoices
    assert "Unbezahlt" in invoices
    assert "Unfakturiert" in invoices


@freeze_time("2017-07-09", tick=True)
def test_manual_reservation_payment_with_extras_discount(
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_item'
    resource.price_per_item = 10.00
    resource.payment_method = 'manual'
    resource.extras_pricing_method = 'one_off'
    resource.discount_method = 'extras'
    resource.definition = textwrap.dedent("""
        Shipping =
            (x) Yes (6 CHF)
            ( ) No (0 CHF)
        Discount =
            (x) Yes (50%)
            ( ) No
    """)

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9),
            datetime(2017, 7, 9)
        ),
        whole_day=True,
        quota=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    reserve(quota=2, whole_day=True)

    page = client.get('/resource/tageskarte/form')
    page.form['email'] = 'info@example.org'

    page.form['shipping'] = 'Yes'
    page.form['discount'] = 'No'
    result = page.form.submit().follow()
    assert '20.00' in result
    assert '6.00' in result
    assert '26.00' in result
    assert '-3.00' not in result

    page.form['shipping'] = 'Yes'
    page.form['discount'] = 'Yes'
    result = page.form.submit().follow()
    assert '20.00' in result
    assert '6.00' in result
    assert '23.00' in result
    assert '-3.00' in result

    ticket = page.form.submit().follow().form.submit().follow()
    assert 'RSV-' in ticket.text

    # mark it as paid
    client.login_editor()
    page = client.get('/tickets/ALL/open').click("Annehmen").follow()

    assert page.pyquery('.payment-state').text() == "Offen"

    client.post(page.pyquery('.mark-as-paid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Bezahlt"

    client.post(page.pyquery('.mark-as-unpaid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Offen"

    payments = client.get('/payments')
    assert "RSV-" in payments
    assert "Manuell" in payments
    assert "info@example.org" in payments
    assert "23.00" in payments
    assert "Offen" in payments

    invoices = client.get('/invoices')
    assert "RSV-" in invoices
    assert "info@example.org" in invoices
    assert "23.00" in invoices
    assert "Unbezahlt" in invoices
    assert "Unfakturiert" in invoices


@freeze_time("2017-07-09", tick=True)
def test_manual_reservation_payment_with_everything_discount(
    client: Client
) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    resource.pricing_method = 'per_item'
    resource.price_per_item = 10.00
    resource.payment_method = 'manual'
    resource.extras_pricing_method = 'one_off'
    resource.discount_method = 'everything'
    resource.definition = textwrap.dedent("""
        Shipping =
            (x) Yes (6 CHF)
            ( ) No (0 CHF)
        Discount =
            (x) Yes (50%)
            ( ) No
    """)

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations = scheduler.allocate(
        dates=(
            datetime(2017, 7, 9),
            datetime(2017, 7, 9)
        ),
        whole_day=True,
        quota=4
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    reserve(quota=2, whole_day=True)

    page = client.get('/resource/tageskarte/form')
    page.form['email'] = 'info@example.org'

    page.form['shipping'] = 'Yes'
    page.form['discount'] = 'No'
    result = page.form.submit().follow()
    assert '20.00' in result
    assert '6.00' in result
    assert '26.00' in result
    assert '-13.00' not in result

    page.form['shipping'] = 'Yes'
    page.form['discount'] = 'Yes'
    result = page.form.submit().follow()
    assert '20.00' in result
    assert '6.00' in result
    assert '13.00' in result
    assert '-13.00' in result

    ticket = page.form.submit().follow().form.submit().follow()
    assert 'RSV-' in ticket.text

    # mark it as paid
    client.login_editor()
    page = client.get('/tickets/ALL/open').click("Annehmen").follow()

    assert page.pyquery('.payment-state').text() == "Offen"

    client.post(page.pyquery('.mark-as-paid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Bezahlt"

    client.post(page.pyquery('.mark-as-unpaid').attr('ic-post-to'))
    page = client.get(page.request.url)

    assert page.pyquery('.payment-state').text() == "Offen"

    payments = client.get('/payments')
    assert "RSV-" in payments
    assert "Manuell" in payments
    assert "info@example.org" in payments
    assert "13.00" in payments
    assert "Offen" in payments

    invoices = client.get('/invoices')
    assert "RSV-" in invoices
    assert "info@example.org" in invoices
    assert "13.00" in invoices
    assert "Unbezahlt" in invoices
    assert "Unfakturiert" in invoices


def test_allocation_rules_on_rooms(client: Client) -> None:
    client.login_admin()

    resources_page = client.get('/resources')

    page = resources_page.click('Raum')
    page.form['title'] = 'Room'
    page.form.submit()

    def count_allocations() -> int:
        s = '2000-01-01'
        e = '2050-01-31'

        return len(client.get(f'/resource/room/slots?start={s}&end={e}').json)

    def run_cronjob() -> None:
        client.get('/resource/room/process-rules')

    page = client.get('/resource/room').click(
        "Verfügbarkeitszeiträume").click("Verfügbarkeitszeitraum")
    page.form['title'] = 'Täglich'
    page.form['extend'] = 'daily'
    page.form['start'] = '2019-01-01'
    page.form['end'] = '2019-01-02'
    page.form['as_whole_day'] = 'yes'

    page.select_checkbox('except_for', "Sa")
    page.select_checkbox('except_for', "So")

    page = page.form.submit().follow()

    assert 'Verfügbarkeitszeitraum aktiv, 2 Verfügbarkeiten erstellt' in page
    assert count_allocations() == 2

    # running the cronjob once will add a new allocation
    with freeze_time('2019-01-02 22:00:00'):
        run_cronjob()

    assert count_allocations() == 3

    # running it a second time will not as there needs to be enough time
    # between the two calls for the second one to succeed
    with freeze_time('2019-01-02 22:00:00'):
        run_cronjob()

    assert count_allocations() == 3

    with freeze_time('2019-01-03 22:00:00'):
        run_cronjob()

    assert count_allocations() == 4

    # the next two times won't change anything as those are Saturday, Sunday
    with freeze_time('2019-01-04 22:00:00'):
        run_cronjob()

    assert count_allocations() == 4

    with freeze_time('2019-01-05 22:00:00'):
        run_cronjob()

    assert count_allocations() == 4

    # then the cronjob should pick up again
    with freeze_time('2019-01-06 22:00:00'):
        run_cronjob()

    assert count_allocations() == 5

    with freeze_time('2019-01-07 22:00:00'):
        run_cronjob()

    assert count_allocations() == 6

    # deleting the rule will delete all associated slots (but not others)
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('room')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    scheduler.allocate(
        dates=(datetime(2018, 1, 1), datetime(2018, 1, 1)),
        whole_day=True,
    )

    transaction.commit()

    assert count_allocations() == 7


def test_allocation_rules_edit(client: Client) -> None:
    client.login_admin()

    resources_page = client.get('/resources')

    page = resources_page.click('Raum')
    page.form['title'] = 'Room'
    page.form.submit()

    def count_allocations() -> int:
        s = '2000-01-01'
        e = '2050-01-31'

        return len(client.get(f'/resource/room/slots?start={s}&end={e}').json)

    def run_cronjob() -> None:
        client.get('/resource/room/process-rules')

    page = client.get('/resource/room').click(
        "Verfügbarkeitszeiträume").click("Verfügbarkeitszeitraum")
    page.form['title'] = 'Täglich'
    page.form['extend'] = 'daily'
    page.form['start'] = '2019-01-01'
    page.form['end'] = '2019-01-02'
    page.form['as_whole_day'] = 'yes'

    page.select_checkbox('except_for', "Sa")
    page.select_checkbox('except_for', "So")

    page = page.form.submit().follow()

    assert 'Verfügbarkeitszeitraum aktiv, 2 Verfügbarkeiten erstellt' in page
    assert count_allocations() == 2

    # Modifying the rule applies changes where possible, but
    # existing reserved slots remain unaffected.
    edit_page = client.get('/resource/room')
    edit_page = edit_page.click('Verfügbarkeitszeiträume').click('Bearbeiten')
    form = edit_page.form
    form['title'] = 'Renamed room'

    edit_page = form.submit().follow()

    assert 'Renamed room' in edit_page


def test_allocation_rules_copy_paste(client: Client) -> None:
    client.login_admin()

    resources_page = client.get('/resources')

    page = resources_page.click('Raum')
    page.form['title'] = 'Room 1'
    page.form.submit()

    page = resources_page.click('Raum')
    page.form['title'] = 'Room 2'
    page.form.submit()

    def count_allocations(room: int) -> int:
        return len(client.get(
            f'/resource/room-{room}/slots?start=2000-01-01&end=2050-01-31'
        ).json)

    def run_cronjob() -> None:
        client.get('/resource/room/process-rules')

    page = client.get('/resource/room-1').click(
        "Verfügbarkeitszeiträume").click("Verfügbarkeitszeitraum")
    page.form['title'] = 'Täglich'
    page.form['extend'] = 'daily'
    page.form['start'] = '2019-01-01'
    page.form['end'] = '2019-01-02'
    page.form['as_whole_day'] = 'yes'

    page.select_checkbox('except_for', "Sa")
    page.select_checkbox('except_for', "So")

    page = page.form.submit().follow()

    assert 'Verfügbarkeitszeitraum aktiv, 2 Verfügbarkeiten erstellt' in page
    assert count_allocations(1) == 2
    assert count_allocations(2) == 0

    # Copy the rule
    edit_page = client.get('/resource/room-1').click('Verfügbarkeitszeiträume')
    copy_links = edit_page.html.find_all('a', string='Kopieren')
    client.post(copy_links[0].attrs['ic-post-to'])
    edit_page = client.get(edit_page.request.path)
    assert 'in die Zwischenablage kopiert' in edit_page

    # Paste the rule in the other room
    edit_page = client.get('/resource/room-2').click('Verfügbarkeitszeiträume')
    paste_links = edit_page.html.find_all('a', string='Einfügen')
    client.post(paste_links[0].attrs['ic-post-to'])
    edit_page = client.get(edit_page.request.path)
    assert 'wurde eingefügt' in edit_page
    assert count_allocations(1) == 2
    assert count_allocations(2) == 2


def test_allocation_rules_on_daypasses(client: Client) -> None:
    client.login_admin()

    resources_page = client.get('/resources')

    page = resources_page.click('Tageskarte', index=0)
    page.form['title'] = 'Daypass'
    page.form.submit()

    page = resources_page.click('Tageskarte', index=0)
    page.form['title'] = 'Daypass'
    page = page.form.submit()
    assert "Eine Resource mit diesem Namen existiert bereits" in page

    def count_allocations() -> int:
        s = '2000-01-01'
        e = '2050-01-31'

        return len(client.get(
            f'/resource/daypass/slots?start={s}&end={e}').json)

    def run_cronjob() -> None:
        client.get('/resource/daypass/process-rules')

    page = client.get('/resource/daypass').click(
        "Verfügbarkeitszeiträume").click("Verfügbarkeitszeitraum")
    page.form['title'] = 'Monatlich'
    page.form['extend'] = 'monthly'
    page.form['start'] = '2019-01-01'
    page.form['end'] = '2019-01-31'
    page.form['daypasses'] = '1'
    page.form['daypasses_limit'] = '1'
    page = page.form.submit().follow()

    assert 'Verfügbarkeitszeitraum aktiv, 31 Verfügbarkeiten erstellt' in page
    assert count_allocations() == 31

    # running the cronjob on an ordinary day will not change anything
    with freeze_time('2019-01-30 22:00:00'):
        run_cronjob()

    assert count_allocations() == 31

    # only run at the end of the month does it work
    with freeze_time('2019-01-31 22:00:00'):
        run_cronjob()

    assert count_allocations() == 59

    # running it a second time on the same day has no effect
    with freeze_time('2019-01-31 22:00:00'):
        run_cronjob()

    assert count_allocations() == 59

    # add another month
    with freeze_time('2019-02-28 22:00:00'):
        run_cronjob()

    assert count_allocations() == 90

    # let's stop the rule, which should leave existing allocations
    page = client.get('/resource/daypass').click("Verfügbarkeitszeiträume")
    page.click('Stop')

    page = client.get('/resource/daypass').click("Verfügbarkeitszeiträume")
    assert "Keine Verfügbarkeitszeiträume" in page
    assert count_allocations() == 90


def test_allocation_rules_with_holidays(client: Client) -> None:
    client.login_admin()

    page = client.get('/holiday-settings')
    page.select_checkbox('cantonal_holidays', "Zug")
    page.form.submit()

    resources_page = client.get('/resources')
    page = resources_page.click('Tageskarte', index=0)
    page.form['title'] = 'Daypass'
    page.form.submit()

    def count_allocations() -> int:
        s = '2000-01-01'
        e = '2050-01-31'

        return len(client.get(
            f'/resource/daypass/slots?start={s}&end={e}').json)

    def run_cronjob() -> None:
        client.get('/resource/daypass/process-rules')

    page = client.get('/resource/daypass').click(
        "Verfügbarkeitszeiträume").click("Verfügbarkeitszeitraum")
    page.form['title'] = 'Jährlich'
    page.form['extend'] = 'yearly'
    page.form['start'] = '2019-01-01'
    page.form['end'] = '2019-12-31'
    page.form['daypasses'] = '1'
    page.form['daypasses_limit'] = '1'
    page.form['on_holidays'] = 'no'
    page = page.form.submit().follow()

    assert 'Verfügbarkeitszeitraum aktiv, 352 Verfügbarkeiten erstellt' in page
    assert count_allocations() == 352

    # running the cronjob on an ordinary day will not change anything
    with freeze_time('2019-01-31 22:00:00'):
        run_cronjob()

    assert count_allocations() == 352

    # only run at the end of the year does it work
    with freeze_time('2019-12-31 22:00:00'):
        run_cronjob()

    assert count_allocations() == 705


def test_allocation_rules_with_school_holidays(client: Client) -> None:
    client.login_admin()

    page = client.get('/holiday-settings')
    page.form['school_holidays'] = (
        '01.03.2019 - 15.03.2019\n'
        '06.03.2020 - 20.03.2020'
    )
    page.form.submit()

    resources = client.get('/resources')
    page = resources.click('Tageskarte', index=0)
    page.form['title'] = 'Daypass'
    page.form.submit()

    def count_allocations() -> int:
        s = '2000-01-01'
        e = '2050-01-31'

        return len(client.get(
            f'/resource/daypass/slots?start={s}&end={e}').json)

    def run_cronjob() -> None:
        client.get('/resource/daypass/process-rules')

    page = client.get('/resource/daypass').click(
        "Verfügbarkeitszeiträume").click("Verfügbarkeitszeitraum")
    page.form['title'] = 'Jährlich'
    page.form['extend'] = 'yearly'
    page.form['start'] = '2019-01-01'
    page.form['end'] = '2019-12-31'
    page.form['daypasses'] = '1'
    page.form['daypasses_limit'] = '1'
    page.form['during_school_holidays'] = 'no'
    page = page.form.submit().follow()

    assert 'Verfügbarkeitszeitraum aktiv, 350 Verfügbarkeiten erstellt' in page
    assert count_allocations() == 350

    # running the cronjob on an ordinary day will not change anything
    with freeze_time('2019-01-31 22:00:00'):
        run_cronjob()

    assert count_allocations() == 350

    # only run at the end of the year does it work
    with freeze_time('2019-12-31 22:00:00'):
        run_cronjob()

    assert count_allocations() == 701


@freeze_time("2019-08-01", tick=True)
def test_zipcode_block(client: Client) -> None:
    client.login_admin()

    # enable zip-code blocking
    page = client.get('/resource/tageskarte/edit')
    page.form['definition'] = '\n'.join((
        'PLZ *= ___',
    ))
    page.select_radio('payment_method', "Keine Kreditkarten-Zahlungen")
    page.select_radio('pricing_method', "Kostenlos")
    page.select_radio('extras_pricing_method', "Pro Eintrag")
    page.select_radio('discount_method', "Nur den Preis pro Eintrag/Stunde")
    page.form['zipcode_block_use'] = True
    page.form['zipcode_days'] = 1
    page.form['zipcode_field'] = 'PLZ'
    page.form['zipcode_list'] = '\n'.join((
        '1234',
        '5678',
    ))
    page.form.submit().follow()

    # create a blocked and an unblocked allocation
    transaction.begin()

    scheduler = (
        ResourceCollection(client.app.libres_context)  # type: ignore[union-attr]
        .by_name('tageskarte')
        .get_scheduler(client.app.libres_context)
    )

    allocations = [
        *scheduler.allocate(
            dates=(datetime(2019, 8, 2), datetime(2019, 8, 2)),
            whole_day=True,
            quota=2,
            quota_limit=2
        ),
        *scheduler.allocate(
            dates=(datetime(2019, 8, 3), datetime(2019, 8, 3)),
            whole_day=True,
            quota=2,
            quota_limit=2
        )
    ]

    transaction.commit()

    allocations[0] = client.app.session().merge(allocations[0])
    allocations[1] = client.app.session().merge(allocations[1])

    # reserve both allocations
    client = client.spawn()
    client.bound_reserve(allocations[0])(quota=1)
    client.bound_reserve(allocations[1])(quota=1)

    # one of them should now show a warning
    page = client.get('/resource/tageskarte/form')
    assert str(page).count('Postleitzahlen') == 1

    # the confirmation should fail
    page.form['email'] = 'info@example.org'
    page.form['plz'] = '0000'
    page = page.form.submit()

    # unless we have the right PLZ
    page.form['plz'] = '1234'
    page.form.submit().follow()

    # or unless we are admin
    client = client.spawn()
    client.login_admin()

    allocations[0] = client.app.session().merge(allocations[0])
    allocations[1] = client.app.session().merge(allocations[1])

    client.bound_reserve(allocations[0])(quota=1)
    client.bound_reserve(allocations[1])(quota=1)

    # we won't get a warning
    page = client.get('/resource/tageskarte/form')
    assert str(page).count('Postleitzahlen') == 0

    # and we can confirm
    page.form['email'] = 'info@example.org'
    page.form['plz'] = '0000'
    page.form.submit().follow()


def test_find_your_spot_link(client: Client) -> None:
    client.login_admin()

    resources = client.get('/resources')
    # This will be a group without a room so no find-your-spot link
    new = resources.click('Gegenstand')
    new.form['title'] = 'Item'
    new.form['group'] = 'Items'
    new.form.submit().follow()

    # A group with two rooms and a find-your-spot link
    new = resources.click('Raum')
    new.form['title'] = 'Meeting 1'
    new.form['group'] = 'Rooms'
    new.form.submit().follow()

    new = resources.click('Raum')
    new.form['title'] = 'Meeting 2'
    new.form['group'] = 'Rooms'
    new.form.submit().follow()

    # And a group with one room and therefore also a find-your-spot link
    new = resources.click('Gegenstand')
    new.form['title'] = 'Item 2'
    new.form['group'] = 'Something'
    new.form.submit().follow()

    new = resources.click('Raum')
    new.form['title'] = 'Room'
    new.form['group'] = 'Something'
    new.form.submit().follow()

    page = client.get('/resources')

    assert page.pyquery('#Allgemein .find-your-spot-link')
    assert not page.pyquery('#Items .find-your-spot-link')
    assert page.pyquery('#Rooms .find-your-spot-link')
    assert page.pyquery('#Something .find-your-spot-link')


def test_resource_recipient_overview(client: Client) -> None:
    resources = ResourceCollection(client.app.libres_context)
    gymnasium = resources.add('Gymnasium', 'Europe/Zurich', type='room')
    dailypass = resources.add('Dailypass', 'Europe/Zurich', type='daypass')
    meeting = resources.add('Meeting', 'Europe/Zurich', type='room')

    recipients = ResourceRecipientCollection(client.app.session())
    recipients.add(
        name='John',
        medium='email',
        address='john@example.org',
        new_reservations=True,
        daily_reservations=True,
        send_on=['FR', 'SU'],
        resources=[
            gymnasium.id.hex,
            dailypass.id.hex,
            meeting.id.hex
        ]
    )

    # John will still have the meeting room id as his resource, but this should
    # not make any problems
    resources.delete(meeting)

    transaction.commit()
    client.login_admin()

    page = client.get('/resource-recipients')
    assert "John" in page
    assert "john@example.org" in page
    assert "Erhält Benachrichtigungen für neue Reservationen." in page
    assert "für Reservationen des Tages an folgenden Tagen:" in page
    assert "Fr , So" in page
    assert "Gymnasium" in page
    assert "Dailypass" in page
    assert "Meeting" not in page


@freeze_time("2024-04-08", tick=True)
def test_reserve_fractions_of_hours_total_correct_in_price(
    client: Client
) -> None:
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource
    resource.pricing_method = 'per_hour'
    resource.price_per_hour = 10.00
    resource.payment_method = 'manual'

    scheduler = resource.get_scheduler(client.app.libres_context)
    allocations = scheduler.allocate(
        dates=(
            datetime(2024, 4, 9, 10, 0),
            datetime(2024, 4, 9, 14, 0)  # 4-hour slot
        ),
        partly_available=True,
        whole_day=False
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # Reserve 1.5 hours (10:00 to 11:30)
    result = reserve(start='10:00', end='11:30')
    assert result.json == {'success': True}

    # Check price on confirmation page
    form_page = client.get('/resource/tageskarte/form')
    form_page.form['email'] = 'tester@example.org'

    confirmation_page = form_page.form.submit().follow()

    # Expected price: 1.5 hours * 10.00 CHF/hour = 15.00 CHF
    total_amount_dt = confirmation_page.pyquery('dt:contains("Totalbetrag")')
    assert total_amount_dt, 'Total amount dt not found'
    total_amount_dd = total_amount_dt.next('dd')
    assert total_amount_dd, 'Total amount dd not found'
    assert total_amount_dd.text().strip() == '15.00'
    assert '10:00 - 11:30' in confirmation_page

    # Check price in reservations JSON
    reservations_url = '/resource/tageskarte/reservations'
    reservations_data = client.get(reservations_url).json['reservations']
    assert len(reservations_data) == 1
    assert reservations_data[0]['price']['amount'] == 15.00


@freeze_time("2015-08-28", tick=True)
def test_my_reservations_view(client: Client) -> None:
    # prepate the required data
    resources = ResourceCollection(client.app.libres_context)
    resource = resources.by_name('tageskarte')
    assert resource is not None
    scheduler = resource.get_scheduler(client.app.libres_context)

    allocations = scheduler.allocate(
        dates=(datetime(2015, 8, 28), datetime(2015, 8, 28)),
        whole_day=True
    )

    reserve = client.bound_reserve(allocations[0])
    transaction.commit()

    # create a reservation
    assert reserve().json == {'success': True}
    formular = client.get('/resource/tageskarte/form')
    formular.form['email'] = 'info@example.org'
    formular.form.submit().follow().form.submit()

    client2 = client.spawn()

    # by default this view is disabled
    client.get('/resources/my-reservations', status=404)
    client.get('/resources/my-reservations-json', status=404)
    client.get('/resources/my-reservations-pdf', status=404)
    client.get('/resources/my-reservations-subscribe', status=404)
    client.get('/resources/my-reservations-ical', status=403)

    # let's enable it
    admin = client.spawn()
    admin.login_admin()
    settings = admin.get('/').click('Einstellungen').click('Kunden-Login')
    settings.form['citizen_login_enabled'].checked = True
    settings.form.submit().follow()

    # now we don't have access yet
    client.get('/resources/my-reservations-json', status=403)
    client.get('/resources/my-reservations-pdf', status=403)
    login = client.get('/resources/my-reservations?date=20150828').follow()
    login.form['email'] = 'info@example.org'
    confirm = login.form.submit().follow()
    assert len(os.listdir(client.app.maildir)) == 2
    message = client.get_email(1)['TextBody']
    token = re.search(r'&token=([^)]+)', message).group(1)  # type: ignore[union-attr]
    confirm.form['token'] = token
    reservations = confirm.form.submit().follow()
    assert reservations.request.path_qs == (
        '/resources/my-reservations?date=20150828'
    )
    reservations = client.get(
        '/resources/my-reservations-json?start=2015-08-28&end=2015-08-29'
    )
    assert reservations.status_code == 200

    # accessing the PDF without a date filter is not allowed
    client.get('/resources/my-reservations-pdf', status=400)
    pdf = client.get(
        '/resources/my-reservations-pdf?start=2015-08-28&end=2015-08-29'
    )
    _, pdf_content = extract_pdf_info(BytesIO(pdf.body))
    assert '28. August 2015 - 29. August 2015' in pdf_content
    assert 'Ganztägig' in pdf_content
    assert 'Noch nicht akzeptiert' in pdf_content

    # if we only include accepted reservations the PDF is empty
    pdf = client.get(
        '/resources/my-reservations-pdf'
        '?start=2015-08-28&end=2015-08-29&accepted=1'
    )
    _, pdf_content = extract_pdf_info(BytesIO(pdf.body))
    assert '28. August 2015 - 29. August 2015' in pdf_content
    assert 'Keine Daten verfügbar' in pdf_content

    # let's also check out the ticket specific pdf
    tickets = client.get('/tickets/ALL/all/my-tickets')
    pdf = tickets.click(href='reservations-pdf')
    _, pdf_content = extract_pdf_info(BytesIO(pdf.body))
    assert 'Ganztägig' in pdf_content

    subscribe = client.get('/resources/my-reservations-subscribe')
    assert 'webcal://' in subscribe
    ical_url = re.search(  # type: ignore[union-attr]
        r'webcal://localhost(/[^"]+)', subscribe.text
    ).group(1)

    # accept the reservation so the ical file contains an event
    ticket = admin.get('/tickets/ALL/open').click('Annehmen').follow()
    ticket.click('Alle Reservationen annehmen')

    assert client.get(ical_url).status_code == 200

    # someone else can open the same ical link without authentication
    assert client2.get(ical_url).status_code == 200
    # but not the other views
    client2.get('/resources/my-reservations-json', status=403)
    client2.get('/resources/my-reservations-pdf', status=403)
