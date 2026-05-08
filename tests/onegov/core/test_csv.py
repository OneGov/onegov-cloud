from __future__ import annotations

import pytest
import tempfile
from io import BytesIO
from onegov.core import utils
from onegov.core.csv import (
    CSVFile,
    convert_excel_to_csv,
    convert_list_of_dicts_to_csv,
    convert_list_of_list_of_dicts_to_xlsx,
    convert_list_of_dicts_to_xlsx,
    convert_xls_to_csv,
    convert_xlsx_to_csv,
    detect_encoding,
    match_headers,
    normalize_header,
    parse_header, normalize_sheet_titles, remove_first_word,
    avoid_duplicate_name, list_duplicates_index,
)
from onegov.core.errors import (
    AmbiguousColumnsError,
    DuplicateColumnNamesError,
    EmptyLineInFileError,
    MissingColumnsError,
)
from openpyxl import load_workbook


def test_parse_header() -> None:
    assert parse_header("   Firtst name;  LastNAME; Designation")\
           == ['firtst name', 'lastname', 'designation']
    assert parse_header("a") == ['a']
    assert parse_header("") == []
    assert parse_header("a,b,c;d") == ['a', 'b', 'c;d']
    assert parse_header("a,b,c,c") == ['a', 'b', 'c', 'c']
    assert parse_header(
        "a,b,c,c,c,c,c,b,b,a", rename_duplicate_column_names=True
    ) == ['a', 'b', 'c', 'c_1', 'c_2', 'c_3', 'c_4', 'b_1', 'b_2', 'a_1']


def test_normalize_header() -> None:
    assert normalize_header("") == ""
    assert normalize_header("Wääh") == "waah"
    assert normalize_header(" a   b\tc ") == "a b c"


def test_detect_encoding() -> None:
    assert detect_encoding(BytesIO('jöö'.encode('ISO-8859-2'))) == 'cp1252'
    assert detect_encoding(BytesIO('jöö'.encode('utf-8'))) == 'utf-8'

    assert detect_encoding(
        BytesIO('abcdefabcdefabcdefabcdefö'.encode('utf-8'))
    ) == 'utf-8'


def test_simple_csv_file() -> None:
    data = (
        b'Datum, Reale Temperatur, Gef\xfchlte Temperatur\n'
        b'01.01.2015, 5, k\xfchl\n'
        b'02.01.2015, 0, kalt'
    )

    csv = CSVFile(
        BytesIO(data), ['datum', 'reale_temperatur', 'gefuhlte_temperatur']
    )

    assert list(csv.headers.keys()) == [
        'datum', 'reale_temperatur', 'gefuhlte_temperatur'
    ]
    assert list(csv.lines) == [
        csv.rowtype(
            rownumber=2,
            datum='01.01.2015',
            reale_temperatur='5',
            gefuhlte_temperatur='kühl'
        ),
        csv.rowtype(
            rownumber=3,
            datum='02.01.2015',
            reale_temperatur='0',
            gefuhlte_temperatur='kalt'
        ),
    ]


def test_wacky_csv_file() -> None:
    data = (
        b'Datum,   Temperatur%, Datum\n'
        b'01.01.2015, 5, 01.01.2014\n'
        b'02.01.2015, 0, 02.01.2014'
    )
    csv = CSVFile(
        BytesIO(data),
        ['datum', 'temperatur'],
        rename_duplicate_column_names=True
    )

    assert list(csv.headers.keys()) == ['datum', 'temperatur', 'datum_1']
    assert list(csv.lines)[0].datum == '01.01.2015'
    assert list(csv.lines)[1].temperatur == '0'


def test_convert_xlsx_to_csv_wrong_format() -> None:
    with pytest.raises(IOError):
        convert_xlsx_to_csv(BytesIO())

    with pytest.raises(IOError):
        convert_xlsx_to_csv(BytesIO(b'abcd'))

    path = utils.module_path('tests.onegov.core', 'fixtures/excel.xls')
    with open(path, 'rb') as file:
        with pytest.raises(IOError):
            convert_xlsx_to_csv(file)


def test_convert_xls_to_csv_wrong_format() -> None:
    with pytest.raises(IOError):
        convert_xls_to_csv(BytesIO())

    with pytest.raises(IOError):
        convert_xls_to_csv(BytesIO(b'abcd'))

    path = utils.module_path('tests.onegov.core', 'fixtures/excel.xlsx')
    with open(path, 'rb') as file:
        with pytest.raises(IOError):
            convert_xls_to_csv(file)


@pytest.mark.parametrize("excel_file", [
    utils.module_path('tests.onegov.core', 'fixtures/excel.xls'),
    utils.module_path('tests.onegov.core', 'fixtures/excel.xlsx'),
])
def test_convert_to_csv(excel_file: str) -> None:
    with open(excel_file, 'rb') as f:
        headers = ['ID', 'Namä', 'Date', 'Bool', 'Leer', 'Formel']
        csv = CSVFile(convert_excel_to_csv(f), headers)

        assert list(csv.headers.keys()) == headers
        assert list(csv.lines) == [
            csv.rowtype(
                rownumber=2,
                id='1',
                nama='Döner',
                date='2015-12-31T00:00:00',
                bool='1',
                leer='',
                formel='2'
            ),
            csv.rowtype(
                rownumber=3,
                id='2',
                nama='“Cola”',
                date='2014-12-31T12:00:00',
                bool='0',
                leer='',
                formel='4'
            ),
            csv.rowtype(
                rownumber=4,
                id='3.1',
                nama='',
                date='',
                bool='',
                leer='',
                formel=''
            )
        ]

        with pytest.raises(NotImplementedError):
            csv = CSVFile(convert_excel_to_csv(f, 'Sheet 2'))
        with pytest.raises(KeyError):
            convert_excel_to_csv(f, 'Sheet 3')

        convert_excel_to_csv(f, None)
        convert_excel_to_csv(f, '')


def test_empty_line_csv_file() -> None:
    data = (
        b'Datum, Reale Temperatur, Gef\xfchlte Temperatur\n'
        b'\n'
        b'02.01.2015, 0, kalt'
    )

    csv = CSVFile(
        BytesIO(data), ['datum', 'reale_temperatur', 'gefuhlte_temperatur']
    )

    assert list(csv.headers) == [
        'datum', 'reale_temperatur', 'gefuhlte_temperatur'
    ]
    with pytest.raises(EmptyLineInFileError):
        list(csv.lines)

    # accept empty lines at the end
    data = (
        b'Datum, Reale Temperatur, Gef\xfchlte Temperatur\n'
        b'02.01.2015, 0, kalt'
        b'\n'
        b'\n'
    )

    csv = CSVFile(
        BytesIO(data), ['datum', 'reale_temperatur', 'gefuhlte_temperatur']
    )

    assert list(csv.headers) == [
        'datum', 'reale_temperatur', 'gefuhlte_temperatur'
    ]
    assert list(csv.lines)


def test_match_headers_duplicate() -> None:
    with pytest.raises(DuplicateColumnNamesError):
        match_headers(['first_name', 'first_name'], expected=[])


def test_match_headers_order() -> None:
    matches = match_headers(
        headers=['firtst name', 'lastname'],
        expected=('first_name', 'last_name')
    )
    assert matches == ['first_name', 'last_name']

    matches = match_headers(
        headers=['firtst name', 'lastname'],
        expected=('last_name', 'first_name')
    )
    assert matches == ['first_name', 'last_name']


def test_match_headers_case() -> None:
    assert match_headers(['a', 'b'], expected=('A', 'B')) == ['A', 'B']
    assert match_headers(['a', 'b'], expected=('b', 'a')) == ['a', 'b']


def test_match_headers_missing() -> None:
    with pytest.raises(MissingColumnsError) as e:
        match_headers(['a', 'b'], expected=('a', 'c'))
    assert e.value.columns == ['c']

    with pytest.raises(MissingColumnsError) as e:
        match_headers(['ab', 'ba'], expected=('a', 'b'))
    assert e.value.columns == ['a', 'b']

    with pytest.raises(MissingColumnsError) as e:
        match_headers(['first', 'second'], expected=('first', 'third'))
    assert e.value.columns == ['third']

    assert match_headers(['a1', 'b2'], expected=('a1', 'c2')) == ['a1', 'c2']

    assert match_headers(['first', 'second'], expected=('first', 'sekond'))\
           == ['first', 'sekond']

    assert match_headers(['a', 'b', 'c'], expected=('a', 'c'))\
           == ['a', 'b', 'c']


def test_match_headers_ambiguous() -> None:
    with pytest.raises(AmbiguousColumnsError) as e:
        match_headers(['abcd', 'bcde'], expected=('bcd',))

    assert list(e.value.columns.keys()) == ['bcd']
    assert set(e.value.columns['bcd']) == {'abcd', 'bcde'}


def test_convert_list_of_dicts_to_csv() -> None:
    data = [
        {
            'first_name': 'Dick',
            'last_name': 'Cheney'
        },
        {
            'first_name': 'Donald',
            'last_name': 'Rumsfeld'
        }
    ]

    # without providing a list of fields, the order of fields is random
    csv = convert_list_of_dicts_to_csv(data)
    header, dick, donald = csv.splitlines()

    assert 'first_name' in header
    assert 'last_name' in header

    assert 'Dick' in dick
    assert 'Cheney' in dick

    assert 'Donald' in donald
    assert 'Rumsfeld' in donald

    # we can change this by being explicit
    csv = convert_list_of_dicts_to_csv(data, ('first_name', 'last_name'))
    header, dick, donald = csv.splitlines()

    assert header == 'first_name,last_name'
    assert dick == 'Dick,Cheney'
    assert donald == 'Donald,Rumsfeld'

    csv = convert_list_of_dicts_to_csv(data, ('last_name', 'first_name'))
    header, dick, donald = csv.splitlines()

    assert header == 'last_name,first_name'
    assert dick == 'Cheney,Dick'
    assert donald == 'Rumsfeld,Donald'

    # or by providing a key
    csv = convert_list_of_dicts_to_csv(data, key=lambda f: f, reverse=True)
    header, dick, donald = csv.splitlines()

    assert header == 'last_name,first_name'
    assert dick == 'Cheney,Dick'
    assert donald == 'Rumsfeld,Donald'

    csv = convert_list_of_dicts_to_csv(data, key=lambda f: f, reverse=False)
    header, dick, donald = csv.splitlines()

    assert header == 'first_name,last_name'
    assert dick == 'Dick,Cheney'
    assert donald == 'Donald,Rumsfeld'


def test_convert_list_of_dicts_to_csv_escaping() -> None:
    data = [
        {
            'value': ',;"'
        }
    ]

    # without providing a list of fields, the order of fields is random
    csv = convert_list_of_dicts_to_csv(data)
    header, row = csv.splitlines()

    assert header == 'value'
    assert row == '",;"""'


def test_convert_list_of_dicts_to_xlsx() -> None:
    data = [
        {
            'first_name': 'Dick',
            'last_name': 'Cheney'
        },
        {
            'first_name': 'Donald',
            'last_name': 'Rumsfeld'
        }
    ]

    xlsx = convert_list_of_dicts_to_xlsx(data, ('first_name', 'last_name'))

    with tempfile.NamedTemporaryFile() as f:
        f.write(xlsx)

        workbook = load_workbook(f)
        assert workbook.active is not None
        rows = tuple(workbook.active.rows)

        assert rows[0][0].value == 'first_name'
        assert rows[0][1].value == 'last_name'
        assert rows[1][0].value == 'Dick'
        assert rows[1][1].value == 'Cheney'
        assert rows[2][0].value == 'Donald'
        assert rows[2][1].value == 'Rumsfeld'


def test_convert_irregular_list_of_dicts_to_csv() -> None:
    data = [
        {
            'name': 'Batman',
            'role': 'Superhero',
            'identity': 'Bruce Wayne'
        },
        {
            'name': 'Joker',
            'role': 'Supervillain',
            'gang': 'Injustice League'
        }
    ]

    # fields in random order
    csv = convert_list_of_dicts_to_csv(data)
    header, batman, joker = csv.splitlines()

    for column in ('name', 'role', 'identity', 'gang'):
        assert column in header

    for value in ('Batman', 'Superhero', 'Bruce Wayne'):
        assert value in batman

    for value in ('Joker', 'Supervillain', 'Injustice League'):
        assert value in joker

    # fields in defined order
    csv = convert_list_of_dicts_to_csv(
        data, ('name', 'role', 'identity', 'gang'))

    header, batman, joker = csv.splitlines()
    assert header == 'name,role,identity,gang'
    assert batman == 'Batman,Superhero,Bruce Wayne,'
    assert joker == 'Joker,Supervillain,,Injustice League'


def test_convert_multiple_list_of_dicts_to_xlsx() -> None:
    data = [
        {
            'first_name': 'Dick',
            'last_name': 'Cheney',
            'plz': '3434',
        },
        {
            'first_name': 'Donald',
            'last_name': 'Rumsfeld',
            'plz': '3434',
        }
    ]
    data2 = [
        {
            'first_name': 'Dick',
            'last_name': 'Cheney',
            'plz': '3434',
        },
        {
            'first_name': 'Donald',
            'last_name': 'Rumsfeld',
            'plz': '3434',
        }
    ]
    xlsx = convert_list_of_list_of_dicts_to_xlsx([data, data2],
                                                 titles_list=["first",
                                                              "second",
                                                              "third"])

    with tempfile.NamedTemporaryFile(delete=False) as f:
        f.write(xlsx)
        for i in range(2):  # Loop over each tab
            rows = tuple(load_workbook(f).worksheets[i].rows)
            assert rows[0][0].value == 'first_name'
            assert rows[0][1].value == 'last_name'
            assert rows[0][2].value == 'plz'
            assert rows[1][0].value == 'Dick'
            assert rows[1][1].value == 'Cheney'
            assert rows[1][2].value == '3434'
            assert rows[2][0].value == 'Donald'
            assert rows[2][1].value == 'Rumsfeld'
            assert rows[2][2].value == '3434'


def test_xlsx_title_validation() -> None:
    # this is 36 chars, excel cannot load more than 31 chars
    title1 = "Schulhaus Schönwetter Gruppenraum 01"
    title2 = "Schulhaus Schönwetter Gruppenraum 02"
    titles = [title1, title2]
    norm = normalize_sheet_titles(titles)

    assert norm == ["schoenwetter-gruppenraum-01",
                    "schoenwetter-gruppenraum-02"]

    # if there are duplicates in the titles
    title1 = "Schulhaus Schönwetter Gruppenraum 01"
    title2 = "Schulhaus Schönwetter Gruppenraum 02"
    title3 = "Schulhaus Schönwetter Gruppenraum 02"
    titles = [title1, title2, title3]

    norm = normalize_sheet_titles(titles)

    # duplicates should be handled appropriately
    assert norm == ["schoenwetter-gruppenraum-01",
                    "schoenwetter-gruppenraum-02",
                    "schoenwetter-gruppenraum-02_1"]

    # even with very long names
    title1 = "Schulhaus Schönwetter Gruppenraum 01"
    title2 = "Schulhaus Schönwetter Gruppenraum 02"
    title3 = "Schulhaus Schönwetter Gruppenraum 02"
    titles = [title1, title2, title3]


def test_remove_first_word() -> None:
    titles = ["raum-zweiter-stock-mit-langem-namen-01",
              "raum-zweiter-stock-mit-langem-namen-02"]

    trimmed_titles = [remove_first_word(t) for t in titles]

    assert trimmed_titles[0] == "zweiter-stock-mit-langem-namen-01"
    assert trimmed_titles[1] == "zweiter-stock-mit-langem-namen-02"


def test_check_duplicates() -> None:
    titles = ["raum-zweiter-stock-mit-langem-namen-01",
              "raum-zweiter-stock-mit-langem-namen-011",
              "raum-zweiter-stock-mit-langem-namen-01",
              "raum-zweiter-stock-mit-langem-namen-02",
              "raum-zweiter-stock-mit-langem-namen-02"]

    duplicate_index = list_duplicates_index(titles)
    assert len(duplicate_index) == 2
    assert duplicate_index == [2, 4]


def test_avoid_duplicates() -> None:
    titles = ["Schulhaus-Schönwetter-Gruppenraum",
              "raum-zweiter-stock-mit-langem-namen-01",
              "raum-zweiter-stock-mit-langem-namen-01"]

    duplicate_index = list_duplicates_index(titles)

    for index in duplicate_index:
        item = titles[index]
        titles[index] = avoid_duplicate_name(titles, item)

    assert titles == ["Schulhaus-Schönwetter-Gruppenraum",
                      "raum-zweiter-stock-mit-langem-namen-01",
                      "raum-zweiter-stock-mit-langem-namen-01_1"]

    titles = ["raum-zweiter-stock-mit-langem-namen-01",
              "raum-zweiter-stock-mit-langem-namen-011",
              "raum-zweiter-stock-mit-langem-namen-01",
              "raum-zweiter-stock-mit-langem-namen-02",
              "raum-zweiter-stock-mit-langem-namen-02"]

    duplicate_index = list_duplicates_index(titles)  # [2, 4]

    for index in duplicate_index:
        item = titles[index]
        titles[index] = avoid_duplicate_name(titles, item)

    assert titles == ["raum-zweiter-stock-mit-langem-namen-01",
                      "raum-zweiter-stock-mit-langem-namen-011",
                      "raum-zweiter-stock-mit-langem-namen-01_2",
                      "raum-zweiter-stock-mit-langem-namen-02",
                      "raum-zweiter-stock-mit-langem-namen-02_1"]
