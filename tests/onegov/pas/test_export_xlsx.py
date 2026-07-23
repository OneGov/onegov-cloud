from __future__ import annotations

import csv
import transaction

from datetime import date
from decimal import Decimal
from io import BytesIO
from io import StringIO
from onegov.pas.models import (
    Attendence,
    PASParliamentarian,
    PASParliamentarianRole,
    Party,
    RateSet
)
from onegov.pas.models.commission import PASCommission
from onegov.pas.models.settlement_run import SettlementRun
from openpyxl import load_workbook


from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from tests.shared import Client
    from .conftest import TestPasApp


def test_export_xlsx(client: Client[TestPasApp]) -> None:

    client.login_admin()
    session = client.app.session()

    transaction.begin()
    # Setup test data
    rate_set = RateSet(
        year=2024,
        cost_of_living_adjustment=Decimal('2.0'),  # 2%
        plenary_none_member_halfday=Decimal('100.03'),
        commission_normal_member_initial=Decimal('50'),
        study_normal_member_halfhour=Decimal('20'),
        shortest_all_member_halfhour=Decimal('15')
    )
    session.add(rate_set)

    parliamentarian1 = PASParliamentarian(
        first_name='Peter',
        last_name='Muster',
        gender='male',
        personnel_number='PN123'
    )
    parliamentarian2 = PASParliamentarian(
        first_name='Petra',
        last_name='Musterfrau',
        gender='female',
        personnel_number='PN456'
    )
    session.add_all([parliamentarian1, parliamentarian2])

    party = Party(name='Test Party')
    session.add(party)

    role1 = PASParliamentarianRole(
        parliamentarian=parliamentarian1,
        role='member',
        party=party,
        start=date(2024, 1, 1)
    )
    role2 = PASParliamentarianRole(
        parliamentarian=parliamentarian2,
        role='member',
        party=party,
        start=date(2024, 1, 1)
    )
    session.add_all([role1, role2])

    commission = PASCommission(name='Finance Commission', type='normal')
    session.add(commission)

    attendances = [
        Attendence(
            parliamentarian=parliamentarian1,
            date=date(2024, 1, 10),
            duration=205,
            type='plenary',
        ),
        Attendence(
            parliamentarian=parliamentarian1,
            date=date(2024, 1, 11),
            duration=205,
            type='plenary',
        ),
        Attendence(
            parliamentarian=parliamentarian1, date=date(2024, 1, 15),
            duration=120, type='commission', commission=commission
        ),
        Attendence(
            parliamentarian=parliamentarian2, date=date(2024, 2, 5),
            duration=60, type='study', commission=commission
        ),
        Attendence(
            parliamentarian=parliamentarian2, date=date(2024, 2, 10),
            duration=30, type='shortest'
        ),
    ]
    session.add_all(attendances)

    settlement_run = SettlementRun(
        name='Q1 2024',
        start=date(2024, 1, 1),
        end=date(2024, 3, 31),
        active=True
    )
    session.add(settlement_run)
    transaction.commit()

    settl_id = session.query(SettlementRun).with_entities(
        SettlementRun.id).scalar()
    # Navigate to the settlement run page and click the export link
    page = client.get('/settlement-runs')
    page = client.get(f'/settlement-run/{settl_id}')
    assert page.status_code == 200

    export_link = page.pyquery(
        'a:contains("Abschlussliste (XLSX)")').attr('href')
    response = client.get(export_link)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.body), data_only=True)
    overview = workbook['Übersicht']
    details = workbook['Details']
    assert Decimal(str(overview.cell(row=2, column=6).value)) == Decimal(
        '200.10'
    )
    plenary_rows = [
        row
        for row in details.iter_rows(min_row=2, values_only=True)
        if row[0] in ('10.01.2024', '11.01.2024')
    ]
    assert [Decimal(str(row[8])) for row in plenary_rows] == [
        Decimal('100.05'),
        Decimal('100.05'),
    ]
    overview_rows = {
        row[0]: row for row in overview.iter_rows(min_row=2, values_only=True)
    }
    assert overview_rows['Musterfrau'][6] == 90
    assert Decimal(str(overview_rows['Musterfrau'][7])) == Decimal('55')

    details_rows = list(details.iter_rows(min_row=2, values_only=True))
    details_base_total = sum(
        (Decimal(str(row[8])) for row in details_rows),
        Decimal('0'),
    )
    overview_base_total = sum(
        (
            Decimal(str(row[5] or 0))
            + Decimal(str(row[7] or 0))
            + Decimal(str(row[8] or 0))
            for row in overview_rows.values()
        ),
        Decimal('0'),
    )
    assert overview_base_total == details_base_total

    export_link = page.pyquery(
        'a:contains("Buchungen Abrechnungslauf (Kontrollliste)")'
    ).attr('href')
    response = client.get(export_link)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.body), data_only=True)
    worksheet = workbook['Buchungen Abrechnungslauf']
    assert [worksheet.cell(row=row, column=6).value for row in (2, 3)] == [
        3.42,
        3.42,
    ]
    assert [worksheet.cell(row=row, column=7).value for row in (2, 3)] == [
        100.05,
        100.05,
    ]
    assert [worksheet.cell(row=row, column=8).value for row in (2, 3)] == [
        102.05,
        102.05,
    ]
    control_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
    assert len(control_rows) == len(details_rows)
    assert (
        sum(
            (Decimal(str(row[6])) for row in control_rows),
            Decimal('0'),
        )
        == details_base_total
    )

    export_link = page.pyquery('a:contains("KR-Entschädigungen (CSV)")').attr(
        'href'
    )
    response = client.get(export_link)

    assert response.status_code == 200
    fibu_rows = list(
        csv.reader(
            StringIO(response.body.decode('utf-8-sig')),
            delimiter=';',
        )
    )
    assert len(fibu_rows) == len(control_rows)
    assert sum(row[2] == '2421' for row in fibu_rows) == 1
    assert sum(
        (Decimal(row[12]) for row in fibu_rows),
        Decimal('0'),
    ) == sum(
        (Decimal(str(row[7])) for row in control_rows),
        Decimal('0'),
    )
