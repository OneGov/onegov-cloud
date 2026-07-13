from __future__ import annotations

import sys

import click
import transaction

from collections import OrderedDict

from onegov.core.cli import command_group
from onegov.core.cli import abort
from onegov.people.collections import PersonCollection
from onegov.people.models import Person
from openpyxl import load_workbook
from openpyxl import Workbook


from typing import IO
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from collections.abc import Callable
    from onegov.core.framework import Framework
    from onegov.core.request import CoreRequest


cli = command_group()


@cli.command('clear')
@click.option('--dry-run/--no-dry-run', default=False)
def clear(dry_run: bool) -> Callable[[CoreRequest, Framework], None]:
    """ Deletes all people.

    Does not support agencies and memberships at the moment.

    Example:

        onegov-people --select '/org/govikon' clear

    """

    def _clear(request: CoreRequest, app: Framework) -> None:
        if not click.confirm('Do you really want to remove all people?'):
            abort('Deletion process aborted')

        session = app.session()

        click.secho('Deleting people', fg='yellow')
        count = 0
        for person in session.query(Person):
            session.delete(person)
            count += 1

        if dry_run:
            transaction.abort()
            click.secho('Aborting transaction', fg='yellow')

        click.secho(f'Deleted {count} person(s)', fg='green')

    return _clear


EXPORT_FIELDS = OrderedDict((
    ('Anrede', 'salutation'),
    ('Akademischer Titel', 'academic_title'),
    ('Vorname', 'first_name'),
    ('Nachname', 'last_name'),
    ('Funktion', 'function'),
    ('E-Mail', 'email'),
    ('Telefon', 'phone'),
    ('Direktnummer', 'phone_direct'),
    ('Geboren', 'born'),
    ('Beruf', 'profession'),
    ('Politische Partei', 'political_party'),
    ('Fraktion', 'parliamentary_group'),
    ('Organisation', 'organisation'),
    ('Unterorganisation', 'sub_organisation'),
    ('Webseite', 'website'),
    ('Webseite 2', 'website_2'),
    ('Standort Adresse', 'location_address'),
    ('Standort Postleitzahl und Ort', 'location_code_city'),
    ('Postadresse', 'postal_address'),
    ('Postleitzahl und Ort', 'postal_code_city'),
    ('Link zum Bild', 'picture_url'),
    ('Notizen', 'notes'),
))


@cli.command('strip-whitespace-from-names')
@click.option('--dry-run/--no-dry-run', default=False)
def strip_whitespace_from_names(
    dry_run: bool
) -> Callable[[CoreRequest, Framework], None]:
    """ Strips leading/trailing whitespace from first_name, last_name,
    and function of all people.

    Example:

        `onegov-people --select /onegov_org/* strip-whitespace-from-names`
        `onegov-people --select /onegov_town6/* strip-whitespace-from-names`

    """

    def _strip(request: CoreRequest, app: Framework) -> None:
        session = app.session()
        count = 0
        for person in session.query(Person):
            first_name = person.first_name.strip()
            last_name = person.last_name.strip()
            function = (
                person.function.strip() if person.function else person.function
            )
            if (first_name, last_name, function) != (
                    person.first_name, person.last_name, person.function):
                person.first_name = first_name
                person.last_name = last_name
                person.function = function
                count += 1

        if dry_run:
            transaction.abort()
            click.secho('Aborting transaction', fg='yellow')

        click.secho(
            f'{app.schema}: Stripped whitespace from {count} person(s)',
            fg='green'
        )

    return _strip


@cli.command('export')
@click.argument('filename', type=click.Path(exists=False))
def export_xlsx(filename: str) -> Callable[[CoreRequest, Framework], None]:
    """ Exports all people to an excel file.

    Does not support agencies and memberships at the moment.

    Example:

        onegov-people --select '/onegov_org/govikon' export people.xlsx

    """

    def _export(request: CoreRequest, app: Framework) -> None:
        session = app.session()

        book = Workbook()
        book.remove(book.active)  # type:ignore[arg-type]

        click.secho('Exporting people', fg='yellow')
        sheet = book.create_sheet('Personen')
        sheet.append(tuple(EXPORT_FIELDS.keys()))
        count = 0
        for person in session.query(Person):
            sheet.append(
                tuple(getattr(person, attr) for attr in EXPORT_FIELDS.values())
            )
            count += 1

        book.save(filename)

        click.secho(f'Exported {count} person(s)', fg='green')

    return _export


@cli.command('import')
@click.argument('file', type=click.File('rb'))
def import_xlsx(file: IO[bytes]) -> Callable[[CoreRequest, Framework], None]:
    """ Imports people from an excel file.

    Does not support agencies and memberships at the moment.

    Example:

        onegov-people --select '/onegov_org/govikon' import people.xlsx

    """

    def _import(request: CoreRequest, app: Framework) -> None:
        session = app.session()

        book = load_workbook(file)

        click.secho('Importing people', fg='yellow')
        sheet = book['Personen']

        # FIXME: We should probably do this check at runtime eventually
        if TYPE_CHECKING:
            from openpyxl.worksheet.worksheet import Worksheet
            assert isinstance(sheet, Worksheet)

        count = 0
        for index, row in enumerate(sheet.rows):
            values = tuple(cell.value for cell in row)
            if not index:
                if values != tuple(EXPORT_FIELDS.keys()):
                    click.echo('Error in column headers')
                    click.echo('\nExpected - Current')
                    for exp, cur in zip(tuple(EXPORT_FIELDS.keys()), values):
                        color = 'green' if exp == cur else 'red'
                        click.secho(f'{exp} - {cur}', fg=color)
                    sys.exit('\nAborting import')
            else:
                session.add(
                    Person(**dict(zip(EXPORT_FIELDS.values(), values)))
                )
                count += 1

        click.secho(f'Imported {count} person(s)', fg='green')

    return _import


# Column headers that identify each Horw export format
_HORW_PERSONEN_HEADER = (
    'Name', 'Vorname', 'E-Mail', 'Internet', 'Funktion',
    'Strasse', 'Nr.', 'PLZ', 'Ort', 'Telefon Arbeit',
    'Fax Arbeit', 'Mobile Arbeit', 'Oberinstanzen', 'Instanzen', 'Behörden',
)

_HORW_BEHOERDEN_HEADER = (
    'Behörde', 'Strasse', 'Nr.', 'Postfach', 'PLZ', 'Ort',
    'Telefon', 'E-Mail', 'URL / Homepage', 'Kategorie',
    'Funktion', 'Name', 'Vorname', 'Präsentation', 'E-Mail',
    'Internet', 'Strasse', 'Nr.', 'Postfach', 'PLZ', 'Ort',
    'Telefon Arbeit', 'Mobile Arbeit', 'Telefon Privat', 'Mobile Privat',
    'Partei', 'Fraktion Einwohnerrat', 'Kategorie',
)

# Full-string replacements applied to the raw Instanzen value before
# comma-splitting. Use when the comma is part of the sub-org name.
_HORW_SUB_ORG_STRICT_MAP: dict[str, str] = {
    'Schulen der Gemeinde Horw, Rektorat': 'Gemeindeschule, Rektorat',
}

# Per-item replacements applied to each sub-org after comma-splitting.
_HORW_SUB_ORG_ITEM_MAP: dict[str, str] = {
    'AHV-Zweigstelle (Einwohnerdienste)': 'Einwohnerdienste',
}


def _v(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    return s or None


def _join_address(*parts: object) -> str | None:
    joined = ' '.join(s for p in parts if (s := _v(p)))
    return joined or None


def _read_excel_rows(path: str) -> list[tuple[object, ...]]:
    """Read all rows from an .xls or .xlsx file as tuples of cell values."""
    if path.lower().endswith('.xls'):
        import xlrd
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(0)
        return [tuple(sheet.row_values(i)) for i in range(sheet.nrows)]
    else:
        workbook = load_workbook(path, data_only=True)
        worksheet = workbook.active
        if TYPE_CHECKING:
            from openpyxl.worksheet.worksheet import Worksheet
            assert isinstance(worksheet, Worksheet)
        return [
            tuple(cell.value for cell in row) for row in worksheet.rows
        ]


def _parse_horw_personen_row(v: tuple[object, ...]) -> dict[str, object]:
    # Name | Vorname | E-Mail | Internet | Funktion | Strasse | Nr. |
    # PLZ | Ort | Telefon Arbeit | Fax Arbeit | Mobile Arbeit |
    # Oberinstanzen | Instanzen | Behörden
    return {
        'last_name': _v(v[0]) or '',
        'first_name': _v(v[1]) or '',
        'email': _v(v[2]),
        'website': _v(v[3]),
        'function': _v(v[4]),
        'location_address': _join_address(v[5], v[6]),
        'location_code_city': _join_address(v[7], v[8]),
        'phone': _v(v[9]),
        # v[10] = Fax Arbeit (no model field)
        'phone_direct': _v(v[11]) if len(v) > 11 else None,
        '_horw_org': _v(v[12]) if len(v) > 12 else None,     # Oberinstanzen
        '_horw_sub_org': _v(v[13]) if len(v) > 13 else None,  # Instanzen
        # v[14] = Behörden (not used for person fields)
    }


def _parse_horw_behoerden_row(v: tuple[object, ...]) -> dict[str, object]:
    # Cols 0-9: Behörde/agency data (skipped)
    # 10: Funktion | 11: Name | 12: Vorname | 13: Präsentation |
    # 14: E-Mail | 15: Internet | 16: Strasse | 17: Nr. | 18: Postfach |
    # 19: PLZ | 20: Ort | 21: Telefon Arbeit | 22: Mobile Arbeit |
    # 23: Telefon Privat | 24: Mobile Privat |
    # 25: Partei | 26: Fraktion Einwohnerrat | 27: Kategorie
    function = ' '.join(filter(None, (_v(v[10]), _v(v[0]))))
    return {
        'last_name': _v(v[11]) or '',
        'first_name': _v(v[12]) or '',
        'function': function or None,
        'notes': _v(v[13]),
        'email': _v(v[14]),
        'website': _v(v[15]),
        'location_address': _join_address(v[16], v[17]),
        'location_code_city': _join_address(v[19], v[20]),
        'phone': _v(v[21]),
        'phone_direct': _v(v[22]),
        # v[23] = Telefon Privat, v[24] = Mobile Privat (skip)
        'political_party': _v(v[25]) if len(v) > 25 else None,
        'parliamentary_group': _v(v[26]) if len(v) > 26 else None,
        # v[27] = Kategorie (skip)
    }


def _upsert_horw_person(
    people: PersonCollection,
    first_name: str,
    last_name: str,
    extra: dict[str, object],
) -> None:
    """Find existing person by email or name and update, or create new."""
    person = None
    email = extra.get('email')
    if email:
        person = (
            people.query()
            .filter(people.model_class.email == email)
            .first()
        )
    if person is None:
        person = (
            people.query()
            .filter(people.model_class.first_name == first_name)
            .filter(people.model_class.last_name == last_name)
            .first()
        )
    if person is not None:
        for key, value in extra.items():
            if key == 'function':
                existing = person.function or ''
                new = str(value) if value else ''
                if new and new not in existing:
                    person.function = (
                        f'{existing}; {new}' if existing else new
                    )
            else:
                setattr(person, key, value)
    else:
        people.add(first_name=first_name, last_name=last_name, **extra)


@cli.command('import-horw')
@click.argument('file', type=click.Path(exists=True))
@click.option('--dry-run', is_flag=True, default=False)
def import_horw(
    file: str,
    dry_run: bool,
) -> Callable[[CoreRequest, Framework], None]:
    """ Imports people from a Horw municipality Excel export.

    Detects the file format automatically from the header row. Supports
    both the Personen export (one person per row) and the Behörden export
    (one membership per row, people deduplicated by email then name).

    Example:

        onegov-people --select /onegov_town6/horw import-horw export.xls

    """

    def _import(request: CoreRequest, app: Framework) -> None:
        session = app.session()
        people = PersonCollection(session)

        all_rows = _read_excel_rows(file)
        if not all_rows:
            click.secho('Empty file', fg='red')
            return

        header = all_rows[0]

        if header == _HORW_PERSONEN_HEADER:
            parse_row = _parse_horw_personen_row
            is_personen = True
            deduplicate = False
            click.secho('Detected format: Personen', fg='yellow')
        elif header == _HORW_BEHOERDEN_HEADER:
            parse_row = _parse_horw_behoerden_row
            is_personen = False
            deduplicate = True
            click.secho('Detected format: Behörden', fg='yellow')
        else:
            click.secho('Unknown file format. Header columns:', fg='red')
            for col in header:
                click.secho(f'  {col!r}', fg='red')
            return

        # Build valid org → sub-orgs lookup from the configured hierarchy
        hierarchy = getattr(
            getattr(app, 'org', None), 'organisation_hierarchy', None
        ) or []
        valid_orgs: dict[str, set[str]] = {}
        for item in hierarchy:
            if isinstance(item, dict):
                for top, subs in item.items():
                    valid_orgs[top] = set(subs)
            elif isinstance(item, str):
                valid_orgs[item] = set()

        seen: set[tuple[str, str]] = set()
        count = 0
        errors = 0

        for row_num, values in enumerate(all_rows[1:], start=2):
            fields = parse_row(values)

            first_name = fields['first_name']
            last_name = fields['last_name']
            assert isinstance(first_name, str)
            assert isinstance(last_name, str)

            if not first_name and not last_name:
                continue

            email = str(fields.get('email') or '')
            key = (email, f'{last_name} {first_name}')
            if deduplicate and key in seen:
                continue
            seen.add(key)

            extra = {
                k: v for k, v in fields.items()
                if k not in ('first_name', 'last_name', '_horw_org',
                             '_horw_sub_org')
                and v is not None
            }

            if is_personen:
                raw_org = fields.get('_horw_org')
                raw_sub_org = fields.get('_horw_sub_org')
                if isinstance(raw_org, str) and raw_org:
                    top_orgs = [s for s in (
                        p.strip() for p in raw_org.split(',')
                    ) if s]
                    # validate top-level orgs; collect valid ones in order
                    valid_tops: list[str] = []
                    for top in top_orgs:
                        if valid_orgs and top not in valid_orgs:
                            click.secho(
                                f'Row {row_num} ({last_name} {first_name}): '
                                f'org {top!r} not in hierarchy',
                                fg='red')
                            errors += 1
                        else:
                            valid_tops.append(top)
                    # assign each sub-org to its parent top-level org
                    subs_by_top: dict[str, list[str]] = {
                        t: [] for t in valid_tops
                    }
                    if (
                        valid_tops
                        and isinstance(raw_sub_org, str)
                        and raw_sub_org
                    ):
                        raw_sub_org = _HORW_SUB_ORG_STRICT_MAP.get(
                            raw_sub_org, raw_sub_org
                        )
                        all_valid_subs: set[str] = set().union(
                            *(valid_orgs.get(t, set()) for t in valid_tops)
                        )
                        # try the full value first (handles names with commas),
                        # fall back to comma-splitting with per-item mapping
                        if raw_sub_org in all_valid_subs:
                            sub_candidates = [raw_sub_org]
                        else:
                            sub_candidates = list(dict.fromkeys(
                                _HORW_SUB_ORG_ITEM_MAP.get(s, s)
                                for s in (
                                    p.strip() for p in raw_sub_org.split(',')
                                ) if s
                            ))
                        for sub in sub_candidates:
                            parent = next(
                                (t for t in valid_tops
                                 if sub in valid_orgs.get(t, set())),
                                None
                            )
                            if parent is None:
                                click.secho(
                                    f'Row {row_num} ({last_name} '
                                    f'{first_name}): sub-org {sub!r} '
                                    f'not in hierarchy',
                                    fg='red')
                                errors += 1
                            else:
                                subs_by_top[parent].append(sub)
                    # build flat list: each org immediately followed
                    # by its subs
                    orgs: list[str] = []
                    for top in valid_tops:
                        orgs.append(top)
                        orgs.extend(f'-{s}' for s in subs_by_top[top])
                    if orgs:
                        extra['organisations_multiple'] = orgs

            _upsert_horw_person(people, first_name, last_name, extra)
            count += 1

        if dry_run:
            transaction.abort()
            click.secho(
                f'Dry run: would import {count} person(s)', fg='yellow')
        else:
            click.secho(f'Imported {count} person(s)', fg='green')
        if errors:
            click.secho(f'{errors} org/sub-org error(s)', fg='red')

    return _import


@cli.command('list')
def list_people() -> Callable[[CoreRequest, Framework], None]:

    def _list(request: CoreRequest, app: Framework) -> None:
        session = app.session()
        properties = [
            ('Funktion', 'function'),
            ('Standortadresse', 'location_address', 'location_code_city'),
            ('Postadresse', 'postal_address', 'postal_code_city'),
            ('E-Mail', 'email'),
            ('Telefon', 'phone'),
            ('Telefon direkt', 'phone_direct'),
            ('Telefon intern', 'phone_internal'),
        ]
        for p in session.query(Person):
            click.secho(f'{p.title}', fg='green')
            for label, *prop in properties:
                value = ', '.join(
                    value
                    for attr in prop
                    if (raw_value := getattr(p, attr, None)) is not None
                    if (value := str(raw_value))
                )
                if value:
                    click.secho(f'  {label}: {value}')

            click.secho()

    return _list
