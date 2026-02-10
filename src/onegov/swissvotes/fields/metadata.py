from __future__ import annotations

from decimal import Decimal
from onegov.form.fields import UploadField
from onegov.form.validators import FileSizeLimit
from onegov.swissvotes import _
from onegov.swissvotes.models import ColumnMapperMetadata
from openpyxl import load_workbook
from wtforms.validators import ValidationError

from typing import Any
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from collections.abc import Sequence
    from onegov.core.types import FileDict as StrictFileDict
    from onegov.form.types import FormT
    from onegov.form.types import Filter
    from onegov.form.types import PricingRules
    from onegov.form.types import Validators
    from onegov.form.types import Widget
    from typing import Self
    from wtforms.form import BaseForm
    from wtforms.meta import DefaultMeta
    from wtforms.meta import _SupportsGettextAndNgettext


class SwissvoteMetadataField(UploadField):
    """ An upload field expecting Swissvotes metadata (XLSX). """

    def __init__(
        self,
        label: str | None = None,
        validators: Validators[FormT, Self] | None = None,
        filters: Sequence[Filter] = (),
        description: str = '',
        id: str | None = None,
        default: Sequence[StrictFileDict] | None = (),
        widget: Widget[Self] | None = None,
        render_kw: dict[str, Any] | None = None,
        name: str | None = None,
        _form: BaseForm | None = None,
        _prefix: str = '',
        _translations: _SupportsGettextAndNgettext | None = None,
        _meta: DefaultMeta | None = None,
        # onegov specific kwargs that get popped off
        *,
        fieldset: str | None = None,
        depends_on: Sequence[Any] | None = None,
        pricing: PricingRules | None = None,
    ) -> None:
        kwargs: dict[str, Any] = {}
        kwargs.setdefault('validators', [])
        kwargs['allowed_mimetypes'] = {
            'application/excel',
            'application/octet-stream',
            'application/vnd.ms-excel',
            'application/vnd.ms-office',
            (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
            'application/zip'
        }
        kwargs['validators'].append(FileSizeLimit(10 * 1024 * 1024))
        kwargs.setdefault('render_kw', {})
        kwargs['render_kw']['force_simple'] = True

        super().__init__(
            label=label,
            validators=validators,
            filters=filters,
            description=description,
            id=id,
            default=default,
            widget=widget,
            render_kw=render_kw,
            name=name,
            _form=_form,
            _prefix=_prefix,
            _translations=_translations,
            _meta=_meta,
        )

    def post_validate(
        self,
        form: BaseForm,
        validation_stopped: bool
    ) -> None:
        """ Make sure the given XLSX is valid (all expected columns are
        present all cells contain reasonable values).

        Converts the XLSX to a list of metadata dictionaries objects,
        available as ``data``.

        """

        super().post_validate(form, validation_stopped)
        if validation_stopped:
            return

        assert self.file is not None
        errors = []
        data: dict[Decimal, dict[str, Any]] = {}
        mapper = ColumnMapperMetadata()

        try:
            workbook = load_workbook(self.file, data_only=True)
        except Exception as exception:
            raise ValidationError(_('Not a valid XLSX file.')) from exception

        if len(workbook.worksheets) < 1:
            raise ValidationError(_('No data.'))

        if 'Metadaten zu Scans' not in workbook.sheetnames:
            raise ValidationError(_("Sheet 'Metadaten zu Scans' is missing."))

        sheet = workbook['Metadaten zu Scans']

        if TYPE_CHECKING:
            from openpyxl.worksheet.worksheet import Worksheet
            assert isinstance(sheet, Worksheet)

        if sheet.max_row <= 1:
            raise ValidationError(_('No data.'))

        headers = [column.value for column in next(sheet.rows)]
        missing = set(mapper.columns.values()) - set(headers)  # type:ignore
        if missing:
            raise ValidationError(_(
                'Some columns are missing: ${columns}.',
                mapping={'columns': ', '.join(missing)}
            ))

        value: Any | None
        for index in range(2, sheet.max_row + 1):
            metadata: dict[str, Any] = {}
            all_columns_empty = True
            column_errors = []
            for (
                attribute, column, type_, nullable, precision, scale
            ) in mapper.items():
                cell = sheet.cell(index, headers.index(column) + 1)
                try:
                    if cell.value is None:
                        value = None
                    elif type_ == 'TEXT':
                        if (
                            cell.data_type == 'n'
                            and int(cell.value) == cell.value  # type:ignore
                        ):
                            value = str(int(cell.value))  # type:ignore
                        else:
                            value = str(cell.value)
                        value = '' if value == '.' else value
                    elif type_ == 'INTEGER':
                        if cell.data_type == 's':
                            value = cell.value
                            value = '' if value == '.' else value
                            value = int(
                                value  # type:ignore[arg-type]
                            ) if value else None
                        else:
                            value = int(cell.value)  # type:ignore[arg-type]
                    elif type_ and type_.startswith('NUMERIC'):
                        if isinstance(cell.value, str):
                            value = cell.value
                            value = '' if value == '.' else value
                            value = Decimal(str(value)) if value else None
                        else:
                            value = Decimal(str(cell.value))
                        if value is not None:
                            value = Decimal(
                                format(value, f'{precision}.{scale}f')
                            )

                    all_columns_empty = all_columns_empty and value is None

                except Exception:
                    column_errors.append((
                        index, column, f"'{value}' ≠ {type_ and type_.lower()}"
                    ))

                else:
                    if not nullable and value is None:
                        column_errors.append((index, column, '∅'))
                    mapper.set_value(metadata, attribute, value)

            if not all_columns_empty:
                errors.extend(column_errors)
                if not column_errors:
                    bfs_number = metadata['bfs_number']
                    filename = metadata['filename']
                    data.setdefault(bfs_number, {})[filename] = metadata

        if errors:
            raise ValidationError(_(
                'Some cells contain invalid values: ${errors}.',
                mapping={
                    'errors': '; '.join(
                        '{}:{} {}'.format(*error) for error in errors
                    )
                }
            ))

        self.data = data
