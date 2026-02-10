from __future__ import annotations

from dateutil.parser import parse
from decimal import Decimal
from onegov.form.fields import UploadField
from onegov.form.validators import FileSizeLimit
from onegov.swissvotes import _
from onegov.swissvotes.models import ColumnMapperDataset
from onegov.swissvotes.models import SwissVote
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from wtforms.validators import ValidationError


from typing import Any
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from collections.abc import Sequence
    from onegov.core.types import FileDict as StrictFileDict
    from onegov.form.types import FormT
    from onegov.form.types import Filter
    from onegov.form.types import PricingRules
    from onegov.form.types import Validators
    from onegov.form.types import Widget
    from typing import Self
    from wtforms.form import BaseForm
    from wtforms.meta import _SupportsGettextAndNgettext
    from wtforms.meta import DefaultMeta


class SwissvoteDatasetField(UploadField):
    """ An upload field expecting a Swissvotes dataset (XLSX). """

    def __init__(
        self,
        label: str | None = None,
        validators: Validators[FormT, Self] | None = None,
        filters: Sequence[Filter] = (),
        description: str = '',
        id: str | None = None,
        default: Sequence[StrictFileDict] = (),
        widget: Widget[Self] | None = None,
        render_kw: dict[str, Any] | None = None,
        name: str | None = None,
        _form: BaseForm | None = None,
        _prefix: str = '',
        _translations: _SupportsGettextAndNgettext | None = None,
        _meta: DefaultMeta | None = None,
        # onegov specific kwargs that get popped off
        *,
        fieldset: str | None = None,
        depends_on: Sequence[Any] | None = None,
        pricing: PricingRules | None = None,
    ) -> None:

        if validators:
            validator_list = list(validators)
        else:
            validator_list = []
        validator_list.append(FileSizeLimit(10 * 1024 * 1024))

        if not render_kw:
            render_kw = {}
        render_kw['force_simple'] = True

        mimetypes = {
            'application/excel',
            'application/octet-stream',
            'application/vnd.ms-excel',
            'application/vnd.ms-office',
            (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
            'application/zip'
        }

        super().__init__(
            label=label,
            validators=validator_list,
            filters=filters,
            description=description,
            id=id,
            default=default,
            widget=widget,
            render_kw=render_kw,
            name=name,
            allowed_mimetypes=mimetypes,
            _form=_form,
            _prefix=_prefix,
            _translations=_translations,
            _meta=_meta,
            fieldset=fieldset,
            depends_on=depends_on,
            pricing=pricing
        )

    data: list[SwissVote]  # type:ignore[assignment]

    def post_validate(
        self,
        form: BaseForm,
        validation_stopped: bool
    ) -> None:
        """ Make sure the given XLSX is valid (all expected columns are
        present all cells contain reasonable values).

        Converts the XLSX to a list of SwissVote objects, available as
        ``data``.

        """

        if validation_stopped:
            return

        assert self.file is not None
        errors = []
        data = []
        mapper = ColumnMapperDataset()

        try:
            workbook = load_workbook(self.file, data_only=True)
        except Exception as exception:
            raise ValidationError(_('Not a valid XLSX file.')) from exception

        if len(workbook.worksheets) < 1:
            raise ValidationError(_('No data.'))

        if 'DATA' not in workbook.sheetnames:
            raise ValidationError(_('Sheet DATA is missing.'))

        sheet = workbook['DATA']

        if TYPE_CHECKING:
            from openpyxl.worksheet.worksheet import Worksheet
            assert isinstance(sheet, Worksheet)

        if sheet.max_row <= 1:
            raise ValidationError(_('No data.'))

        headers = [column.value for column in next(sheet.rows)]
        missing = set(mapper.columns.values()) - set(headers)  # type:ignore
        if missing:
            raise ValidationError(_(
                'Some columns are missing: ${columns}.',
                mapping={'columns': ', '.join(missing)}
            ))

        super().post_validate(form, validation_stopped)

        value: Any | None
        for index in range(2, sheet.max_row + 1):
            vote = SwissVote()
            all_columns_empty = True
            column_errors = []
            for (
                attribute, column, type_, nullable, precision, scale
            ) in mapper.items():
                cell = sheet.cell(index, headers.index(column) + 1)
                try:
                    if cell.value is None:
                        value = None
                    elif type_ == 'TEXT':
                        if (
                            cell.data_type == 'n'
                            and int(cell.value) == cell.value  # type:ignore
                        ):
                            value = str(int(cell.value))  # type:ignore
                        else:
                            value = str(cell.value)
                        value = '' if value == '.' else value
                    elif type_ == 'DATE':
                        if cell.data_type == 's':
                            value = parse(
                                cell.value,  # type:ignore[arg-type]
                                dayfirst=True
                            ).date()
                        elif cell.data_type == 'n':
                            value = from_excel(cell.value).date()
                        elif cell.data_type == 'd':
                            assert hasattr(cell.value, 'date')
                            value = cell.value.date()
                        else:
                            raise ValidationError('Not a valid date format')
                    elif type_ == 'INTEGER':
                        if cell.data_type == 's':
                            value = cell.value
                            value = '' if value == '.' else value
                            value = int(
                                value  # type:ignore[arg-type]
                            ) if value else None
                        else:
                            value = int(cell.value)  # type:ignore
                    elif type_ and type_.startswith('NUMERIC'):
                        if isinstance(cell.value, str):
                            value = cell.value
                            value = '' if value == '.' else value
                            value = Decimal(str(value)) if value else None
                        else:
                            value = Decimal(str(cell.value))
                        if value is not None:
                            value = Decimal(
                                format(value, f'{precision}.{scale}f')
                            )

                    all_columns_empty = all_columns_empty and value is None

                except Exception:
                    errors.append((
                        index, column, f"'{value}' ≠ {type_ and type_.lower()}"
                    ))

                else:
                    if not nullable and value is None:
                        column_errors.append((index, column, '∅'))
                    mapper.set_value(vote, attribute, value)

            if not all_columns_empty:
                errors.extend(column_errors)
                data.append(vote)

        if errors:
            raise ValidationError(_(
                'Some cells contain invalid values: ${errors}.',
                mapping={
                    'errors': '; '.join(
                        '{}:{} {}'.format(*error) for error in errors
                    )
                }
            ))

        self.data = data
