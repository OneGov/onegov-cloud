from __future__ import annotations

import csv
import openpyxl
from dataclasses import dataclass
from datetime import datetime, date
from onegov.core.csv import CSVFile, convert_excel_to_csv, detect_encoding
from onegov.pas.models import (
    PASCommission,
    PASCommissionMembership,
    PASParliamentarian,
    PASParliamentarianRole,
    PASParliamentaryGroup,
    Party,
)
from pathlib import Path
from tempfile import NamedTemporaryFile


from typing import (
    Any,
    Any as Incomplete,
    TypeVar,
    BinaryIO,
    cast,
    Protocol,
    ParamSpec,
    Self,
)
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from _typeshed import StrOrBytesPath
    from collections.abc import Callable
    from onegov.parliament.models.commission_membership import MembershipRole
    from sqlalchemy.orm import Session
    from types import TracebackType

    class Row(Protocol):
        adress_anrede: str
        akademischer_titel: str
        anrede: str
        austritt_kommission: str
        bemerkungen: str
        beruf: str
        brief_anrede: str
        burgerort: str
        count: str
        e_mail_1: str
        e_mail_2: str
        eintritt_kommission: str
        fraktion: str
        geburtsdatum: str
        geschlecht: str
        id: str
        index: str
        nachname: str
        partei: str
        personalnummer: str
        privat_adresse: str
        privat_adresszusatz: str
        privat_ort: str
        privat_plz: str
        rolle_kommission: str
        rownumber: int
        spedition_kr_vorlagen: str
        telefon_geschaft: str
        telefon_mobile: str
        telefon_privat: str
        versand_adresse: str
        versand_adresszusatz: str
        versand_ort: str
        versand_plz: str
        versandart: str
        vertragsnummer: str
        vorname: str
        wahlkreis: str
        webseite: str
        zusatzinformationen: str

T = TypeVar('T')
P = ParamSpec('P')

# For commission import, Callable keys on the Row object of CSV files.
EXPECTED_HEADERS = [
    'adress_anrede', 'akademischer_titel', 'anrede', 'austritt_kommission',
    'bemerkungen', 'beruf', 'brief_anrede', 'burgerort', 'count',
    'e_mail_1', 'e_mail_2', 'eintritt_kommission', 'fraktion',
    'geburtsdatum', 'geschlecht', 'id', 'index', 'nachname', 'partei',
    'personalnummer', 'privat_adresse', 'privat_adresszusatz',
    'privat_ort', 'privat_plz', 'rolle_kommission', 'rownumber',
    'spedition_kr_vorlagen', 'telefon_geschaft', 'telefon_mobile',
    'telefon_privat', 'versand_adresse', 'versand_adresszusatz',
    'versand_ort', 'versand_plz', 'versandart', 'vertragsnummer',
    'vorname', 'wahlkreis', 'webseite', 'zusatzinformationen'
]


class ImportFile:
    """Provides a unified interface for both CSV and Excel files."""

    def __init__(self, file: BinaryIO) -> None:
        self.ENCODINGS = ['utf-8', 'iso-8859-1']
        self.file = file
        self._detect_type()
        self.rows: list[Incomplete] = []
        self._parse()

    def _detect_type(self) -> None:
        """Detect if file is CSV or Excel based on content."""
        header = self.file.read(4)
        self.file.seek(0)
        self.is_excel = header.startswith(b'PK')

    def _parse(self) -> None:
        """Parse the file contents based on detected type."""
        if self.is_excel:
            self._parse_excel()
        else:
            self._parse_csv()

    def _parse_csv(self) -> None:
        encoding = detect_encoding(self.file)
        self.file.seek(0)
        csv_file = CSVFile(
            self.file,
            encoding=encoding,
        )
        self.rows = list(csv_file.lines)

    def _parse_excel(self) -> None:
        """Parse the file contents for excel files."""
        csv_file_like = convert_excel_to_csv(self.file)
        encoding = detect_encoding(csv_file_like)
        csv_file_like.seek(0)
        csv_file = CSVFile(
            csv_file_like,
            encoding=encoding,
        )
        self.rows = list(csv_file.lines)

    def __enter__(self: Self) ->  Self:
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: TracebackType | None,
    ) -> None:
        self.file.close()


def preprocess_headers(
    headers: list[str], header_renames: dict[str, str]
) -> list[str]:
    """Helper function to process headers using rename mapping."""
    return [header_renames.get(header, header) for header in headers]


def preprocess_csv_headers(
    csv_path_abs: StrOrBytesPath, expected: list[str] | None = None
) -> StrOrBytesPath:
    """
    Preprocesses a CSV file to rename specific headers to avoid issues
    with the CSV parser. Creates a temporary file with the modified headers.
    Renames:
        "1. E-Mail" to "E_Mail_1"
        "2. E-Mail" to "E_Mail_2"
    Returns:
        Path to the temporary CSV file with preprocessed headers.
    """

    with NamedTemporaryFile(
            mode='w+t',
            suffix='.csv',
            delete=False,
            encoding='utf-8',
            newline='',
    ) as temp_file, open(csv_path_abs, encoding='utf-8', newline='') as infile:
        temp_file_path = temp_file.name
        reader = csv.reader(infile)
        writer = csv.writer(temp_file)
        header_row = next(reader)

        match_expected_headers_or_fail(
            expected,
            header_row
        )

        header_renames = {
            '1. E-Mail': 'E_Mail_1',
            '2. E-Mail': 'E_Mail_2',
        }
        modified_header_row = [
            header_renames.get(header, header) for header in header_row
        ]
        writer.writerow(modified_header_row)
        writer.writerows(reader)
    return temp_file_path


@dataclass
class HeaderValidationResult:
    is_valid: bool
    missing_headers: list[str]
    unexpected_headers: list[str]
    original_headers: list[str]


def validate_headers(
    current_headers: list[str], expected_headers: list[str]
) -> HeaderValidationResult:
    current_set = set(current_headers)
    expected_set = set(expected_headers)
    missing = list(expected_set - current_set)
    unexpected = list(current_set - expected_set)
    return HeaderValidationResult(
        is_valid=not (missing or unexpected),
        missing_headers=sorted(missing),
        unexpected_headers=sorted(unexpected),
        original_headers=current_headers,
    )


def preprocess_excel_headers(
    excel_path: StrOrBytesPath,
    expected: list[str] | None = None
) -> StrOrBytesPath:
    """
    Preprocess Excel headers using a context manager for temporary file
    handling.

    Args:
        excel_path: Path to the input Excel file
        expected: Optional list of expected headers to validate against

    Returns:
        Path to the processed temporary file

    Raises:
        Exception: If header processing or validation fails
    """
    header_renames = {
        '1. E-Mail': 'E_Mail_1',
        '2. E-Mail': 'E_Mail_2',
    }

    with NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        temp_file_path = temp_file.name

        try:
            wb = openpyxl.load_workbook(excel_path)  # type: ignore[arg-type]

            for sheet in wb.worksheets:
                if sheet.max_row > 0:
                    header_row = [str(cell.value) for cell in sheet[1]]
                    match_expected_headers_or_fail(expected, header_row)
                    modified_headers = preprocess_headers(
                        header_row,
                        header_renames,
                    )

                    # Update headers in the first row
                    for col, header in enumerate(modified_headers, 1):
                        sheet.cell(row=1, column=col, value=header)

            wb.save(temp_file_path)
            wb.close()
            return temp_file_path

        except Exception as e:
            # Clean up the temporary file if an error occurs
            Path(temp_file_path).unlink(missing_ok=True)
            raise e from None  # Re-raise the exception with a clean traceback


def match_expected_headers_or_fail(
    expected: Incomplete,
    header_row: Incomplete
) -> None:
    # Validate headers if expected headers are provided
    if expected is not None:
        validation_results = validate_headers(header_row, expected)
        if not validation_results.is_valid:
            error_msg = []
            if validation_results.missing_headers:
                error_msg.append(
                    f"Missing "
                    f"headers: {', '.join(validation_results.missing_headers)}"
                )
            if validation_results.unexpected_headers:
                error_msg.append(
                    f"Unexpected "
                    f"headers: "
                    f"{', '.join(validation_results.unexpected_headers)}"
                )
            error_msg.append(
                f"Original headers:"
                f" {', '.join(validation_results.original_headers)}"
            )
            raise ValueError('\n'.join(error_msg))


def with_open_excel_or_csv(
    func: Callable[..., T]
) -> Callable[..., T]:
    """
    Decorator to handle opening and parsing import files.
    Handles both CSV and Excel files, preprocessing headers as needed.
    """

    def wrapper(
        filename: str,
        *args: tuple[Any, ...],
        expected_headers: list[str],
    ) -> T:

        if filename.lower().endswith(('.xls', '.xlsx')):
            preprocessed_filename = preprocess_excel_headers(
                filename, expected_headers
            )
        else:
            preprocessed_filename = preprocess_csv_headers(
                filename,
                expected_headers,
            )

        with open(preprocessed_filename, 'rb') as f, ImportFile(
            f
        ) as import_file:
            try:
                result = func(
                    import_file,
                    *args,
                    expected_headers=expected_headers,
                )
            finally:
                if preprocessed_filename != filename:
                    import os

                    os.remove(preprocessed_filename)
            return result
    return wrapper


@with_open_excel_or_csv
def import_commissions(
    import_file: ImportFile,
    session: Session,
    commission_file_abs_path: str,
    **kwargs: Any
) -> None:
    """Imports all data from commission file."""

    def parse_date(date_str: str | None) -> date | None:
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str, '%d.%m.%Y').date()
        except ValueError:
            return None

    # reverse transl. mapping
    role_translations = {
        'mitglied': 'member',
        'gast': 'guest',
        'erweitertes mitglied': 'extended_member',
        'präsident': 'president',
        'präsidentin': 'president',
    }

    commission_name = Path(commission_file_abs_path).stem

    # if the file originally contains underscores, make them spaces for display
    commission_name = commission_name.replace('_', ' ')

    # Get or create commission
    commission = session.query(PASCommission).filter_by(
        name=commission_name
    ).first()

    if not commission:
        commission = PASCommission(
            name=commission_name,
            type='normal'
        )
        session.add(commission)

    # First pass - create parties and parliamentary groups
    for row in import_file.rows:
        # Create party if needed
        party = session.query(Party).filter_by(name=row.partei).first()
        if not party:
            party = Party(name=row.partei)
            session.add(party)

        # Create parliamentary group if needed
        group = session.query(PASParliamentaryGroup).filter_by(
            name=row.fraktion
        ).first()
        if not group:
            group = PASParliamentaryGroup(name=row.fraktion)
            session.add(group)

    session.flush()

    # Second pass - create parliamentarians and memberships
    for row in import_file.rows:
        # Get party and group
        party = session.query(Party).filter_by(name=row.partei).one()
        group = session.query(PASParliamentaryGroup).filter_by(
            name=row.fraktion
        ).one()

        # Create parliamentarian if needed
        parliamentarian = session.query(PASParliamentarian).filter_by(
            personnel_number=row.personalnummer
        ).first()

        if not parliamentarian:
            parliamentarian = PASParliamentarian(
                personnel_number=row.personalnummer,
                contract_number=row.vertragsnummer,
                first_name=row.vorname,
                last_name=row.nachname,
                gender='female' if row.geschlecht == 'Weiblich' else 'male',
                shipping_method='a',
                shipping_address=row.versand_adresse,
                shipping_address_addition=row.versand_adresszusatz,
                shipping_address_zip_code=row.versand_plz,
                shipping_address_city=row.versand_ort,
                private_address=row.privat_adresse,
                private_address_addition=row.privat_adresszusatz,
                private_address_zip_code=row.privat_plz,
                private_address_city=row.privat_ort,
                date_of_birth=parse_date(row.geburtsdatum),
                place_of_origin=row.burgerort,
                occupation=row.beruf,
                academic_title=row.akademischer_titel,
                salutation=row.anrede,
                salutation_for_address=row.adress_anrede,
                salutation_for_letter=row.brief_anrede,
                forwarding_of_bills=row.spedition_kr_vorlagen,
                phone_private=row.telefon_privat,
                phone_mobile=row.telefon_mobile,
                phone_business=row.telefon_geschaft,
                email_primary=row.e_mail_1,
                email_secondary=row.e_mail_2,
                website=row.webseite,
                remarks=row.bemerkungen
            )
            session.add(parliamentarian)

            # Create roles linking to party and group
            parliamentarian.roles.append(PASParliamentarianRole(
                party=party,
                party_role='member',
                parliamentary_group=group,
                parliamentary_group_role='member',
                district=row.wahlkreis,
                additional_information=row.zusatzinformationen
            ))

        # Translate the German role to English
        try:
            role = role_translations.get(
                row.rolle_kommission.lower(),
                'member',
            )
            # Create commission membership
            membership = PASCommissionMembership(
                commission=commission,
                parliamentarian=parliamentarian,
                role=cast('MembershipRole', role),
                start=parse_date(row.eintritt_kommission),
                end=parse_date(row.austritt_kommission)
            )
            session.add(membership)
        except AttributeError:
            pass

    session.flush()
