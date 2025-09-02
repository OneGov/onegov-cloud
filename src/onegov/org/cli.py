""" Provides commands used to initialize org websites. """
from __future__ import annotations
import base64
import json

import click
import html
import isodate
import re
import shutil
import sys
import textwrap
from functools import wraps

from bs4 import BeautifulSoup
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO

import pytz
import locale
import requests
import transaction
import yaml
from onegov.core.orm.utils import QueryChain
from libres.modules.errors import (InvalidEmailAddress, AlreadyReservedError,
                                   TimerangeTooLong)
from markupsafe import Markup
from onegov.chat import MessageCollection
from onegov.core.cli import command_group, pass_group_context, abort
from onegov.core.crypto import random_token
from onegov.core.utils import Bunch
from onegov.directory import DirectoryEntry
from onegov.directory.models.directory import DirectoryFile
from onegov.event import Event, Occurrence, EventCollection
from onegov.event.collections.events import EventImportItem
from onegov.file import File
from onegov.file.collection import FileCollection
from onegov.form import FormCollection, FormDefinition
from onegov.org import log
from onegov.org.formats import DigirezDB
from onegov.org.forms.event import TAGS
from onegov.org.management import LinkMigration
from onegov.org.models.file import ImageFileCollection
from onegov.org.models.page import Page
from onegov.org.models import ExtendedDirectory
from onegov.org.models import Organisation, TicketNote, TicketMessage
from onegov.org.models.resource import Resource
from onegov.org.models import (
    MeetingCollection,
    Meeting,
    MeetingItem,
    MeetingItemCollection,
    PoliticalBusiness,
    PoliticalBusinessCollection,
    PoliticalBusinessParticipationCollection,
    RISCommissionCollection,
    RISCommissionMembership,
    RISCommissionMembershipCollection,
    RISParliamentarian,
    RISParliamentarianCollection,
    RISParliamentarianRole,
    RISParliamentarianRoleCollection,
    RISParliamentaryGroupCollection,
)
from onegov.page.collection import PageCollection
from onegov.parliament.collections import ParliamentarianCollection
from onegov.reservation import ResourceCollection
from onegov.ticket import TicketCollection
from onegov.town6.upgrade import migrate_homepage_structure_for_town6
from onegov.town6.upgrade import migrate_theme_options
from onegov.user.models import TAN
from onegov.user import UserCollection, User
from openpyxl import load_workbook
from operator import add as add_op
from pathlib import Path
from sqlalchemy import func, and_, or_
from sqlalchemy.dialects.postgresql import array
from uuid import uuid4
from elasticsearch.exceptions import ConnectionError as ESConnectionError


from typing import IO, Any, TYPE_CHECKING, TypedDict
if TYPE_CHECKING:
    from collections.abc import Callable, Iterator, Sequence
    from depot.fields.upload import UploadedFile
    from onegov.core.cli.core import GroupContext
    from onegov.core.csv import DefaultRow
    from onegov.core.upgrade import UpgradeContext
    from onegov.org.app import OrgApp
    from onegov.org.request import OrgRequest
    from onegov.ticket import Ticket
    from sqlalchemy.orm import Query, Session

    from translationstring import TranslationString
    from uuid import UUID
    from onegov.org.models.political_business import (
        PoliticalBusinessStatus,
        PoliticalBusiness
    )

cli = command_group()


@cli.command(context_settings={'creates_path': True})
@click.argument('name')
@click.option(
    '--locale',
    default='de_CH',
    type=click.Choice(['de_CH', 'fr_CH', 'it_CH'])
)
@pass_group_context
def add(
    group_context: GroupContext,
    name: str,
    locale: str
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Adds an org with the given name to the database. For example:

    .. code-block:: bash

        onegov-org --select '/onegov_org/evilcorp' add "Evilcorp"

    """

    def add_org(request: OrgRequest, app: OrgApp) -> None:

        if app.session().query(Organisation).first():
            abort('{} already contains an organisation'.format(
                group_context.selector))

        app.settings.org.create_new_organisation(app, name, locale=locale)

        click.echo('{} was created successfully'.format(name))

    return add_org


@cli.command(name='import-digirez', context_settings={'singular': True})
@click.argument('accessdb', type=click.Path(exists=True, resolve_path=True))
@click.option('--min-date', default=None,
              help="Min date in the form '2016-12-31'")
@click.option('--ignore-booking-conflicts', default=False, is_flag=True,
              help='Ignore booking conflicts (TESTING ONlY)')
def import_digirez(
    accessdb: str,
    min_date: str,
    ignore_booking_conflicts: bool
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Imports a Digirez reservation database into onegov.org.

    Example:
    .. code-block:: bash

        onegov-org --select '/orgs/govikon' import-digirez room_booking.mdb

    """

    if not shutil.which('mdb-export'):
        abort("Could not find 'mdb-export', please install mdbtools!")

    min_date_ = min_date and isodate.parse_date(min_date) or date.today()
    min_dt = datetime(min_date_.year, min_date_.month, min_date_.day)

    db = DigirezDB(accessdb)
    db.open()

    # XXX this is currently specific to what Gemeinde Rüti has, we could put
    # this into an external file however.
    form_definition = textwrap.dedent("""
        Name *= ___
        Telefon =___
        Zweck *= ___
        Einrichtungen =
            [ ] Konzertbestuhlung
            [ ] Bankettbestuhlung
            [ ] Flipchart
            [ ] Hellraumprojektor
            [ ] Kinoleinwand (Saal)
            [ ] Mobile Leinwand
            [ ] Lautsprecheranlage
            [ ] Podestelemente
            [ ] Flügel (Saal)
            [ ] Getränke
            [ ] Küche (Saal)
            [ ] Office (Keller)
        Bemerkungen = ...
    """)

    floor_hours = {
        'Amthaus': (8, 24),
        'Löwen': (0, 24),
        'Gemeinschaftszentrum': (8, 24)
    }

    title_map = {
        'Jugend- raum': 'Jugendraum',
        'Spielgruppen- Raum': 'Spielgruppenraum',
        'Vereins- raum': 'Vereinsraum'
    }

    def get_formdata(
        member: DefaultRow,
        booking: DefaultRow
    ) -> dict[str, Any]:

        einrichtungen = []

        if booking.kbest == '1':
            einrichtungen.append('Konzertbestuhlung')

        if booking.bbest == '1':
            einrichtungen.append('Bankettbestuhlung')

        if booking.flip == '1':
            einrichtungen.append('Flipchart')

        if booking.ohproj == '1':
            einrichtungen.append('Hellraumprojektor')

        if booking.klw == '1':
            einrichtungen.append('Kinoleinwand (Saal)')

        if booking.mlw == '1':
            einrichtungen.append('Mobile Leinwand')

        if booking.lspr == '1':
            einrichtungen.append('Lautsprecheranlage')

        if booking.pod == '1':
            einrichtungen.append('Podestelemente')

        if booking.flg == '1':
            einrichtungen.append('Flügel (Saal)')

        if booking.getr == '1':
            einrichtungen.append('Getränke')

        if booking.getr == '1':
            einrichtungen.append('Küche (Saal)')

        if booking.ofk == '1':
            einrichtungen.append('Office (Keller)')

        return {
            'name': booking.anspname or (
                f'{member.member_last_name} {member.member_first_name}'),
            'telefon': booking.ansptel or member.member_phone,
            'zweck': booking.title,
            'bemerkungen': booking.description,
            'einrichtungen': einrichtungen
        }

    def unescape_dictionary(d: dict[str, Any]) -> dict[str, Any]:
        for key, value in d.items():
            if isinstance(value, str):
                d[key] = html.unescape(value)
            if isinstance(value, dict):
                d[key] = unescape_dictionary(value)
            if isinstance(value, list):
                d[key] = [html.unescape(v) for v in value]

        return d

    email_fixes = (
        (re.compile(r'2da-capo-rueti.ch$'), '@da-capo.rueti.ch'),
        (re.compile(r'@mzol$'), '@mzol.ch'),
        (re.compile(r'^jo$'), 'info@rueti.ch'),
        (re.compile(r'@pro-vitalis$'), '@pro-vitalis.ch')
    )

    def run_import(request: OrgRequest, app: OrgApp) -> None:

        # create all resources first, fails if at least one exists already
        click.echo('Creating resources')

        resources = ResourceCollection(app.libres_context)
        floors = {f.id: f.floor_name for f in db.records.floors}
        resources_by_room = {}

        for room in db.records.room:
            resource = resources.add(
                title=title_map.get(room.room_name, room.room_name),
                timezone='Europe/Zurich',
                type='room',
                group=floors[room.floor_id],
                definition=form_definition,
                meta={'access': 'private'}
            )

            resources_by_room[room.id] = resource

        # gather all information needed to create the allocations/reservations
        relevant_bookings = (
            b for b in db.records.room_booking
            if isodate.parse_datetime(b.hour_end) > min_dt
        )

        # group by room id in addition to the multi_id, as digirez supports
        # reservations which span multiple rooms
        reservations = defaultdict(list)

        for booking in relevant_bookings:
            group = f'{booking.room_id}-{booking.multi_id}'
            reservations[group].append(booking)

        # get the max_date per resource
        max_dates = {}

        for reservation_group in reservations:
            resource_id = resources_by_room[reservation_group.split('-')[0]].id

            if resource_id not in max_dates:
                max_dates[resource_id] = min_dt

            for booking in reservations[reservation_group]:
                date = isodate.parse_datetime(booking.hour_end)

                if max_dates[resource_id] < date:
                    max_dates[resource_id] = date

        # get the user associated with the migration (the first admin)
        users = UserCollection(app.session())
        user = users.query().filter(User.role == 'admin').first()

        if not user:
            abort('No admin user found')

        # get the member mails
        members = {}

        for member in db.records.members:
            members[member.member_id] = member

        # create an allocation for all days between min/max date
        click.echo('Creating allocations')

        for resource in resources_by_room.values():

            # true if the resource does not have any bookings
            if resource.id not in max_dates:
                continue

            assert resource.group is not None
            start_hour, end_hour = floor_hours[resource.group]
            days = (max_dates[resource.id].date() - min_dt.date()).days

            first_day_start = datetime(
                min_dt.year, min_dt.month, min_dt.day,
                start_hour
            )

            first_day_end = datetime(
                min_dt.year, min_dt.month, min_dt.day,
                end_hour - 1, 59, 59, 999999
            )

            whole_day = (start_hour, end_hour) == (0, 24)

            for day in range(days + 1):
                start = first_day_start + timedelta(days=day)
                end = first_day_end + timedelta(days=day)

                resource.scheduler.allocate(
                    dates=(start, end),
                    partly_available=True,
                    whole_day=whole_day,
                )

        # create the reservations
        click.echo('Creating reservations')

        booking_conflicts = 0

        for reservation_group, bookings in reservations.items():
            resource = resources_by_room[reservation_group.split('-')[0]]
            scheduler = resource.scheduler
            found_conflict = False

            session_id = uuid4()

            for booking in bookings:
                assert bookings[0].anspmail == booking.anspmail
                assert bookings[0].member_id == booking.member_id
                assert bookings[0].anspname == booking.anspname
                assert bookings[0].ansptel == booking.ansptel
                assert bookings[0].title == booking.title
                assert bookings[0].description == booking.description

                email = booking.anspmail
                email = email or members[booking.member_id].member_email

                for expression, replacement in email_fixes:
                    email = expression.sub(replacement, email)

                try:
                    token_uuid = scheduler.reserve(
                        email=email,
                        dates=(
                            isodate.parse_datetime(booking.hour_start),
                            isodate.parse_datetime(booking.hour_end)
                        ),
                        session_id=session_id,
                        single_token_per_session=True,
                        data={'accepted': True}  # accepted through ticket
                    )
                    token = token_uuid.hex
                except InvalidEmailAddress:
                    abort(f'{email} is an invalid e-mail address')
                except AlreadyReservedError:
                    booking_conflicts += 1
                    found_conflict = True

                    click.echo(
                        f'Booking conflict in {resource.title} '
                        f'at {booking.hour_start}'
                    )

            if found_conflict:
                continue

            assert resource.form_class is not None
            forms = FormCollection(app.session())
            form_data = get_formdata(members[booking.member_id], booking)
            form_data = unescape_dictionary(form_data)
            form = resource.form_class(data=form_data)

            if not form.validate():
                abort(f'{form_data} failed the form check with {form.errors}')

            submission = forms.submissions.add_external(
                form=form,
                state='pending',
                id=token_uuid
            )

            scheduler.queries.confirm_reservations_for_session(session_id)
            scheduler.approve_reservations(token_uuid)

            forms.submissions.complete_submission(submission)

            with forms.session.no_autoflush:
                ticket = TicketCollection(request.session).open_ticket(
                    handler_code='RSV', handler_id=token
                )
                ticket.accept_ticket(user)
                ticket.close_ticket()

        if not ignore_booking_conflicts and booking_conflicts:
            abort(
                f'There were {booking_conflicts} booking conflicts, aborting'
            )

    return run_import


@cli.command(context_settings={'default_selector': '*'})
@click.option('--dry-run', default=False, is_flag=True,
              help='Do not write any changes into the database.')
@pass_group_context
def fix_tags(
    group_context: GroupContext,
    dry_run: bool
) -> Callable[[OrgRequest, OrgApp], None]:

    def fixes_german_tags_in_db(request: OrgRequest, app: OrgApp) -> None:
        session = request.session

        de_transl = app.translations.get('de_CH')
        assert de_transl is not None

        defined_tags = list(TAGS)
        defined_tag_ids = [str(s) for s in defined_tags]

        def translate(text: TranslationString) -> str:
            return text.interpolate(de_transl.gettext(text))

        form_de_to_en = {translate(text): str(text) for text in defined_tags}

        predefined = {
            'Theater / Tanz': ('Dancing', 'Theater'),
            'Cultur': ('Culture',),
            'Vortrag / Lesung': ('Talk', 'Reading')
        }

        msg_log = []

        def replace_with_predefined(
            tags: list[str]
        ) -> list[str]:

            new_tags = tags.copy()
            for t in tags:
                predef = predefined.get(t)
                if predef:
                    new_tags.remove(t)
                    new_tags.extend(predef)
                    if dry_run:
                        msg_log.append(
                            f'{t} -> {", ". join(predef)}')
            return new_tags

        undefined_msg_ids = set()

        def handle_occurrence_tags(occurrence: Event | Occurrence) -> None:
            tags = occurrence.tags.copy()
            tags = replace_with_predefined(tags)
            for tag in occurrence.tags:
                if tag in predefined:
                    continue
                if tag not in defined_tag_ids:
                    if tag in form_de_to_en:
                        tags.remove(tag)
                        tags.append(form_de_to_en[tag])
                        msg_log.append(
                            f'{tag} -> {form_de_to_en[tag]}')
                    else:
                        undefined_msg_ids.add(tag)
            if tags != occurrence.tags:
                if not dry_run:
                    occurrence.tags = tags

        for event_ in session.query(Event):
            handle_occurrence_tags(event_)

        for occurrence in session.query(Occurrence):
            handle_occurrence_tags(occurrence)

        if dry_run:
            click.echo('\n'.join(set(msg_log)))

        assert not undefined_msg_ids, (
            f'Define {", ".join(undefined_msg_ids)}'
            f' in org/forms/event.py'
        )

    return fixes_german_tags_in_db


def close_ticket(ticket: Ticket, user: User, request: OrgRequest) -> None:
    if ticket.state == 'open':
        ticket.accept_ticket(user)
        TicketMessage.create(
            ticket,
            request,
            'opened'
        )

    TicketMessage.create(
        ticket,
        request,
        'closed'
    )
    ticket.close_ticket()


@cli.command('fetch')
@pass_group_context
@click.option('--source', multiple=True)
@click.option('--tag', multiple=True)
@click.option('--location', multiple=True)
@click.option('--create-tickets', is_flag=True, default=False)
@click.option('--state-transfers', multiple=True,
              help='Usage: local:remote, e.g. published:withdrawn')
@click.option('--published-only', is_flag=True, default=False,
              help='Only add event is they are published on remote')
@click.option('--delete-orphaned-tickets', is_flag=True)
def fetch(
    group_context: GroupContext,
    source: Sequence[str],
    tag: Sequence[str],
    location: Sequence[str],
    create_tickets: bool,
    state_transfers: Sequence[str],
    published_only: bool,
    delete_orphaned_tickets: bool
) -> Callable[[OrgRequest, OrgApp], None]:
    r""" Fetches events from other instances.

    Only fetches events from the same namespace which have not been imported
    themselves.

    Example
    .. code-block:: bash

        onegov-org --select '/veranstaltungen/zug' fetch \
            --source menzingen --source steinhausen \
            --tag Sport --tag Konzert \
            --location Zug

    Additional parameters:

    - ``--state-transfers published:withdrawn``

        Will update the local event.state from published to withdrawn
        automatically. If there are any tickets associated with the event,
        the will be closed automatically.

    - ``--pusblished-only``

        When passing the remote items to the EventCollection, only add
        events if they are published.

    - ``--delete-orphaned-tickets``

        Delete Tickets, TicketNotes and TicketMessasges if an
        event gets deleted automatically.

    The following example will close tickets automatically for
    submitted and published events that were withdrawn on the remote.

    .. code-block:: bash
        onegov-event --select '/veranstaltungen/zug' fetch \
            --source menzingen --source steinhausen \
            --published-only \
            --create-tickets \
            --state-transfers published:withdrawn \
            --state-transfers submitted:withdrawm

    """

    def vector_add(a: Sequence[int], b: Sequence[int]) -> list[int]:
        return list(map(add_op, a, b))

    if not len(source):
        abort('Provide at least one source')

    valid_state_transfers = {}
    valid_choices = ('initiated', 'submitted', 'published', 'withdrawn')
    if len(state_transfers):
        for string in state_transfers:
            local, remote = string.split(':')
            assert local, remote
            assert local in valid_choices
            assert remote in valid_choices
            valid_state_transfers[local] = remote

    def _fetch(request: OrgRequest, app: OrgApp) -> None:

        def event_file(reference: UploadedFile) -> BytesIO:
            # XXX use a proper depot manager for this
            assert app.depot_storage_path is not None
            path = Path(app.depot_storage_path) / reference['path'] / 'file'
            with open(path):
                content = BytesIO(path.open('rb').read())
            return content

        try:
            result = [0, 0, 0]

            for key in source:
                remote_schema = f'{app.namespace}-{key}'
                local_schema = app.session_manager.current_schema
                assert local_schema is not None
                assert remote_schema in app.session_manager.list_schemas()

                app.session_manager.set_current_schema(remote_schema)
                remote_session = app.session_manager.session()
                assert remote_session.info['schema'] == remote_schema

                query = remote_session.query(Event)
                query = query.filter(
                    or_(
                        Event.meta['source'].astext.is_(None),
                        Event.meta['source'].astext == ''
                    )
                )
                if tag:
                    query = query.filter(
                        Event._tags.has_any(array(tag))  # type:ignore
                    )
                if location:
                    query = query.filter(
                        or_(*(
                            Event.location.op('~')(f'\\y{term}\\y')
                            for term in location
                        ))
                    )

                def remote_events(
                    query: Query[Event] = query,
                    key: str = key
                ) -> Iterator[EventImportItem]:

                    for event_ in query:
                        event_._es_skip = True
                        yield EventImportItem(
                            event=Event(  # type:ignore[misc]
                                state=event_.state,
                                title=event_.title,
                                start=event_.start,
                                end=event_.end,
                                timezone=event_.timezone,
                                recurrence=event_.recurrence,
                                content=event_.content,
                                location=event_.location,
                                tags=event_.tags,
                                source=f'fetch-{key}-{event_.name}',
                                coordinates=event_.coordinates,
                                organizer=event_.organizer,
                                organizer_email=event_.organizer_email,
                                organizer_phone=event_.organizer_phone,
                                price=event_.price,
                            ),
                            image=(
                                event_file(event_.image.reference)
                                if event_.image else None
                            ),
                            image_filename=(
                                event_.image.name
                                if event_.image else None
                            ),
                            pdf=(
                                event_file(event_.pdf.reference)
                                if event_.pdf else None
                            ),
                            pdf_filename=(
                                event_.pdf.name
                                if event_.pdf else None
                            )
                        )

                # be sure to switch back to the local schema, lest we
                # accidentally update things on the remote
                app.session_manager.set_current_schema(local_schema)
                local_session = app.session_manager.session()
                local_events = EventCollection(local_session)

                # the responsible user is the first admin that was added
                local_admin = (
                    local_session.query(User)
                    .filter_by(role='admin')
                    .order_by(User.created).first()
                )

                if create_tickets and not local_admin:
                    abort('Can not create tickets, no admin is registered')

                def ticket_for_event(
                    event_id: UUID,
                    local_session: Session = local_session
                ) -> Ticket | None:
                    return TicketCollection(local_session).by_handler_id(
                        event_id.hex)

                added, updated, purged = local_events.from_import(
                    remote_events(),
                    to_purge=[f'fetch-{key}'],
                    publish_immediately=False,
                    valid_state_transfers=valid_state_transfers,
                    published_only=published_only
                )

                for event_ in added:
                    event_.submit()
                    if not create_tickets:
                        event_.publish()
                        continue

                    assert local_admin is not None
                    with local_session.no_autoflush:
                        tickets = TicketCollection(local_session)
                        new_ticket = tickets.open_ticket(
                            handler_code='EVN', handler_id=event_.id.hex,
                            source=event_.meta['source'],
                            user=local_admin.username
                        )
                        new_ticket.muted = True
                        TicketNote.create(new_ticket, request, (
                            f'Importiert von Instanz {key}'

                        ), owner=local_admin.username)

                helper_request: OrgRequest = Bunch(  # type:ignore
                    current_username=local_admin and local_admin.username,
                    session=local_session)

                for event_id in purged:
                    ticket = ticket_for_event(event_id)
                    if ticket:
                        if not delete_orphaned_tickets:
                            if local_admin is None:
                                abort(
                                    'Can not close orphaned ticket, '
                                    'no admin is registered'
                                )
                            close_ticket(ticket, local_admin, helper_request)
                        else:
                            messages = MessageCollection(
                                local_session,
                                channel_id=ticket.number
                            )
                            for msg in messages.query():
                                local_session.delete(msg)
                            local_session.delete(ticket)

                result = vector_add(
                    result,
                    (len(added), len(updated), len(purged))
                )

            click.secho(
                f'Events successfully fetched '
                f'({result[0]} added, {result[1]} updated, '
                f'{result[2]} deleted)',
                fg='green'
            )

        except Exception:
            log.exception('Error fetching events')
            raise

    return _fetch


@cli.command('fix-directory-files')
@pass_group_context
def fix_directory_files(
    group_context: GroupContext
) -> Callable[[OrgRequest, OrgApp], None]:
    """
    Not sure of this doubles the files, but actually the file
    reference remains, so it shouldn't

    This command will become obsolete as soon as the type of files in
    submissions are set correctly with type 'directory'.

    """
    def execute(request: OrgRequest, app: OrgApp) -> None:
        count = 0
        for entry in request.session.query(DirectoryEntry).all():
            for field in entry.directory.file_fields:
                field_data = entry.content['values'][field.id]
                if field_data and field_data.get('data', '').startswith('@'):
                    file_id = field_data['data'].lstrip('@')
                    file = request.session.query(File).filter_by(
                        id=file_id).first()
                    if file and file.type != 'directory':
                        new = DirectoryFile(  # type:ignore[misc]
                            id=random_token(),
                            name=file.name,
                            note=file.note,
                            reference=file.reference
                        )
                        entry.files.append(new)
                        entry.content['values'][field.id][
                            'data'] = f'@{new.id}'
                        entry.content.changed()  # type:ignore[attr-defined]
                        count += 1
        if count:
            click.secho(
                f'{app.schema} - {count} files adapted with type `directory`',
                fg='green'
            )
    return execute


@cli.command('migrate-town', context_settings={'singular': True})
@pass_group_context
def migrate_town(
    group_context: GroupContext
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Migrates the database from an old town to the new town like in the
    upgrades.

    """

    def migrate_to_new_town(request: OrgRequest, app: OrgApp) -> None:
        context: UpgradeContext = Bunch(session=app.session())  # type:ignore
        migrate_theme_options(context)
        migrate_homepage_structure_for_town6(context)

    return migrate_to_new_town


@cli.command('migrate-links', context_settings={'singular': True})
@pass_group_context
@click.argument('old-uri')
@click.option('--dry-run', is_flag=True, default=False)
def migrate_links_cli(
    group_context: GroupContext,
    old_uri: str,
    dry_run: bool
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Migrates url's in pages. Supports domains and full urls. Most of
    the urls are located in meta and content fields.
    """

    if '.' not in old_uri:
        click.secho('Domain must contain a dot')
        sys.exit(1)

    def execute(request: OrgRequest, app: OrgApp) -> None:
        if old_uri.startswith('http'):
            new_uri = request.host_url
        else:
            new_uri = request.domain
        migration = LinkMigration(request, old_uri=old_uri, new_uri=new_uri)
        total, grouped_count = migration.migrate_site_collection(
            test=dry_run
        )

        if total:
            click.secho(f'Total replaced: {total}')
            click.secho('Replaced links by group:')
            for group, counts in grouped_count.items():
                for field, count in counts.items():
                    if count:
                        click.secho(f'{group}.{field}: {count}')
        else:
            click.secho('Nothing found')

    return execute


@cli.command(context_settings={'default_selector': '*'})
@pass_group_context
def migrate_publications(
    group_context: GroupContext,
    dry_run: bool
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Marks signed files for publication. """

    def mark_as_published(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        files = session.query(File).filter_by(signed=True).all()
        for file in files:
            file.publication = True
        if files:
            click.echo(
                f'{session.info["schema"]}: '
                f'Marked {len(files)} signed files for publication'
            )

    return mark_as_published


@cli.command(name='delete-invisible-links')
def delete_invisible_links() -> Callable[[OrgRequest, OrgApp], None]:
    """ Deletes all the data associated with a period, including:

    Example:
    .. code-block:: bash

        onegov-org --select /foo/bar delete-invisible-links

    """

    def delete_invisible_links(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        query = QueryChain(
            (session.query(Page),
             session.query(Resource),
             session.query(ExtendedDirectory),
             session.query(FormDefinition))
        )  # type:ignore
        models = query.all()

        click.echo(click.style(
            {session.info['schema']},
            fg='yellow'
        ))

        invisible_links = []
        for page in models:
            # Find links with no text, only br tags and/or whitespaces
            for field in page.content_fields_containing_links_to_files:
                if not page.content.get(field):
                    continue
                soup = BeautifulSoup(page.content.get(field), 'html.parser')
                for link in soup.find_all('a'):
                    if not any(
                        tag.name != 'br' and (
                            tag.name or not tag.isspace()
                        ) for tag in link.contents
                    ):
                        invisible_links.append(link)
                        if all(tag.name == 'br' for tag in link.contents):
                            link.replace_with(
                                BeautifulSoup('<br/>', 'html.parser')
                            )
                        else:
                            link.decompose()

                # Save the modified HTML back to page.text
                if page.content[field] != str(soup):
                    page.content[field] = str(soup)

        click.echo(
            click.style(
                f'{session.info["schema"]}: '
                f'Deleted {len(invisible_links)} invisible links',
                fg='yellow'
            )
        )

    return delete_invisible_links


@cli.command(name='get-resources-and-forms')
@click.argument('option_file', type=click.File('rb'))
def get_resources_and_forms(
    option_file: IO[bytes]
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Get the resources and forms from the option file. """

    def print_resources_and_forms(request: OrgRequest, app: OrgApp) -> None:
        book = load_workbook(option_file)
        sheet = book['Reservationen']

        class OptionDict(TypedDict):
            options: set[str]
            type: str

        resources: dict[str, dict[str, OptionDict]] = {}
        for index, row in enumerate(sheet.rows):
            if index <= 3:
                continue
            for i, cell in enumerate(row):
                value = str(cell.value)
                if i == 0:
                    resource_name = value
                    if resource_name not in resources:
                        resources[resource_name] = {}
                if i == 13:  # Option name
                    option = value
                    if resources[resource_name].get(option) is None:
                        resources[resource_name][option] = {
                            'type': 'text',
                            'options': set()
                        }
                if i == 14:  # Option answer
                    answer = value
                    resources[resource_name][option]['options'].add(answer)
                if i == 15:  # Option price
                    if int(value) > 0:
                        resources[resource_name][option]['options'].remove(
                            answer)
                        answer = answer.replace(' (', ', ').replace(
                            ')', '') + f' ({value} CHF)'
                        resources[resource_name][option]['options'].add(
                            answer)
                        resources[resource_name][option]['type'] = 'radio'

        for resource in resources.keys():
            click.secho(resource, fg='blue')
            for option in resources[resource]:
                if resources[resource][option]['type'] == 'text':
                    click.secho(f'{option} = ___', fg='cyan')
                else:
                    click.secho(f'{option} =', fg='cyan')
                    for answer in resources[resource][option]['options']:
                        click.echo(f'    ( ) {answer}',)

    return print_resources_and_forms


@cli.command(name='import-reservations', context_settings={'singular': True})
@click.argument('reservation_file', type=click.File('rb'))
@click.argument('option_file', type=click.File('rb'))
@click.argument('mapping_yaml', type=click.File('rb'))
@click.option('--dry-run', is_flag=True, default=False)
@click.option('--masked', is_flag=True, default=False)
def import_reservations(
    reservation_file: IO[bytes],
    option_file: IO[bytes],
    mapping_yaml: IO[bytes],
    dry_run: bool,
    masked: bool
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Imports reservations from a Excel file (needs to be .xlsx).
    Creates no resources or allocations, so the availabilty periods need to
    be set in the resource settings.
    """

    def import_reservations(request: OrgRequest, app: OrgApp) -> None:

        class Reservation(TypedDict):
            state: str
            general: dict[str, Any]
            fields: dict[str, Any]
            dates: list[Any]

        yaml_file = mapping_yaml.read()
        yaml_dict = yaml.safe_load(yaml_file)
        shared_fields = yaml_dict.get('shared_fields', {})
        resource_options = yaml_dict.get('resource_options', {})

        book = load_workbook(reservation_file)
        sheet = book['Reservationsdaten']

        book_options = load_workbook(option_file)
        sheet_options = book_options['Reservationen']

        reservations: dict[
            str, dict[str, Reservation]] = {}
        count = 0
        last_reservation_id = ''
        for index, row in enumerate(sheet.rows):
            if index <= 3:  # Skip the first 3 rows, they are headers
                continue

            reservation: Reservation = {
                'state': '',
                'general': {},
                'fields': {},
                'dates': []
            }
            row_empty: bool = True
            id = ''
            resource_name = ''

            for i, cell in enumerate(row):
                value = cell.value
                if i == 0:
                    if value is None:
                        continue
                    row_empty = False
                    resource_name = str(value)
                elif i == 3:
                    id = str(value)
                    if last_reservation_id == id:
                        reservation = reservations[resource_name][id]
                elif i == 6 or i == 8:
                    reservation['dates'].append(value)
                elif i == 22:
                    if value is None:
                        value = 'info@seantis.ch'
                    reservation['general']['email'] = value
                    reservation['fields']['email'] = value
                elif i == 37:
                    reservation['state'] = str(value)
                elif i in shared_fields:
                    value = str(value or '')
                    key = shared_fields[i]
                    if reservation['fields'].get(key) is None or '':
                        reservation['fields'][key] = value
                    elif value not in reservation['fields'][key]:
                        if i == 17 or i == 18:
                            reservation['fields'][key] += f' {value}'
                        else:
                            reservation['fields'][key] += f', {value}'

            if not row_empty:
                # Check if the dates spread across multiple days
                start = reservation['dates'][-2]
                end = reservation['dates'][-1]
                if start.day != end.day:
                    days = (end - start).days + 1
                    for day in range(days):
                        reservation['dates'].insert(
                            -1, start + timedelta(days=day)
                        ) if day != 0 else None
                        end_of_day = datetime.combine(
                            start, end.time()
                        )
                        reservation['dates'].insert(
                            -1, end_of_day + timedelta(days=day)
                        ) if day != days - 1 else None
                count += 1
                if resource_name not in reservations:
                    reservations[resource_name] = {}

                if id not in reservations[resource_name] and (
                    id != last_reservation_id
                ):
                    reservations[resource_name][id] = reservation
                last_reservation_id = id or ''

        options_count = 0
        options_not_found = set()
        for index, row in enumerate(sheet_options.rows):
            row_empty = True
            if index <= 3:
                continue
            for i, cell in enumerate(row):
                value = cell.value
                if value is None:
                    continue
                row_empty = False
                options_count += 1
                if i == 0:
                    resource_name = str(value)
                elif i == 2:
                    id = str(value)
                elif resource_name in reservations:
                    if id in reservations[resource_name]:
                        if i == 13:  # Option name
                            key = resource_options[resource_name][
                                'options'].get(value)
                            if key is None:
                                options_not_found.add(
                                    f'{resource_name}: {value}')
                        if i == 14:  # Option answer
                            if key is not None:
                                reservation = reservations[resource_name][id]
                                if reservation['fields'].get(key) is None:
                                    reservation['fields'][key] = value
        for option in options_not_found:
            click.secho(f'Option not found in the mapping file: {option}',
                        fg='yellow')

        if dry_run:
            res_show = json.dumps(reservations, indent=4, default=str)
            click.secho(f'Reservations: {res_show}', fg='green')

        # Create reservations
        if not dry_run:
            resources = ResourceCollection(app.libres_context)
            for resource_name in reservations.keys():

                real_resource_name = resource_options.get(resource_name, '')
                if not real_resource_name:
                    click.echo(
                        f'Resource {resource_name} not found in the mapping '
                        'file'
                    )
                    continue
                real_resource_name = real_resource_name.get('name', '')
                resource = resources.by_name(real_resource_name.lower())

                if not resource:
                    click.echo(
                        f'Resource {resource} not found in the database'
                    )
                    continue

                click.echo(
                    f'Importing reservations for {resource.title}'
                )
                scheduler = resource.scheduler

                for id, reservation in reservations[resource_name].items():
                    if reservation['state'] == 'annulliert':
                        click.secho(
                            f'{id}: Reservation state is deleted, '
                            f'skipping {resource.title}', fg='yellow')
                        continue
                    found_conflict = False
                    session_id = uuid4()
                    for n in range(int(len(reservation['dates'])/2)):
                        email = reservation['general'].get('email')
                        start = reservation['dates'][n*2]
                        end = reservation['dates'][n*2+1]

                        try:
                            token_uuid = scheduler.reserve(
                                email=str(email),
                                dates=(start, end),
                                session_id=session_id,
                                single_token_per_session=True,
                                data={'accepted': True}
                            )
                            token = token_uuid.hex
                        except InvalidEmailAddress:
                            click.secho(f'{id}: {email} is an invalid e-mail '
                                        'address')
                        except AlreadyReservedError:
                            found_conflict = True
                            click.secho(
                                f'{id}: Booking conflict in {resource.title} '
                                f'at {start} - {end}', fg='red')
                        except TimerangeTooLong:
                            click.secho(
                                f'{id}: Timerange too long: {start} - {end}',
                                fg='yellow')
                            rules = resource.content['rules']
                            relevant_rules = []
                            for rule in rules:
                                rule = rule['options']
                                if rule['start'] <= start.date(
                                ) and rule['end'] >= end.date(
                                ) and start.weekday(
                                ) not in rule['except_for'] and start.time(
                                ) <= rule['start_time'] and end.time(
                                ) >= rule['end_time']:
                                    relevant_rules.append(rule)
                            for i, rule in enumerate(relevant_rules):
                                start = datetime.combine(
                                    start.date(), rule['start_time'])
                                end = datetime.combine(
                                    end.date(), rule['end_time'])
                                click.secho(
                                    f'{id}: Trying to reserve {resource.title}'
                                    f' at {start} - {end} with rule {i+1}',
                                    fg='cyan')
                                try:
                                    token_uuid = scheduler.reserve(
                                        email=str(email),
                                        dates=(start, end),
                                        session_id=session_id,
                                        single_token_per_session=True,
                                        data={'accepted': True}
                                    )
                                    token = token_uuid.hex
                                except Exception as e:
                                    click.secho(
                                        f'{id}: Error reserving '
                                        f'{resource.title} at {start} - {end} '
                                        f'- {e}', fg='red')
                                    found_conflict = True
                            if not relevant_rules:
                                click.secho(
                                    f'{id}: No rules found for '
                                    f'{resource.title} at {start}', fg='red')
                        except Exception as e:
                            click.secho(
                                f'{id}: Error {e} reserving {resource.title} '
                                f'at {start} - {end} outside of availability'
                                'period', fg='red')
                            continue

                    if found_conflict:
                        continue

                    assert resource.form_class is not None
                    forms = FormCollection(app.session())

                    form_data = {}
                    for key, value in reservation['fields'].items():
                        form_data[key] = str(value)

                    form = resource.form_class(data=form_data)

                    if not form.validate():
                        form_data_show = json.dumps(
                            form_data, indent=4, default=str)
                        click.secho(f'{id}: {form_data_show} failed the form '
                            f'check with {form.errors}', fg='red')
                        continue

                    submission = forms.submissions.add_external(
                        form=form,
                        state='pending',
                        id=token_uuid
                    )

                    scheduler.queries.confirm_reservations_for_session(
                        session_id)
                    scheduler.approve_reservations(token_uuid)

                    forms.submissions.complete_submission(submission)

                    users = UserCollection(app.session())
                    user = users.query().filter(
                        User.username == 'info@seantis.ch').first()

                    if not user:
                        abort('info@seantis.ch not found in users')

                    with forms.session.no_autoflush:
                        ticket = TicketCollection(request.session).open_ticket(
                            handler_code='RSV', handler_id=token
                        )
                        ticket.accept_ticket(user)
                        if reservation['state'] != 'unbestätigt':
                            ticket.close_ticket()

                    click.secho(f'{id}: Sucessfully imported reservation '
                                f'at {start} - {end}',
                                fg='green')

    return import_reservations


@cli.command(context_settings={'default_selector': '*'})
@click.argument('year', type=click.IntRange(1900, date.today().year))
@click.argument('month', type=click.IntRange(1, 12))
@pass_group_context
def mtan_statistics(
    group_context: GroupContext,
    year: int,
    month: int,
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Generate mTAN SMS statistics for the given year and month. """

    if date(year, month, 1) >= date.today().replace(day=1):
        abort('Year and month needs to be fully in the past')

    def mtan_statistics(request: OrgRequest, app: OrgApp) -> None:
        mtan_count: int = request.session.query(
            func.count(TAN.id)
        ).filter(and_(
            func.extract('year', TAN.created) == year,
            func.extract('month', TAN.created) == month,
            TAN.meta['mobile_number'].isnot(None)
        )).scalar()
        if mtan_count:
            org_name = app.org.name if hasattr(app, 'org') else None
            title = f'{org_name} ({app.schema})' if org_name else app.schema
            click.echo(
                f'{title}: '
                f'Sent {mtan_count} mTAN SMS'
            )

    return mtan_statistics


def ul(inner: str) -> str:
    return f'<ul>{inner}</ul>'


def ol(inner: str) -> str:
    return f'<ol>{inner}</ol>'


def li(inner: str) -> str:
    return f'<li>{inner}</li>'


def p(inner: str) -> str:
    return f'<p>{inner}</p>'


def b(inner: str) -> str:
    return f'<b>{inner}</b>'


def em(inner: str) -> str:
    return f'<em>{inner}</em>'


def br() -> str:
    return '<br/>'


def a(href: str, text: str) -> str:
    return f'<a href="{href}">{text}</a>'


def h(level: int, text: str) -> str:
    return f'<h{level}>{text}</h{level}>'


def img(src: str, alt: str) -> str:
    return f'<img class="lazyload-alt" src="{src}" alt="{alt}">'


@cli.command(name='import-news')
@click.argument('path', type=click.Path(exists=True, resolve_path=True))
@click.option('--start-date', type=click.DateTime(formats=['%Y-%m-%d']),
              default=None)
@click.option('--end-date', type=click.DateTime(formats=['%Y-%m-%d']),
              default=None)
@click.option('--overwrite-content', is_flag=True, default=False)
@click.option('--local', is_flag=True, default=False)
@click.option('--dry-run', is_flag=True, default=False)
def import_news(
    path: str,
    start_date: datetime | None,
    end_date: datetime | None,
    overwrite_content: bool,
    dry_run: bool,
    local: bool
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Imports news from archive of json files

    Example:
    .. code-block:: bash

        onegov-org --select '/foo/bar' import-news /path/to/news
    """

    # Read all json files in the given directory
    def read_json_files(path: str) -> Iterator[tuple[dict[str, Any], str]]:
        for file in Path(path).iterdir():

            if file.suffix == '.json':
                with open(file) as f:
                    yield (json.load(f), file.name)

    def import_news(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        news_parent = session.query(Page).filter(Page.type == 'news').filter(
            Page.parent_id == None).first()
        news_collection = PageCollection(request.session)
        image_collection = ImageFileCollection(request.session)

        if not news_parent:
            click.echo('No news parent found')
            return

        news = read_json_files(path)
        import_counter = 0
        overwrite_counter = 0

        def content_to_markup(element: dict[str, Any],
                              image_counter: int) -> str:
            element_markup = ''
            children_markup = ''
            bold = element.get('bold', False)
            italic = element.get('italic', False)

            if 'children' in element:
                for child in element['children']:
                    children_markup += content_to_markup(
                        child, image_counter) + ' '

            if element['type'] == 'Text':
                text = element.get('text', '')
                if bold:
                    text = b(text)
                if italic:
                    text = em(text)
                element_markup = text

            if element['type'] == 'Break':
                element_markup = br()

            if element['type'] == 'Link':
                element_markup = a(element['url'], element.get('text', ''))

            if element['type'] == 'Paragraph':
                element_markup = p(children_markup)

            elif element['type'] == 'Heading':
                heading = element.get('text', '')
                level = element.get('level', 1)
                if bold:
                    heading = b(heading)
                if italic:
                    heading = em(heading)
                element_markup = h(level, heading)

            if element['type'] == 'ListItem':
                if bold:
                    element_markup = li(b(children_markup))
                else:
                    element_markup = li(children_markup)

            elif element['type'] == 'UnorderedList':
                element_markup = ul(children_markup)

            elif element['type'] == 'OrderedList':
                element_markup = ol(children_markup)

            elif element['type'] == 'EmbeddedImage' and element['content']:
                base64_string = element['content']
                base64_data = base64_string.split(',', 1)[1]
                image_type = base64_string.split(';')[0].split(
                    '/')[1]
                image_bytes = base64.b64decode(base64_data)
                image_bytes_io = BytesIO(image_bytes)

                try:
                    image = image_collection.add(
                        filename=(f'{article_number}-{image_counter}'
                                  f'.{image_type}'),
                        content=image_bytes_io
                    )
                    if note := element['caption']:
                        image.note = note
                    url = request.link(image)
                    if local:
                        url = '/' + '/'.join(url.split('/')[3:])
                    else:
                        url = '/' + '/'.join(url.split('/')[5:])

                    if image_counter == 0:
                        added.content['page_image'] = url  # type:ignore
                        added.meta['page_image'] = url  # type:ignore
                    else:
                        alt = element['alt'] if element[
                            'alt'] else ''
                        element_markup = img(url, alt)
                    image_counter += 1
                except Exception as e:
                    log.error(
                        f'Error importing image {image_counter}'
                        f' for {article_name}: {e}')
                    if e.args[0].type == 'image':
                        image = e.args[0]
                        url = request.link(image)
                        if local:
                            url = '/' + '/'.join(url.split('/')[3:])
                        else:
                            url = '/' + '/'.join(url.split('/')[5:])
                        if image_counter == 0:
                            added.content['page_image'] = url  # type:ignore
                            added.meta['page_image'] = url  # type:ignore
                        else:
                            alt = element['alt'] if element[
                                'alt'] else ''
                            element_markup = img(url, alt)

            return element_markup

        for news_item, article_name in news:
            content = ''
            image_counter = 0
            article_number = article_name.split('.')[0]

            article_date = datetime.strptime(article_name.split('_')[0],
                                             '%Y-%m-%d')
            publication_start = datetime.strptime(
                        news_item['metadata']['published'],
                        '%Y-%m-%dT%H:%M:%S')
            publication_start = publication_start.replace(
                        tzinfo=pytz.UTC)

            if start_date and end_date and not (
                start_date <= article_date <= end_date
                ) or start_date and (
                    article_date < start_date) or end_date and (
                        article_date > end_date):
                continue

            # Check if article already exists based on title and date
            if added := session.query(Page).filter(
                Page.title == news_item['metadata']['title']
            ).filter(Page.publication_start == publication_start).first():
                if overwrite_content:
                    click.echo(f'Overwriting {article_name}')

                    for element in news_item['elements']:
                        content += content_to_markup(element, image_counter)

                    added.content['text'] = Markup(content)  # nosec: B704
                    overwrite_counter += 1
                else:
                    click.secho((f'Skipped {article_name} with '
                                f'title {news_item["metadata"]["title"]}'
                                f' as it already exists'),
                                fg='yellow')
            else:
                click.echo(f'Importing {article_name}')
                import_counter += 1

                if not dry_run:
                    if news_item['metadata'].get('lead', ''):
                        lead = ' '.join([
                            t.get('text', '') for t in news_item[
                                'metadata']['lead']['children']])
                    else:
                        lead = ''

                    added = news_collection.add(
                        parent=news_parent,
                        title=news_item['metadata']['title'],
                        lead=lead,
                        meta={'trait': 'news'},
                        type='news',
                        publication_start=publication_start,
                    )

                    for element in news_item['elements']:
                        content += content_to_markup(element, image_counter)

                    added.content['text'] = Markup(content)  # nosec: B704

        click.echo(f'{import_counter} news items imported')
        click.echo(f'{overwrite_counter} news items overwritten')

    return import_news


def content_to_markup(element: dict[str, Any]) -> str:
    element_markup = ''
    children_markup = ''
    bold = element.get('bold', False)
    italic = element.get('italic', False)

    if 'children' in element:
        for child in element['children']:
            children_markup += content_to_markup(
                child) + ' '

    if element['type'] == 'Text':
        text = element.get('text', '')
        if bold:
            text = b(text)
        if italic:
            text = em(text)
        element_markup = text

    if element['type'] == 'Break':
        element_markup = br()

    if element['type'] == 'Link':
        element_markup = a(element['url'], element.get('text', ''))

    if element['type'] == 'Paragraph':
        element_markup = p(children_markup)

    elif element['type'] == 'Heading':
        heading = element.get('text', '')
        level = element.get('level', 1)
        if bold:
            heading = b(heading)
        if italic:
            heading = em(heading)
        element_markup = h(level, heading)

    return element_markup


@cli.command(name='import-meetings')
@click.argument('path', type=click.Path(exists=True, resolve_path=True))
@click.option('--dry-run', is_flag=True, default=False)
def import_meetings(
    path: str,
    dry_run: bool,
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Imports meetings from archive of json files

    Example:
    .. code-block:: bash

        onegov-org --select '/foo/bar' import-meetings /path/to/items
    """

    # Read all json files in the given directory
    def read_json_files(path: str) -> Iterator[tuple[dict[str, Any], str]]:
        for file in Path(path).iterdir():

            if file.suffix == '.json':
                with open(file) as f:
                    yield (json.load(f), file.name)

    def create_meetings(request: OrgRequest, app: OrgApp) -> None:
        meeting_collection = MeetingCollection(request.session)
        file_coll = FileCollection(request.session)
        meeting_item_collection = MeetingItemCollection(request.session)

        meetings = read_json_files(path)
        import_counter = 0
        overwrite_counter = 0

        start_dt: datetime | None
        end_dt: datetime | None
        for meeting, article_name in meetings:
            date = meeting['elements'][0]['children'][0]['text']
            # Set locale to German for month parsing
            try:
                locale.setlocale(locale.LC_TIME, 'de_CH.UTF-8')
            except locale.Error:
                locale.setlocale(locale.LC_TIME, 'de_DE.UTF-8')

            date_str = date  # e.g. "7. Nov. 2024, 17.00 Uhr - 20.05 Uhr"

            # Regex to extract parts
            match = re.match(
                r'(\d{1,2})\. (\w+)\.? (\d{4}), '
                r'(\d{1,2})\.(\d{2}) Uhr - (\d{1,2})\.(\d{2}) Uhr',
                date_str
            )
            if match:
                day, month_str, year, sh, sm, eh, em = match.groups()
                # Parse month name to month number
                # month string can be:
                # month_str = month_str.strip('.')
                if month_str == 'März':
                    month_str = 'Mär'
                elif month_str == 'Juni':
                    month_str = 'Jun'
                elif month_str == 'Juli':
                    month_str = 'Jul'
                elif month_str == 'Sept':
                    month_str = 'Sep'

                try:
                    month = datetime.strptime(month_str, '%b').month
                    start_dt = datetime(
                        int(year), month, int(day), int(sh), int(sm)
                    )
                    end_dt = datetime(
                        int(year), month, int(day), int(eh), int(em)
                    )
                    # Add timezone
                    start_dt = pytz.timezone('Europe/Zurich').localize(
                        start_dt)
                    end_dt = pytz.timezone('Europe/Zurich').localize(end_dt)
                except ValueError as e:
                    click.secho(f'Error parsing date: {e}', fg='red')
                    start_dt = end_dt = None
                    meta_date = (f'{day} {month_str} {year} '
                                  f'{sh}:{sm} - {eh}:{em}')
            else:
                start_dt = end_dt = None

            click.echo(f'Importing {article_name}')
            import_counter += 1

            if not dry_run:
                location = Markup(  # nosec: B704
                    content_to_markup(meeting['elements'][1]))
                documents = False
                desc = False
                meeting_items = False
                meetings_items_list = []
                description = ''

                files = []
                for element in meeting['elements']:
                    if element['type'] == 'Heading':
                        if element['text'] == 'Informationen':
                            desc = True
                        elif element['text'] == 'Dokumente':
                            documents = True
                            desc = False
                        elif element['text'] == 'Traktanden':
                            meeting_items = True
                            desc = False
                            documents = False
                        else:
                            desc = False
                            documents = False
                            meeting_items = False
                    if desc:
                        if element['type'] == 'Paragraph':
                            if element.get('children') and (
                                element['children'][0].get(
                                    'text') != 'Beschreibung'):
                                description += content_to_markup(element)
                    if documents:
                        if element['type'] == 'Table':
                            for row in element['rows']:
                                try:
                                    link = row['cells'][0]['url']
                                except KeyError:
                                    click.echo(
                                        f'No link found in row'
                                        f' {row["cells"][0]}'
                                    )
                                    continue
                                resp = requests.get(link, timeout=(5, 10))
                                if resp.status_code == 200:
                                    if existing_file := file_coll.by_content(
                                        BytesIO(resp.content)
                                    ).first():
                                        files.append(existing_file)
                                        click.echo(
                                            f'File {existing_file} already '
                                            'exists, skipping.'
                                    )
                                    else:
                                        file = file_coll.add(
                                            filename=row['cells'][0]['text'],
                                            content=BytesIO(resp.content)
                                        )
                                        files.append(file)
                    if meeting_items:
                        if element['type'] == 'Table':
                            for row in element['rows']:
                                cells = row['cells']
                                if cells[1]['type'] != 'Link':
                                    continue
                                political_business_id = cells[2]['url'].split(
                                    '/')[-1]
                                meetings_items_list.append(
                                    {
                                        'number': cells[0]['text'],
                                        'title': cells[1]['text'],
                                        'political_business_id':
                                        political_business_id
                                    }
                                )

                added = meeting_collection.add(
                    title='Sitzung des Stadtparlaments',
                    type='generic',
                    start_datetime=start_dt,
                    end_datetime=end_dt,
                    address=location,
                )

                for meeting in meetings_items_list:
                    meeting_item_collection.add(
                        number=meeting['number'],
                        title=meeting['title'],
                        meeting_id=added.id,
                        meta={
                            'meta_date': meta_date,
                        },
                        political_business_link_id=meeting[
                                'political_business_id'],
                    )
                if files:
                    added.files = files

        click.echo(f'{import_counter} meetings imported')
        click.echo(f'{overwrite_counter} meetings overwritten')

    return create_meetings


@cli.command(name='import-commissions')
@click.argument('path', type=click.Path(exists=True, resolve_path=True))
@click.option('--dry-run', is_flag=True, default=False)
def import_commissions(
    path: str,
    dry_run: bool,
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Imports commissions from archive of json files

    Example:
    .. code-block:: bash

        onegov-org --select '/foo/bar' import-commissions /path/to/items
    """

    # Read all json files in the given directory
    def read_json_files(path: str) -> Iterator[tuple[dict[str, Any], str]]:
        for file in Path(path).iterdir():

            if file.suffix == '.json':
                with open(file) as f:
                    yield (json.load(f), file.name)

    def create_commissions(request: OrgRequest, app: OrgApp) -> None:
        commission_collection = RISCommissionCollection(request.session)

        commissions = read_json_files(path)
        import_counter = 0
        overwrite_counter = 0

        for commission, article_name in commissions:
            content = ''

            click.echo(f'Importing {article_name}')
            import_counter += 1

            if not dry_run:

                content = ''
                people_ids = []
                for element in commission['elements']:
                    if element['type'] != 'Heading' and (
                        element['type'] != 'Table'):
                        content += content_to_markup(element)
                    if element['type'] == 'Table':
                        for row in element['rows']:
                            link = row['cells'][0]['url']
                            function = row['cells'][1]['text']
                            id = link.split('/')[-1]
                            people_ids.append((id, function))
                        break

                commission_collection.add(
                    poly_type='ris_commission',
                    name=commission['metadata']['title'],
                    content={'description': Markup(content)},  # nosec: B704
                    meta={
                        'commission_id': article_name.split('_')[1].replace(
                            '.json', ''
                        ),
                        'people_ids': people_ids},
                )

        click.echo(f'{import_counter} commissions imported')
        click.echo(f'{overwrite_counter} commissions overwritten')

    return create_commissions


@cli.command(name='import-parliamentary_groups')
@click.argument('path', type=click.Path(exists=True, resolve_path=True))
@click.option('--dry-run', is_flag=True, default=False)
def import_parliamentary_groups(
    path: str,
    dry_run: bool,
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Imports parliamentary_groups from archive of json files

    Example:
    .. code-block:: bash

        onegov-org --select '/foo/bar' import-parliamentary_groups /path
    """

    # Read all json files in the given directory
    def read_json_files(path: str) -> Iterator[tuple[dict[str, Any], str]]:
        for file in Path(path).iterdir():

            if file.suffix == '.json':
                with open(file) as f:
                    yield (json.load(f), file.name)

    def create_parliamentary_groups(request: OrgRequest, app: OrgApp) -> None:
        parliamentary_group_collection = RISParliamentaryGroupCollection(
            request.session)

        parliamentary_groups = read_json_files(path)
        import_counter = 0
        overwrite_counter = 0

        for parliamentary_group, article_name in parliamentary_groups:
            content = ''

            click.echo(f'Importing {article_name}')
            import_counter += 1

            if not dry_run:

                content = ''
                people_ids = []
                for element in parliamentary_group['elements']:
                    if element['type'] != 'Heading' and (
                        element['type'] != 'Table'):
                        content += content_to_markup(element)
                    if element['type'] == 'Table':
                        for row in element['rows']:
                            link = row['cells'][0]['url']
                            id = link.split('/')[-1]
                            people_ids.append(id)
                        break

                parliamentary_group_collection.add(
                    name=parliamentary_group['metadata']['title'],
                    content={'description': Markup(content)},  # nosec: B704
                    meta={
                        'parliamentary_group_id': article_name.split('_')[1],
                        'people_ids': people_ids},
                )

        click.echo(f'{import_counter} parliamentary_groups imported')
        click.echo(f'{overwrite_counter} parliamentary_groups overwritten')

    return create_parliamentary_groups


def handle_es_connection_error(func: Any) -> Callable[[OrgRequest, OrgApp],
                                                      None]:
    @wraps(func)
    def wrapper(request: OrgRequest, app: OrgApp) -> None:
        try:
            return func(request, app)
        except ESConnectionError:
            click.echo('Ignoring Elasticsearch Connection error.')
    return wrapper


@cli.command(name='import-political-business')
@click.argument('path', type=click.Path(exists=True, resolve_path=True))
@click.option('--dry-run', is_flag=True, default=False)
def import_political_business(
    path: str,
    dry_run: bool,
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Imports political_businesses from archive of json files

    Example:
    .. code-block:: bash

        onegov-org --select '/foo/bar' import-political_businesses ./json-path

    Note this assumes the filename is named 'unknown_2416648.json' where the
    number 2416648 can be used to find the object in the original url.
    """

    # Read all json files in the given directory
    def read_json_files(path: str) -> Iterator[tuple[dict[str, Any], str]]:
        for file in Path(path).iterdir():

            if file.suffix == '.json':
                with open(file) as f:
                    yield (json.load(f), file.name)

    german_to_english_business_type_map: dict[str, str] = {
        'Anfrage': 'inquiry',
        'Antrag': 'proposal',

        'Auftrag': 'mandate',
        'Bericht': 'report',
        'Bericht und Antrag': 'report and proposal',
        'Beschluss': 'decision',
        'Botschaft': 'message',
        'Dringliche Interpellation': 'urgent interpellation',
        'Einladung': 'invitation',
        'Interpellation': 'interpelleation',  # Match existing enum typo
        'Kommissionsbericht': 'commission report',
        'Mitteilung': 'communication',
        'Motion': 'motion',
        'Postulat': 'postulate',
        'Resolution': 'resolution',
        'Verordnung': 'regulation',
        'Verschiedenes': 'miscellaneous',
        'Wahlen': 'elections',
    }

    german_to_english_status_map: dict[str, PoliticalBusinessStatus] = {
        'Abgeschrieben': 'abgeschrieben',
        'Beantwortet': 'beantwortet',
        'Erheblich erklärt': 'erheblich_erklaert',
        'Erledigt': 'erledigt',
        'Nicht erheblich erklärt': 'nicht_erheblich_erklaert',
        'Nicht zustandegekommen': 'nicht_zustandegekommen',
        'Pendent Exekutive': 'pendent_exekutive',
        'Pendent Legislative': 'pendent_legislative',
        'Rückzug': 'rueckzug',
        'Umgewandelt': 'umgewandelt',
        'Zurückgewiesen': 'zurueckgewiesen',
        'Überwiesen': 'ueberwiesen',
    }

    def parse_german_date(date_str: str | None) -> date | None:
        # Set locale to German for month parsing
        locale.setlocale(locale.LC_TIME, 'de_CH.UTF-8')
        return (datetime.strptime(date_str, '%d. %B %Y').date()
                if date_str else None)

    @handle_es_connection_error
    def create_political_businesses(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        political_business_collection = PoliticalBusinessCollection(session)

        political_businesses = read_json_files(path)
        import_counter = 0
        overwrite_counter = 0
        parliamentarian: RISParliamentarian | None
        for political_business, article_name in political_businesses:
            click.echo(f'Importing {article_name}')
            import_counter += 1
            if not dry_run:
                elements = political_business.get('elements', [])
                data_fields: dict[str, str | None] = {
                    'Nummer': None,
                    'Geschäftsart': None,
                    'Status': None,
                    'Datum': None
                }
                people_ids: list[tuple[str, str]] = []
                parliamentary_group_ids: list[str] = []

                i = 0
                while i < len(elements):
                    element = elements[i]
                    if (
                        element.get('type') == 'Paragraph'
                        and element.get('children')
                    ):
                        child_text_element = element['children'][0]
                        label_text = child_text_element.get('text', '').strip()

                        if label_text in data_fields.keys():
                            if i + 1 < len(elements):
                                value_element = elements[i+1]
                                if (
                                    value_element.get('type') == 'Paragraph'
                                    and value_element.get('children')
                                ):
                                    val_child = value_element['children'][0]
                                    data_fields[label_text] = val_child.get(
                                        'text', '').strip()
                                i += 1  # Consumed value element
                        elif label_text == 'Verfasser/Beteiligte':
                            if i + 1 < len(elements):
                                participants_el = elements[i+1]
                                if (
                                    participants_el.get('type') == 'Paragraph'
                                    and participants_el.get('children')
                                ):
                                    for c, child in enumerate(
                                        participants_el['children']
                                    ):
                                        if (
                                            child.get('type') == 'Link'
                                            and 'url' in child
                                            and c > 0
                                        ):
                                            url_parts = child['url'].split('/')
                                            type = participants_el['children'][
                                                c-1].get('text', '').strip()
                                            if url_parts:
                                                people_ids.append(
                                                    (url_parts[-1], type))
                                i += 1  # Consumed participants element
                        elif (
                            label_text == 'Fraktionen'
                            or label_text == 'Fraktion'
                        ):
                            if i + 1 < len(elements):
                                fraktionen_el = elements[i+1]
                                if (
                                    fraktionen_el.get('type') == 'Paragraph'
                                    and fraktionen_el.get('children')
                                ):
                                    for child in fraktionen_el['children']:
                                        if (
                                            child.get('type') == 'Link'
                                            and 'url' in child
                                        ):
                                            url_parts = child['url'].split('/')
                                            if url_parts:
                                                parliamentary_group_ids.append(
                                                    url_parts[-1])
                                i += 1  # Consumed fraktionen element
                        else:
                            # Check if elements[i] itself is an unlinked
                            # participant list
                            is_unlinked_list = False
                            if element.get('type') == 'Paragraph' and (
                                element.get('children')):
                                all_children_are_unlinked = True
                                if not element['children']:
                                    all_children_are_unlinked = False
                                else:
                                    for child_node_check in element[
                                        'children']:
                                        text_content_check = (
                                            child_node_check.get('text', ''))
                                        last_paren_open_check = (
                                            text_content_check.rfind('('))
                                        last_paren_close_check = (
                                            text_content_check.rfind(')'))
                                        is_valid_format = (
                                            child_node_check.get('type') ==
                                                'Text' and
                                            text_content_check and
                                            last_paren_open_check != -1 and
                                            last_paren_close_check == len(
                                                text_content_check) - 1 and
                                            last_paren_open_check <
                                                last_paren_close_check - 1 and
                                                # role not empty
                                            text_content_check[
                                                :last_paren_open_check].strip()
                                                # name not empty
                                        )
                                        if not is_valid_format:
                                            all_children_are_unlinked = False
                                            break
                                if all_children_are_unlinked:
                                    is_unlinked_list = True

                            if is_unlinked_list:
                                for child_node in element['children']:
                                    text_content = child_node.get('text', '')
                                    last_paren_open = text_content.rfind('(')
                                    last_paren_close = text_content.rfind(')')

                                    if (
                                        last_paren_open != -1
                                        and last_paren_close == (
                                            len(text_content) - 1
                                        )
                                        and last_paren_open < last_paren_close
                                    ):

                                        full_name = text_content[
                                            :last_paren_open].strip()
                                        role = text_content[
                                            last_paren_open+1:last_paren_close
                                            ].strip()

                                        name_parts = full_name.strip().split()
                                        if not name_parts:
                                            click.secho(f'Warning: Empty name '
                                                        'found for role '
                                                        f'{role} in '
                                                        f'{article_name}. '
                                                        'Skipping.',
                                                        fg='yellow')
                                            continue

                                        last_name = name_parts[-1]
                                        first_name = ' '.join(name_parts[:-1])

                                        if not first_name:
                                            # Handle single word name
                                            first_name = last_name

                                        click.echo(
                                            'Creating new parliamentarian: '
                                            f'{first_name} {last_name} for '
                                            f'{article_name} (unlinked)')
                                        parliamentarian = RISParliamentarian(
                                            first_name=first_name,
                                            last_name=last_name)
                                        session.add(parliamentarian)
                                        try:
                                            with session.begin_nested():
                                                # Ensure ID is available
                                                session.flush()
                                        except Exception as e:
                                            click.secho(
                                                'Error creating '
                                                f'parliamentarian {first_name}'
                                                f' {last_name}: {e}', fg='red')
                                            parliamentarian = None
                                            # Failed to create

                                        if (
                                            parliamentarian
                                            and parliamentarian.id
                                        ):
                                            people_ids.append(
                                                (str(parliamentarian.id),
                                                 role))
                                            if not parliamentarian.meta:
                                                parliamentarian.meta = {}
                                            parliamentarian.meta[
                                                'parliamentarian_id'] = str(
                                                    parliamentarian.id)
                    i += 1

                german_business_type = data_fields.get('Geschäftsart')
                english_business_type = None
                if german_business_type:
                    value = german_to_english_business_type_map.get(
                        german_business_type
                    )
                    english_business_type = value
                    if english_business_type is None:
                        click.secho(f'Warning: Unknown business type '
                                    f'"{german_business_type}" in '
                                    f'{article_name}. Setting to None.',
                                    fg='yellow')

                german_status = data_fields.get('Status')
                english_status = None
                if german_status:
                    english_status = german_to_english_status_map.get(
                        german_status
                    )
                    if english_status is None:
                        click.secho(f'Warning: Unknown status '
                                    f'"{german_status}" in {article_name}. '
                                    f'Setting to None.', fg='yellow')
                try:
                    if '_' not in article_name:
                        continue
                    pol_business_id = article_name.split('_')[1].split('.')[0]
                    political_business_collection.add(
                        title=political_business['metadata']['title'],
                        number=data_fields.get('Nummer'),
                        political_business_type=english_business_type,
                        status=english_status,
                        entry_date=parse_german_date(data_fields.get('Datum')),
                        content={},
                        meta={
                            'people_ids': people_ids,
                            'parliamentary_group_ids': parliamentary_group_ids,
                            'self_id': pol_business_id
                        }
                    )
                except ESConnectionError:
                    click.echo('Elasticsearch connection error')

        click.echo(f'{import_counter} political_businesses imported')
        click.echo(f'{overwrite_counter} political_businesses overwritten')

    return create_political_businesses


@cli.command(name='import-parliamentarians')
@click.argument('path', type=click.Path(exists=True, resolve_path=True))
@click.option('--dry-run', is_flag=True, default=False)
def import_parliamentarians(
    path: str,
    dry_run: bool,
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Imports parliamentarians from archive of json files

    Example:
    .. code-block:: bash

        onegov-org --select '/foo/bar' import-parliamentarians /path/to/items
    """

    # Read all json files in the given directory
    def read_json_files(path: str) -> Iterator[tuple[dict[str, Any], str]]:
        for file in Path(path).iterdir():

            if file.suffix == '.json':
                with open(file) as f:
                    yield (json.load(f), file.name)

    def create_parliamentarians(request: OrgRequest, app: OrgApp) -> None:
        parliamentarian_collection = RISParliamentarianCollection(
            request.session
        )

        parliamentarians = read_json_files(path)
        import_counter = 0

        for parliamentarian, article_name in parliamentarians:

            click.echo(f'Importing {article_name}')
            import_counter += 1

            if not dry_run:

                polit_business_ids = []
                name = parliamentarian['metadata']['title']
                first_name, last_name = name.split(' ', 1)
                contact = {}
                address = True
                table_as_dict: dict[str, Any] = {
                    'headers': [],
                    'rows': []
                }
                for e, element in enumerate(parliamentarian['elements']):
                    if element['type'] == 'Paragraph':
                        for i, child in enumerate(element['children']):
                            text = child.get('text', '')
                            if text == 'Personalien' or not text:
                                continue
                            if address:
                                if text.startswith('im Stadtparlament'):
                                    contact['date'] = text
                                elif re.search(r'^[^\d\s]*[^\d]+.*\d', text):
                                    contact['address'] = text
                                elif re.match(r'^\d{4}\s+.+', text):
                                    contact['zip_code'] = text.split(' ')[0]
                                    contact['city'] = ' '.join(text.split(
                                        ' ')[1:])
                                elif '@' in text:
                                    contact['email'] = text
                                elif text.startswith('Tel. P'):
                                    text = element['children'][i+1].get(
                                        'text', '')
                                    contact['phone_private'] = text
                                elif text.startswith('Tel.'):
                                    text = element['children'][i+1].get(
                                        'text', '')
                                    contact['phone_business'] = text
                                elif text.startswith('Mobile'):
                                    text = element['children'][i+1].get(
                                        'text', '')
                                    contact['phone_mobile'] = text
                                elif text.startswith('Funktion'):
                                    address = False
                                    text = parliamentarian['elements'][e+1][
                                        'children'][0].get('text', '')
                                    contact['function'] = text
                                elif re.match(r'^[^\d]*$', text):
                                    if text not in name:
                                        contact['addition'] = text
                            else:
                                if text.startswith('Partei'):
                                    text = parliamentarian['elements'][e+1][
                                        'children'][0].get('text', '')
                                    contact['party'] = text
                                elif text.startswith('Beruf'):
                                    text = parliamentarian['elements'][e+1][
                                        'children'][0].get('text', '')
                                    contact['profession'] = text

                    if element['type'] == 'Heading':
                        if element['text'] == 'Politische Vorstösse':
                            table = parliamentarian['elements'][e+1]
                            if table['type'] == 'Table':
                                for row in table['rows']:
                                    link = row['cells'][3]['url']
                                    id = link.split('/')[-1]
                                    polit_business_ids.append(id)
                        if element['text'] == 'Interessenbindungen':
                            try:
                                table = parliamentarian['elements'][e+1]
                            except IndexError:
                                click.secho(
                                    'No table found for Interessenbindungen',
                                    fg='red')
                                continue
                            if table['type'] == 'Table':
                                headers = [cell['text'] for cell in table[
                                    'headers']]
                                table_as_dict['headers'] = headers
                                for row in table['rows']:
                                    row_as_dict = {}
                                    for i, cell in enumerate(row['cells']):
                                        row_as_dict[headers[i]] = cell.get(
                                            'text', '')
                                    table_as_dict['rows'].append(row_as_dict)

                parliamentarian_collection.add(
                    first_name=first_name,
                    last_name=last_name,
                    phone_mobile=contact.get('phone_mobile', ''),
                    phone_private=contact.get('phone_private', ''),
                    phone_business=contact.get('phone_business', ''),
                    email_primary=contact.get('email', ''),
                    private_address=contact.get('address', ''),
                    private_address_addition=contact.get('addition', ''),
                    private_address_zip_code=contact.get('zip_code', ''),
                    private_address_city=contact.get('city', ''),
                    party=contact.get('party', ''),
                    occupation=contact.get('profession', ''),
                    # FIXME: I don't think this was correct, but I'll
                    #        leave it here just in case
                    # function=contact.get('function', ''),
                    meta={
                        'parliamentarian_id': article_name.split(
                            '_')[1].replace('.json', ''),
                        'polit_business_ids': polit_business_ids},
                    content={
                        'info': contact.get('date', ''),
                        'interests': table_as_dict
                        },
                )

        click.echo(f'{import_counter} parliamentarians imported')

    return create_parliamentarians


@cli.command(name='create-commission-memberships')
def create_memberships(
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Creates Commission Memberships

    onegov-org --select '/foo/bar' create-commission-memberships

    """

    def connect_ids(request: OrgRequest, app: OrgApp) -> None:

        session = request.session
        commission_memberships = RISCommissionMembershipCollection(session)
        people = RISParliamentarianCollection(session)
        commissions = RISCommissionCollection(session)

        for commission in commissions.query():
            connect_ids = {}
            for person_id, function in commission.meta.get('people_ids', []):
                for person in people.query():
                    if person.meta.get('parliamentarian_id') == person_id:
                        connect_ids[person.id] = function

            if not connect_ids:
                click.secho(
                    f'No people found for commission {commission.name}',
                     fg='yellow')
                continue

            # Check if membership already exists
            existing_membership = commission_memberships.query().filter(
                RISCommissionMembership.parliamentarian_id.in_(connect_ids.keys(

                )),
                RISCommissionMembership.commission_id == commission.id
            ).first()

            if existing_membership:
                click.secho(f'Membership already exists for {commission.name}',
                            fg='yellow')
                continue

            # Create new membership
            for person_id, function in connect_ids.items():
                commission_memberships.add(
                    parliamentarian_id=person_id,
                    commission_id=commission.id,
                    function=function,
                    role='member'
                )
                click.echo(
                    f'Created membership for {person_id} in {commission.name}')

    return connect_ids


@cli.command(name='create-political-business-participants')
def create_polical_business_participants(
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Creates Political Business Participants

    onegov-org --select '/foo/bar' create-political-business-participants

    """

    def connect_ids(request: OrgRequest, app: OrgApp) -> None:

        session = request.session
        business_participants = PoliticalBusinessParticipationCollection(
            session)
        people = RISParliamentarianCollection(session)
        political_businesses = PoliticalBusinessCollection(session)

        for political_business in political_businesses.query():
            connect_ids = {}
            for person_id, function in political_business.meta.get(
                'people_ids', []):
                for person in people.query():
                    if person.meta.get('parliamentarian_id') == person_id:
                        connect_ids[person.id] = function

            if not connect_ids:
                click.secho(
                    'No people found for political business '
                    f'{political_business.title}', fg='yellow')
                continue

            # Check if participation already exists
            existing_membership = business_participants.query().filter(
                RISCommissionMembership.parliamentarian_id.in_(
                    connect_ids.keys()),
                RISCommissionMembership.commission_id == political_business.id
            ).first()

            if existing_membership:
                click.secho(
                    'participation already exists for '
                    f'{political_business.title}', fg='yellow')
                continue

            # Create new participation
            for person_id, function in connect_ids.items():
                business_participants.add(
                    parliamentarian_id=person_id,
                    political_business_id=political_business.id,
                    participant_type=function,
                    # FIXME: I don't think this was correct
                    # role='member'
                )
                click.echo(
                    f'Created participation for {person_id} in '
                    f'{political_business.title}')

    return connect_ids


@cli.command(name='create-parliamentarian-roles')
def create_parliamentarian_roles(
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Creates Parliamentarian Roles

    onegov-org --select '/foo/bar' create-parliamentarian-roles

    """

    def connect_ids(request: OrgRequest, app: OrgApp) -> None:

        session = request.session
        parliamentarian_roles = RISParliamentarianRoleCollection(session)
        people = RISParliamentarianCollection(session)
        parliamentary_groups = RISParliamentaryGroupCollection(session)

        for parliamentary_group in parliamentary_groups.query():
            connect_ids = [
                person.id
                for person_id in parliamentary_group.meta.get('people_ids', [])
                for person in people.query()
                if person.meta.get('parliamentarian_id') == person_id
            ]

            if not connect_ids:
                click.secho(
                    'No people found for parliamentary group '
                    f'{parliamentary_group.name}', fg='yellow')
                continue

            # Check if participation already exists
            existing_membership = parliamentarian_roles.query().filter(
                RISCommissionMembership.parliamentarian_id.in_(connect_ids),
                RISCommissionMembership.commission_id == parliamentary_group.id
            ).first()

            if existing_membership:
                click.secho(
                    'participation already exists for '
                    f'{parliamentary_group.name}', fg='yellow')
                continue

            # Create new participation
            for i, person_id in enumerate(connect_ids):
                role = 'president' if i == 0 else 'member'
                parliamentarian_roles.add(
                    parliamentarian_id=person_id,
                    parliamentary_group_id=parliamentary_group.id,
                    role=role
                )
                click.echo(
                    f'Created participation for {person_id} in '
                    f'{parliamentary_group.name}')

    return connect_ids


@cli.command(name='connect-political-business-meeting-items')
def connect_political_business_meeting_items(
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Connects Political Business and Meeting Items

    onegov-org --select '/foo/bar' connect-political-business-meeting-items

    """

    def connect_ids(request: OrgRequest, app: OrgApp) -> None:

        session = request.session
        meeting_items = MeetingItemCollection(session)
        political_businesses = PoliticalBusinessCollection(session)

        for political_business in political_businesses.query():
            self_id = political_business.meta.get('self_id')
            if not self_id:
                click.secho(
                    'No self_id found for political business '
                    f'{political_business.title}', fg='yellow')
                continue
            meeting_item = meeting_items.query().filter(
                MeetingItem.political_business_link_id == self_id
            ).first()
            if meeting_item is not None:
                meeting_item.political_business_id = political_business.id
        transaction.commit()
    return connect_ids


@cli.command(name='ris-shipping-to-private-address')
def ris_shipping_to_private_address(
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Sets the RIS shipping address to the private address

    onegov-org --select /foo/bar ris-shipping-to-private-address

    """

    def set_private_address(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        parliamentarians = RISParliamentarianCollection(session)

        for p in parliamentarians.query():

            if not p.private_address and p.shipping_address:
                p.private_address = p.shipping_address
                p.shipping_address = None
            if not p.private_address_addition and p.shipping_address_addition:
                p.private_address_addition = p.shipping_address_addition
                p.shipping_address_addition = None
            if not p.private_address_zip_code and p.shipping_address_zip_code:
                p.private_address_zip_code = p.shipping_address_zip_code
                p.shipping_address_zip_code = None
            if not p.private_address_city and p.shipping_address_city:
                p.private_address_city = p.shipping_address_city
                p.shipping_address_city = None

        transaction.commit()

    return set_private_address


@cli.command(name='ris-set-end-date-for-inactive-parliamentarians')
def ris_set_end_date_for_inactive_parliamentarians(
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Sets the end date for inactive parliamentarians

    onegov-org --select /foo/bar ris-set-end-date-for-inactive-parliamentarians

    """

    def set_end_date(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        parliamentarians = RISParliamentarianCollection(session)

        for p in parliamentarians.query():
            if p.active:
                continue

            p.roles.append(RISParliamentarianRole(
                end=date(2024, 12, 31),
                parliamentarian_id=p.id,
                role='member'
            ))

        transaction.commit()

    return set_end_date


@cli.command(name='ris-resolve-parliamentarian-doublette')
def ris_resolve_parliamentarian_doublette(
) -> Callable[[OrgRequest, OrgApp], None]:
    """
    ogc-2394
    replace participants reference for business and delete
    parliamentarian (id_1)
    """

    def resolve_doublette(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        parliamentarians = RISParliamentarianCollection(session)
        businesses = PoliticalBusinessCollection(session)
        id = 'c0293891-7694-4da8-b846-844c7d1c7378'
        id_1 = 'dc83ffc4-2683-490f-ae30-1a0ab95fc0cc'
        business_id = '61964b73-f92e-40b4-8157-23c5048ca0d6'
        changed = False

        parliamentarian = (
            parliamentarians.query()
            .filter(RISParliamentarian.id == id).first())

        if not parliamentarian:
            return

        click.echo(f'parliamentarian: {parliamentarian.last_name} '
                   f'{parliamentarian.first_name} {parliamentarian.active}')
        parliamentarian_1 = (
            parliamentarians.query()
            .filter(RISParliamentarian.id == id_1).first())

        if not parliamentarian_1:
            return

        click.echo(f'parliamentarian_1: {parliamentarian_1.last_name} '
                   f'{parliamentarian_1.first_name} '
                   f'{parliamentarian_1.active}')
        business = businesses.query().filter(
            PoliticalBusiness.id == business_id).first()

        if not business:
            click.echo(f'Business with id {business_id} not found')
            return

        if business.participants:
            for participation in business.participants:
                if str(participation.parliamentarian_id) == id_1:
                    click.echo(f'replace {participation.parliamentarian_id} '
                               f'with {parliamentarian.id}')
                    participation.parliamentarian_id = id  # type: ignore[assignment]
                    changed = True

        transaction.commit()

        if changed:
            click.echo(f'delete {id_1}')
            session.delete(parliamentarian_1)
            transaction.commit()

    return resolve_doublette


@cli.command(name='ris-rename-imported-participation-types-to-english')
def ris_rename_imported_participation_types_to_english(
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Renames imported participation types to English

    onegov-org --select /foo/bar
        ris-rename-imported-participation-types-to-english
    """
    map = {
        'Erstunterzeichner/-in': 'First signatory',
        'Mitunterzeichner/-in': 'Co-signatory',
        'Erstunterzeichner/in': 'First signatory',
        'Mitunterzeichner/in': 'Co-signatory',
        'Vorstösser/in': 'First signatory',
        'Vorstösser/-in': 'First signatory',
    }

    def rename_participation_types(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        collection = PoliticalBusinessParticipationCollection(session)

        for participation in collection.query():
            if participation.participant_type in map:
                old_type = participation.participant_type
                participation.participant_type = map[old_type]
                click.echo(
                    f'Renamed {old_type} to {participation.participant_type}')

        transaction.commit()

    return rename_participation_types


@cli.command(name='ris-rebuild-political-business-links-to-meetings')
def ris_rebuild_political_business_links_to_meetings(
) -> Callable[[OrgRequest, OrgApp], None]:
    """ Rebuilds political business links to meetings. """

    def rebuild_political_business_links(
        request: OrgRequest, app: OrgApp
    ) -> None:
        session = request.session
        businesses = PoliticalBusinessCollection(session)
        meetings = MeetingCollection(session)
        meeting_items = MeetingItemCollection(session)
        already_ok_counter = 0
        no_business_link_id_counter = 0
        assigned_counter = 0
        no_business_found_counter = 0
        multiple_meetings_found_counter = 0
        single_meeting_found_counter = 0
        no_meeting_found_counter = 0

        for meeting_item in meeting_items.query():
            if meeting_item.political_business_id:
                already_ok_counter += 1
                continue

            if not meeting_item.political_business_link_id:
                # possible case, especially for new entries (after migration)
                click.secho(
                    f'No political business link found for meeting '
                    f'item {meeting_item.id}', fg='yellow')
                no_business_link_id_counter += 1
                continue

            if meeting_item.political_business_link_id:
                # click.secho(f'Attempt to link political business ..')
                business = (
                    businesses.query()
                    .filter(PoliticalBusiness.meta['self_id'] ==
                            meeting_item.political_business_link_id)
                ).first()
                if business:
                    click.secho(f'Assign political business id to '
                                f'{meeting_item.title}', fg='green')
                    meeting_item.political_business_id = business.id
                    assigned_counter += 1
                else:
                    no_business_found_counter += 1
                    meeting = meetings.query().filter(
                        Meeting.id == meeting_item.meeting_id).first()
                    if meeting:
                        click.secho(
                            f'No political business found for '
                            f'business link id '
                            f'{meeting_item.political_business_link_id} '
                            f'for meeting from {meeting.start_datetime}',
                            fg='red')
                    else:
                        click.secho(
                            f'No political business found for business '
                            f'link id '
                            f'{meeting_item.political_business_link_id}',
                            fg='red')
        transaction.commit()

        for business in businesses.query():
            collected_meetings = []

            for meeting_item in business.meeting_items:
                meeting = meetings.query().get(meeting_item.meeting_id)
                collected_meetings.append(meeting)

            business.meetings = collected_meetings  # type: ignore[assignment]
            if len(collected_meetings) > 1:
                multiple_meetings_found_counter += 1
                click.secho(f'Multiple meetings found for political business '
                            f'{business.title}', fg='green')
            elif len(collected_meetings) == 0:
                no_meeting_found_counter += 1
                click.secho(f'No meeting found for political business '
                            f'{business.title}', fg='yellow')
            else:
                single_meeting_found_counter += 1
                click.secho(f'Set meeting for political business '
                            f'{business.title}', fg='green')

        # echo counter results
        click.secho('')
        click.secho(f'Meeting items with political business link id '
                    f'{already_ok_counter}', fg='green')
        click.secho(f'Meeting items with no political business link id '
                    f'{no_business_link_id_counter}', fg='red')
        click.secho(f'No business found for {no_business_found_counter} '
                    f'meeting items', fg='yellow')
        click.secho(f'Political business links assigned to meeting items '
                    f'{assigned_counter}', fg='green')
        click.secho(f'Of totally {meeting_items.query().count()} meeting '
                    f'items', fg='green')

        click.secho('')
        click.secho(f'Multiple meetings found for '
                    f'{multiple_meetings_found_counter} political businesses',
                    fg='green')
        click.secho(f'Single meeting found for {single_meeting_found_counter} '
                    f'political businesses', fg='green')
        click.secho(f'No meeting found for {no_meeting_found_counter} '
                    f'political businesses', fg='yellow')
        click.secho(f'Total number of political businesses '
                    f'{businesses.query().count()}', fg='green')

        transaction.commit()

    return rebuild_political_business_links


@cli.command(name='ris-make-imported-files-general-file')
def ris_make_imported_files_general_file(
) -> Callable[[OrgRequest, OrgApp], None]:
    """
    onegov-org --select /onegov_town6/wil ris-make-imported-files-general-file
    """

    def make_general_file(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        businesses = PoliticalBusinessCollection(session)
        meetings = MeetingCollection(session)

        counter = 0
        for business in businesses.query():
            for file in business.files:
                if file.type == 'generic':
                    file.type = 'general'
                    counter += 1
        click.secho(f'Set {counter} political business files to '
                    f'type "general"', fg='green')
        transaction.commit()

        counter = 0
        for meeting in meetings.query():
            for file in meeting.files:
                if file.type == 'generic':
                    file.type = 'general'
                    counter += 1
        click.secho(f'Set {counter} meeting files to type "general"',
                    fg='green')

    return make_general_file


@cli.command(name='ris-wil-meetings-fix-audio-links')
def ris_wil_meetings_shorten_audio_links(
) -> Callable[[OrgRequest, OrgApp], None]:

    def ris_wil_meetings_fix_audio_links(
        request: OrgRequest,
        app: OrgApp
    ) -> None:

        meetings = MeetingCollection(request.session)
        recapp_counter = 0
        http_counter = 0
        for meeting in meetings.query():
            if not meeting.audio_link:
                continue

            if (meeting.audio_link ==
                    'https://wil.recapp.ch/viewer/default/timeline'):
                meeting.audio_link = 'https://wil.recapp.ch'
                recapp_counter += 1
                continue

            if 'http://wil.recapp.ch' in meeting.audio_link:
                meeting.audio_link = (
                    meeting.audio_link.replace('http', 'https'))
                http_counter += 1
                continue

            if meeting.audio_link == 'https://wil.recapp.ch':
                continue

            if 'http://verbalix.stadtwil.ch' in meeting.audio_link:
                meeting.audio_link = (
                    meeting.audio_link.replace('http', 'https'))
                http_counter += 1
                continue

            if meeting.audio_link == 'https://verbalix.stadtwil.ch/':
                continue

            if 'https://verbalix.stadtwil.ch/' in meeting.audio_link:
                continue

            click.secho(
                f'audio link for meeting {meeting.title} vom '
                f'{meeting.start_datetime}: {meeting.audio_link}', fg='yellow')

        click.secho(f'Fixed recapp audio links for {recapp_counter} '
                    f'meetings', fg='green')
        click.secho(f'Fixed verbalix http audio links for {http_counter} '
                    f'meetings', fg='green')
        transaction.commit()

    return ris_wil_meetings_fix_audio_links


@cli.command(name='add-images-to-parliamentarians')
@click.argument('path', type=click.Path(exists=True, resolve_path=True))
def add_images_to_parliamentarians(
    path: str,
) -> Callable[[OrgRequest, OrgApp], None]:
    """
    Adds images to parliamentarians in the system
    """

    def read_json_files(path: str) -> Iterator[tuple[dict[str, Any], str]]:
        for file in Path(path).iterdir():

            if file.suffix == '.json':
                with open(file) as f:
                    yield (json.load(f), file.name)

    def add_images(request: OrgRequest, app: OrgApp) -> None:
        session = request.session
        parliamentarian_collection = ParliamentarianCollection(session)

        people = read_json_files(path)
        import_counter = 0
        article_counter = 0
        for person, article_name in people:
            article_counter += 1
            article_number = article_name.split('_')[1].replace('.json', '')
            elements = person.get('elements', [])

            name = person.get('metadata').get('title', '')  # type: ignore
            click.secho(f'Importing images for {article_name}, {name}',
                        fg='cyan')
            first_name, last_name = name.split(' ', 1)

            for element in elements:
                if element['type'] == 'EmbeddedImage' and element['content']:
                    print('Found image')
                    base64_string = element['content']
                    alt = element.get('alt', article_number)
                    base64_data = base64_string.split(',', 1)[1]
                    image_type = base64_string.split(';')[0].split(
                        '/')[1]
                    image_bytes = base64.b64decode(base64_data)
                    image_bytes_io = BytesIO(image_bytes)

                    found_person = False
                    person = parliamentarian_collection.query().filter_by(
                        first_name=first_name,
                        last_name=last_name
                    )
                    if len(person.all()) > 0:
                        found_person = True
                        person = person.first()
                        if not person.picture:
                            person.picture = (
                                image_bytes_io,
                                (f'{alt}.{image_type}')
                            )
                            import_counter += 1
                        else:
                            click.secho(f'Parliamentarian {first_name} {last_name} already has a picture',
                                        fg='yellow')
                    else:
                        for p in parliamentarian_collection.query():
                            if article_number == p.meta.get(
                                'parliamentarian_id'):
                                found_person = True
                                if not p.picture:
                                    p.picture = (
                                        image_bytes_io,
                                        (f'{alt}.{image_type}')
                                    )
                                    import_counter += 1
                                else:
                                    click.secho(f'Parliamentarian {first_name} {last_name} already has a picture',
                                                fg='yellow')
                                break
                    if found_person:
                        print(f'Found parliamentarian: {first_name} {last_name}')
                    else:
                        print(f'No matching parliamentarian found for {first_name} {last_name}')
                    continue
        click.secho(f'Imported {import_counter} images from '
                    f'{article_counter} files',
                    fg='green')

        transaction.commit()

    return add_images
