""" Offers tools to deal with csv (and xls, xlsx) files. """
from __future__ import annotations

import codecs

import openpyxl
import re
import sys
import tempfile
import xlrd  # type:ignore[import-untyped]

from collections import namedtuple, OrderedDict
from csv import DictWriter, Sniffer, QUOTE_ALL
from csv import Error as CsvError
from csv import reader as csv_reader
from csv import writer as csv_writer
from datetime import datetime
from editdistance import eval as distance
from functools import lru_cache
from io import BytesIO, StringIO
from itertools import permutations
from onegov.core import errors
from ordered_set import OrderedSet
from unidecode import unidecode
from xlsxwriter.workbook import Workbook
from onegov.core.utils import normalize_for_url


from typing import overload, Any, Generic, IO, TypeVar, TYPE_CHECKING
if TYPE_CHECKING:
    from _typeshed import SupportsRichComparison
    from collections.abc import (
        Callable, Collection, Iterable, Iterator, Sequence)
    from csv import Dialect
    from openpyxl.worksheet.worksheet import Worksheet
    from typing import Protocol, TypeAlias

    _T = TypeVar('_T')
    _T_co = TypeVar('_T_co', covariant=True)
    _SupportsRichComparisonT = TypeVar(
        '_SupportsRichComparisonT',
        bound=SupportsRichComparison
    )

    class _RowType(Protocol[_T_co]):
        def __call__(self, *, rownumber: int, **kwargs: str) -> _T_co: ...

    class DefaultRow(Protocol):
        @property
        def rownumber(self) -> int: ...
        def __getattr__(self, name: str) -> str: ...

    KeyFunc: TypeAlias = Callable[[_T], SupportsRichComparison]
    DefaultCSVFile: TypeAlias = 'CSVFile[DefaultRow]'
    _RowT = TypeVar('_RowT', default=DefaultRow)
else:
    _RowT = TypeVar('_RowT')


VALID_CSV_DELIMITERS = ',;\t'
WHITESPACE = re.compile(r'\s+')
INVALID_XLSX_TITLE = re.compile(r'[\\*?:/\[\]]')

small_chars = 'fijlrt:,;.+i '
large_chars = 'GHMWQ_'

max_width = 75


class CSVFile(Generic[_RowT]):
    """ Provides access to a csv file.

    :param csvfile:
        The csv file to be accessed. Must be an open file (not a path), opened
        in binary mode. For example::

            with open(path, 'rb') as f:
                csv = CSVFile(f)

    :param expected_headers:
        The expected headers if known. Expected headers are headers which
        *must* exist in the CSV file. There may be additional headers.

        If the headers are slightly misspelled, a matching algorithm tries to
        guess the correct header, without accidentally matching the wrong
        headers.

        See :func:`match_headers` for more information.

        If the no expected_headers are passed, no checks are done, but the
        headers are still available. Headers matching is useful if a user
        provides the CSV and it might be wrong.

        If it is impossible for misspellings to occurr, the expected headers
        don't have to be specified.

    :param dialect:
        The CSV dialect to expect. By default, the dialect will be guessed
        using Python's heuristic.

    :param encoding:
        The CSV encoding to expect. By default, the encoding will be guessed
        and will either be UTF-8 or CP1252.

    :param rename_duplicate_column_names:
        It is possible to rename duplicate column names to deal with super
        wacky files. If this option is set and a duplicate header is found,
        a suffix is appended to the column name rather than throwing a
        DuplicateColumnNamesError.

    :param rowtype:
        An alternative rowtype for the resulting rows. This should be a
        callable that receives a `rownumber` key/value and all the other
        keys/values found in the csv. The keys are normalized and are valid
        Python identifiers usable as attribute names.

        Defaults to a namedtuple created using the found headers.

    Once the csv file is open, the records can be acceessed as follows::

        with open(path, 'rb') as f:
            csv = CSVFile(f)

            for line in csv.lines:
                csv.my_field  # access the column with the 'my_field' header

    """

    rowtype: _RowType[_RowT]

    @overload
    def __init__(
        self: DefaultCSVFile,
        csvfile: IO[bytes],
        expected_headers: Collection[str] | None = None,
        dialect: type[Dialect] | Dialect | str | None = None,
        encoding: str | None = None,
        rename_duplicate_column_names: bool = False,
        rowtype: None = None
    ): ...

    @overload
    def __init__(
        self: CSVFile[_RowT],
        csvfile: IO[bytes],
        expected_headers: Collection[str] | None = None,
        dialect: type[Dialect] | Dialect | str | None = None,
        encoding: str | None = None,
        rename_duplicate_column_names: bool = False,
        *,
        rowtype: _RowType[_RowT]
    ): ...

    def __init__(
        self,
        csvfile: IO[bytes],
        expected_headers: Collection[str] | None = None,
        dialect: type[Dialect] | Dialect | str | None = None,
        encoding: str | None = None,
        rename_duplicate_column_names: bool = False,
        rowtype: _RowType[_RowT] | None = None
    ):

        # guess the encoding if not already provided
        encoding = encoding or detect_encoding(csvfile)
        if encoding is None:
            raise errors.InvalidFormatError()

        self.csvfile = codecs.getreader(encoding)(csvfile)

        # sniff the dialect if not already provided
        try:
            self.csvfile.seek(0)
            self.dialect = dialect or sniff_dialect(self.csvfile.read(1024))
        except (CsvError, errors.InvalidFormatError):
            self.csvfile.seek(0)
            self.dialect = sniff_dialect(self.csvfile.read())

        # match the headers
        self.csvfile.seek(0)

        # if no expected headers expect, we just take what we can get
        headers = parse_header(
            self.csvfile.readline(),
            self.dialect,
            rename_duplicate_column_names
        )
        expected_headers = expected_headers or headers

        self.headers = OrderedDict(
            (h, c) for c, h in enumerate(match_headers(
                headers=headers,
                expected=expected_headers
            ))
        )

        # create an output type
        assert 'rownumber' not in expected_headers, """
            rownumber can't be used as a header
        """

        self.rowtype = rowtype or namedtuple(  # type:ignore  # noqa: PYI024
            'CSVFileRow', ['rownumber', *(
                self.as_valid_identifier(k)
                for k in self.headers.keys()
            )]
        )

    @staticmethod
    @lru_cache(maxsize=128)
    def as_valid_identifier(value: str) -> str:
        result = normalize_header(value)
        for invalid in '- .%/,;()':
            result = result.replace(invalid, '_')
        while result and result[0] in '_0123456789':
            result = result[1:]
        return result

    def __iter__(self) -> Iterator[_RowT]:
        yield from self.lines

    @property
    def lines(self) -> Iterator[_RowT]:
        self.csvfile.seek(0)

        encountered_empty_line = False

        for ix, line in enumerate(csv_reader(self.csvfile, self.dialect)):

            # raise an empty line error if we found one somewhere in the
            # middle -> at the end they don't count
            if not line:
                encountered_empty_line = True
                continue

            if line and encountered_empty_line:
                raise errors.EmptyLineInFileError()

            # the first line is the header
            if ix == 0:
                continue

            yield self.rowtype(
                rownumber=ix + 1,  # row numbers are for customers, not coders
                **{
                    self.as_valid_identifier(header): line[column].strip()
                    for header, column in self.headers.items()
                }
            )


def detect_encoding(csvfile: IO[bytes]) -> str:
    """ Since encoding detection is hard to get right (and work correctly
    every time), we limit ourselves here to UTF-8 or CP1252, whichever works
    first. CP1252 is basically the csv format you get if you use windows and
    excel and it is a superset of ISO-8859-1/LATIN1.

    """
    csvfile.seek(0)

    try:
        for line in csvfile.readlines():
            line.decode('utf-8')

        return 'utf-8'
    except UnicodeDecodeError:
        return 'cp1252'


def sniff_dialect(csv: str) -> type[Dialect]:
    """ Takes the given csv string and returns the dialect or raises an error.
    Works just like Python's built in sniffer, just that it is a bit more
    conservative and doesn't just accept any kind of character as csv
    delimiter.

    """
    if not csv:
        raise errors.EmptyFileError()

    try:
        dialect = Sniffer().sniff(csv, VALID_CSV_DELIMITERS)
    except CsvError as exception:

        # sometimes we can get away with an extra pass just over the first line
        # (the header tends to contain fewer special cases)
        if '\n' in csv:
            return sniff_dialect(csv[:csv.find('\n')])

        raise errors.InvalidFormatError() from exception

    return dialect


def normalize_header(header: str) -> str:
    """ Normalizes a header value to be as uniform as possible.

    This includes:
        * stripping the whitespace around it
        * lowercasing everything
        * transliterating unicode (e.g. 'ä' becomes 'a')
        * removing duplicate whitespace inside it
    """

    header = header.strip()
    header = header.lower()
    header = unidecode(header)
    header = WHITESPACE.sub(' ', header)

    return header


def convert_xlsx_to_csv(
    xlsx: IO[bytes],
    sheet_name: str | None = None
) -> BytesIO:
    """ Takes an XLS file and returns a csv file using the given worksheet
    name or the first worksheet found.

    """

    xlsx.seek(0)

    try:
        excel = openpyxl.load_workbook(xlsx, data_only=True)
    except Exception as exception:
        raise OSError('Could not read XLSX file') from exception

    sheet: Worksheet
    if sheet_name:
        try:
            sheet = excel[sheet_name]
        except KeyError as exception:
            raise KeyError(
                'Could not find the given sheet in this excel file!'
            ) from exception
    else:
        sheet = excel.worksheets[0]

    # FIXME: We should probably do this check at runtime eventually since
    # Workbook[name] might return a Worksheet, ReadOnlyWorksheet or a
    # a WriteOnlyWorksheet. Workbook.worksheet[index] might additionaly return
    # a Chartsheet.
    if TYPE_CHECKING:
        assert isinstance(sheet, Worksheet)

    text_output = StringIO()
    writecsv = csv_writer(text_output, quoting=QUOTE_ALL)

    for row in range(1, sheet.max_row + 1):
        values = []

        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)

            if cell.value is None:
                value = ''
            elif cell.data_type == 's':
                value = cell.value  # type:ignore[assignment]
            elif cell.data_type == 'n':
                if (int_value := int(cell.value)) == cell.value:  # type:ignore
                    value = str(int_value)
                else:
                    value = str(cell.value)
            elif cell.data_type == 'd':
                value = cell.value.isoformat()  # type:ignore[union-attr]
            elif cell.data_type == 'b':
                value = '1' if cell.value else '0'
            else:
                raise NotImplementedError

            values.append(value)

        if any(values):
            writecsv.writerow(values)

    text_output.seek(0)
    # FIXME: We can use StringIOWrapper around a BytesIO, then we don't
    #        need to convert at the end!
    output = BytesIO()

    for line in text_output.readlines():
        output.write(line.encode('utf-8'))

    return output


def convert_xls_to_csv(
    xls: IO[bytes],
    sheet_name: str | None = None
) -> BytesIO:
    """ Takes an XLS file and returns a csv file using the given worksheet
    name or the first worksheet found.

    """

    xls.seek(0)

    try:
        excel = xlrd.open_workbook(file_contents=xls.read())
    except Exception as exception:
        raise OSError('Could not read XLS file') from exception

    if sheet_name:
        try:
            sheet = excel.sheet_by_name(sheet_name)
        except xlrd.XLRDError as exception:
            raise KeyError(
                'Could not find the given sheet in this excel file!'
            ) from exception
    else:
        sheet = excel.sheet_by_index(0)

    text_output = StringIO()
    writecsv = csv_writer(text_output, quoting=QUOTE_ALL)

    for rownum in range(sheet.nrows):
        values = []

        for cell in sheet.row(rownum):

            if cell.ctype == xlrd.XL_CELL_TEXT:
                value = cell.value
            elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                value = ''
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                if cell.value.is_integer():
                    value = str(int(cell.value))
                else:
                    value = str(cell.value)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                value = xlrd.xldate_as_tuple(cell.value, excel.datemode)
                value = datetime(*value).isoformat()
            elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                value = str(cell.value)
            else:
                raise NotImplementedError

            values.append(value)

        writecsv.writerow(values)

    text_output.seek(0)
    # FIXME: We can use StringIOWrapper around a BytesIO, then we don't
    #        need to convert at the end!
    output = BytesIO()

    for line in text_output.readlines():
        output.write(line.encode('utf-8'))

    return output


def convert_excel_to_csv(
    file: IO[bytes],
    sheet_name: str | None = None
) -> BytesIO:
    """ Takes an XLS/XLSX file and returns a csv file using the given worksheet
    name or the first worksheet found.

    """

    try:
        return convert_xlsx_to_csv(file, sheet_name)
    except OSError:
        return convert_xls_to_csv(file, sheet_name)


def character_width(char: str) -> float:
    # those numbers have been acquired by chasing unicorns
    # and fairies in the magic forest of Excel
    #
    # tweak them as needed, but know that there's no correct answer,
    # as each excel version on each platform or os-version will render
    # the fonts used at different widths
    if char in small_chars:
        return 0.75
    elif char in large_chars:
        return 1.2
    else:
        return 1


def estimate_width(text: str) -> float:
    if not text:
        return 0

    width = max(
        sum(character_width(c) for c in line)
        for line in text.splitlines()
    )

    return min(width, max_width)


@overload
def get_keys_from_list_of_dicts(
    rows: Iterable[dict[_SupportsRichComparisonT, Any]],
    key: None = None,
    reverse: bool = False
) -> tuple[_SupportsRichComparisonT, ...]: ...


@overload
def get_keys_from_list_of_dicts(
    rows: Iterable[dict[_T, Any]],
    key: KeyFunc[_T],
    reverse: bool = False
) -> tuple[_T, ...]: ...


def get_keys_from_list_of_dicts(
    rows: Iterable[dict[Any, Any]],
    key: KeyFunc[Any] | None = None,
    reverse: bool = False
) -> tuple[Any, ...]:
    """ Returns all keys of a list of dicts in an ordered tuple.

    If the list of dicts is irregular, the keys found in later rows are
    added at the end of the list.

    Note that the order of keys is otherwise defined by the order of the keys
    of the dictionaries. So if ordered dictionaries are used, the order is
    defined. If regular dictionaries are used, the order is undefined.

    Alternatively, a key and a reverse flag may be provided which will be
    used to order the fields. If the list of fields is specified, the key and
    the reverse flag is ignored.

    """
    fields_set: OrderedSet[str] = OrderedSet()

    for dictionary in rows:
        fields_set.update(dictionary.keys())

    if key:
        fields = tuple(sorted(fields_set, key=key, reverse=reverse))
    else:
        fields = tuple(fields_set)

    return fields


def convert_list_of_dicts_to_csv(
    rows: Iterable[dict[str, Any]],
    fields: Sequence[str] | None = None,
    key: KeyFunc[str] | None = None,
    reverse: bool = False
) -> str:
    """ Takes a list of dictionaries and returns a csv.

    If no fields are provided, all fields are included in the order of the keys
    of the first dict. With regular dictionaries this is random. Use an ordered
    dict or provide a list of fields to have a fixed order.

    Alternatively, a key and a reverse flag may be provided which will be
    used to order the fields. If the list of fields is specified, the key and
    the reverse flag is ignored.

    The function returns a string created in memory. Therefore this function
    is limited to small-ish datasets.

    """

    if not rows:
        return ''

    fields = fields or get_keys_from_list_of_dicts(rows, key, reverse)

    output = StringIO()
    writer = DictWriter(output, fieldnames=fields)

    writer.writeheader()

    for row in rows:
        writer.writerow({field: row.get(field, '') for field in fields})

    output.seek(0)
    return output.read()


def convert_list_of_dicts_to_xlsx(
    rows: Iterable[dict[str, Any]],
    fields: Sequence[str] | None = None,
    key: KeyFunc[str] | None = None,
    reverse: bool = False
) -> bytes:
    """ Takes a list of dictionaries and returns a xlsx.

    This behaves the same way as :func:`convert_list_of_dicts_to_csv`.

    """

    with tempfile.NamedTemporaryFile() as file:
        workbook = Workbook(file.name, options={'constant_memory': True})
        cellformat = workbook.add_format({'text_wrap': True})

        worksheet = workbook.add_worksheet()

        fields_ = fields or get_keys_from_list_of_dicts(rows, key, reverse)

        # write the header
        worksheet.write_row(0, 0, fields_, cellformat)

        # keep track of the maximum character width
        column_widths = [estimate_width(field) for field in fields_]

        def values(row: dict[str, Any]) -> Iterator[str]:
            for ix, field in enumerate(fields_):
                value = row.get(field, '')
                column_widths[ix] = max(
                    column_widths[ix],
                    estimate_width(str(value))
                )

                if isinstance(value, str):
                    value = value.replace('\r', '')

                yield value

        # write the rows
        for r, row in enumerate(rows, start=1):
            worksheet.write_row(r, 0, values(row), cellformat)

        # set the column widths
        for col, width in enumerate(column_widths):
            worksheet.set_column(col, col, width)

        workbook.close()

        file.seek(0)
        return file.read()


def convert_list_of_list_of_dicts_to_xlsx(
    row_list: Sequence[Iterable[dict[str, Any]]],
    titles_list: Sequence[str],
    key_list: Sequence[KeyFunc[str] | None] | None = None,
    reverse: bool = False
) -> bytes:
    """
    Like to :func:`convert_list_of_dicts_to_xlsx`, but operates on a list
    instead of in a single item.

    """
    titles_list = normalize_sheet_titles(titles_list)

    with tempfile.NamedTemporaryFile() as file:
        workbook = Workbook(file.name, options={'constant_memory': True})
        if key_list is None:
            key_list = [None] * len(titles_list)
        for rows, title, key in zip(row_list, titles_list, key_list):

            cellformat = workbook.add_format({'text_wrap': True})
            worksheet = workbook.add_worksheet(title)
            fields = get_keys_from_list_of_dicts(rows, key, reverse)

            # write the header
            worksheet.write_row(0, 0, fields, cellformat)

            # keep track of the maximum character width
            column_widths = [estimate_width(field) for field in fields]

            def values(
                row: dict[str, Any],
                fields: tuple[str, ...] = fields,
                column_widths: list[float] = column_widths
            ) -> Iterator[str]:
                for ix, field in enumerate(fields):
                    value = row.get(field, '')
                    column_widths[ix] = max(
                        column_widths[ix],
                        estimate_width(str(value))
                    )

                    if isinstance(value, str):
                        value = value.replace('\r', '')

                    yield value

            # write the list_of_rows
            for r, row in enumerate(rows, start=1):
                worksheet.write_row(r, 0, values(row), cellformat)

            # set the column widths
            for col, width in enumerate(column_widths):
                worksheet.set_column(col, col, width)

        workbook.close()
        file.seek(0)
        return file.read()


def normalize_sheet_titles(titles: Sequence[str]) -> list[str]:
    """
    Ensuring the title of the xlsx is valid.
    """

    def valid_characters_or_raise(title: str) -> None:
        m = INVALID_XLSX_TITLE.search(title)
        if m:
            msg = f'Invalid character {m.group(0)} found in xlsx sheet title'
            raise ValueError(msg)

    titles = [normalize_for_url(title) for title in titles]
    duplicate_idxs = list_duplicates_index(titles)

    # change name of the duplicate sheet names
    for index in duplicate_idxs:
        current_title = titles[index]
        valid_characters_or_raise(current_title)
        titles[index] = avoid_duplicate_name(titles, current_title)

    # change name of the duplicate sheet names
    for index, item in enumerate(titles):
        if len(item) > 31:
            while len(titles[index]) > 31:
                titles[index] = remove_first_word(titles[index])

    return titles


def avoid_duplicate_name(titles: Sequence[str], title: str) -> str:
    """
    Naive check to see whether name already exists.
    If name does exist suggest a name using an incrementer
    Duplicates are case-insensitive
    """
    # Check for an absolute match in which case we need to find an alternative
    match = [n for n in titles if n.lower() == title.lower()]
    if match:
        titles = ','.join(titles)
        sheet_title_regex = re.compile(
            f'(?P<title>{re.escape(title)})(?P<count>\\d*),?', re.IGNORECASE
        )
        matches = sheet_title_regex.findall(titles)
        if matches:
            # use name, but append with the next highest integer
            counts = [int(idx) for (t, idx) in matches if idx.isdigit()]
            highest = 0
            if counts:
                highest = max(counts)
            title = f'{title}_{highest + 1}'
    return title


def remove_first_word(title: str) -> str:
    """
    Removes all chars from beginning up until and including the first "-".
    """
    return re.sub(r'^.*?-', '', title)


def has_duplicates(a_list: Sequence[Any]) -> bool:
    return len(a_list) != len(set(a_list))


def list_duplicates_index(a: Sequence[Any]) -> list[int]:
    """
    Returns a list of indexes of duplicates in a list.
    for example::

        a = [1, 2, 3, 2, 1, 5, 6, 5, 5, 5]
        list_duplicates_index(a) == [3, 4, 7, 8, 9]
    """
    return [idx for idx, item in enumerate(a) if item in a[:idx]]


def parse_header(
    csv: str,
    dialect: type[Dialect] | Dialect | str | None = None,
    rename_duplicate_column_names: bool = False
) -> list[str]:
    """ Takes the first line of the given csv string and returns the headers.

    Headers are normalized (stripped and normalized) and expected to be
    unique. The dialect is sniffed, if not provided.

    :return: A list of headers in the order of appearance.

    """
    try:
        dialect = dialect or sniff_dialect(csv)
    except errors.InvalidFormatError:
        # FIXME: Is this actually legal? typeshed seems to think not
        dialect = None  # let the csv reader try, it might still work
    except errors.EmptyFileError:
        return []

    headers = next(csv_reader(
        csv.splitlines(),
        dialect=dialect  # type: ignore[arg-type]
    ))
    headers = [normalize_header(h) for h in headers if h]

    if rename_duplicate_column_names:
        indexes: dict[str, list[int]] = {}
        for i, item in enumerate(headers):
            indexes.setdefault(item, []).append(i)
        for value in indexes.values():
            for suffix, index in enumerate(value[1:]):
                headers[index] += '_{}'.format(suffix + 1)

    return headers


def match_headers(
    headers: Collection[str],
    expected: Collection[str]
) -> list[str]:
    """ Takes a list of normalized headers and matches them up against a
    list of expected headers.

    The headers may differ from the expected headers. This function tries to
    match them up using the Levenshtein distance. It does so somewhat carefully
    by calculating a sane distance using the input.

    The passed headers are expected to be have been normalized by
    :func:`normalize_header`, since we usually will pass the result of
    :func:`parse_header` to this function.

    For example::

        match_headers(
            headers=['firstname', 'lastname'],
            expected=['last_name', 'first_name']
        )

    Results in::

        ['first_name', 'last_name']

    If no match is possible, an
    :class:`~onegov.core.errors.MissingColumnsError`, or
    :class:`~onegov.core.errors.AmbiguousColumnsError` or
    :class:`~onegov.core.errors.DuplicateColumnNamesError` error is raised.

    :return: The matched headers in the order of appearance.

    """

    # for this to work we need unique header names
    if len(headers) != len(set(headers)):
        raise errors.DuplicateColumnNamesError()

    # we calculate a 'sane' levenshtein distance by comparing the
    # distances between all headers, permutations, as well as the lengths
    # of all expected headers. This makes sure we don't end up with matches
    # that make no sense (like ['first', 'second'] matching ['first', 'third'])
    sane_distance = getattr(sys, 'maxsize', 0) or sys.maxsize

    if len(headers) > 1:
        sane_distance = min((
            sane_distance,
            min(distance(a, b) for a, b in permutations(headers, 2))
        ))

    if len(expected) > 1:
        sane_distance = min((
            sane_distance,
            min(distance(a, b) for a, b in permutations(expected, 2))
        ))

    sane_distance = min((
        sane_distance,
        min(len(c) for c in expected)
    ))

    mapping = {}
    missing = []
    ambiguous = {}

    for column in expected:
        normalized = normalize_header(column)
        distances = {h: distance(normalized, h) for h in headers}

        closest = min(distances.values())

        if closest >= sane_distance:
            missing.append(column)
            continue

        matches = tuple(c for c in distances if distances[c] == closest)

        if len(matches) > 1:
            ambiguous[column] = matches
            continue

        mapping[matches[0]] = column

    if missing:
        raise errors.MissingColumnsError(missing)

    if ambiguous:
        raise errors.AmbiguousColumnsError(ambiguous)

    # return the mapping or the original header (we allow for extra headers,
    # we just don't allow for missing headers)
    return [mapping.get(h, h) for h in headers]
